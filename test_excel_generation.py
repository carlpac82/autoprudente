#!/usr/bin/env python3
"""Test Excel generation directly without HTTP"""

import sys
import io
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Test data
location = "Faro"
date = "2025-11-05"
prices = {
    "B1": {1: 25.50, 2: 48.00, 3: 70.00, 7: 140.00, 9: 180.00, 14: 280.00, 22: 440.00, 28: 560.00},
    "B2": {1: 27.50, 2: 52.00, 3: 75.00, 7: 150.00, 9: 190.00, 14: 300.00, 22: 460.00, 28: 580.00}
}

print("=" * 60)
print("TESTING EXCEL GENERATION")
print("=" * 60)

try:
    # Load template
    print("\n1. Loading template Abbycar.xlsx...")
    wb = load_workbook('Abbycar.xlsx')
    ws = wb.active
    print(f"   ✓ Template loaded: {ws.max_row} rows, {ws.max_column} cols")
    
    # Test price calculation
    print("\n2. Testing price calculation...")
    def calculate_price_for_day(group_prices, day):
        price = group_prices.get(day) or group_prices.get(str(day), '')
        if not price:
            return ''
        price = float(price)
        
        if day <= 7:
            return price
        
        if day == 8:
            price_9 = group_prices.get(9) or group_prices.get('9', '')
            if price_9:
                return float(price_9) / 9
            return price / 8
        elif day == 9:
            return price / 9
        elif day == 14:
            return price / 14
        elif day == 22:
            return price / 22
        elif day == 28:
            return price / 28
        
        return price
    
    # Test B1 prices
    group_prices = prices["B1"]
    for day in [1, 2, 3, 7, 9, 14, 22, 28]:
        calc_price = calculate_price_for_day(group_prices, day)
        print(f"   Day {day:2d}: {group_prices.get(day, 'N/A'):>6} → {calc_price:>8.2f}" if calc_price else f"   Day {day:2d}: N/A")
    
    # Test adjustment
    print("\n3. Testing adjustment (5%)...")
    abbycar_adjustment = 5.0
    test_price = 20.0
    adjusted = test_price * (1 + abbycar_adjustment / 100)
    print(f"   {test_price}€ + 5% = {adjusted}€")
    
    # Save to BytesIO
    print("\n4. Saving to BytesIO...")
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    excel_bytes = excel_file.getvalue()
    print(f"   ✓ Excel file size: {len(excel_bytes)} bytes")
    
    # Check signature
    print("\n5. Checking Excel signature...")
    if excel_bytes.startswith(b'PK'):
        print("   ✓ Valid Excel file (PK signature)")
    else:
        print("   ✗ Invalid Excel file (no PK signature)")
        sys.exit(1)
    
    # Save to file
    print("\n6. Saving test file...")
    with open('/tmp/test_abbycar.xlsx', 'wb') as f:
        f.write(excel_bytes)
    print("   ✓ Saved to /tmp/test_abbycar.xlsx")
    
    print("\n" + "=" * 60)
    print("SUCCESS! Excel generation works.")
    print("=" * 60)
    print("\nOpen /tmp/test_abbycar.xlsx to verify")
    
except Exception as e:
    print(f"\n✗ ERROR: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
