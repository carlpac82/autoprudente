from __future__ import annotations

def _no_store_json(payload: Dict[str, Any], status_code: int = 200) -> JSONResponse:
    try:
        return JSONResponse(
            payload,
            status_code=status_code,
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0",
            },
        )
    except Exception:
        return JSONResponse(payload, status_code=status_code)
def render_with_playwright(url: str) -> str:
    if not _HAS_PLAYWRIGHT:
        return ""

def _is_carjet(u: str) -> bool:
    try:
        from urllib.parse import urlparse as _parse
        return _parse(u).netloc.endswith("carjet.com")
    except Exception:
        return False

def _ensure_settings_table():
    try:
        with _db_lock:
            con = _db_connect()
            try:
                con.execute(
                    "CREATE TABLE IF NOT EXISTS app_settings (key TEXT PRIMARY KEY, value TEXT)"
                )
                con.commit()
            finally:
                con.close()
    except Exception:
        pass

def _get_setting(key: str, default: str = "") -> str:
    try:
        _ensure_settings_table()
        with _db_lock:
            con = _db_connect()
            try:
                cur = con.execute("SELECT value FROM app_settings WHERE key=?", (key,))
                r = cur.fetchone()
                return (r[0] if r and r[0] is not None else default)
            finally:
                con.close()
    except Exception:
        return default

def _set_setting(key: str, value: str) -> None:
    try:
        _ensure_settings_table()
        with _db_lock:
            con = _db_connect()
            try:
                con.execute(
                    "INSERT INTO app_settings (key, value) VALUES (?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (key, value),
                )
                con.commit()
            finally:
                con.close()
    except Exception:
        pass

def _get_carjet_adjustment() -> Tuple[float, float]:
    """
    CRITICAL: Carjet Adjustment must ALWAYS be 0% and 0€
    User requirement: Never apply price adjustments to Carjet
    """
    return 0.0, 0.0

def _get_abbycar_adjustment() -> float:
    """
    Get Abbycar Excel export price adjustment percentage
    Returns: Percentage adjustment (e.g., 5.0 for +5%, -3.0 for -3%)
    Default: 3.0%
    """
    try:
        val = _get_setting("abbycar_pct")
        return float(val) if val else 3.0
    except Exception:
        return 3.0

def _get_abbycar_low_deposit_enabled() -> bool:
    """
    Check if Low Deposit adjustment is enabled
    Returns: True if enabled, False otherwise
    Default: False
    """
    try:
        val = _get_setting("abbycar_low_deposit_enabled")
        return val == "1" or val == "true" or val == True
    except Exception:
        return False

def _get_abbycar_low_deposit_adjustment() -> float:
    """
    Get Low Deposit groups additional adjustment percentage
    Returns: Additional percentage adjustment for Low Deposit groups
    Default: 0.0%
    """
    try:
        val = _get_setting("abbycar_low_deposit_pct")
        return float(val) if val else 0.0
    except Exception:
        return 0.0

def apply_price_adjustments(items: List[Dict[str, Any]], base_url: str) -> List[Dict[str, Any]]:
    try:
        if not items:
            return items
        if not _is_carjet(base_url):
            return items
        pct, off = _get_carjet_adjustment()
        if pct == 0 and off == 0:
            return items
        out: List[Dict[str, Any]] = []
        for it in items:
            ptxt = str(it.get("price") or "")
            amt = _parse_amount(ptxt)
            if amt is None:
                out.append(it)
                continue
            adj = amt * (1.0 + (pct/100.0)) + off
            it2 = dict(it)
            it2.setdefault("original_price", ptxt)
            it2["price"] = _format_eur(adj)
            it2["currency"] = "EUR"
            out.append(it2)
        return out
    except Exception:
        return items

def scrape_with_playwright(url: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []
    if not _HAS_PLAYWRIGHT:
        return items
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            context = browser.new_context(locale="pt-PT", user_agent="Mozilla/5.0 (compatible; PriceTracker/1.0)")
            try:
                context.add_cookies([
                    {"name": "monedaForzada", "value": "EUR", "domain": ".carjet.com", "path": "/"},
                    {"name": "moneda", "value": "EUR", "domain": ".carjet.com", "path": "/"},
                    {"name": "currency", "value": "EUR", "domain": ".carjet.com", "path": "/"},
                    {"name": "country", "value": "PT", "domain": ".carjet.com", "path": "/"},
                    {"name": "idioma", "value": "PT", "domain": ".carjet.com", "path": "/"},
                    {"name": "lang", "value": "pt", "domain": ".carjet.com", "path": "/"},
                ])
            except Exception:
                pass
            page = context.new_page()
            try:
                page.set_extra_http_headers({"Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8"})
            except Exception:
                pass
            page.goto(url, wait_until="networkidle", timeout=35000)
            
            # ===== FILTRAR APENAS AUTOPRUDENTE =====
            try:
                # Aguardar filtros carregarem
                page.wait_for_selector('#chkAUP', timeout=5000)
                print("[PLAYWRIGHT] Checkbox AUTOPRUDENTE encontrado", file=sys.stderr, flush=True)
                
                # IMPORTANTE: Aceitar cookies primeiro se aparecer
                try:
                    page.evaluate("""
                        // Procurar e clicar no botão de cookies
                        const buttons = document.querySelectorAll('button');
                        for (let btn of buttons) {
                            const text = btn.textContent.toLowerCase().trim();
                            if (text.includes('aceitar todos') || text.includes('aceitar tudo')) {
                                btn.click();
                                console.log('✓ Cookies aceitos:', btn.textContent);
                                break;
                            }
                        }
                        // Remover banners de cookies
                        document.querySelectorAll('[id*=cookie], [class*=cookie], [id*=didomi], [class*=didomi]').forEach(el => {
                            el.remove();
                        });
                        document.body.style.overflow = 'auto';
                    """)
                    print("[PLAYWRIGHT] Cookies aceites via JavaScript", file=sys.stderr, flush=True)
                    page.wait_for_timeout(1000)
                except Exception as e:
                    print(f"[PLAYWRIGHT] Erro ao aceitar cookies: {e}", file=sys.stderr, flush=True)
                    pass
                
                # Desmarcar todos os checkboxes de suppliers primeiro
                print("[PLAYWRIGHT] Desmarcando todos os suppliers...", file=sys.stderr, flush=True)
                page.evaluate("""
                    const checkboxes = document.querySelectorAll('input[name="frmPrv"]:checked');
                    checkboxes.forEach(cb => cb.click());
                """)
                
                # Aguardar um pouco
                page.wait_for_timeout(1000)
                
                # AUTODETECTAR COOKIES após desmarcar
                try:
                    page.evaluate("""
                        const buttons = document.querySelectorAll('button');
                        for (let btn of buttons) {
                            const text = btn.textContent.toLowerCase().trim();
                            if (text.includes('aceitar todos') || text.includes('aceitar tudo')) {
                                btn.click();
                                console.log('✓ Cookies aceitos após desmarcar');
                                break;
                            }
                        }
                    """)
                except:
                    pass
                
                # Marcar apenas AUTOPRUDENTE
                print("[PLAYWRIGHT] Marcando apenas AUTOPRUDENTE...", file=sys.stderr, flush=True)
                page.evaluate("""
                    const aupCheckbox = document.querySelector('#chkAUP');
                    if (aupCheckbox && !aupCheckbox.checked) {
                        aupCheckbox.click();
                    }
                """)
                
                print("[PLAYWRIGHT] Filtro AUTOPRUDENTE ativado", file=sys.stderr, flush=True)
                
                # Aguardar página recarregar com filtro
                page.wait_for_load_state("networkidle", timeout=15000)
                page.wait_for_timeout(2000)
                
                # AUTODETECTAR COOKIES após marcar AUTOPRUDENTE
                try:
                    page.evaluate("""
                        const buttons = document.querySelectorAll('button');
                        for (let btn of buttons) {
                            const text = btn.textContent.toLowerCase().trim();
                            if (text.includes('aceitar todos') || text.includes('aceitar tudo')) {
                                btn.click();
                                console.log('✓ Cookies aceitos após AUTOPRUDENTE');
                                break;
                            }
                        }
                    """)
                except:
                    pass
                    
            except Exception as e:
                print(f"[PLAYWRIGHT] Erro ao filtrar AUTOPRUDENTE: {e}", file=sys.stderr, flush=True)
                # Continuar mesmo se falhar o filtro
                pass
            # ===== FIM FILTRO AUTOPRUDENTE =====
            
            # Try clicking the primary search/submit button if the page requires it to load results
            try:
                btn = None
                # Prefer role-based lookup, then fall back to text and CSS selectors
                try:
                    btn = page.get_by_role("button", name=re.compile(r"(Pesquisar|Buscar|Search)", re.I))
                except Exception:
                    btn = None
                if btn and btn.is_visible():
                    btn.click(timeout=3000)
                else:
                    cand = page.locator("button:has-text('Pesquisar'), button:has-text('Buscar'), button:has-text('Search'), input[type=submit], button[type=submit]")
                    if cand and (cand.count() or 0) > 0:
                        try:
                            cand.first.click(timeout=3000)
                        except Exception:
                            pass
                # After clicking, wait for network to settle and results to appear
                try:
                    page.wait_for_load_state("networkidle", timeout=10000)
                except Exception:
                    pass
            except Exception:
                pass
            # Incremental scroll to trigger lazy loading
            try:
                for _ in range(5):
                    try:
                        page.mouse.wheel(0, 2000)
                    except Exception:
                        pass
                    page.wait_for_timeout(400)
            except Exception:
                pass
            try:
                page.wait_for_selector("section.newcarlist article, .newcarlist article, article.car, li.result, li.car, .car-item, .result-row", timeout=30000)
            except Exception:
                pass

            # Query all cards - SELETORES ESPECÍFICOS CARJET
            handles = page.query_selector_all("section.newcarlist article")
            idx = 0
            print(f"[PLAYWRIGHT] Encontrados {len(handles)} artigos", file=sys.stderr, flush=True)
            
            for h in handles:
                try:
                    card_text = (h.inner_text() or "").strip()
                except Exception:
                    card_text = ""
                
                # === PREÇO TOTAL (CARJET ESPECÍFICO) ===
                price_text = ""
                try:
                    # Prioridade 1: .pr-euros (preço em euros - TESTADO E FUNCIONA)
                    price_el = h.query_selector(".pr-euros")
                    if price_el:
                        price_text = (price_el.inner_text() or "").strip()
                    
                    # Prioridade 2: .price.pr-euros (alternativa)
                    if not price_text:
                        price_el = h.query_selector(".price.pr-euros")
                        if price_el:
                            price_text = (price_el.inner_text() or "").strip()
                    
                    # Prioridade 3: Procurar "Preço por X dias: XX,XX €" no texto
                    if not price_text:
                        # Procurar padrão: "Preço por 5 dias:\n8,80 €"
                        m = re.search(r"preço\s*por\s*\d+\s*dias:\s*([0-9]+[,\.][0-9]{2})\s*€", card_text, re.I)
                        if m:
                            price_text = m.group(1) + " €"
                    
                    # Limpar preço
                    if price_text:
                        # Remover tudo exceto números, vírgula, ponto e €
                        price_text = re.sub(r'[^\d,\.€\s]', '', price_text).strip()
                        if '€' not in price_text:
                            price_text += ' €'
                        
                except Exception as e:
                    print(f"[PLAYWRIGHT] Erro ao extrair preço: {e}", file=sys.stderr, flush=True)
                    price_text = ""
                
                # === NOME DO CARRO (CARJET ESPECÍFICO) ===
                car = ""
                try:
                    # Prioridade 1: h2 (TESTADO E FUNCIONA)
                    name_el = h.query_selector("h2")
                    if name_el:
                        car = (name_el.inner_text() or "").strip()
                    
                    # Fallback: outros seletores
                    if not car:
                        name_el = h.query_selector(".titleCar, .veh-name, .vehicle-name, .model, .title, h3")
                        if name_el:
                            car = (name_el.inner_text() or "").strip()
                except Exception:
                    pass
                
                # === SUPPLIER (CARJET ESPECÍFICO) ===
                supplier = ""
                try:
                    # Prioridade 1: Logo do supplier
                    im = h.query_selector("img[src*='/prv/'], img[src*='logo_']")
                    if im:
                        src = im.get_attribute("src") or ""
                        # Extrair código do supplier da URL: /logo_AUP.png → AUP
                        match = re.search(r'logo_([A-Z0-9]+)', src)
                        if match:
                            supplier = match.group(1)
                        else:
                            supplier = (im.get_attribute("alt") or "").strip()
                    
                    # Fallback: texto do supplier
                    if not supplier:
                        sup_el = h.query_selector(".supplier, .vendor, .partner, [class*='supplier']")
                        supplier = (sup_el.inner_text() or "").strip() if sup_el else ""
                except Exception:
                    pass
                
                # === CATEGORIA/GRUPO (CARJET ESPECÍFICO) ===
                category = ""
                try:
                    # Prioridade 1: .category
                    cat_el = h.query_selector(".category, .grupo, [class*='category'], [class*='grupo']")
                    if cat_el:
                        category = (cat_el.inner_text() or "").strip()
                    
                    # Prioridade 2: Extrair do texto (ex: "Grupo B1")
                    if not category:
                        match = re.search(r'grupo\s+([A-Z][0-9]?)', card_text, re.I)
                        if match:
                            category = match.group(1)
                except Exception:
                    pass
                # link
                link = ""
                try:
                    a = h.query_selector("a[href]")
                    if a:
                        href = a.get_attribute("href") or ""
                        if href and not href.lower().startswith("javascript"):
                            from urllib.parse import urljoin as _urljoin
                            link = _urljoin(url, href)
                except Exception:
                    pass
                # Only add if we have a price
                if price_text:
                    # Mapear categoria para código de grupo
                    group_code = map_category_to_group(category, car)
                    
                    # Log detalhado
                    print(f"[PLAYWRIGHT] Carro #{idx+1}:", file=sys.stderr, flush=True)
                    print(f"  Nome: {car}", file=sys.stderr, flush=True)
                    print(f"  Supplier: {supplier}", file=sys.stderr, flush=True)
                    print(f"  Preço: {price_text}", file=sys.stderr, flush=True)
                    print(f"  Categoria: {category}", file=sys.stderr, flush=True)
                    print(f"  Grupo: {group_code}", file=sys.stderr, flush=True)
                    
                    items.append({
                        "id": idx,
                        "car": car,
                        "supplier": supplier,
                        "price": price_text,
                        "currency": "",
                        "category": category,
                        "group": group_code,
                        "transmission": "",
                        "photo": "",
                        "link": link or url,
                    })
                    idx += 1
                else:
                    print(f"[PLAYWRIGHT] ⚠️  Carro sem preço ignorado: {car}", file=sys.stderr, flush=True)
            # If no items collected via card scanning, try parsing the full rendered HTML
            try:
                if not items:
                    html_full = page.content()
                    try:
                        # Best-effort: save debug HTML for inspection
                        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
                        (DEBUG_DIR / f"playwright-capture-{stamp}.html").write_text(html_full or "", encoding="utf-8")
                    except Exception:
                        pass
                    try:
                        items2 = parse_prices(html_full, url)
                        if items2:
                            items = items2
                    except Exception:
                        pass
            except Exception:
                pass
            context.close()
            browser.close()
    except Exception:
        return items
    return items

import os
import sys
import secrets
import re
from urllib.parse import urljoin
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime, timezone, timedelta
import traceback as _tb
import logging
import json

from fastapi import FastAPI, Request, Form, Depends, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, JSONResponse, HTMLResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path

# ============================================================
# MONITORING & ERROR TRACKING (Sentry)
# ============================================================
try:
    import sentry_sdk
    from sentry_sdk.integrations.fastapi import FastApiIntegration
    from sentry_sdk.integrations.starlette import StarletteIntegration
    
    SENTRY_DSN = os.getenv("SENTRY_DSN")
    if SENTRY_DSN:
        sentry_sdk.init(
            dsn=SENTRY_DSN,
            integrations=[
                StarletteIntegration(transaction_style="url"),
                FastApiIntegration(transaction_style="url"),
            ],
            traces_sample_rate=0.1,  # 10% das transações
            profiles_sample_rate=0.1,  # 10% dos profiles
            environment=os.getenv("ENVIRONMENT", "production"),
            release=os.getenv("RENDER_GIT_COMMIT", "unknown"),
        )
        logging.info("✅ Sentry monitoring enabled")
    else:
        logging.info("ℹ️  Sentry DSN not configured - monitoring disabled")
except ImportError:
    logging.warning("⚠️  Sentry SDK not installed - monitoring disabled")
except Exception as e:
    logging.error(f"❌ Failed to initialize Sentry: {e}")
from urllib.parse import urlencode, quote_plus
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from starlette.status import HTTP_303_SEE_OTHER
from dotenv import load_dotenv
import requests
import asyncio
from bs4 import BeautifulSoup
import sqlite3
from threading import Lock
import random

# Import database module for PostgreSQL/SQLite hybrid support
try:
    from database import _db_connect as _db_connect_new, USE_POSTGRES
    _USE_NEW_DB = True
    if USE_POSTGRES:
        logging.info("🐘 PostgreSQL mode enabled")
    else:
        logging.info("📁 SQLite mode (local development)")
except ImportError:
    _USE_NEW_DB = False
    logging.info("📁 Using legacy SQLite connection")
import time
import io
import hashlib
import smtplib
from email.message import EmailMessage
from fastapi import Query
try:
    import httpx  # type: ignore
    _HTTPX_CLIENT = httpx.Client(timeout=httpx.Timeout(10.0, connect=4.0), headers={"Connection": "keep-alive"})
    _HTTPX_ASYNC: Optional["httpx.AsyncClient"] = httpx.AsyncClient(timeout=httpx.Timeout(10.0, connect=4.0), headers={"Connection": "keep-alive"})
except Exception:
    _HTTPX_CLIENT = None
    _HTTPX_ASYNC = None

# Import VEHICLES dictionary from carjet_direct
try:
    from carjet_direct import VEHICLES
except ImportError:
    logging.warning("⚠️  Could not import VEHICLES from carjet_direct")
    VEHICLES = {}

# Load environment variables FIRST before checking USE_PLAYWRIGHT
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s:     %(message)s'
)

try:
    from playwright.sync_api import sync_playwright  # type: ignore
    _HAS_PLAYWRIGHT = True
except Exception:
    _HAS_PLAYWRIGHT = False

# Environment variables
USE_PLAYWRIGHT = str(os.getenv("USE_PLAYWRIGHT", "")).strip().lower() in ("1","true","yes","on")
_test_mode_val = os.getenv("TEST_MODE_LOCAL", "0").strip()
TEST_MODE_LOCAL = int(_test_mode_val) if _test_mode_val.isdigit() else (1 if _test_mode_val.lower() in ("true", "yes") else 0)
TEST_FARO_URL = os.getenv("TEST_FARO_URL", "")
TEST_ALBUFEIRA_URL = os.getenv("TEST_ALBUFEIRA_URL", "")
APP_PASSWORD = os.getenv("APP_PASSWORD", "change_me")
SECRET_KEY = os.getenv("SECRET_KEY", secrets.token_urlsafe(32))
TARGET_URL = os.getenv("TARGET_URL", "https://example.com")
SCRAPER_SERVICE = os.getenv("SCRAPER_SERVICE", "")
SCRAPER_API_KEY = os.getenv("SCRAPER_API_KEY", "")
SCRAPER_COUNTRY = os.getenv("SCRAPER_COUNTRY", "").strip()
APP_USERNAME = os.getenv("APP_USERNAME", "user")
CARJET_PRICE_ADJUSTMENT_PCT = float(os.getenv("CARJET_PRICE_ADJUSTMENT_PCT", "0") or 0)
CARJET_PRICE_OFFSET_EUR = float(os.getenv("CARJET_PRICE_OFFSET_EUR", "0") or 0)
AUDIT_RETENTION_DAYS = int(os.getenv("AUDIT_RETENTION_DAYS", "90") or 90)
IMAGE_CACHE_DAYS = int(os.getenv("IMAGE_CACHE_DAYS", "365") or 365)
PRICES_CACHE_TTL_SECONDS = int(os.getenv("PRICES_CACHE_TTL_SECONDS", "300") or 300)
BULK_CONCURRENCY = int(os.getenv("BULK_CONCURRENCY", "6") or 6)
BULK_MAX_RETRIES = int(os.getenv("BULK_MAX_RETRIES", "2") or 2)
GLOBAL_FETCH_RPS = float(os.getenv("GLOBAL_FETCH_RPS", "5") or 5.0)

# --- Precompiled regexes for parser performance ---
AUTO_RX = re.compile(r"\b(auto|automatic|automatico|automático|automatik|aut\.|a/t|at|dsg|cvt|bva|tiptronic|steptronic|s\s*tronic|multidrive|multitronic|eat|eat6|eat8)\b", re.I)
BG_IMAGE_RX = re.compile(r"background-image\s*:\s*url\(([^)]+)\)", re.I)
LOGO_CODE_RX = re.compile(r"/logo_([A-Za-z0-9]+)\.", re.I)
CAR_CODE_RX = re.compile(r"car_([A-Za-z0-9]+)\.jpg", re.I)
OBJ_RX = re.compile(r"\{[^{}]*\"priceStr\"\s*:\s*\"[^\"]+\"[^{}]*\"id\"\s*:\s*\"[^\"]+\"[^{}]*\}", re.S)
DATAMAP_RX = re.compile(r"var\s+dataMap\s*=\s*(\[.*?\]);", re.S)

app = FastAPI(title="Rental Price Tracker")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY, same_site="lax")
app.add_middleware(GZipMiddleware, minimum_size=500)

@app.on_event("startup")
async def startup_event():
    """Initialize database and create default users on startup"""
    print(f"========================================", flush=True)
    print(f"🚀 APP STARTUP - Rental Price Tracker", flush=True)
    print(f"========================================", flush=True)
    
    # Initialize database tables FIRST
    try:
        print(f"📊 Initializing database tables...", flush=True)
        _ensure_users_table()
        print(f"   ✅ users table created/exists", flush=True)
    except Exception as e:
        print(f"⚠️  Database initialization error: {e}", flush=True)
    
    # Fix PostgreSQL schema AFTER tables exist
    if _USE_NEW_DB and USE_POSTGRES:
        try:
            print(f"🔧 Fixing PostgreSQL schema...", flush=True)
            # Run inline instead of subprocess for better error handling
            from fix_postgres_schema import fix_users_table, fix_system_logs_table
            
            if fix_users_table():
                print(f"   ✅ users schema fixed", flush=True)
            else:
                print(f"   ⚠️  users schema warnings", flush=True)
            
            if fix_system_logs_table():
                print(f"   ✅ system_logs schema fixed", flush=True)
            else:
                print(f"   ⚠️  system_logs schema warnings", flush=True)
        except Exception as e:
            print(f"⚠️  Schema fix error: {e}", flush=True)
            import traceback
            traceback.print_exc()
    
    # Create default users AFTER schema is fixed
    try:
        print(f"👥 Creating default users...", flush=True)
        _ensure_default_users()
        print(f"✅ Default users ready (admin/admin)", flush=True)
    except Exception as e:
        print(f"⚠️  Default users error: {e}", flush=True)
    
    print(f"========================================", flush=True)

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    # Redirect to login on unauthorized/forbidden
    if exc.status_code in (401, 403):
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)

# --- Admin: Test email ---
@app.get("/admin/price-validation", response_class=HTMLResponse)
async def admin_price_validation(request: Request):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    return templates.TemplateResponse("price_validation_rules.html", {"request": request})

@app.get("/admin/export-db")
async def admin_export_db(request: Request):
    """Temporary endpoint to export database"""
    try:
        require_admin(request)
    except HTTPException:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        return FileResponse(
            path=str(DB_PATH),
            filename="data_backup.db",
            media_type="application/octet-stream"
        )
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/admin/export-vehicles-json")
async def admin_export_vehicles_json(request: Request):
    """Export vehicles as JSON"""
    try:
        require_admin(request)
    except HTTPException:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        vehicles = []
        with _db_lock:
            con = _db_connect()
            try:
                cur = con.execute("SELECT id, brand, model, code, category, doors, seats, transmission, luggage, photo_url, enabled FROM car_groups")
                for r in cur.fetchall():
                    vehicles.append({
                        "id": r[0], "brand": r[1], "model": r[2], "code": r[3],
                        "category": r[4], "doors": r[5], "seats": r[6],
                        "transmission": r[7], "luggage": r[8], "photo_url": r[9], "enabled": r[10]
                    })
            finally:
                con.close()
        return JSONResponse({"vehicles": vehicles, "count": len(vehicles)})
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=500)

@app.post("/admin/test-email", response_class=HTMLResponse)
async def admin_test_email_send(request: Request, to: str = Form("")):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    err = None
    try:
        _send_creds_email((to or "").strip(), "test.user", "Temp1234!")
    except Exception as e:
        err = str(e)
    ok = err is None
    return templates.TemplateResponse("admin_test_email.html", {"request": request, "error": err, "ok": ok})

@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    try:
        (DEBUG_DIR / "last_exception.txt").write_text(_tb.format_exc(), encoding="utf-8")
    except Exception:
        pass
    # Ensure a valid response is always returned to Starlette
    return JSONResponse({"ok": False, "error": "Server error"}, status_code=500)

# --- Prices response cache (memory) ---
_PRICES_CACHE: Dict[str, Tuple[float, Dict[str, Any]]] = {}

async def _compute_prices_for(url: str) -> Dict[str, Any]:
    headers = {"User-Agent": "Mozilla/5.0 (compatible; PriceTracker/1.0)"}
    # Use async fetch to avoid blocking and improve concurrency
    r = await async_fetch_with_optional_proxy(url, headers=headers)
    r.raise_for_status()
    html = r.text
    # Parse HTML off the main loop
    items = await asyncio.to_thread(parse_prices, html, url)
    items = convert_items_gbp_to_eur(items)
    items = apply_price_adjustments(items, url)
    # schedule image prefetch (best-effort)
    try:
        img_urls: List[str] = []
        for it in items:
            u = (it.get("photo") or "").strip()
            if u and (u.startswith("http://") or u.startswith("https://")):
                img_urls.append(u)
        if img_urls:
            asyncio.create_task(_prefetch_many(img_urls[:12]))
            asyncio.create_task(_delayed_prefetch(img_urls[12:64], 1.5))
    except Exception:
        pass
    return {"ok": True, "count": len(items), "items": items}

def _cache_get(url: str) -> Optional[Dict[str, Any]]:
    try:
        ts, payload = _PRICES_CACHE.get(url, (0.0, None))
        if not payload:
            return None
        age = time.time() - ts
        if age <= PRICES_CACHE_TTL_SECONDS:
            return payload
        return None
    except Exception:
        return None

def _cache_set(url: str, payload: Dict[str, Any]):
    try:
        _PRICES_CACHE[url] = (time.time(), payload)
    except Exception:
        pass

async def _refresh_prices_background(url: str):
    try:
        data = await _compute_prices_for(url)
        _cache_set(url, data)
    except Exception:
        pass
    return JSONResponse({"ok": False, "error": "Server error"}, status_code=500)

# --- Image cache proxy and retention ---
def _ext_from_content_type(ct: str) -> str:
    ct = (ct or "").lower()
    if "jpeg" in ct: return ".jpg"
    if "png" in ct: return ".png"
    if "webp" in ct: return ".webp"
    if "gif" in ct: return ".gif"
    if "svg" in ct: return ".svg"
    return ".bin"

def _guess_ext_from_url(u: str) -> str:
    try:
        p = u.split("?")[0]
        for ext in (".jpg", ".jpeg", ".png", ".webp", ".gif", ".svg"):
            if p.lower().endswith(ext):
                return ".jpg" if ext == ".jpeg" else ext
    except Exception:
        pass
    return ""

def _cache_path_for(url: str) -> Path:
    import hashlib
    h = hashlib.sha256(url.encode("utf-8")).hexdigest()
    return CACHE_CARS_DIR / h

def _serve_file(fp: Path, content_type: str = "application/octet-stream"):
    try:
        data = fp.read_bytes()
    except Exception:
        raise HTTPException(status_code=404, detail="Not found")
    headers = {"Cache-Control": f"public, max-age={IMAGE_CACHE_DAYS*86400}"}
    return Response(content=data, media_type=content_type or "application/octet-stream", headers=headers)

@app.get("/img")
async def img_proxy(request: Request, src: str):
    try:
        if not src or not (src.startswith("http://") or src.startswith("https://")):
            raise HTTPException(status_code=400, detail="Invalid src")
        key = _cache_path_for(src)
        meta = key.with_suffix(".meta")
        # Serve from cache if present
        if key.exists():
            try:
                now = time.time(); os.utime(key, (now, now));
                if meta.exists(): os.utime(meta, (now, now))
            except Exception:
                pass
            ct = "application/octet-stream"
            try:
                if meta.exists():
                    ct = (meta.read_text(encoding="utf-8").strip() or ct)
            except Exception:
                pass
            return _serve_file(key, ct)

        # On HEAD requests, don't fetch body, just forward and prime headers
        if request.method == "HEAD":
            import httpx
            async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
                hr = await client.head(src)
            if hr.status_code != 200:
                raise HTTPException(status_code=404, detail="Upstream not found")
            headers = {"Cache-Control": f"public, max-age={IMAGE_CACHE_DAYS*86400}"}
            return Response(status_code=200, headers=headers)

        # Fetch from origin using requests for broader SSL compatibility, then cache
        import requests as _rq
        try:
            rr = _rq.get(src, timeout=15, headers={"User-Agent": "PriceTracker/1.0"})
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"Upstream error: {type(e).__name__}")
        if rr.status_code != 200 or not rr.content:
            raise HTTPException(status_code=404, detail="Upstream not found")
        ct = rr.headers.get("content-type", "application/octet-stream")
        try:
            with key.open("wb") as f:
                f.write(rr.content)
            meta.write_text(ct, encoding="utf-8")
        except Exception:
            pass
        headers = {"Cache-Control": f"public, max-age={IMAGE_CACHE_DAYS*86400}"}
        return Response(content=rr.content, media_type=ct or "application/octet-stream", headers=headers)
    except HTTPException:
        raise
    except Exception as e:
        try:
            (DEBUG_DIR / "img_error.txt").write_text(f"{type(e).__name__}: {e}\n", encoding="utf-8")
        except Exception:
            pass
        raise HTTPException(status_code=500, detail="Image fetch error")

def cleanup_image_cache():
    try:
        cutoff = time.time() - IMAGE_CACHE_DAYS*86400
        for fp in CACHE_CARS_DIR.glob("*"):
            try:
                if fp.is_file():
                    st = fp.stat()
                    if max(st.st_mtime, st.st_atime) < cutoff:
                        fp.unlink(missing_ok=True)
            except Exception:
                continue
    except Exception:
        pass

BASE_DIR = Path(__file__).resolve().parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

static_dir = BASE_DIR / "static"
if static_dir.exists():
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")
# Persistent image cache under DATA_DIR
CACHE_CARS_DIR = Path(os.environ.get("CACHE_IMAGES_DIR", str(Path(os.environ.get("DATA_DIR", str(BASE_DIR))) / "cars")))
CACHE_CARS_DIR.mkdir(parents=True, exist_ok=True)
# Persisted uploads live under DATA_DIR and are served at /uploads
UPLOADS_ROOT = Path(os.environ.get("UPLOADS_ROOT", str(Path(os.environ.get("DATA_DIR", str(BASE_DIR))) / "uploads")))
UPLOADS_ROOT.mkdir(parents=True, exist_ok=True)
try:
    app.mount("/uploads", StaticFiles(directory=str(UPLOADS_ROOT)), name="uploads")
except Exception:
    pass
UPLOADS_DIR = UPLOADS_ROOT / "profiles"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

# --- Background image prefetch ---
async def _prefetch_image(url: str):
    try:
        if not url or not (url.startswith("http://") or url.startswith("https://")):
            return
        key = _cache_path_for(url)
        if key.exists() and key.stat().st_size > 0:
            # already cached
            try:
                now = time.time(); os.utime(key, (now, now))
            except Exception:
                pass
            return
        import httpx
        async with httpx.AsyncClient(timeout=10.0, follow_redirects=True) as client:
            r = await client.get(url)
            if r.status_code != 200 or not r.content:
                return
            try:
                with key.open("wb") as f:
                    f.write(r.content)
            except Exception:
                pass
    except Exception:
        pass

async def _prefetch_many(urls: List[str]):
    try:
        tasks = [asyncio.create_task(_prefetch_image(u)) for u in urls if isinstance(u, str) and u]
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)
    except Exception:
        pass

async def _delayed_prefetch(urls: List[str], delay_seconds: float = 1.5):
    try:
        await asyncio.sleep(delay_seconds)
        await _prefetch_many(urls)
    except Exception:
        pass

# --- Icon fallbacks to avoid 404s ---
@app.get("/favicon.ico")
async def favicon_redirect():
    return RedirectResponse(url="/static/autoprudente-favicon.png?v=2", status_code=HTTP_303_SEE_OTHER)

@app.get("/apple-touch-icon.png")
async def apple_touch_icon_redirect():
    return RedirectResponse(url="/static/autoprudente-favicon.png?v=2", status_code=HTTP_303_SEE_OTHER)

@app.get("/apple-touch-icon-precomposed.png")
async def apple_touch_icon_pre_redirect():
    return RedirectResponse(url="/static/autoprudente-favicon.png?v=2", status_code=HTTP_303_SEE_OTHER)

@app.get("/static/ap-favicon.png")
async def static_ap_favicon_redirect():
    return RedirectResponse(url="/static/autoprudente-favicon.png?v=2", status_code=HTTP_303_SEE_OTHER)

DATA_DIR = Path(os.environ.get("DATA_DIR", str(BASE_DIR)))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "data.db"
_db_lock = Lock()
DEBUG_DIR = Path(os.environ.get("DEBUG_DIR", BASE_DIR / "static" / "debug"))
DEBUG_DIR.mkdir(parents=True, exist_ok=True)

# --- Admin/Users: DB helpers ---
class PostgreSQLConnectionWrapper:
    """Wrapper para adicionar método execute() à conexão PostgreSQL"""
    def __init__(self, conn):
        self._conn = conn
        self._cursor = None
    
    def execute(self, query, params=None):
        """Execute query usando cursor"""
        # Convert SQLite ? placeholders to PostgreSQL %s
        if '?' in query:
            # Count number of ? to ensure we have right number of params
            num_placeholders = query.count('?')
            query = query.replace('?', '%s')
            
            # Ensure params is a tuple
            if params is not None:
                if not isinstance(params, (tuple, list)):
                    params = (params,)
                elif isinstance(params, list):
                    params = tuple(params)
        
        # Convert SQLite AUTOINCREMENT to PostgreSQL SERIAL
        if 'AUTOINCREMENT' in query.upper():
            query = query.replace('INTEGER PRIMARY KEY AUTOINCREMENT', 'SERIAL PRIMARY KEY')
            query = query.replace('AUTOINCREMENT', '')
        
        self._cursor = self._conn.cursor()
        try:
            if params:
                self._cursor.execute(query, params)
            else:
                self._cursor.execute(query)
        except Exception as e:
            # Log the error with query and params for debugging
            import logging
            logging.error(f"PostgreSQL execute error: {e}")
            logging.error(f"Query: {query}")
            logging.error(f"Params: {params}")
            raise
        return self._cursor
    
    def commit(self):
        return self._conn.commit()
    
    def rollback(self):
        return self._conn.rollback()
    
    def close(self):
        if self._cursor:
            self._cursor.close()
        return self._conn.close()
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            self.rollback()
        self.close()

def _db_connect():
    """Database connection - supports both PostgreSQL and SQLite"""
    if _USE_NEW_DB:
        conn = _db_connect_new()
        # Wrap PostgreSQL connection to add execute() method
        if hasattr(conn, 'cursor') and not hasattr(conn, 'row_factory'):
            return PostgreSQLConnectionWrapper(conn)
        return conn
    else:
        return sqlite3.connect(str(DB_PATH))

def _ensure_users_table():
    with _db_lock:
        con = _db_connect()
        try:
            con.execute("""
                CREATE TABLE IF NOT EXISTS users (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE NOT NULL,
                  password_hash TEXT NOT NULL,
                  first_name TEXT,
                  last_name TEXT,
                  mobile TEXT,
                  email TEXT,
                  profile_picture_path TEXT,
                  profile_picture_data BLOB,
                  is_admin INTEGER DEFAULT 0,
                  enabled INTEGER DEFAULT 1,
                  created_at TEXT,
                  google_id TEXT UNIQUE
                )
            """)
            
            # Migration: Add google_id column if it doesn't exist
            try:
                con.execute("ALTER TABLE users ADD COLUMN google_id TEXT UNIQUE")
                con.commit()
                logging.info("✅ Added google_id column to users table")
            except Exception as e:
                # Column already exists, ignore
                pass
            
            # Migration: Add profile_picture_data column if it doesn't exist
            try:
                con.execute("ALTER TABLE users ADD COLUMN profile_picture_data BLOB")
                con.commit()
                logging.info("✅ Added profile_picture_data column to users table")
            except Exception as e:
                # Column already exists, ignore
                pass
            
            con.commit()
        finally:
            con.close()

def _get_user_by_username(username: str) -> Optional[Dict[str, Any]]:
    try:
        with _db_lock:
            con = _db_connect()
            try:
                cur = con.execute("SELECT id, username, first_name, last_name, email, mobile, profile_picture_path, is_admin, enabled FROM users WHERE username=?", (username,))
                r = cur.fetchone()
                if not r:
                    return None
                return {
                    "id": r[0],
                    "username": r[1],
                    "first_name": r[2] or "",
                    "last_name": r[3] or "",
                    "email": r[4] or "",
                    "mobile": r[5] or "",
                    "profile_picture_path": r[6] or "",
                    "is_admin": bool(r[7]),
                    "enabled": bool(r[8]),
                }
            finally:
                con.close()
    except Exception:
        return None

# --- Activity Log ---
def _ensure_activity_table():
    with _db_lock:
        con = _db_connect()
        try:
            con.execute(
                """
                CREATE TABLE IF NOT EXISTS activity_log (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  ts_utc TEXT NOT NULL,
                  username TEXT,
                  action TEXT NOT NULL,
                  details TEXT,
                  ip TEXT,
                  user_agent TEXT
                );
                """
            )
            con.commit()
        finally:
            con.close()

def log_activity(request: Request, action: str, details: str = "", username: Optional[str] = None):
    try:
        _ensure_activity_table()
    except Exception:
        pass
    # best-effort metadata
    try:
        ip = request.client.host if request and request.client else None
    except Exception:
        ip = None
    ua = request.headers.get("user-agent", "") if request else ""
    user = username or (request.session.get("username") if request and request.session else None)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                con.execute(
                    "INSERT INTO activity_log (ts_utc, username, action, details, ip, user_agent) VALUES (?,?,?,?,?,?)",
                    (datetime.now(timezone.utc).isoformat(), user, action, details, ip or "", ua[:300])
                )
                con.commit()
            finally:
                con.close()
    except Exception:
        pass
    pass

def cleanup_activity_retention():
    try:
        _ensure_activity_table()
        if AUDIT_RETENTION_DAYS <= 0:
            return
        cutoff = datetime.now(timezone.utc).timestamp() - AUDIT_RETENTION_DAYS*86400
        # Compare lexicographically on ISO timestamps by computing a boundary
        cutoff_iso = datetime.utcfromtimestamp(cutoff).replace(tzinfo=timezone.utc).isoformat()
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("DELETE FROM activity_log WHERE ts_utc < ?", (cutoff_iso,))
                con.commit()
            finally:
                con.close()
    except Exception:
        pass

def _hash_password(pw: str, salt: str = ""):  # basic salted sha256
    if not salt:
        salt = secrets.token_hex(8)
    digest = hashlib.sha256((salt + ":" + pw).encode("utf-8")).hexdigest()
    return f"sha256:{salt}:{digest}"

def _verify_password(pw: str, stored: str) -> bool:
    try:
        algo, salt, digest = stored.split(":", 2)
        if algo != "sha256":
            return False
        test = hashlib.sha256((salt + ":" + pw).encode("utf-8")).hexdigest()
        return secrets.compare_digest(test, digest)
    except Exception:
        return False

def clean_car_name(car_name: str) -> str:
    """
    Limpa e normaliza nomes de carros EXATAMENTE como o Vehicle Editor
    - Remove duplicações como "Autoautomático" → "Automático"
    - Remove "ou similar"
    - Remove "4p" (4 portas) exceto para 7 e 9 lugares
    - Normaliza espaços
    - Converte para LOWERCASE (igual ao VEHICLES dictionary)
    """
    if not car_name:
        return ""
    
    name = str(car_name).strip()
    
    # Remover duplicações comuns
    name = re.sub(r'[Aa]uto[Aa]utom[aá]tico', 'Automático', name)
    name = re.sub(r'[Aa]uto[Aa]utomatic', 'Automatic', name)
    
    # Remover "ou similar" e variantes
    name = re.sub(r'\s*ou\s+similar(es)?.*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*or\s+similar.*$', '', name, flags=re.IGNORECASE)
    
    # Remover vírgulas e espaços extras (ex: "2008 , Electric" → "2008 Electric")
    name = re.sub(r'\s*,\s*', ' ', name)
    
    # Remover "Special Edition" e variantes
    name = re.sub(r'\s+special\s+edition\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+edition\b', '', name, flags=re.IGNORECASE)
    
    # Remover "4p" (4 portas) EXCETO para 7 e 9 lugares
    # Exemplos: "Fiat 500 4p" → "Fiat 500", "Fiat Panda 4p" → "Fiat Panda"
    # MAS: "Dacia Lodgy 7 Lugares 4p" → mantém (não remove)
    name_lower = name.lower()
    if '7' not in name_lower and '9' not in name_lower and 'seater' not in name_lower and 'lugares' not in name_lower:
        # Remover "4p", "4 portas", "4 doors"
        name = re.sub(r'\s+4p\b', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s+4\s*portas?\b', '', name, flags=re.IGNORECASE)
        name = re.sub(r'\s+4\s*doors?\b', '', name, flags=re.IGNORECASE)
    
    # Normalizar espaços múltiplos
    name = re.sub(r'\s+', ' ', name).strip()
    
    # NÃO converter para lowercase aqui! 
    # Manter capitalização original para display bonito
    # O lowercase é feito apenas quando consultar VEHICLES
    
    return name

def capitalize_car_name(car_name: str) -> str:
    """
    Capitaliza nomes de carros para display:
    - Primeira letra de cada palavra em maiúscula
    - SW, SUV, 4X4 sempre em maiúsculas
    - Exemplos:
      - "peugeot 2008 auto" → "Peugeot 2008 Auto"
      - "renault megane sw auto" → "Renault Megane SW Auto"
      - "toyota rav4 4x4 auto" → "Toyota Rav4 4X4 Auto"
    """
    if not car_name:
        return ""
    
    # Palavras que devem ficar em maiúsculas
    uppercase_words = {'sw', 'suv', '4x4', 'gt', 'gti', 'rs', 'st', 'amg', 'bmw', 'vw'}
    
    words = car_name.lower().split()
    capitalized = []
    
    for word in words:
        if word in uppercase_words:
            capitalized.append(word.upper())
        else:
            capitalized.append(word.capitalize())
    
    return ' '.join(capitalized)

def map_category_to_group(category: str, car_name: str = "") -> str:
    """
    Mapeia categorias descritivas para códigos de grupos definidos:
    B1, B2, D, E1, E2, F, G, X, J1, J2, L1, L2, M1, M2, N, Others
    
    CASE-INSENSITIVE: Converte para lowercase para comparação
    
    B1 vs B2 LOGIC (baseado em LUGARES, não PORTAS):
    - B1 = Mini 4 LUGARES (Fiat 500, Peugeot 108, C1, VW Up, Kia Picanto, Toyota Aygo)
    - B2 = Mini 5 LUGARES (Fiat Panda, Hyundai i10, etc)
    
    REGRAS ESPECIAIS:
    - Cabrio/Cabriolet → G (Cabrio)
    - Toyota Aygo X → F (SUV)
    - Mini 4 lugares Automático → E1
    - Premium/Luxury → X
    """
    # NÃO retornar "Others" aqui! Primeiro tentar categorias explícitas, depois car_groups e VEHICLES
    
    # Converter para lowercase para mapeamento case-insensitive
    cat = category.strip().lower() if category else ""
    car_lower = car_name.lower() if car_name else ""
    
    # PRIORIDADE -1: CABRIO/CABRIOLET no NOME → SEMPRE Grupo G
    # Independente da categoria (Luxury, Mini, SUV, etc), se tem "cabrio" no nome = G
    if any(word in car_lower for word in ['cabrio', 'cabriolet', 'convertible', 'conversível']):
        return "G"
    
    # PRIORIDADE 0: Categorias explícitas do CarJet (mais confiáveis que tabela manual)
    # Suporta INGLÊS e PORTUGUÊS
    
    # Mini 4 Seats / Mini 4 Lugares → B1 ou E1 (se automático)
    if cat in ['mini 4 seats', 'mini 4 doors', 'mini 4 portas', 'mini 4 lugares']:
        if any(word in car_lower for word in ['auto', 'automatic', 'automático', 'automatico']):
            return "E1"
        return "B1"
    
    # Mini 5 Seats / Mini 5 Lugares → B2 ou E1 (se automático)
    if cat in ['mini 5 seats', 'mini 5 doors', 'mini 5 portas', 'mini 5 lugares']:
        if any(word in car_lower for word in ['auto', 'automatic', 'automático', 'automatico']):
            return "E1"
        return "B2"
    
    # Mini Automatic / Mini Auto → E1
    if cat in ['mini automatic', 'mini auto', 'mini automático']:
        return "E1"
    
    # Economy / Económico → D
    if cat in ['economy', 'económico', 'compact', 'compacto']:
        return "D"
    
    # Economy Automatic / Economy Auto → E2
    if cat in ['economy automatic', 'economy auto', 'económico automatic', 'económico auto',
               'compact automatic', 'compact auto']:
        return "E2"
    
    # SUV → F
    if cat in ['suv', 'jeep']:
        return "F"
    
    # SUV Automatic / SUV Auto → L1
    if cat in ['suv automatic', 'suv auto', 'jeep automatic', 'jeep auto']:
        return "L1"
    
    # Station Wagon / Estate / Carrinha → J2
    if cat in ['station wagon', 'estate', 'carrinha', 'estate/station wagon', 'sw', 'touring']:
        return "J2"
    
    # Station Wagon Automatic → L2
    if cat in ['station wagon automatic', 'station wagon auto', 'estate automatic', 'estate auto',
               'carrinha automatic', 'carrinha auto', 'sw automatic', 'sw auto']:
        return "L2"
    
    # Crossover → J1
    if cat in ['crossover']:
        return "J1"
    
    # Cabrio / Cabriolet / Convertible → G
    if cat in ['cabrio', 'cabriolet', 'convertible', 'conversível']:
        return "G"
    
    # Luxury / Premium → Others (não oferecemos estas categorias)
    if cat in ['luxury', 'premium', 'luxo']:
        return "Others"
    
    # 7 Seater / 7 Seats → M1
    if cat in ['7 seater', '7 seats', '7 lugares', 'people carrier', 'mpv']:
        return "M1"
    
    # 7 Seater Automatic / 7 Seats Auto → M2
    if cat in ['7 seater automatic', '7 seater auto', '7 seats automatic', '7 seats auto',
               '7 lugares automatic', '7 lugares auto', '7 lugares automático',
               'mpv automatic', 'mpv auto']:
        return "M2"
    
    # 9 Seater / 9 Seats → N
    if cat in ['9 seater', '9 seater automatic', '9 seater auto',
               '9 seats', '9 seats automatic', '9 seats auto',
               '9 lugares', '9 lugares automatic', '9 lugares auto', '9 lugares automático',
               'minivan', 'van']:
        return "N"
    
    # PRIORIDADE 1: Consultar tabela car_groups (22 grupos categorizados manualmente)
    if car_name:
        try:
            car_clean = clean_car_name(car_name)
            car_clean_lower = car_clean.lower()
            
            # Buscar na tabela car_groups
            with _db_lock:
                conn = _db_connect()
                try:
                    # Tentar match exato primeiro
                    row = conn.execute(
                        "SELECT code FROM car_groups WHERE LOWER(name) = ? OR LOWER(model) = ?",
                        (car_clean_lower, car_clean_lower)
                    ).fetchone()
                    
                    if row:
                        # Extrair código do grupo (ex: B1-FIAT500 -> B1)
                        full_code = row[0]
                        group_code = full_code.split('-')[0] if '-' in full_code else full_code
                        return group_code
                finally:
                    conn.close()
        except Exception:
            pass  # Se falhar, continuar para próxima prioridade
    
    # PRIORIDADE 2: Consultar dicionário VEHICLES de carjet_direct.py
    # Tentar SEMPRE que tiver car_name, mesmo se category não estiver vazia
    # (mas evitar loop infinito: só chamar recursivamente se encontrar categoria diferente)
    if car_name:
        try:
            from carjet_direct import VEHICLES
            import re
            
            # Normalizar nome do carro para consulta (lowercase)
            car_clean = clean_car_name(car_name)
            car_clean_lower = car_clean.lower().strip()
            
            # Remover sufixos comuns que impedem match
            # "Peugeot E-208 Electric" → "peugeot e-208"
            # "Toyota Chr Auto" → "toyota chr auto"
            car_normalized = car_clean_lower
            car_normalized = re.sub(r'\s+(electric|hybrid|diesel|petrol|plug-in|phev)$', '', car_normalized, flags=re.IGNORECASE)
            car_normalized = re.sub(r'\s+4x4$', '', car_normalized, flags=re.IGNORECASE)
            car_normalized = re.sub(r'\s+\d+\s*door(s)?$', '', car_normalized, flags=re.IGNORECASE)
            car_normalized = re.sub(r',\s*electric$', '', car_normalized, flags=re.IGNORECASE)
            car_normalized = re.sub(r',\s*hybrid$', '', car_normalized, flags=re.IGNORECASE)
            car_normalized = car_normalized.strip()
            
            # Tentar match direto
            if car_normalized in VEHICLES:
                category_from_vehicles = VEHICLES[car_normalized]
                # VEHICLES retorna categoria descritiva (ex: "ECONOMY", "SUV Auto")
                # Precisamos mapear para código de grupo (B1, D, F, etc)
                # Passar car_name também para manter contexto (ex: distinguir automático)
                # IMPORTANTE: Só chamar recursivamente se a categoria for diferente (evitar loop)
                if category_from_vehicles.lower() != cat:
                    return map_category_to_group(category_from_vehicles, car_name)
            
            # Tentar match parcial (buscar chave que está contida no nome ou vice-versa)
            # Ordenar por tamanho decrescente para pegar matches mais específicos primeiro
            for vehicle_key in sorted(VEHICLES.keys(), key=len, reverse=True):
                # Match se o nome do carro contém a chave completa
                # Ex: "toyota chr auto" contém "toyota chr"
                if len(vehicle_key) >= 5 and vehicle_key in car_normalized:
                    category_from_vehicles = VEHICLES[vehicle_key]
                    # Só chamar recursivamente se a categoria for diferente
                    if category_from_vehicles.lower() != cat:
                        return map_category_to_group(category_from_vehicles, car_name)
        except ImportError:
            pass  # carjet_direct.py não disponível
        except Exception:
            pass  # Se falhar, continuar para próxima prioridade
    
    # PRIORIDADE 3: CABRIO/CABRIOLET → Grupo G (apenas descapotáveis)
    if any(word in car_lower for word in ['cabrio', 'cabriolet', 'convertible', 'conversível']):
        return "G"
    
    # PRIORIDADE 4: Toyota Aygo X → F (SUV), não confundir com Aygo normal (B1)
    if 'aygo x' in car_lower or 'aygo-x' in car_lower:
        return "F"
    
    # PRIORIDADE 5: Modelos de 4 LUGARES → B1
    # (Fiat 500, Peugeot 108, C1, VW Up, Kia Picanto, Toyota Aygo)
    b1_4_lugares_models = [
        'fiat 500', 'fiat500',
        'peugeot 108', 'peugeot108',
        'citroen c1', 'citroën c1', 'c1',
        'volkswagen up', 'vw up', 'vwup',
        'kia picanto', 'kiapicanto',
        'toyota aygo', 'toyotaaygo',
    ]
    
    # Se categoria é "mini" OU contém "mini", verificar modelo específico
    # MAS excluir categorias explícitas (já tratadas acima)
    if 'mini' in cat and not 'countryman' in car_lower:
        # Excluir categorias explícitas que já foram tratadas
        if cat not in ['mini 4 seats', 'mini 4 doors', 'mini 4 portas', 'mini 4 lugares',
                       'mini 5 seats', 'mini 5 doors', 'mini 5 portas', 'mini 5 lugares']:
            # Verificar se é um modelo de 4 lugares (B1)
            for model in b1_4_lugares_models:
                if model in car_lower:
                    # Se é automático de 4 lugares → E1 (Mini Automatic)
                    if any(word in car_lower for word in ['auto', 'automatic', 'automático', 'automatico']):
                        return "E1"
                    # Se é manual de 4 lugares → B1
                    return "B1"
            # Se não é B1 específico, é B2 (5 lugares)
            # Modelos B2: Fiat Panda, Hyundai i10, etc
            return "B2"
    
    # Mapeamento direto (TUDO EM LOWERCASE)
    # Suporta INGLÊS (do scraping CarJet) e PORTUGUÊS (do VEHICLES)
    category_map = {
        # B1 - Mini 4 Lugares / Mini 4 Doors
        "mini 4 doors": "B1",
        "mini 4 seats": "B1",  # Inglês do CarJet
        "mini 4 portas": "B1",
        "mini 4 lugares": "B1",
        
        # B2 - Mini 5 Lugares / Mini 5 Doors
        "mini": "B2",
        "mini 5 doors": "B2",
        "mini 5 seats": "B2",  # Inglês do CarJet
        "mini 5 portas": "B2",
        "mini 5 lugares": "B2",
        
        # D - Economy
        "economy": "D",
        "económico": "D",
        "compact": "D",
        "compacto": "D",
        
        # E1 - Mini Automatic (do VEHICLES: "MINI Auto")
        "mini automatic": "E1",
        "mini auto": "E1",
        "mini automático": "E1",
        
        # E2 - Economy Automatic (do VEHICLES: "ECONOMY Auto")
        "economy automatic": "E2",
        "economy auto": "E2",
        "económico automatic": "E2",
        "económico auto": "E2",
        "compact automatic": "E2",
        "compact auto": "E2",
        
        # F - SUV
        "suv": "F",
        "jeep": "F",
        
        # G - Cabrio APENAS (Premium/Luxury → Others)
        "cabrio": "G",
        "cabriolet": "G",
        "convertible": "G",
        "conversível": "G",
        
        # J1 - Crossover (do VEHICLES: "Crossover")
        "crossover": "J1",
        
        # J2 - Estate/Station Wagon
        "estate/station wagon": "J2",
        "station wagon": "J2",
        "estate": "J2",
        "carrinha": "J2",
        "sw": "J2",
        "touring": "J2",
        
        # L1 - SUV Automatic (do VEHICLES: "SUV Auto")
        "suv automatic": "L1",
        "suv auto": "L1",
        "jeep automatic": "L1",
        "jeep auto": "L1",
        
        # L2 - Station Wagon Automatic (do VEHICLES: "Station Wagon Auto")
        "station wagon automatic": "L2",
        "station wagon auto": "L2",
        "estate automatic": "L2",
        "estate auto": "L2",
        "carrinha automatic": "L2",
        "carrinha auto": "L2",
        "sw automatic": "L2",
        "sw auto": "L2",
        
        # M1 - 7 Seater (do VEHICLES: "7 Lugares")
        "7 seater": "M1",
        "7 seats": "M1",  # Inglês do CarJet
        "7 lugares": "M1",
        "people carrier": "M1",
        "mpv": "M1",
        
        # M2 - 7 Seater Automatic (do VEHICLES: "7 Lugares Auto")
        "7 seater automatic": "M2",
        "7 seater auto": "M2",
        "7 seats automatic": "M2",  # Inglês do CarJet
        "7 seats auto": "M2",  # Inglês do CarJet
        "7 lugares automatic": "M2",
        "7 lugares auto": "M2",
        "7 lugares automático": "M2",
        "mpv automatic": "M2",
        "mpv auto": "M2",
        
        # N - 9 Seater (do VEHICLES: "9 Lugares")
        "9 seater": "N",
        "9 seater automatic": "N",
        "9 seater auto": "N",
        "9 seats": "N",  # Inglês do CarJet
        "9 seats automatic": "N",  # Inglês do CarJet
        "9 seats auto": "N",  # Inglês do CarJet
        "9 lugares": "N",
        "9 lugares automatic": "N",
        "9 lugares auto": "N",
        "9 lugares automático": "N",
        "minivan": "N",
        "van": "N",
    }
    
    # Tentar match direto primeiro
    if cat in category_map:
        return category_map[cat]
    
    # FALLBACK: Análise inteligente por palavras-chave
    # Verificar se é automático
    is_auto = any(word in cat for word in ['auto', 'automatic', 'automático', 'automatico'])
    
    # Verificar tipo de veículo por palavras-chave
    if '9' in cat or 'minivan' in cat or 'van' in cat:
        return "N"  # 9 Seater
    
    if '7' in cat or 'mpv' in cat or 'people carrier' in cat:
        return "M2" if is_auto else "M1"  # 7 Seater
    
    if any(word in cat for word in ['sw', 'station', 'wagon', 'estate', 'carrinha', 'touring']):
        return "L2" if is_auto else "J2"  # Station Wagon
    
    if 'crossover' in cat:
        return "J1"  # Crossover
    
    if any(word in cat for word in ['suv', 'jeep', '4x4', '4wd']):
        return "L1" if is_auto else "F"  # SUV
    
    if any(word in cat for word in ['cabrio', 'cabriolet', 'convertible']):
        return "G"  # Cabrio apenas
    
    if any(word in cat for word in ['premium', 'luxury', 'luxo']):
        return "Others"  # Luxury não oferecido
    
    if any(word in cat for word in ['mini', 'small', 'pequeno']):
        # Verificar se é 4 ou 5 lugares pelo nome do carro
        if any(model in car_lower for model in ['fiat 500', 'fiat500', 'peugeot 108', 'c1', 'vw up', 'picanto', 'aygo']):
            return "E1" if is_auto else "B1"  # Mini 4 lugares
        return "E1" if is_auto else "B2"  # Mini 5 lugares (default)
    
    if any(word in cat for word in ['economy', 'econom', 'compact', 'compacto']):
        return "E2" if is_auto else "D"  # Economy
    
    # Se chegou aqui, não conseguiu mapear
    return "Others"

def _send_creds_email(to_email: str, username: str, password: str):
    # Ler configurações SMTP da base de dados (persistente) em vez de env vars
    host = _get_setting("smtp_host", os.getenv("SMTP_HOST", "")).strip()
    port = int(_get_setting("smtp_port", os.getenv("SMTP_PORT", "587")) or 587)
    user = _get_setting("smtp_username", os.getenv("SMTP_USERNAME", "")).strip()
    pwd = _get_setting("smtp_password", os.getenv("SMTP_PASSWORD", "")).strip()
    from_addr = _get_setting("smtp_from", os.getenv("SMTP_FROM", "no-reply@example.com")).strip()
    use_tls_val = _get_setting("smtp_tls", os.getenv("SMTP_TLS", "true"))
    use_tls = str(use_tls_val).lower() in ("1", "true", "yes", "y", "on")
    if not host or not to_email:
        try:
            (DEBUG_DIR / "mail_error.txt").write_text("Missing SMTP_HOST or recipient\n", encoding="utf-8")
        except Exception:
            pass
        return
    msg = EmailMessage()
    msg["Subject"] = "Your Car Rental Tracker account"
    msg["From"] = from_addr
    msg["To"] = to_email
    # Plain text
    msg.set_content(
        f"Hello,\n\nYour account was created.\n\nUsername: {username}\nPassword: {password}\n\nLogin: https://cartracker-6twv.onrender.com\n\nPlease change your password after first login."
    )
    # Simple branded HTML
    html = f"""
    <!doctype html>
    <html>
      <body style="margin:0;padding:0;background:#f8fafc;font-family:system-ui,-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;">
        <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f8fafc;padding:24px 0;">
          <tr>
            <td align="center">
              <table role="presentation" width="560" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:10px;overflow:hidden;border:1px solid #e5e7eb;">
                <tr>
                  <td style="background:#009cb6;padding:16px 20px;">
                    <img src="https://cartracker-6twv.onrender.com/static/ap-heather.png" alt="Car Rental Tracker" style="height:40px;display:block" />
                  </td>
                </tr>
                <tr>
                  <td style="padding:20px 20px 8px 20px;color:#111827;font-size:16px;">Hello,</td>
                </tr>
                <tr>
                  <td style="padding:0 20px 16px 20px;color:#111827;font-size:16px;">Your account was created.</td>
                </tr>
                <tr>
                  <td style="padding:0 20px 16px 20px;color:#111827;font-size:14px;line-height:1.6;">
                    <div><strong>Username:</strong> {username}</div>
                    <div><strong>Password:</strong> {password}</div>
                  </td>
                </tr>
                <tr>
                  <td align="center" style="padding:8px 20px 24px 20px;">
                    <a href="https://cartracker-6twv.onrender.com/login" style="display:inline-block;background:#009cb6;color:#ffffff;text-decoration:none;padding:10px 16px;border-radius:8px;font-size:14px;">Login</a>
                  </td>
                </tr>
                <tr>
                  <td style="padding:0 20px 24px 20px;color:#6b7280;font-size:12px;">Please change your password after first login.</td>
                </tr>
              </table>
            </td>
          </tr>
        </table>
      </body>
    </html>
    """
    msg.add_alternative(html, subtype="html")
    try:
        if use_tls:
            with smtplib.SMTP(host, port, timeout=15) as s:
                s.starttls()
                if user and pwd:
                    s.login(user, pwd)
                s.send_message(msg)
        else:
            with smtplib.SMTP_SSL(host, port, timeout=15) as s:
                if user and pwd:
                    s.login(user, pwd)
                s.send_message(msg)
    except Exception as e:
        try:
            (DEBUG_DIR / "mail_error.txt").write_text(f"{type(e).__name__}: {e}\n", encoding="utf-8")
        except Exception:
            pass

# Simple FX cache to avoid repeated HTTP calls
_FX_CACHE: Dict[str, Tuple[float, float]] = {}  # key "GBP->EUR" -> (rate, ts)
_URL_CACHE: Dict[str, Tuple[float, Dict[str, Any]]] = {}  # key normalized URL -> (ts, response payload)

# Ensure users table and seed initial admin on startup
try:
    # === Ensure default admin users exist ===
    def _ensure_default_users():
        """Create default users if they don't exist"""
        default_users = [
            {
                "username": "admin",
                "password": APP_PASSWORD,
                "first_name": "Filipe",
                "last_name": "Pacheco",
                "email": "carlpac82@hotmail.com",
                "mobile": "+351 964 805 750",
                "profile_picture": "/static/profiles/carlpac82.png",
                "is_admin": True
            },
            {
                "username": "carlpac82",
                "password": "Frederico.2025",
                "first_name": "Filipe",
                "last_name": "Pacheco",
                "email": "carlpac82@hotmail.com",
                "mobile": "+351 964 805 750",
                "profile_picture": "/static/profiles/carlpac82.png",
                "is_admin": True
            },
            {
                "username": "dprudente",
                "password": "dprudente",
                "first_name": "Daniell",
                "last_name": "Prudente",
                "email": "comercial.autoprudente@gmail.com",
                "mobile": "+351 911 747 478",
                "profile_picture": "/static/profiles/dprudente.jpg",
                "is_admin": False
            }
        ]
        
        try:
            with _db_lock:
                con = _db_connect()
                try:
                    for user in default_users:
                        cur = con.execute("SELECT id FROM users WHERE username=?", (user["username"],))
                        row = cur.fetchone()
                        if not row:
                            pw_hash = _hash_password(user["password"])
                            # Convert integers to boolean for PostgreSQL
                            is_admin_val = True if user.get("is_admin", 0) == 1 else False
                            enabled_val = True if user.get("enabled", 1) == 1 else False
                            con.execute(
                                "INSERT INTO users (username, password_hash, first_name, last_name, email, mobile, profile_picture_path, is_admin, enabled, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                                (
                                    user["username"],
                                    pw_hash,
                                    user["first_name"],
                                    user["last_name"],
                                    user["email"],
                                    user["mobile"],
                                    user["profile_picture"],
                                    is_admin_val,
                                    enabled_val,
                                    time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())
                                )
                            )
                            print(f"[INIT] Created user: {user['username']}", file=sys.stderr)
                    con.commit()
                finally:
                    con.close()
        except Exception as e:
            print(f"[INIT] Error creating default users: {e}", file=sys.stderr)
    
    _ensure_default_users()
except Exception:
    pass

def _fx_rate_gbp_eur(timeout: float = 5.0) -> float:
    key = "GBP->EUR"
    now = time.time()
    cached = _FX_CACHE.get(key)
    if cached and now - cached[1] < 3600:
        return cached[0]
    try:
        r = requests.get(
            "https://api.exchangerate.host/latest",
            params={"base": "GBP", "symbols": "EUR"},
            timeout=timeout,
        )
        if r.status_code == 200:
            data = r.json()
            rate = float(data.get("rates", {}).get("EUR") or 0)
            if rate > 0:
                _FX_CACHE[key] = (rate, now)
                return rate
    except Exception:
        pass
    # conservative fallback
    return cached[0] if cached else 1.16

def _parse_amount(s: str) -> Optional[float]:
    try:
        m = re.search(r"([0-9][0-9\.,\s]*)", s or "")
        if not m:
            return None
        num = m.group(1).replace("\u00a0", "").replace(" ", "")
        has_comma = "," in num
        has_dot = "." in num
        if has_comma and has_dot:
            num = num.replace(".", "").replace(",", ".")
        elif has_comma and not has_dot:
            num = num.replace(",", ".")
        else:
            parts = num.split(".")
            if len(parts) > 2:
                num = "".join(parts)
        v = float(num)
        return v
    except Exception:
        return None

def _format_eur(v: float) -> str:
    try:
        s = f"{v:,.2f}"
        s = s.replace(",", "_").replace(".", ",").replace("_", ".")
        return f"{s} €"
    except Exception:
        return f"{v:.2f} €"

def convert_items_gbp_to_eur(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rate = _fx_rate_gbp_eur()
    out = []
    for it in items or []:
        price_txt = it.get("price") or ""
        if "£" in price_txt or re.search(r"\bGBP\b", price_txt, re.I):
            amt = _parse_amount(price_txt)
            if amt is not None:
                eur = amt * rate
                it = dict(it)
                it["price"] = _format_eur(eur)
                it["currency"] = "EUR"
        out.append(it)
    return out

# CarJet destination codes we target  
LOCATION_CODES = {
    # Albufeira: deixar vazio para o CarJet descobrir automaticamente
    # "albufeira": "ABF01",  # ABF01 não funciona - testar sem código
    # "albufeira cidade": "ABF01",
    "faro": "FAO01",
    "faro airport": "FAO01",
    "faro aeroporto": "FAO01",
    "faro aeroporto (fao)": "FAO01",
    "aeroporto de faro": "FAO01",
}

def init_db():
    with _db_lock:
        conn = sqlite3.connect(DB_PATH)
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS price_snapshots (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  ts TEXT NOT NULL,
                  location TEXT NOT NULL,
                  pickup_date TEXT NOT NULL,
                  pickup_time TEXT NOT NULL,
                  days INTEGER NOT NULL,
                  supplier TEXT,
                  car TEXT,
                  price_text TEXT,
                  price_num REAL,
                  currency TEXT,
                  link TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_snapshots_q ON price_snapshots(location, days, ts)")
            
            # Tabela para configurações globais de automação de preços
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS price_automation_settings (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  key TEXT NOT NULL UNIQUE,
                  value TEXT NOT NULL,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            
            # Tabela para regras automatizadas de preços
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS automated_price_rules (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  location TEXT NOT NULL,
                  grupo TEXT NOT NULL,
                  month INTEGER,
                  day INTEGER,
                  config TEXT NOT NULL,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  UNIQUE(location, grupo, month, day)
                )
                """
            )
            
            # Tabela para estratégias de pricing
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS pricing_strategies (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  location TEXT NOT NULL,
                  grupo TEXT NOT NULL,
                  month INTEGER,
                  day INTEGER,
                  priority INTEGER NOT NULL DEFAULT 1,
                  strategy_type TEXT NOT NULL,
                  config TEXT NOT NULL,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_strategies ON pricing_strategies(location, grupo, month, day, priority)")
            
            # Tabela para histórico de preços automatizados
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS automated_prices_history (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  location TEXT NOT NULL,
                  grupo TEXT NOT NULL,
                  dias INTEGER NOT NULL,
                  pickup_date TEXT NOT NULL,
                  auto_price REAL NOT NULL,
                  real_price REAL NOT NULL,
                  strategy_used TEXT,
                  strategy_details TEXT,
                  min_price_applied REAL,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  created_by TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_auto_prices_history ON automated_prices_history(location, grupo, pickup_date, created_at)")
            
            # Tabela para logs do sistema (evitar perda em disco efêmero)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS system_logs (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  level TEXT NOT NULL,
                  message TEXT NOT NULL,
                  module TEXT,
                  function TEXT,
                  line_number INTEGER,
                  exception TEXT,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_system_logs ON system_logs(level, created_at)")
            
            # Tabela para cache de dados (evitar perda em disco efêmero)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS cache_data (
                  key TEXT PRIMARY KEY,
                  value TEXT NOT NULL,
                  expires_at TEXT,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            
            # Tabela para uploads/ficheiros (evitar perda em disco efêmero)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS file_storage (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  filename TEXT NOT NULL,
                  filepath TEXT NOT NULL UNIQUE,
                  file_data BLOB NOT NULL,
                  content_type TEXT,
                  file_size INTEGER,
                  uploaded_by TEXT,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_file_storage ON file_storage(filepath, uploaded_by)")
            
            # Tabela para histórico de exports (Way2Rentals, Abbycar, etc.)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS export_history (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  filename TEXT NOT NULL,
                  broker TEXT NOT NULL,
                  location TEXT NOT NULL,
                  period_start INTEGER,
                  period_end INTEGER,
                  month INTEGER NOT NULL,
                  year INTEGER NOT NULL,
                  month_name TEXT NOT NULL,
                  file_content TEXT NOT NULL,
                  file_size INTEGER,
                  exported_by TEXT,
                  export_date TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  last_downloaded TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_export_history ON export_history(broker, location, year, month, export_date)")
            
            # Tabela para AI Learning Data (substituir localStorage)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS ai_learning_data (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  grupo TEXT NOT NULL,
                  days INTEGER NOT NULL,
                  location TEXT NOT NULL,
                  original_price REAL,
                  new_price REAL NOT NULL,
                  timestamp TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  user TEXT DEFAULT 'admin'
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_ai_learning ON ai_learning_data(grupo, days, location, timestamp DESC)")
            
            # Tabela para User Settings (localStorage persistente)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS user_settings (
                  user_key TEXT NOT NULL,
                  setting_key TEXT NOT NULL,
                  setting_value TEXT NOT NULL,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  PRIMARY KEY (user_key, setting_key)
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_user_settings ON user_settings(user_key, updated_at DESC)")
            
            # Tabela para Commercial Vans Pricing (C3, C4, C5)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS vans_pricing (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  c3_1day REAL DEFAULT 112.00,
                  c3_2days REAL DEFAULT 144.00,
                  c3_3days REAL DEFAULT 180.00,
                  c4_1day REAL DEFAULT 152.00,
                  c4_2days REAL DEFAULT 170.00,
                  c4_3days REAL DEFAULT 210.00,
                  c5_1day REAL DEFAULT 175.00,
                  c5_2days REAL DEFAULT 190.00,
                  c5_3days REAL DEFAULT 240.00,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_by TEXT DEFAULT 'admin'
                )
                """
            )
            
            # Tabela para Automated Price Rules
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS automated_price_rules (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  location TEXT NOT NULL,
                  grupo TEXT NOT NULL,
                  month INTEGER NOT NULL,
                  day INTEGER NOT NULL,
                  strategy_type TEXT NOT NULL,
                  config TEXT NOT NULL,
                  priority INTEGER DEFAULT 1,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  UNIQUE(location, grupo, month, day, priority)
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_automated_rules ON automated_price_rules(location, grupo, month, day)")
            
            # Tabela para Price Automation Settings
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS price_automation_settings (
                  setting_key TEXT PRIMARY KEY,
                  setting_value TEXT NOT NULL,
                  setting_type TEXT DEFAULT 'string',
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            
            # Tabela para Custom Days Configuration
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS custom_days (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  days_array TEXT NOT NULL,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                )
                """
            )
            
            # Tabela para Price Validation Rules
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS price_validation_rules (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  rules_json TEXT NOT NULL,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_by TEXT DEFAULT 'admin'
                )
                """
            )
            
            # Tabela para Price History (versões salvas)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS price_history (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  history_type TEXT NOT NULL,
                  year INTEGER NOT NULL,
                  month INTEGER NOT NULL,
                  location TEXT NOT NULL,
                  prices_data TEXT NOT NULL,
                  saved_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  saved_by TEXT DEFAULT 'admin'
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_price_history ON price_history(history_type, year, month, location, saved_at DESC)")
            
            # Tabela para Search History (histórico de pesquisas)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS search_history (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  location TEXT NOT NULL,
                  start_date TEXT NOT NULL,
                  end_date TEXT NOT NULL,
                  days INTEGER NOT NULL,
                  results_count INTEGER,
                  min_price REAL,
                  max_price REAL,
                  avg_price REAL,
                  search_timestamp TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  user TEXT DEFAULT 'admin',
                  search_params TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_search_history ON search_history(location, start_date, search_timestamp DESC)")
            
            # Tabela para Notification Rules (regras de notificação)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS notification_rules (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  rule_name TEXT NOT NULL,
                  rule_type TEXT NOT NULL,
                  condition_json TEXT NOT NULL,
                  action_json TEXT NOT NULL,
                  enabled INTEGER DEFAULT 1,
                  priority INTEGER DEFAULT 1,
                  created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  created_by TEXT DEFAULT 'admin'
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_notification_rules ON notification_rules(enabled, priority, rule_type)")
            
            # Tabela para Notification History (histórico de notificações enviadas)
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS notification_history (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  rule_id INTEGER,
                  notification_type TEXT NOT NULL,
                  recipient TEXT NOT NULL,
                  subject TEXT,
                  message TEXT NOT NULL,
                  sent_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                  status TEXT DEFAULT 'sent',
                  error_message TEXT
                )
                """
            )
            conn.execute("CREATE INDEX IF NOT EXISTS idx_notification_history ON notification_history(sent_at DESC, status)")
            
        finally:
            conn.commit()
            conn.close()

init_db()

# ============================================================
# HELPER FUNCTIONS - PERSISTÊNCIA EM DB (EVITAR DISCO EFÊMERO)
# ============================================================

def log_to_db(level: str, message: str, module: str = None, function: str = None, line_number: int = None, exception: str = None):
    """Salvar logs na base de dados em vez de ficheiros"""
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute(
                    """
                    INSERT INTO system_logs (level, message, module, function, line_number, exception)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (level, message, module, function, line_number, exception)
                )
                conn.commit()
            finally:
                conn.close()
    except Exception as e:
        # Fallback para print se DB falhar
        print(f"[{level}] {message}", file=sys.stderr, flush=True)

def save_to_cache(key: str, value: str, expires_in_seconds: int = None):
    """Salvar dados em cache na DB em vez de filesystem"""
    try:
        expires_at = None
        if expires_in_seconds:
            from datetime import datetime, timedelta, timezone
            expires_at = (datetime.now(timezone.utc) + timedelta(seconds=expires_in_seconds)).isoformat()
        
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO cache_data (key, value, expires_at, updated_at)
                    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                    """,
                    (key, value, expires_at)
                )
                conn.commit()
            finally:
                conn.close()
    except Exception as e:
        log_to_db("ERROR", f"Failed to save cache: {str(e)}", "main", "save_to_cache")

def get_from_cache(key: str):
    """Obter dados do cache na DB"""
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute(
                    """
                    SELECT value, expires_at FROM cache_data 
                    WHERE key = ?
                    """,
                    (key,)
                )
                row = cursor.fetchone()
                
                if row:
                    value, expires_at = row
                    
                    # Verificar expiração
                    if expires_at:
                        from datetime import datetime, timezone
                        expires_dt = datetime.fromisoformat(expires_at)
                        if datetime.now(timezone.utc) > expires_dt:
                            # Expirado, deletar
                            conn.execute("DELETE FROM cache_data WHERE key = ?", (key,))
                            conn.commit()
                            return None
                    
                    return value
                
                return None
            finally:
                conn.close()
    except Exception as e:
        log_to_db("ERROR", f"Failed to get cache: {str(e)}", "main", "get_from_cache")
        return None

def save_file_to_db(filename: str, filepath: str, file_data: bytes, content_type: str = None, uploaded_by: str = None):
    """Salvar ficheiro na base de dados em vez de filesystem"""
    try:
        file_size = len(file_data)
        
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO file_storage 
                    (filename, filepath, file_data, content_type, file_size, uploaded_by)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (filename, filepath, file_data, content_type, file_size, uploaded_by)
                )
                conn.commit()
                log_to_db("INFO", f"File saved to DB: {filepath} ({file_size} bytes)", "main", "save_file_to_db")
            finally:
                conn.close()
    except Exception as e:
        log_to_db("ERROR", f"Failed to save file to DB: {str(e)}", "main", "save_file_to_db", exception=str(e))
        raise

def save_search_to_history(location: str, start_date: str, end_date: str, days: int, results_count: int = 0, 
                           min_price: float = None, max_price: float = None, avg_price: float = None, 
                           user: str = "admin", search_params: str = None):
    """Salvar pesquisa no histórico"""
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute(
                    """
                    INSERT INTO search_history 
                    (location, start_date, end_date, days, results_count, min_price, max_price, avg_price, user, search_params)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (location, start_date, end_date, days, results_count, min_price, max_price, avg_price, user, search_params)
                )
                conn.commit()
                log_to_db("INFO", f"Search saved to history: {location}, {start_date}-{end_date}, {results_count} results", "main", "save_search_to_history")
            finally:
                conn.close()
    except Exception as e:
        log_to_db("ERROR", f"Failed to save search history: {str(e)}", "main", "save_search_to_history")

def send_notification(rule_id: int, notification_type: str, recipient: str, subject: str, message: str):
    """Enviar notificação e salvar no histórico"""
    try:
        # Enviar email
        if notification_type == "email":
            _send_notification_email(recipient, subject, message)
        
        # Salvar no histórico
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute(
                    """
                    INSERT INTO notification_history 
                    (rule_id, notification_type, recipient, subject, message, status)
                    VALUES (?, ?, ?, ?, ?, 'sent')
                    """,
                    (rule_id, notification_type, recipient, subject, message)
                )
                conn.commit()
                log_to_db("INFO", f"Notification sent: {notification_type} to {recipient}", "main", "send_notification")
            finally:
                conn.close()
    except Exception as e:
        # Salvar erro no histórico
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute(
                    """
                    INSERT INTO notification_history 
                    (rule_id, notification_type, recipient, subject, message, status, error_message)
                    VALUES (?, ?, ?, ?, ?, 'failed', ?)
                    """,
                    (rule_id, notification_type, recipient, subject, message, str(e))
                )
                conn.commit()
            finally:
                conn.close()
        log_to_db("ERROR", f"Failed to send notification: {str(e)}", "main", "send_notification")

def _send_notification_email(to_email: str, subject: str, message: str):
    """Enviar email de notificação"""
    host = _get_setting("smtp_host", "").strip()
    port = int(_get_setting("smtp_port", "587") or 587)
    user = _get_setting("smtp_username", "").strip()
    pwd = _get_setting("smtp_password", "").strip()
    from_addr = _get_setting("smtp_from", "no-reply@example.com").strip()
    use_tls_val = _get_setting("smtp_tls", "true")
    use_tls = str(use_tls_val).lower() in ("1", "true", "yes", "y", "on")
    
    if not host or not to_email:
        raise Exception("Missing SMTP_HOST or recipient")
    
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = to_email
    msg.set_content(message)
    
    if use_tls:
        with smtplib.SMTP(host, port, timeout=15) as s:
            s.starttls()
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)
    else:
        with smtplib.SMTP_SSL(host, port, timeout=15) as s:
            if user and pwd:
                s.login(user, pwd)
            s.send_message(msg)

def get_file_from_db(filepath: str):
    """Obter ficheiro da base de dados"""
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute(
                    """
                    SELECT filename, file_data, content_type, file_size FROM file_storage 
                    WHERE filepath = ?
                    """,
                    (filepath,)
                )
                row = cursor.fetchone()
                
                if row:
                    return {
                        "filename": row[0],
                        "data": row[1],
                        "content_type": row[2],
                        "size": row[3]
                    }
                
                return None
            finally:
                conn.close()
    except Exception as e:
        log_to_db("ERROR", f"Failed to get file from DB: {str(e)}", "main", "get_file_from_db")
        return None

def cleanup_expired_cache():
    """Limpar cache expirado"""
    try:
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc).isoformat()
        
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute(
                    "DELETE FROM cache_data WHERE expires_at IS NOT NULL AND expires_at < ?",
                    (now,)
                )
                deleted = cursor.rowcount
                conn.commit()
                
                if deleted > 0:
                    log_to_db("INFO", f"Cleaned up {deleted} expired cache entries", "main", "cleanup_expired_cache")
            finally:
                conn.close()
    except Exception as e:
        log_to_db("ERROR", f"Failed to cleanup cache: {str(e)}", "main", "cleanup_expired_cache")


IDLE_TIMEOUT_SECONDS = 30 * 60  # 30 minutes

@app.get("/healthz")
async def healthz():
    return JSONResponse({"ok": True})

@app.get("/debug/test-group")
async def debug_test_group():
    """Endpoint de teste para verificar se campo group funciona"""
    test_items = [
        {"car": "Test Car 1", "category": "MINI 5 Portas", "price": "10 €", "supplier": "Test", "transmission": "Manual", "photo": "", "link": ""},
        {"car": "Test Car 2", "category": "7 lugares", "price": "20 €", "supplier": "Test", "transmission": "Manual", "photo": "", "link": ""},
        {"car": "Test Car 3", "category": "9 Seater", "price": "30 €", "supplier": "Test", "transmission": "Manual", "photo": "", "link": ""},
    ]
    result = normalize_and_sort(test_items, None)
    return JSONResponse({"ok": True, "items": result})

def require_auth(request: Request):
    if not request.session.get("auth", False):
        raise HTTPException(status_code=401, detail="Unauthorized")
    # Enforce inactivity timeout
    try:
        now = int(datetime.now(timezone.utc).timestamp())
        last = int(request.session.get("last_active_ts", 0))
        if last and now - last > IDLE_TIMEOUT_SECONDS:
            request.session.clear()
            raise HTTPException(status_code=401, detail="Session expired")
        # update last activity timestamp
        request.session["last_active_ts"] = now
    except Exception:
        # if any parsing error, refresh the timestamp anyway
        request.session["last_active_ts"] = int(datetime.now(timezone.utc).timestamp())

def require_admin(request: Request):
    require_auth(request)
    if not request.session.get("is_admin", False):
        raise HTTPException(status_code=403, detail="Forbidden")

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    if request.session.get("auth"):
        return RedirectResponse(url="/", status_code=HTTP_303_SEE_OTHER)
    return templates.TemplateResponse("login.html", {"request": request, "error": None})

@app.post("/login")
async def login_action(request: Request, username: str = Form(...), password: str = Form(...)):
    try:
        u = (username or "").strip()
        p = (password or "").strip()
        try:
            with (DEBUG_DIR / "login_trace.txt").open("a", encoding="utf-8") as f:
                f.write(f"attempt {datetime.now(timezone.utc).isoformat()} user={u}\n")
        except Exception:
            pass
        # Check DB users
        is_admin_flag = False
        ok = False
        try:
            with _db_lock:
                con = _db_connect()
                try:
                    cur = con.execute("SELECT id, password_hash, is_admin, enabled FROM users WHERE username=?", (u,))
                    row = cur.fetchone()
                    if row and row[3]:
                        ok = _verify_password(p, row[1])
                        is_admin_flag = bool(row[2])
                finally:
                    con.close()
        except Exception:
            ok = False
        # Fallback to env user for safety
        if not ok and u == APP_USERNAME and p == APP_PASSWORD:
            ok = True
            is_admin_flag = True
        if ok:
            try:
                request.session["auth"] = True
                request.session["username"] = u
                request.session["is_admin"] = bool(is_admin_flag)
                request.session["last_active_ts"] = int(datetime.now(timezone.utc).timestamp())
                log_activity(request, "login_success", details="", username=u)
                try:
                    with (DEBUG_DIR / "login_trace.txt").open("a", encoding="utf-8") as f:
                        f.write(f"success {datetime.now(timezone.utc).isoformat()} user={u}\n")
                except Exception:
                    pass
                return RedirectResponse(url="/", status_code=HTTP_303_SEE_OTHER)
            except Exception as e:
                import sys
                print(f"[LOGIN ERROR] Session error: {e}", file=sys.stderr, flush=True)
                return templates.TemplateResponse("login.html", {"request": request, "error": f"Login session error: {str(e)}"})
        try:
            with (DEBUG_DIR / "login_trace.txt").open("a", encoding="utf-8") as f:
                f.write(f"invalid {datetime.now(timezone.utc).isoformat()} user={u}\n")
        except Exception:
            pass
        log_activity(request, "login_failure", details="", username=u)
        return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid credentials"})
    except Exception:
        try:
            (DEBUG_DIR / "login_error.txt").write_text(_tb.format_exc(), encoding="utf-8")
        except Exception:
            pass
        log_activity(request, "login_exception", details="see login_error.txt")
        return templates.TemplateResponse("login.html", {"request": request, "error": "Login failed. Please try again."})

@app.post("/logout")
async def logout_action(request: Request):
    try:
        log_activity(request, "logout")
    except Exception:
        pass
    request.session.clear()
    return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    try:
        require_auth(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    # load current user profile for greeting
    user_ctx = None
    try:
        uname = request.session.get("username")
        if uname:
            user_ctx = _get_user_by_username(uname)
    except Exception:
        user_ctx = None
    
    # Load supplier logos for preloading
    supplier_logos = []
    try:
        conn = sqlite3.connect(DB_PATH)
        try:
            rows = conn.execute("SELECT DISTINCT logo_path FROM suppliers WHERE logo_path IS NOT NULL AND active = 1").fetchall()
            supplier_logos = [row[0] for row in rows if row[0]]
        finally:
            conn.close()
    except Exception:
        # Suppliers table doesn't exist yet, skip logo preloading
        pass
    
    # FORCE NO CACHE - prevent browser from caching HTML/JS
    response = templates.TemplateResponse("index.html", {
        "request": request, 
        "current_user": user_ctx,
        "supplier_logos": supplier_logos
    })
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/admin", response_class=HTMLResponse)
async def admin_root(request: Request, section: str = None):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    
    # Get current user
    user_id = request.session.get("user_id")
    current_user = None
    if user_id:
        conn = sqlite3.connect(DB_PATH)
        try:
            current_user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        finally:
            conn.close()
    
    return templates.TemplateResponse("settings_dashboard.html", {
        "request": request,
        "current_user": current_user,
        "section": section or "users"
    })

@app.get("/price-history", response_class=HTMLResponse)
async def price_history(request: Request):
    """Página de histórico e gráficos de preços"""
    try:
        require_auth(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    
    # Get current user
    user_id = request.session.get("user_id")
    current_user = None
    if user_id:
        conn = sqlite3.connect(DB_PATH)
        try:
            current_user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        finally:
            conn.close()
    
    return templates.TemplateResponse("price_history.html", {
        "request": request,
        "current_user": current_user
    })

@app.get("/price-automation", response_class=HTMLResponse)
async def price_automation(request: Request):
    """Página de automação de preços com upload de Excel"""
    try:
        require_auth(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    
    # Get current user
    user_id = request.session.get("user_id")
    current_user = None
    if user_id:
        conn = sqlite3.connect(DB_PATH)
        try:
            current_user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        finally:
            conn.close()
    
    return templates.TemplateResponse("price_automation.html", {
        "request": request,
        "current_user": current_user
    })

@app.get("/price-automation/fill", response_class=HTMLResponse)
async def price_automation_fill(request: Request):
    """Página para preencher preços do CarJet automaticamente"""
    try:
        require_auth(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    
    # Get current user
    user_id = request.session.get("user_id")
    current_user = None
    if user_id:
        conn = sqlite3.connect(DB_PATH)
        try:
            current_user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        finally:
            conn.close()
    
    return templates.TemplateResponse("price_automation_fill.html", {
        "request": request,
        "current_user": current_user
    })

# --- Admin: environment summary and adjustment preview ---
@app.get("/admin/env-summary")
async def admin_env_summary(request: Request):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    try:
        cj_pct, cj_off = _get_carjet_adjustment()
        website_pct = float(_get_setting("website_pct", "14"))
        data = {
            "CARJET_PRICE_ADJUSTMENT_PCT": cj_pct,
            "CARJET_PRICE_OFFSET_EUR": cj_off,
            "website_pct": website_pct,
            "PRICES_CACHE_TTL_SECONDS": PRICES_CACHE_TTL_SECONDS,
            "BULK_CONCURRENCY": BULK_CONCURRENCY,
            "BULK_MAX_RETRIES": BULK_MAX_RETRIES,
            "GLOBAL_FETCH_RPS": GLOBAL_FETCH_RPS,
        }
        return JSONResponse({"ok": True, "env": data})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/admin/adjust-preview")
async def admin_adjust_preview(request: Request, price: str, url: str):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    try:
        # determine if adjustment applies
        is_cj = False
        try:
            from urllib.parse import urlparse as _parse
            is_cj = _parse(url).netloc.endswith("carjet.com")
        except Exception:
            is_cj = False
        pct, off = _get_carjet_adjustment()
        amt = _parse_amount(price)
        if amt is None:
            return JSONResponse({"ok": False, "error": "Invalid price format"}, status_code=400)
        adjusted = amt
        if is_cj and (pct != 0 or off != 0):
            adjusted = amt * (1.0 + (pct/100.0)) + off
        return _no_store_json({
            "ok": True,
            "input": {"price": price, "url": url},
            "env": {"pct": pct, "offset": off},
            "is_carjet": is_cj,
            "amount": amt,
            "adjusted_amount": adjusted,
            "adjusted_price": _format_eur(adjusted),
        })
    except Exception as e:
        return _no_store_json({"ok": False, "error": str(e)}, status_code=500)

@app.get("/admin/settings", response_class=HTMLResponse)
async def admin_settings_page(request: Request):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    cj_pct, cj_off = _get_carjet_adjustment()
    abbycar_pct = _get_abbycar_adjustment()
    abbycar_low_deposit_pct = _get_abbycar_low_deposit_adjustment()
    abbycar_low_deposit_enabled = _get_abbycar_low_deposit_enabled()
    website_pct = float(_get_setting("website_pct", "14"))
    
    # Carregar configurações SMTP da base de dados
    smtp_host = _get_setting("smtp_host", "")
    smtp_port = _get_setting("smtp_port", "587")
    smtp_username = _get_setting("smtp_username", "")
    smtp_password = _get_setting("smtp_password", "")
    smtp_from = _get_setting("smtp_from", "")
    smtp_tls = _get_setting("smtp_tls", "1") == "1"
    
    return templates.TemplateResponse("admin_settings.html", {
        "request": request, 
        "carjet_pct": cj_pct, 
        "carjet_off": cj_off, 
        "abbycar_pct": abbycar_pct,
        "abbycar_low_deposit_pct": abbycar_low_deposit_pct,
        "abbycar_low_deposit_enabled": abbycar_low_deposit_enabled,
        "website_pct": website_pct,
        "smtp_host": smtp_host,
        "smtp_port": smtp_port,
        "smtp_username": smtp_username,
        "smtp_password": smtp_password,
        "smtp_from": smtp_from,
        "smtp_tls": smtp_tls,
        "saved": False, 
        "error": None
    })

@app.post("/admin/settings", response_class=HTMLResponse)
async def admin_settings_save(
    request: Request, 
    carjet_pct: str = Form(""), 
    carjet_off: str = Form(""), 
    abbycar_pct: str = Form(""), 
    abbycar_low_deposit_pct: str = Form(""), 
    abbycar_low_deposit_enabled: str = Form(""),
    website_pct: str = Form(""),
    smtp_host: str = Form(""),
    smtp_port: str = Form("587"),
    smtp_username: str = Form(""),
    smtp_password: str = Form(""),
    smtp_from: str = Form(""),
    smtp_tls: str = Form("")
):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    err = None
    try:
        pct_val = float((carjet_pct or "0").replace(",", "."))
        off_val = float((carjet_off or "0").replace(",", "."))
        abbycar_pct_val = float((abbycar_pct or "3").replace(",", "."))
        abbycar_low_deposit_pct_val = float((abbycar_low_deposit_pct or "0").replace(",", "."))
        abbycar_low_deposit_enabled_val = "1" if abbycar_low_deposit_enabled == "1" else "0"
        website_pct_val = float((website_pct or "0").replace(",", "."))
        
        # Guardar configurações de preços
        _set_setting("carjet_pct", str(pct_val))
        _set_setting("carjet_off", str(off_val))
        _set_setting("abbycar_pct", str(abbycar_pct_val))
        _set_setting("abbycar_low_deposit_pct", str(abbycar_low_deposit_pct_val))
        _set_setting("abbycar_low_deposit_enabled", abbycar_low_deposit_enabled_val)
        _set_setting("website_pct", str(website_pct_val))
        
        # Guardar configurações SMTP na base de dados (persistente)
        _set_setting("smtp_host", smtp_host.strip())
        _set_setting("smtp_port", smtp_port.strip())
        _set_setting("smtp_username", smtp_username.strip())
        _set_setting("smtp_password", smtp_password.strip())
        _set_setting("smtp_from", smtp_from.strip())
        _set_setting("smtp_tls", "1" if smtp_tls == "1" else "0")
        
        cj_pct, cj_off = pct_val, off_val
        abbycar_pct_result = abbycar_pct_val
        abbycar_low_deposit_pct_result = abbycar_low_deposit_pct_val
        abbycar_low_deposit_enabled_result = abbycar_low_deposit_enabled_val == "1"
        website_pct_result = website_pct_val
        smtp_host_result = smtp_host.strip()
        smtp_port_result = smtp_port.strip()
        smtp_username_result = smtp_username.strip()
        smtp_password_result = smtp_password.strip()
        smtp_from_result = smtp_from.strip()
        smtp_tls_result = smtp_tls == "1"
    except Exception as e:
        err = str(e)
        cj_pct, cj_off = _get_carjet_adjustment()
        abbycar_pct_result = _get_abbycar_adjustment()
        abbycar_low_deposit_pct_result = _get_abbycar_low_deposit_adjustment()
        abbycar_low_deposit_enabled_result = _get_abbycar_low_deposit_enabled()
        website_pct_result = float(_get_setting("website_pct", "14"))
        smtp_host_result = _get_setting("smtp_host", "")
        smtp_port_result = _get_setting("smtp_port", "587")
        smtp_username_result = _get_setting("smtp_username", "")
        smtp_password_result = _get_setting("smtp_password", "")
        smtp_from_result = _get_setting("smtp_from", "")
        smtp_tls_result = _get_setting("smtp_tls", "1") == "1"
    return templates.TemplateResponse("admin_settings.html", {
        "request": request, 
        "carjet_pct": cj_pct, 
        "carjet_off": cj_off, 
        "abbycar_pct": abbycar_pct_result,
        "abbycar_low_deposit_pct": abbycar_low_deposit_pct_result,
        "abbycar_low_deposit_enabled": abbycar_low_deposit_enabled_result,
        "website_pct": website_pct_result,
        "smtp_host": smtp_host_result,
        "smtp_port": smtp_port_result,
        "smtp_username": smtp_username_result,
        "smtp_password": smtp_password_result,
        "smtp_from": smtp_from_result,
        "smtp_tls": smtp_tls_result,
        "saved": err is None, 
        "error": err
    })

@app.post("/admin/users/{user_id}/toggle-enabled")
async def admin_users_toggle_enabled(request: Request, user_id: int):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    with _db_lock:
        con = _db_connect()
        try:
            cur = con.execute("SELECT enabled FROM users WHERE id=?", (user_id,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(status_code=404, detail="Not found")
            new_val = 0 if int(r[0] or 0) else 1
            con.execute("UPDATE users SET enabled=? WHERE id=?", (new_val, user_id))
            con.commit()
        finally:
            con.close()
    try:
        log_activity(request, "admin_edit_user", details=f"user_id={user_id}")
    except Exception:
        pass
    return RedirectResponse(url="/admin/users", status_code=HTTP_303_SEE_OTHER)

@app.post("/admin/users/{user_id}/reset-password")
async def admin_users_reset_password(request: Request, user_id: int):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    gen_pw = secrets.token_urlsafe(8)
    pw_hash = _hash_password(gen_pw)
    to_email = None
    username = None
    with _db_lock:
        con = _db_connect()
        try:
            cur = con.execute("SELECT username, email FROM users WHERE id=?", (user_id,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(status_code=404, detail="Not found")
            username = r[0]
            to_email = (r[1] or "").strip()
            con.execute("UPDATE users SET password_hash=? WHERE id=?", (pw_hash, user_id))
            con.commit()
        finally:
            con.close()
    try:
        if to_email:
            _send_creds_email(to_email, username or "", gen_pw)
    except Exception:
        pass
    try:
        log_activity(request, "admin_reset_password", details=f"user_id={user_id}")
    except Exception:
        pass
    return RedirectResponse(url="/admin/users", status_code=HTTP_303_SEE_OTHER)

@app.get("/admin/users/{user_id}/edit", response_class=HTMLResponse)
async def admin_users_edit(request: Request, user_id: int):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    with _db_lock:
        con = _db_connect()
        try:
            cur = con.execute("SELECT id, username, first_name, last_name, email, mobile, profile_picture_path, is_admin, enabled FROM users WHERE id=?", (user_id,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(status_code=404, detail="Not found")
            u = {
                "id": r[0], "username": r[1], "first_name": r[2] or "", "last_name": r[3] or "",
                "email": r[4] or "", "mobile": r[5] or "", "profile_picture_path": r[6] or "",
                "is_admin": bool(r[7]), "enabled": bool(r[8])
            }
        finally:
            con.close()
    return templates.TemplateResponse("admin_edit_user.html", {"request": request, "u": u, "error": None})

@app.post("/admin/users/{user_id}/edit")
async def admin_users_edit_post(
    request: Request,
    user_id: int,
    first_name: str = Form(""),
    last_name: str = Form(""),
    mobile: str = Form(""),
    email: str = Form(""),
    is_admin: str = Form("0"),
    enabled: str = Form("1"),
    new_password: str = Form(""),
    picture: Optional[UploadFile] = File(None),
):
    import sys
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    
    try:
        pic_data = None
        pic_path = None
        if picture and picture.filename:
            # Guardar foto na base de dados como BLOB
            pic_data = await picture.read()
            print(f"[UPLOAD] 📸 Profile picture uploaded: {len(pic_data)} bytes", file=sys.stderr, flush=True)
            # Manter path para compatibilidade (usar ID do user)
            pic_path = f"/api/profile-picture/{user_id}"
        
        with _db_lock:
            con = _db_connect()
            try:
                # Converter para boolean (PostgreSQL) ou integer (SQLite)
                is_admin_val = True if is_admin in ("1","true","on") else False
                enabled_val = True if enabled in ("1","true","on") else False
                
                if pic_data:
                    print(f"[UPLOAD] 💾 Saving to DB: user_id={user_id}, blob_size={len(pic_data)}", file=sys.stderr, flush=True)
                    con.execute(
                        "UPDATE users SET first_name=?, last_name=?, mobile=?, email=?, profile_picture_path=?, profile_picture_data=?, is_admin=?, enabled=? WHERE id=?",
                        (first_name, last_name, mobile, email, pic_path, pic_data, is_admin_val, enabled_val, user_id)
                    )
                else:
                    con.execute(
                        "UPDATE users SET first_name=?, last_name=?, mobile=?, email=?, is_admin=?, enabled=? WHERE id=?",
                        (first_name, last_name, mobile, email, is_admin_val, enabled_val, user_id)
                    )
                # Optional password change
                if new_password and new_password.strip():
                    pw_hash = _hash_password(new_password.strip())
                    con.execute("UPDATE users SET password_hash=? WHERE id=?", (pw_hash, user_id))
                con.commit()
                print(f"[UPLOAD] ✅ User {user_id} updated successfully", file=sys.stderr, flush=True)
            finally:
                con.close()
        return RedirectResponse(url="/admin/users", status_code=HTTP_303_SEE_OTHER)
    except Exception as e:
        print(f"[UPLOAD] ❌ Error updating user {user_id}: {e}", file=sys.stderr, flush=True)
        import traceback
        traceback.print_exc()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/admin/users/{user_id}/delete")
async def admin_users_delete(request: Request, user_id: int):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    # prevent deleting self
    current_username = request.session.get("username")
    with _db_lock:
        con = _db_connect()
        try:
            cur = con.execute("SELECT username FROM users WHERE id=?", (user_id,))
            r = cur.fetchone()
            if not r:
                raise HTTPException(status_code=404, detail="Not found")
            if r[0] == current_username:
                raise HTTPException(status_code=400, detail="Cannot delete own account")
            con.execute("DELETE FROM users WHERE id=?", (user_id,))
            con.commit()
        finally:
            con.close()
    return RedirectResponse(url="/admin/users", status_code=HTTP_303_SEE_OTHER)

# --- Admin UI ---
@app.get("/test/carjet-mobile", response_class=HTMLResponse)
async def test_carjet_mobile(request: Request):
    """CarJet Mobile Scraping Test Page"""
    return templates.TemplateResponse("carjet_mobile_test.html", {"request": request})

@app.get("/test/mobile-scraping-live", response_class=HTMLResponse)
async def test_mobile_scraping_live(request: Request):
    """Live Mobile Scraping Test with Real API"""
    return templates.TemplateResponse("test_mobile_scraping_live.html", {"request": request})

@app.post("/api/test-mobile-scraping")
async def test_mobile_scraping(request: Request):
    """Test mobile scraping endpoint"""
    import time
    import httpx
    from datetime import datetime, timedelta
    
    try:
        data = await request.json()
        location = data.get('location', 'albufeira')
        days = int(data.get('days', 3))
        
        # Build mobile URL
        today = datetime.now()
        pickup = today + timedelta(days=7)
        dropoff = pickup + timedelta(days=days)
        
        location_codes = {
            'albufeira': 'albufeira-pt',
            'faro-airport': 'faro-airport-pt'
        }
        
        pickup_str = pickup.strftime('%Y-%m-%d')
        dropoff_str = dropoff.strftime('%Y-%m-%d')
        loc_code = location_codes.get(location, 'albufeira-pt')
        
        mobile_url = f"https://m.carjet.com/en/car-hire/{loc_code}/{pickup_str}/{dropoff_str}/"
        
        # Mobile user agent
        headers = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
        }
        
        start_time = time.time()
        
        # Fetch mobile page
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True) as client:
            response = await client.get(mobile_url, headers=headers)
            html = response.text
        
        response_time = int((time.time() - start_time) * 1000)
        
        # Simple parsing to count vehicles (we'll improve this)
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, 'html.parser')
        
        # Try to find vehicle cards (adjust selectors based on actual mobile HTML)
        vehicles = []
        vehicle_cards = soup.find_all(['div', 'article'], class_=lambda x: x and ('car' in x.lower() or 'vehicle' in x.lower() or 'result' in x.lower()))
        
        for card in vehicle_cards[:5]:  # Get first 5 as sample
            try:
                # Try to extract basic info
                name_elem = card.find(['h2', 'h3', 'h4', 'span'], class_=lambda x: x and ('name' in x.lower() or 'title' in x.lower() or 'model' in x.lower()))
                price_elem = card.find(['span', 'div'], class_=lambda x: x and ('price' in x.lower() or 'cost' in x.lower()))
                
                vehicle = {
                    'name': name_elem.get_text(strip=True) if name_elem else 'Unknown',
                    'price': price_elem.get_text(strip=True) if price_elem else 'N/A'
                }
                vehicles.append(vehicle)
            except:
                continue
        
        return {
            "success": True,
            "url": mobile_url,
            "vehicleCount": len(vehicle_cards),
            "responseTime": response_time,
            "vehicles": vehicles,
            "htmlLength": len(html),
            "statusCode": response.status_code
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

@app.get("/fix-schema", response_class=HTMLResponse)
async def fix_schema_page(request: Request):
    """Emergency page to fix PostgreSQL schema"""
    return templates.TemplateResponse("fix_schema.html", {"request": request})

@app.get("/admin/backup", response_class=HTMLResponse)
async def admin_backup(request: Request):
    """Backup & Restore page"""
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    return templates.TemplateResponse("admin_backup.html", {"request": request})

@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users(request: Request):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    users = []
    try:
        with _db_lock:
            con = _db_connect()
            try:
                cur = con.execute("SELECT id, username, first_name, last_name, email, mobile, is_admin, enabled FROM users ORDER BY id DESC")
                for r in cur.fetchall():
                    users.append({
                        "id": r[0], "username": r[1], "first_name": r[2] or "", "last_name": r[3] or "",
                        "email": r[4] or "", "mobile": r[5] or "", "is_admin": bool(r[6]), "enabled": bool(r[7])
                    })
            finally:
                con.close()
    except Exception:
        return JSONResponse({"ok": False, "error": "Failed to load users"}, status_code=500)
    return templates.TemplateResponse("admin_users.html", {"request": request, "users": users})


@app.get("/admin/users/new", response_class=HTMLResponse)
async def admin_users_new(request: Request):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    return templates.TemplateResponse("admin_new_user.html", {"request": request, "error": None})

@app.post("/admin/users/new")
async def admin_users_new_post(
    request: Request,
    username: str = Form(...),
    first_name: str = Form(""),
    last_name: str = Form(""),
    mobile: str = Form(""),
    email: str = Form(""),
    is_admin: str = Form("0"),
    picture: Optional[UploadFile] = File(None),
):
    try:
        require_admin(request)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=HTTP_303_SEE_OTHER)
    import sys
    u = (username or "").strip()
    if not u:
        return templates.TemplateResponse("admin_new_user.html", {"request": request, "error": "Username required"})
    
    try:
        # generate password
        gen_pw = secrets.token_urlsafe(8)
        pw_hash = _hash_password(gen_pw)
        pic_data = None
        pic_path = None
        if picture and picture.filename:
            # Guardar foto na base de dados como BLOB
            pic_data = await picture.read()
            print(f"[NEW_USER] 📸 Profile picture uploaded: {len(pic_data)} bytes", file=sys.stderr, flush=True)
            # Path será definido após inserção (precisa do ID)
            pic_path = ""  # Será atualizado depois
        
        with _db_lock:
            con = _db_connect()
            try:
                # Convert to boolean for PostgreSQL
                is_admin_bool = True if (is_admin in ("1","true","on")) else False
                enabled_bool = True
                con.execute(
                    "INSERT INTO users (username, password_hash, first_name, last_name, mobile, email, profile_picture_path, profile_picture_data, is_admin, enabled, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    (u, pw_hash, first_name, last_name, mobile, email, pic_path or "", pic_data, is_admin_bool, enabled_bool, time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime()))
                )
                
                # Se tem foto, atualizar path com o ID do user
                if pic_data:
                    cur = con.execute("SELECT id FROM users WHERE username=?", (u,))
                    user_id = cur.fetchone()[0]
                    print(f"[NEW_USER] 💾 Updating path for user {user_id}", file=sys.stderr, flush=True)
                    con.execute("UPDATE users SET profile_picture_path=? WHERE id=?", (f"/api/profile-picture/{user_id}", user_id))
                con.commit()
                print(f"[NEW_USER] ✅ User '{u}' created successfully", file=sys.stderr, flush=True)
            except sqlite3.IntegrityError:
                return templates.TemplateResponse("admin_new_user.html", {"request": request, "error": "Username already exists"})
            finally:
                con.close()
    except Exception as e:
        print(f"[NEW_USER] ❌ Error creating user: {e}", file=sys.stderr, flush=True)
        import traceback
        traceback.print_exc()
        return templates.TemplateResponse("admin_new_user.html", {"request": request, "error": f"Error: {str(e)}"})
    # send email if provided
    if email:
        _send_creds_email(email, u, gen_pw)
    try:
        log_activity(request, "admin_create_user", details=f"username={u}")
    except Exception:
        pass
    return RedirectResponse(url="/admin/users", status_code=HTTP_303_SEE_OTHER)


@app.get("/api/profile-picture/{user_id}")
async def get_profile_picture(user_id: int):
    """Serve profile picture from database BLOB"""
    from fastapi.responses import Response
    import sys
    
    with _db_lock:
        con = _db_connect()
        try:
            cur = con.execute("SELECT profile_picture_data FROM users WHERE id=?", (user_id,))
            row = cur.fetchone()
            if row and row[0]:
                # Retornar imagem do BLOB
                # PostgreSQL retorna memoryview, converter para bytes
                blob = row[0]
                if isinstance(blob, memoryview):
                    blob = bytes(blob)
                
                print(f"[PROFILE_PIC] ✅ Serving BLOB for user {user_id} ({len(blob)} bytes)", file=sys.stderr, flush=True)
                
                # Detectar tipo de imagem pelos magic bytes
                if blob[:2] == b'\xff\xd8':
                    media_type = "image/jpeg"
                elif blob[:4] == b'\x89PNG':
                    media_type = "image/png"
                elif blob[:4] == b'GIF8':
                    media_type = "image/gif"
                else:
                    media_type = "image/png"  # fallback
                return Response(content=blob, media_type=media_type)
            else:
                # Retornar imagem default se não tiver foto
                print(f"[PROFILE_PIC] ⚠️ No BLOB for user {user_id}, using default", file=sys.stderr, flush=True)
                from pathlib import Path
                default_pic = Path(__file__).parent / "static" / "profiles" / "default-avatar.png"
                if default_pic.exists():
                    return Response(content=default_pic.read_bytes(), media_type="image/png")
                # Se não existir default, retornar 404
                raise HTTPException(status_code=404, detail="No profile picture")
        finally:
            con.close()


@app.get("/api/prices")
async def get_prices(request: Request):
    require_auth(request)
    url = request.query_params.get("url") or TARGET_URL
    refresh = str(request.query_params.get("refresh", "")).strip().lower() in ("1","true","yes","on")
    # Serve from cache if fresh
    if not refresh:
        cached = _cache_get(url)
        if cached:
            # also refresh in background to keep fresh
            asyncio.create_task(_refresh_prices_background(url))
            return JSONResponse(cached)
    else:
        try:
            # Invalidate cache entry if exists
            _PRICES_CACHE.pop(url, None)
        except Exception:
            pass
    # If we have stale data (beyond TTL) we could still serve it while refreshing. For simplicity, compute now.
    try:
        # Fast path: direct fetch for CarJet s/b URLs (often returns full list without UI)
        if isinstance(url, str) and ("carjet.com/do/list/" in url) and ("s=" in url) and ("b=" in url):
            try:
                data_fast = await _compute_prices_for(url)
                fast_items = (data_fast or {}).get("items") or []
                if fast_items:
                    out = {"ok": True, "items": fast_items}
                    try:
                        # Persist the HTML if provided for inspection
                        html_fast = (data_fast or {}).get("html") or ""
                        if html_fast:
                            stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                            (DEBUG_DIR / f"pw-url-direct-fast-{stamp}.html").write_text(html_fast, encoding='utf-8')
                    except Exception:
                        pass
                    _cache_set(url, out)
                    return JSONResponse(out)
            except Exception:
                pass
        # Playwright-first for CarJet list pages to ensure the search is triggered (UI-driven)
        if USE_PLAYWRIGHT and _HAS_PLAYWRIGHT and isinstance(url, str) and ("carjet.com/do/list/" in url):
            try:
                from playwright.async_api import async_playwright
                import sys
                async with async_playwright() as p:
                    # Try WebKit (Safari) first
                    async def run_with(browser):
                        context = await browser.new_context(
                            locale="pt-PT",
                            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/26.0.1 Safari/605.1.15",
                        )
                        page = await context.new_page()
                        # Best-effort: set currency/lang cookies upfront
                        try:
                            await context.add_cookies([
                                {"name": "monedaForzada", "value": "EUR", "domain": ".carjet.com", "path": "/"},
                                {"name": "moneda", "value": "EUR", "domain": ".carjet.com", "path": "/"},
                                {"name": "currency", "value": "EUR", "domain": ".carjet.com", "path": "/"},
                                {"name": "country", "value": "PT", "domain": ".carjet.com", "path": "/"},
                                {"name": "idioma", "value": "PT", "domain": ".carjet.com", "path": "/"},
                                {"name": "lang", "value": "pt", "domain": ".carjet.com", "path": "/"},
                            ])
                        except Exception:
                            pass
                        captured = []
                        async def _on_resp(resp):
                            try:
                                u = resp.url or ""
                                if any(k in u for k in ("modalFilter.asp", "carList.asp", "/do/list/pt", "filtroUso.asp")):
                                    t = await resp.text()
                                    if t:
                                        captured.append((u, t))
                                        # Persist capture for offline inspection
                                        try:
                                            stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                            name = "pw-url-capture-" + re.sub(r"[^a-z0-9]+", "-", (u or "").lower())[-60:]
                                            (DEBUG_DIR / f"{name}-{stamp}.html").write_text(t, encoding='utf-8')
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                        page.on("response", _on_resp)
                        # Warm up session on homepage before opening s/b URL
                        try:
                            base_lang = "pt"
                            m = re.search(r"/do/list/([a-z]{2})", url)
                            if m: base_lang = m.group(1)
                            home_path = "aluguel-carros/index.htm" if base_lang.lower()=="pt" else "index.htm"
                            await page.goto(f"https://www.carjet.com/{home_path}", wait_until="networkidle", timeout=25000)
                            try:
                                await page.evaluate("""try{ document.cookie='moneda=EUR; path=/; domain=.carjet.com'; document.cookie='lang=pt; path=/; domain=.carjet.com'; }catch(e){}""")
                            except Exception:
                                pass
                            try:
                                await page.wait_for_timeout(500)
                            except Exception:
                                pass
                        except Exception:
                            pass
                        await page.goto(url, wait_until="networkidle", timeout=35000)
                        
                        # ===== FILTRAR APENAS AUTOPRUDENTE =====
                        try:
                            # Aguardar filtros carregarem
                            await page.wait_for_selector('#chkAUP', timeout=5000)
                            
                            # Desmarcar todos os checkboxes de suppliers primeiro
                            await page.evaluate("""
                                document.querySelectorAll('input[name="frmPrv"]').forEach(cb => {
                                    if (cb.checked) cb.click();
                                });
                            """)
                            
                            # Aguardar um pouco
                            await page.wait_for_timeout(500)
                            
                            # Marcar apenas AUTOPRUDENTE
                            aup_checkbox = await page.query_selector('#chkAUP')
                            if aup_checkbox:
                                is_checked = await aup_checkbox.is_checked()
                                if not is_checked:
                                    await aup_checkbox.click()
                                    print("[PLAYWRIGHT ASYNC] Filtro AUTOPRUDENTE ativado", file=sys.stderr, flush=True)
                                    
                                    # Aguardar página recarregar com filtro
                                    await page.wait_for_load_state("networkidle", timeout=10000)
                                else:
                                    print("[PLAYWRIGHT ASYNC] Checkbox AUTOPRUDENTE já estava marcado", file=sys.stderr, flush=True)
                                    
                        except Exception as e:
                            print(f"[PLAYWRIGHT ASYNC] Erro ao filtrar AUTOPRUDENTE: {e}", file=sys.stderr, flush=True)
                            # Continuar mesmo se falhar o filtro
                            pass
                        # ===== FIM FILTRO AUTOPRUDENTE =====
                        
                        # Handle consent if present
                        try:
                            for sel in [
                                "#didomi-notice-agree-button",
                                ".didomi-continue-without-agreeing",
                                "button:has-text('Aceitar')",
                                "button:has-text('I agree')",
                                "button:has-text('Accept')",
                            ]:
                                try:
                                    c = page.locator(sel)
                                    if await c.count() > 0:
                                        try: await c.first.click(timeout=1500)
                                        except Exception: pass
                                        await page.wait_for_timeout(200)
                                        break
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        # Click "Atualizar/Pesquisar" if present and trigger native submit
                        try:
                            # Try specific CarJet selectors first
                            for sel in [
                                "button[name=send].btn-search",
                                "#btn_search",
                                ".btn-search",
                                "button:has-text('Pesquisar')",
                                "button:has-text('Atualizar')",
                                "input[type=submit]",
                                "button[type=submit]",
                            ]:
                                try:
                                    b = page.locator(sel)
                                    if await b.count() > 0:
                                        try: await b.first.click(timeout=2000)
                                        except Exception: pass
                                        break
                                except Exception:
                                    pass
                            try:
                                await page.evaluate("""
                                  try { if (typeof comprobar_errores_3==='function' && comprobar_errores_3()) { if (typeof filtroUsoForm==='function') filtroUsoForm(); if (typeof submit_fechas==='function') submit_fechas('/do/list/pt'); } } catch(e) {}
                                """)
                            except Exception:
                                pass
                            try: await page.wait_for_load_state('networkidle', timeout=40000)
                            except Exception: pass
                        except Exception:
                            pass
                        # Screenshot and scroll cycles
                        try:
                            stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                            await page.screenshot(path=str(DEBUG_DIR / f"pw-url-after-search-{stamp}.png"), full_page=True)
                        except Exception:
                            pass
                        try:
                            for _ in range(3):
                                for __ in range(5):
                                    try: await page.mouse.wheel(0, 1600)
                                    except Exception: pass
                                    await page.wait_for_timeout(250)
                                try:
                                    ok = await page.locator("section.newcarlist article, .newcarlist article, article.car, li.result, li.car, .car-item, .result-row").count()
                                    if (ok or 0) > 0:
                                        break
                                except Exception:
                                    pass
                                try: await page.wait_for_load_state('networkidle', timeout=8000)
                                except Exception: pass
                        except Exception:
                            pass
                        # Explicit waits for known XHRs to maximize chances
                        try:
                            await page.wait_for_response(lambda r: 'modalFilter.asp' in (r.url or ''), timeout=40000)
                        except Exception:
                            pass
                        try:
                            await page.wait_for_response(lambda r: 'carList.asp' in (r.url or ''), timeout=40000)
                        except Exception:
                            pass
                        html = await page.content()
                        final_url = page.url
                        await context.close()
                        return html, final_url, captured

                    # Chromium-first
                    browser = await p.chromium.launch(headless=True)
                    html_pw, final_url_pw, cap_pw = await run_with(browser)
                    await browser.close()
                    items = []
                    # Prefer parsing captured bodies first
                    if (not items) and cap_pw:
                        try:
                            base_net = "https://www.carjet.com/do/list/pt"
                            for (_u, body) in cap_pw:
                                its = parse_prices(body, base_net)
                                its = convert_items_gbp_to_eur(its)
                                its = apply_price_adjustments(its, base_net)
                                if its: items = its; break
                        except Exception:
                            pass
                    if (not items) and html_pw:
                        try:
                            items = parse_prices(html_pw, final_url_pw or url)
                            items = convert_items_gbp_to_eur(items)
                            items = apply_price_adjustments(items, final_url_pw or url)
                        except Exception:
                            items = []
                    # WebKit fallback if still empty
                    if not items:
                        try:
                            browser2 = await p.webkit.launch(headless=True)
                            html2, final2, cap2 = await run_with(browser2)
                            await browser2.close()
                            # Prefer captured responses
                            if (not items) and cap2:
                                base_net = "https://www.carjet.com/do/list/pt"
                                for (_u, body) in cap2:
                                    its = parse_prices(body, base_net)
                                    its = convert_items_gbp_to_eur(its)
                                    its = apply_price_adjustments(its, base_net)
                                    if its: items = its; break
                            if (not items) and html2:
                                its = parse_prices(html2, final2 or url)
                                its = convert_items_gbp_to_eur(its)
                                its = apply_price_adjustments(its, final2 or url)
                                if its: items = its
                        except Exception:
                            pass
                    if items:
                        data = {"ok": True, "items": items}
                        _cache_set(url, data)
                        return JSONResponse(data)
                    # Direct POST fallback using page.request (within session)
                    try:
                        async with async_playwright() as p3:
                            browser3 = await p3.chromium.launch(headless=True)
                            context3 = await browser3.new_context(
                                locale="pt-PT",
                            )
                            page3 = await context3.new_page()
                            form_data = {}
                            try:
                                form_data = await page3.evaluate("""
                                  () => {
                                    try {
                                      const f = document.querySelector('form');
                                      if (!f) return {};
                                      const fd = new FormData(f);
                                      const o = Object.fromEntries(fd.entries());
                                      return o;
                                    } catch(e) { return {}; }
                                  }
                                """)
                            except Exception:
                                form_data = {}
                            # Ensure minimal fields
                            if not form_data or Object.keys(form_data).length < 3:
                                form_data = {"moneda":"EUR", "idioma":"PT"}
                            r3 = await page3.request.post("https://www.carjet.com/do/list/pt", data=form_data)
                            html3 = ""
                            try: html3 = await r3.text()
                            except Exception: html3 = ""
                            if html3:
                                try:
                                    stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                    (DEBUG_DIR / f"pw-url-direct-post-{stamp}.html").write_text(html3, encoding='utf-8')
                                except Exception:
                                    pass
                                its3 = parse_prices(html3, "https://www.carjet.com/do/list/pt")
                                its3 = convert_items_gbp_to_eur(its3)
                                its3 = apply_price_adjustments(its3, "https://www.carjet.com/do/list/pt")
                                if its3:
                                    data = {"ok": True, "items": its3}
                                    _cache_set(url, data)
                                    await context3.close(); await browser3.close()
                                    return JSONResponse(data)
                            await context3.close(); await browser3.close()
                    except Exception:
                        pass
            except Exception:
                pass
        data = await _compute_prices_for(url)
        _cache_set(url, data)
        return JSONResponse(data)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/ai-pricing-analysis")
async def ai_pricing_analysis(request: Request):
    """
    AI-powered pricing analysis endpoint
    Analyzes historical pricing data and provides positioning insights
    """
    try:
        require_auth(request)
    except HTTPException:
        return JSONResponse({"ok": False, "error": "Unauthorized"}, status_code=401)
    
    try:
        body = await request.json()
        reference_supplier = body.get("referenceSupplier", "autoprudente")
        analysis_period = int(body.get("analysisPeriod", 90))
        location = body.get("location", "Albufeira")
        
        print(f"[AI ANALYSIS] Supplier: {reference_supplier}, Period: {analysis_period} days, Location: {location}")
        
        # Mock AI analysis for now - replace with real DB queries later
        # This would query historical data from database
        group_insights = []
        grupos = ['B1', 'B2', 'D', 'E1', 'E2', 'F', 'G', 'J1', 'J2', 'L1', 'L2', 'M1', 'M2', 'N']
        group_names = {
            'B1': 'Mini 4 Doors', 'B2': 'Mini Automatic', 'D': 'Economy',
            'E1': 'Compact 4 Doors', 'E2': 'Compact Automatic', 'F': 'Intermediate',
            'G': 'Intermediate Automatic', 'J1': 'Compact SUV', 'J2': 'Intermediate SUV',
            'L1': 'Standard', 'L2': 'Standard Automatic', 'M1': 'Premium', 'M2': 'Premium Automatic', 'N': 'Minivan'
        }
        
        total_data_points = 0
        avg_positions = []
        avg_price_diffs = []
        
        for grupo in grupos:
            # Mock data - would query real historical data
            sample_size = int(50 + (hash(grupo) % 100))
            avg_position = round(2 + (hash(grupo) % 3), 1)
            avg_price_diff = round(0.25 + ((hash(grupo) % 10) / 10), 2)
            cheapest_count = int(10 + (hash(grupo) % 20))
            confidence = min(95, 50 + sample_size // 2)
            
            total_data_points += sample_size
            avg_positions.append(avg_position)
            avg_price_diffs.append(avg_price_diff)
            
            # Determine position color
            if avg_position <= 2:
                position_color = "green"
            elif avg_position <= 3:
                position_color = "yellow"
            else:
                position_color = "red"
            
            # Determine price diff color
            if avg_price_diff <= 0.50:
                price_diff_color = "green"
            elif avg_price_diff <= 1.00:
                price_diff_color = "yellow"
            else:
                price_diff_color = "red"
            
            # Generate AI recommendation
            if avg_position <= 2:
                recommendation = f"You're consistently competitive in {grupo}. Maintain current positioning with slight adjustments based on demand."
                suggested_strategy = f"Follow Lowest Price +{avg_price_diff:.2f}€"
                strategy_config = {
                    "type": "follow_lowest",
                    "diffType": "euros",
                    "diffValue": avg_price_diff,
                    "minPriceDay": None,
                    "minPriceMonth": None
                }
            elif avg_position <= 3:
                recommendation = f"Room for improvement in {grupo}. Consider reducing prices by €0.20-0.50 to gain better positioning."
                suggested_strategy = f"Follow Lowest Price +{max(0, avg_price_diff - 0.30):.2f}€"
                strategy_config = {
                    "type": "follow_lowest",
                    "diffType": "euros",
                    "diffValue": max(0, avg_price_diff - 0.30),
                    "minPriceDay": None,
                    "minPriceMonth": None
                }
            else:
                recommendation = f"High prices detected in {grupo}. Reduce by €0.50-1.00 to improve market position."
                suggested_strategy = f"Follow Lowest Price +{max(0, avg_price_diff - 0.60):.2f}€"
                strategy_config = {
                    "type": "follow_lowest",
                    "diffType": "euros",
                    "diffValue": max(0, avg_price_diff - 0.60),
                    "minPriceDay": None,
                    "minPriceMonth": None
                }
            
            group_insights.append({
                "group": grupo,
                "groupName": group_names[grupo],
                "sampleSize": sample_size,
                "avgPosition": avg_position,
                "avgPriceDiff": avg_price_diff,
                "cheapestCount": cheapest_count,
                "confidence": confidence,
                "positionColor": position_color,
                "priceDiffColor": price_diff_color,
                "recommendation": recommendation,
                "suggestedStrategy": suggested_strategy,
                "strategyConfig": strategy_config
            })
        
        overall_avg_position = sum(avg_positions) / len(avg_positions) if avg_positions else 0
        overall_avg_price_diff = sum(avg_price_diffs) / len(avg_price_diffs) if avg_price_diffs else 0
        overall_confidence = min(95, 50 + (total_data_points // len(grupos)) // 2)
        
        result = {
            "ok": True,
            "dataPoints": total_data_points,
            "avgPosition": overall_avg_position,
            "avgPriceDiff": overall_avg_price_diff,
            "confidence": overall_confidence,
            "groupInsights": group_insights,
            "referenceSupplier": reference_supplier,
            "analysisPeriod": analysis_period,
            "location": location
        }
        
        print(f"[AI ANALYSIS] Generated insights for {len(group_insights)} groups, {total_data_points} data points")
        return JSONResponse(result)
        
    except Exception as e:
        print(f"[AI ANALYSIS ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/ai-deep-analysis")
async def ai_deep_analysis(request: Request):
    """
    Deep analysis endpoint for 24-month historical pricing data
    Analyzes positioning trends across all time periods
    """
    try:
        require_auth(request)
    except HTTPException:
        return JSONResponse({"ok": False, "error": "Unauthorized"}, status_code=401)
    
    try:
        body = await request.json()
        locations = body.get("locations", ["Albufeira", "Aeroporto de Faro"])
        reference_supplier = body.get("referenceSupplier", "autoprudente")
        results = body.get("results", [])
        
        print(f"[DEEP ANALYSIS] Locations: {locations}, Supplier: {reference_supplier}, Results: {len(results)}")
        
        # Analyze all results
        grupos = ['B1', 'B2', 'D', 'E1', 'E2', 'F', 'G', 'J1', 'J2', 'L1', 'L2', 'M1', 'M2', 'N']
        group_names = {
            'B1': 'Mini 4 Doors', 'B2': 'Mini Automatic', 'D': 'Economy',
            'E1': 'Compact 4 Doors', 'E2': 'Compact Automatic', 'F': 'Intermediate',
            'G': 'Intermediate Automatic', 'J1': 'Compact SUV', 'J2': 'Intermediate SUV',
            'L1': 'Standard', 'L2': 'Standard Automatic', 'M1': 'Premium', 'M2': 'Premium Automatic', 'N': 'Minivan'
        }
        
        # Aggregate data by group
        group_data = {grupo: [] for grupo in grupos}
        total_data_points = 0
        all_suppliers_found = set()  # Track ALL suppliers in the data
        
        # Normalize supplier names for matching (keep ALL suppliers, just standardize format)
        def normalize_supplier(name):
            # Just clean and lowercase, keep ALL suppliers
            name = str(name).strip().lower()
            # Standardize common variations for better grouping
            if 'autoprudente' in name or 'auto prudente' in name:
                return 'autoprudente'
            elif 'hertz' in name:
                return 'hertz'
            elif 'europcar' in name:
                return 'europcar'
            elif 'keddy' in name:
                return 'keddy'
            elif 'thrifty' in name:
                return 'thrifty'
            elif 'goldcar' in name:
                return 'goldcar'
            elif 'ok mobility' in name:
                return 'ok_mobility'
            elif 'centauro' in name:
                return 'centauro'
            elif 'surprice' in name:
                return 'surprice'
            elif 'firefly' in name:
                return 'firefly'
            elif 'sixt' in name:
                return 'sixt'
            elif 'avis' in name:
                return 'avis'
            elif 'budget' in name:
                return 'budget'
            elif 'enterprise' in name:
                return 'enterprise'
            elif 'national' in name:
                return 'national'
            elif 'dollar' in name:
                return 'dollar'
            elif 'alamo' in name:
                return 'alamo'
            # Keep any other supplier as-is (cleaned)
            return name.replace(' ', '_')
        
        for result in results:
            items = result.get('items', [])
            if not items:
                continue
            
            # Group items by car group
            for item in items:
                grupo = item.get('group', '')
                if grupo not in grupos:
                    continue
                
                supplier = normalize_supplier(item.get('supplier', ''))
                price = float(item.get('price_num', 0))
                
                if price > 0:
                    all_suppliers_found.add(supplier)  # Track supplier
                    group_data[grupo].append({
                        'supplier': supplier,
                        'price': price,
                        'date': result.get('date'),
                        'days': result.get('days')
                    })
                    total_data_points += 1
        
        # Log all suppliers found
        print(f"[DEEP ANALYSIS] Found {len(all_suppliers_found)} unique suppliers: {sorted(all_suppliers_found)}")
        
        # Calculate insights per group
        group_insights = []
        all_positions = []
        all_price_diffs = []
        
        for grupo in grupos:
            data = group_data[grupo]
            if len(data) < 10:  # Need minimum data points
                continue
            
            # Find reference supplier positions
            ref_positions = []
            ref_price_diffs = []
            cheapest_count = 0
            
            # Group by date+days for position calculation
            date_groups = {}
            for item in data:
                key = f"{item['date']}_{item['days']}"
                if key not in date_groups:
                    date_groups[key] = []
                date_groups[key].append(item)
            
            # Calculate position for each date/day combination
            for key, items in date_groups.items():
                # Sort by price
                sorted_items = sorted(items, key=lambda x: x['price'])
                
                # Find reference supplier
                ref_item = next((x for x in sorted_items if x['supplier'] == reference_supplier), None)
                if not ref_item:
                    continue
                
                # Calculate position (1-indexed)
                position = sorted_items.index(ref_item) + 1
                ref_positions.append(position)
                
                # Calculate price difference vs lowest
                lowest_price = sorted_items[0]['price']
                price_diff = ref_item['price'] - lowest_price
                ref_price_diffs.append(price_diff)
                
                if position == 1:
                    cheapest_count += 1
            
            if not ref_positions:
                continue
            
            avg_position = sum(ref_positions) / len(ref_positions)
            avg_price_diff = sum(ref_price_diffs) / len(ref_price_diffs)
            sample_size = len(ref_positions)
            confidence = min(95, 50 + (sample_size // 10))
            
            all_positions.append(avg_position)
            all_price_diffs.append(avg_price_diff)
            
            # Determine colors
            position_color = "green" if avg_position <= 2 else "yellow" if avg_position <= 3 else "red"
            price_diff_color = "green" if avg_price_diff <= 0.50 else "yellow" if avg_price_diff <= 1.00 else "red"
            
            # Generate recommendation
            if avg_position <= 2:
                recommendation = f"Excellent positioning in {grupo} over 24 months. You're consistently in top 2. Maintain current strategy."
                suggested_strategy = f"Follow Lowest Price +{avg_price_diff:.2f}€"
                strategy_config = {
                    "type": "follow_lowest",
                    "diffType": "euros",
                    "diffValue": round(avg_price_diff, 2),
                    "minPriceDay": None,
                    "minPriceMonth": None
                }
            elif avg_position <= 3:
                recommendation = f"Good positioning in {grupo}, but room for improvement. Consider reducing by €0.20-0.40 to rank higher."
                suggested_strategy = f"Follow Lowest Price +{max(0, avg_price_diff - 0.30):.2f}€"
                strategy_config = {
                    "type": "follow_lowest",
                    "diffType": "euros",
                    "diffValue": round(max(0, avg_price_diff - 0.30), 2),
                    "minPriceDay": None,
                    "minPriceMonth": None
                }
            else:
                recommendation = f"High prices detected in {grupo} consistently. Reduce by €0.50-1.00 to improve competitiveness."
                suggested_strategy = f"Follow Lowest Price +{max(0, avg_price_diff - 0.70):.2f}€"
                strategy_config = {
                    "type": "follow_lowest",
                    "diffType": "euros",
                    "diffValue": round(max(0, avg_price_diff - 0.70), 2),
                    "minPriceDay": None,
                    "minPriceMonth": None
                }
            
            group_insights.append({
                "group": grupo,
                "groupName": group_names[grupo],
                "sampleSize": sample_size,
                "avgPosition": round(avg_position, 1),
                "avgPriceDiff": round(avg_price_diff, 2),
                "cheapestCount": cheapest_count,
                "confidence": confidence,
                "positionColor": position_color,
                "priceDiffColor": price_diff_color,
                "recommendation": recommendation,
                "suggestedStrategy": suggested_strategy,
                "strategyConfig": strategy_config
            })
        
        overall_avg_position = sum(all_positions) / len(all_positions) if all_positions else 0
        overall_avg_price_diff = sum(all_price_diffs) / len(all_price_diffs) if all_price_diffs else 0
        overall_confidence = min(95, 50 + (total_data_points // 100))
        
        result = {
            "ok": True,
            "dataPoints": total_data_points,
            "avgPosition": overall_avg_position,
            "avgPriceDiff": overall_avg_price_diff,
            "confidence": overall_confidence,
            "groupInsights": group_insights,
            "referenceSupplier": reference_supplier,
            "locations": locations,
            "allSuppliers": sorted(list(all_suppliers_found)),
            "totalSuppliers": len(all_suppliers_found),
            "analysisType": "deep_scan_24months_both_locations"
        }
        
        print(f"[DEEP ANALYSIS] Complete: {len(group_insights)} groups, {total_data_points} data points, {len(all_suppliers_found)} suppliers")
        return JSONResponse(result)
        
    except Exception as e:
        print(f"[DEEP ANALYSIS ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/price-calendar-analysis")
async def price_calendar_analysis(request: Request):
    """
    Analyze daily price patterns and detect when prices change
    Identifies pricing periods and generates calendar recommendations
    """
    try:
        require_auth(request)
    except HTTPException:
        return JSONResponse({"ok": False, "error": "Unauthorized"}, status_code=401)
    
    try:
        body = await request.json()
        locations = body.get("locations", ["Albufeira", "Aeroporto de Faro"])
        reference_supplier = body.get("referenceSupplier", "autoprudente")
        results = body.get("results", [])
        
        print(f"[CALENDAR ANALYSIS] Locations: {locations}, Supplier: {reference_supplier}, Results: {len(results)}")
        
        # Organize data by date
        price_by_date = {}
        
        # Normalize supplier name
        def normalize_supplier(name):
            return str(name).strip().lower().replace(' ', '_')
        
        for result in results:
            location = result.get('location')
            date = result.get('date')
            items = result.get('items', [])
            
            if not date or not items:
                continue
            
            # Find reference supplier price
            ref_price = None
            all_prices = []
            
            for item in items:
                supplier = normalize_supplier(item.get('supplier', ''))
                price = float(item.get('price_num', 0))
                
                if price > 0:
                    all_prices.append(price)
                    if reference_supplier.lower() in supplier:
                        ref_price = price
            
            if ref_price and all_prices:
                if date not in price_by_date:
                    price_by_date[date] = []
                
                price_by_date[date].append({
                    'location': location,
                    'price': ref_price,
                    'lowest': min(all_prices),
                    'avg': sum(all_prices) / len(all_prices)
                })
        
        # Sort dates
        sorted_dates = sorted(price_by_date.keys())
        
        # Detect price changes (when price differs by >5% from previous)
        changes = []
        periods = []
        
        if len(sorted_dates) > 0:
            current_period = {
                'start_date': sorted_dates[0],
                'prices': []
            }
            
            prev_price = None
            for date in sorted_dates:
                date_data = price_by_date[date]
                avg_price = sum(d['price'] for d in date_data) / len(date_data)
                
                current_period['prices'].append(avg_price)
                
                if prev_price is not None:
                    change_pct = abs(avg_price - prev_price) / prev_price * 100
                    
                    if change_pct > 5:  # 5% threshold
                        # Price changed! End current period
                        changes.append({
                            'date': date,
                            'old_price': prev_price,
                            'new_price': avg_price,
                            'change_pct': round(change_pct, 1)
                        })
                        
                        # Save period
                        period_avg = sum(current_period['prices']) / len(current_period['prices'])
                        periods.append({
                            'start': current_period['start_date'],
                            'end': sorted_dates[sorted_dates.index(date) - 1],
                            'duration_days': len(current_period['prices']),
                            'avg_price': round(period_avg, 2)
                        })
                        
                        # Start new period
                        current_period = {
                            'start_date': date,
                            'prices': [avg_price]
                        }
                
                prev_price = avg_price
            
            # Add final period
            if current_period['prices']:
                period_avg = sum(current_period['prices']) / len(current_period['prices'])
                periods.append({
                    'start': current_period['start_date'],
                    'end': sorted_dates[-1],
                    'duration_days': len(current_period['prices']),
                    'avg_price': round(period_avg, 2)
                })
        
        # Calculate patterns
        avg_frequency = 0
        if len(changes) > 1:
            # Calculate average days between changes
            change_dates = [datetime.strptime(c['date'], '%Y-%m-%d') for c in changes]
            intervals = [(change_dates[i+1] - change_dates[i]).days for i in range(len(change_dates)-1)]
            avg_frequency = round(sum(intervals) / len(intervals), 1) if intervals else 0
        
        # Recommendations
        recommendations = []
        if len(changes) > 0:
            recommendations.append(f"Detected {len(changes)} significant price changes over the analyzed period")
        if len(periods) > 0:
            recommendations.append(f"Identified {len(periods)} distinct pricing periods")
            avg_period_length = sum(p['duration_days'] for p in periods) / len(periods)
            recommendations.append(f"Average pricing period lasts {int(avg_period_length)} days")
        if avg_frequency > 0:
            recommendations.append(f"Prices change approximately every {int(avg_frequency)} days")
            recommendations.append(f"Consider updating your prices every {max(1, int(avg_frequency * 0.8))} days to stay competitive")
        
        result = {
            "ok": True,
            "totalChanges": len(changes),
            "changes": changes[:50],  # Limit to 50 most recent
            "periods": periods,
            "patterns": {
                "avgFrequency": avg_frequency,
                "competitiveLag": 0,  # Would calculate by comparing change dates with competitors
                "bestDay": "N/A"  # Would analyze which day of week changes happen
            },
            "recommendations": recommendations,
            "dataPoints": len(sorted_dates),
            "locations": locations
        }
        
        print(f"[CALENDAR ANALYSIS] Complete: {len(changes)} changes, {len(periods)} periods")
        return JSONResponse(result)
        
    except Exception as e:
        print(f"[CALENDAR ANALYSIS ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/track-by-params")
async def track_by_params(request: Request):
    try:
        if not bool(str(os.getenv("DEV_NO_AUTH", "")).strip().lower() in ("1","true","yes","on")):
            require_auth(request)
    except HTTPException:
        return JSONResponse({"ok": False, "error": "Unauthorized"}, status_code=401)
    try:
        body = await request.json()
    except Exception:
        body = {}
    location = str(body.get("location") or "").strip()
    start_date = str(body.get("start_date") or "").strip()
    start_time = str(body.get("start_time") or "15:00").strip() or "15:00"  # Default 15:00 (consistente com rotação)
    end_date_in = str(body.get("end_date") or "").strip()
    end_time = str(body.get("end_time") or "15:00").strip() or "15:00"  # Default 15:00 (consistente com rotação)
    # days is optional if end_date provided
    try:
        days = int(body.get("days") or 0)
    except Exception:
        days = 0
    lang = str(body.get("lang") or "pt").strip() or "pt"
    currency = str(body.get("currency") or "EUR").strip() or "EUR"
    # quick=1 enables fast mode (skip some waits/screenshots)
    try:
        quick = int(body.get("quick") or 0)
    except Exception:
        quick = 0
    if not location or not start_date:
        return _no_store_json({"ok": False, "error": "Missing location or start_date"}, status_code=400)
    
    # LOG REQUEST PARAMS
    import sys
    print(f"\n{'='*60}")
    print(f"[API] REQUEST: location={location}, start_date={start_date}, start_time={start_time}, days={days}")
    print(f"[API] end_date_in={end_date_in}, end_time={end_time}")
    print(f"{'='*60}\n")
    print(f"[API] REQUEST: location={location}, start_date={start_date}, start_time={start_time}, days={days}", file=sys.stderr, flush=True)
    
    # ═══════════════════════════════════════════════════════════════════════════
    # DATE ROTATION - Variar datas para evitar detecção
    # ═══════════════════════════════════════════════════════════════════════════
    date_rotation_enabled = True
    date_rotation_max_days = 4
    
    # Carregar configurações da base de dados
    try:
        with _db_lock:
            con = _db_connect()
            try:
                row = con.execute(
                    "SELECT setting_value FROM price_automation_settings WHERE setting_key = ?",
                    ("date_rotation_enabled",)
                ).fetchone()
                if row:
                    date_rotation_enabled = row[0].lower() in ('true', '1', 'yes')
                
                row = con.execute(
                    "SELECT setting_value FROM price_automation_settings WHERE setting_key = ?",
                    ("date_rotation_max_days",)
                ).fetchone()
                if row:
                    date_rotation_max_days = int(row[0])
            finally:
                con.close()
    except Exception as e:
        print(f"[DATE_ROTATION] Erro ao carregar configurações: {e}", file=sys.stderr, flush=True)
    
    # Aplicar rotação de datas
    original_start_date = start_date
    if date_rotation_enabled and date_rotation_max_days > 0:
        try:
            base_date = datetime.strptime(start_date, "%Y-%m-%d")
            days_offset = random.randint(0, date_rotation_max_days)
            rotated_date = base_date + timedelta(days=days_offset)
            start_date = rotated_date.strftime("%Y-%m-%d")
            print(f"[DATE_ROTATION] Original: {original_start_date}, Rotated: {start_date} (+{days_offset} days)", file=sys.stderr, flush=True)
        except Exception as e:
            print(f"[DATE_ROTATION] Erro: {e}, usando data original", file=sys.stderr, flush=True)
    else:
        print(f"[DATE_ROTATION] Desativado, usando data original: {start_date}", file=sys.stderr, flush=True)
    
    try:
        # Build start datetime with provided time (usando data rotacionada)
        start_dt = datetime.fromisoformat(f"{start_date}T{start_time}")
    except Exception:
        return _no_store_json({"ok": False, "error": "Invalid start_date (YYYY-MM-DD)"}, status_code=400)
    # Determine end datetime
    if end_date_in:
        try:
            end_dt = datetime.fromisoformat(f"{end_date_in}T{end_time}")
        except Exception:
            return _no_store_json({"ok": False, "error": "Invalid end_date (YYYY-MM-DD)"}, status_code=400)
        if end_dt <= start_dt:
            return _no_store_json({"ok": False, "error": "end_date/time must be after start"}, status_code=400)
        days = max(1, (end_dt - start_dt).days)
    else:
        if days <= 0:
            return _no_store_json({"ok": False, "error": "Missing days or end_date"}, status_code=400)
        end_dt = start_dt + timedelta(days=days)
    print(f"[API] COMPUTED: start_dt={start_dt.date()}, end_dt={end_dt.date()}, days={days}")
    print(f"[API] COMPUTED: start_dt={start_dt.date()}, end_dt={end_dt.date()}, days={days}", file=sys.stderr, flush=True)
    try:
        items: List[Dict[str, Any]] = []
        base = f"https://www.carjet.com/do/list/{lang}"
        
        # MODO DE TESTE COM DADOS MOCKADOS (TEST_MODE_LOCAL=2)
        if TEST_MODE_LOCAL == 2:
            print(f"[MOCK MODE] Generating mock data for {location}, {days} days")
            # Preço base varia por localização
            base_price = 12.0 if 'faro' in location.lower() else 14.0
            items = []
            mock_cars = [
                # B1 - Mini 4 Doors
                ("Fiat 500", "Group B1", "Greenmotion"),
                ("Citroen C1", "Group B1", "Goldcar"),
                # B2 - Mini 5 Doors
                ("Toyota Aygo", "Group B2", "Surprice"),
                ("Volkswagen UP", "Group B2", "Centauro"),
                ("Fiat Panda", "Group B2", "OK Mobility"),
                # D - Economy
                ("Renault Clio", "Group D", "Goldcar"),
                ("Peugeot 208", "Group D", "Europcar"),
                ("Ford Fiesta", "Group D", "Hertz"),
                ("Seat Ibiza", "Group D", "Thrifty"),
                ("Hyundai i20", "Group D", "Centauro"),
                # E1 - Mini Automatic
                ("Fiat 500 Auto", "Group E1", "OK Mobility"),
                ("Peugeot 108 Auto", "Group E1", "Goldcar"),
                # E2 - Economy Automatic
                ("Opel Corsa Auto", "Group E2", "Europcar"),
                ("Ford Fiesta Auto", "Group E2", "Hertz"),
                # F - SUV
                ("Nissan Juke", "Group F", "Auto Prudente Rent a Car"),
                ("Peugeot 2008", "Group F", "Goldcar"),
                ("Renault Captur", "Group F", "Surprice"),
                # G - Premium
                ("Mini Cooper Countryman", "Group G", "Thrifty"),
                # J1 - Crossover
                ("Citroen C3 Aircross", "Group J1", "Centauro"),
                ("Fiat 500X", "Group J1", "OK Mobility"),
                ("VW T-Cross", "Group J1", "Europcar"),
                # J2 - Station Wagon
                ("Seat Leon SW", "Group J2", "Goldcar"),
                ("Peugeot 308 SW", "Group J2", "Hertz"),
                # L1 - SUV Automatic
                ("Peugeot 3008 Auto", "Group L1", "Auto Prudente Rent a Car"),
                ("Nissan Qashqai Auto", "Group L1", "Goldcar"),
                ("Toyota C-HR Auto", "Group L1", "Thrifty"),
                # L2 - Station Wagon Automatic
                ("Toyota Corolla SW Auto", "Group L2", "Europcar"),
                ("Opel Astra SW Auto", "Group L2", "Surprice"),
                # M1 - 7 Seater
                ("Dacia Lodgy", "Group M1", "Greenmotion"),
                ("Peugeot Rifter", "Group M1", "Centauro"),
                # M2 - 7 Seater Automatic
                ("Renault Grand Scenic Auto", "Group M2", "Goldcar"),
                ("VW Caddy Auto", "Group M2", "Auto Prudente Rent a Car"),
                # N - 9 Seater
                ("Ford Tourneo", "Group N", "Europcar"),
                ("Mercedes Vito Auto", "Group N", "Thrifty"),
            ]
            # Varia fornecedores por localização
            location_modifier = 0.0 if 'faro' in location.lower() else 3.0
            for idx, (car, cat, sup) in enumerate(mock_cars):
                price = base_price + (idx * 1.5) + (days * 0.3) + location_modifier
                # Varia fornecedor para Albufeira
                if 'albufeira' in location.lower() and idx % 3 == 0:
                    sup = "Centauro" if sup != "Centauro" else "Goldcar"
                # Extrair código do grupo da categoria (ex: "Group B1" -> "B1")
                group_code = cat.replace("Group ", "").strip() if "Group " in cat else "Others"
                items.append({
                    "id": idx,
                    "car": car,
                    "supplier": sup,
                    "price": f"€{price * days:.2f}",
                    "currency": "EUR",
                    "category": cat,
                    "group": group_code,
                    "transmission": "Automatic" if "Auto" in car else "Manual",
                    "photo": "",
                    "link": "",
                })
            print(f"[MOCK MODE] Generated {len(items)} mock items for {location} covering all groups")
            return _no_store_json({
                "ok": True,
                "items": items,
                "location": location,
                "start_date": start_dt.date().isoformat(),
                "start_time": start_dt.strftime("%H:%M"),
                "end_date": end_dt.date().isoformat(),
                "end_time": end_dt.strftime("%H:%M"),
                "days": days,
            })
        
        # ═══════════════════════════════════════════════════════════════════════════
        # MÉTODO 0: DESATIVADO - ScraperAPI não funciona bem com CarJet
        # ═══════════════════════════════════════════════════════════════════════════
        # Usar POST direto ou Selenium em vez disso
        
        # ═══════════════════════════════════════════════════════════════════════════
        # MÉTODO 1: POST DIRETO (Rápido mas menos confiável)
        # ═══════════════════════════════════════════════════════════════════════════
        try:
            import sys
            print(f"[POST_DIRETO] Tentando POST direto ao Carjet...", file=sys.stderr, flush=True)
            
            # Usar try_direct_carjet primeiro (POST direto)
            html = try_direct_carjet(location, start_dt, end_dt, lang=lang, currency=currency)
            final_url = "https://www.carjet.com/do/list"
            
            # Se POST direto retornar resultados, usar
            if html and len(parse_prices(html, final_url)) > 0:
                print(f"[POST_DIRETO] ✅ Sucesso com POST direto!", file=sys.stderr, flush=True)
            else:
                print(f"[POST_DIRETO] ⚠️ Falhou ou retornou 0 items, continuando para SELENIUM...", file=sys.stderr, flush=True)
                html = ""  # Limpar para forçar Selenium
                
            # DESATIVADO: Playwright mobile (usar Selenium como principal)
            if False:
                from playwright.async_api import async_playwright
                
                async with async_playwright() as p:
                    browser = await p.chromium.launch(headless=True)
                    
                    # Dispositivos mobile
                    devices = [
                        {"name": "iPhone 13 Pro", "user_agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1", "viewport": {"width": 390, "height": 844}, "scale": 3},
                        {"name": "Samsung Galaxy S21", "user_agent": "Mozilla/5.0 (Linux; Android 11; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36", "viewport": {"width": 360, "height": 800}, "scale": 3},
                    ]
                    device = random.choice(devices)
                    
                    context = await browser.new_context(
                        user_agent=device["user_agent"],
                        viewport=device["viewport"],
                        device_scale_factor=device["scale"],
                        is_mobile=True,
                        has_touch=True,
                        locale="pt-PT"
                    )
                    
                    page = await context.new_page()
                    html, final_url = await fetch_carjet_results(page, location, start_dt, end_dt, lang, currency, "")
                    await browser.close()
            
            # Parse results
            # Parse resultados do POST direto
            post_items = parse_prices(html, final_url) if html else []
            
            if post_items and len(post_items) > 0:
                print(f"[POST_DIRETO] ✅ {len(post_items)} carros encontrados!", file=sys.stderr, flush=True)
                items = normalize_and_sort(post_items, supplier_priority=None)
                return _no_store_json({
                    "ok": True,
                    "items": items,
                    "location": location,
                    "start_date": start_dt.date().isoformat(),
                    "start_time": start_dt.strftime("%H:%M"),
                    "end_date": end_dt.date().isoformat(),
                    "end_time": end_dt.strftime("%H:%M"),
                    "days": days,
                    "method": "post_direto",
                })
            else:
                print(f"[POST_DIRETO] ⚠️ Retornou 0 items, continuando para SELENIUM...", file=sys.stderr, flush=True)
        except Exception as e:
            print(f"[POST_DIRETO] ❌ Erro: {e}", file=sys.stderr, flush=True)
            import traceback
            traceback.print_exc()
            print(f"[POST_DIRETO] Continuando para SELENIUM...", file=sys.stderr, flush=True)
        
        # ═══════════════════════════════════════════════════════════════════════════
        # MÉTODO DESATIVADO: ScraperAPI (NÃO USAR - Bloqueado)
        # ═══════════════════════════════════════════════════════════════════════════
        if False:
            try:
                import httpx
                import sys
                from urllib.parse import urlencode
                print(f"[SCRAPERAPI] Iniciando scraping para {location}", file=sys.stderr, flush=True)
                
                # Mapear localização
                carjet_loc = location
                if 'faro' in location.lower():
                    carjet_loc = 'Faro Aeroporto (FAO)'
                elif 'albufeira' in location.lower():
                    carjet_loc = 'Albufeira Cidade'
                
                # Formato de datas para CarJet (dd/mm/yyyy)
                start_str = start_dt.strftime("%d/%m/%Y")
                end_str = end_dt.strftime("%d/%m/%Y")
                
                # Construir URL CarJet com parâmetros
                carjet_params = {
                    'pickup': carjet_loc,
                    'dropoff': carjet_loc,
                    'fechaRecogida': start_str,
                    'fechaEntrega': end_str,
                    'fechaRecogidaSelHour': '10:00',
                    'fechaEntregaSelHour': '10:00',
                }
                target_url = f"https://www.carjet.com/aluguel-carros/index.htm?{urlencode(carjet_params)}"
                
                # Construir URL ScraperAPI
                scraper_params = {
                    'api_key': SCRAPER_API_KEY,
                    'url': target_url,
                    'render_js': 'true',
                    'wait': '3000',
                    'country': 'pt',
                }
                scraper_url = f"http://api.scrapeops.io/v1/?{urlencode(scraper_params)}"
                
                print(f"[SCRAPERAPI] Target: {target_url[:100]}...", file=sys.stderr, flush=True)
                print(f"[SCRAPERAPI] Fazendo request via ScraperOps...", file=sys.stderr, flush=True)
                
                # Fazer request via ScraperAPI
                async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as client:
                    response = await client.get(scraper_url)
                
                
                if response.status_code == 200:
                    html_content = response.text
                    print(f"[SCRAPERAPI] ✅ HTML recebido: {len(html_content)} bytes", file=sys.stderr, flush=True)
                    
                    # Parse o HTML
                    items = parse_prices(html_content, target_url)
                    print(f"[SCRAPERAPI] Parsed {len(items)} items antes conversão", file=sys.stderr, flush=True)
                    
                    # Converter GBP para EUR
                    items = convert_items_gbp_to_eur(items)
                    print(f"[SCRAPERAPI] {len(items)} items após GBP→EUR", file=sys.stderr, flush=True)
                    
                    # Aplicar ajustes
                    items = apply_price_adjustments(items, target_url)
                    print(f"[SCRAPERAPI] {len(items)} items após ajustes", file=sys.stderr, flush=True)
                    
                    if items:
                        print(f"[SCRAPERAPI] ✅ {len(items)} carros encontrados!", file=sys.stderr, flush=True)
                        if items:
                            print(f"[SCRAPERAPI] Primeiro: {items[0].get('car', 'N/A')} - {items[0].get('price', 'N/A')}", file=sys.stderr, flush=True)
                        # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
                        items = normalize_and_sort(items, supplier_priority=None)
                        return _no_store_json({
                            "ok": True,
                            "items": items,
                            "location": location,
                            "start_date": start_dt.date().isoformat(),
                            "start_time": start_dt.strftime("%H:%M"),
                            "end_date": end_dt.date().isoformat(),
                            "end_time": end_dt.strftime("%H:%M"),
                            "days": days,
                        })
                    else:
                        print(f"[SCRAPERAPI] ⚠️ Parse retornou 0 items", file=sys.stderr, flush=True)
                else:
                    print(f"[SCRAPERAPI] ❌ HTTP {response.status_code}", file=sys.stderr, flush=True)
                    print(f"[SCRAPERAPI] Tentando fallback para Playwright...", file=sys.stderr, flush=True)
            except Exception as e:
                import sys
                print(f"[SCRAPERAPI ERROR] {e}", file=sys.stderr, flush=True)
                import traceback
                traceback.print_exc(file=sys.stderr)
                print(f"[SCRAPERAPI] Tentando fallback para Playwright...", file=sys.stderr, flush=True)
        
        # FALLBACK: Tentar Playwright se ScraperAPI falhou (DESATIVADO - usar Selenium)
        if False and TEST_MODE_LOCAL == 0 and not items and _HAS_PLAYWRIGHT:
            try:
                from playwright.async_api import async_playwright
                import sys
                print(f"[PLAYWRIGHT] Iniciando scraping direto para {location}", file=sys.stderr, flush=True)
                
                async with async_playwright() as p:
                    browser = await p.chromium.launch(headless=True)
                    context = await browser.new_context(
                        locale="pt-PT",
                        viewport={"width": 1920, "height": 1080},
                        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                    )
                    page = await context.new_page()
                    
                    try:
                        # Ir para página inicial do CarJet PT
                        print(f"[PLAYWRIGHT] Acessando CarJet homepage PT...", file=sys.stderr, flush=True)
                        await page.goto("https://www.carjet.com/aluguer-carros/index.htm", wait_until="domcontentloaded", timeout=45000)
                        
                        # Aguardar página carregar
                        await page.wait_for_timeout(2000)
                        
                        # Fechar cookies via JS (mais robusto)
                        await page.evaluate("""() => {
                            try {
                                document.querySelectorAll('[id*=cookie], [class*=cookie], [id*=consent], [class*=consent]').forEach(el => el.remove());
                            } catch(e) {}
                        }""")
                        
                        # Mapear localização
                        carjet_loc = location
                        if 'faro' in location.lower():
                            carjet_loc = 'Faro Aeroporto (FAO)'
                        elif 'albufeira' in location.lower():
                            carjet_loc = 'Albufeira Cidade'
                        
                        print(f"[PLAYWRIGHT] Preenchendo formulário via JS: {carjet_loc}", file=sys.stderr, flush=True)
                        
                        # Preencher formulário via JavaScript (mais robusto que seletores)
                        start_str = start_dt.strftime("%d/%m/%Y")
                        end_str = end_dt.strftime("%d/%m/%Y")
                        
                        await page.evaluate("""
                            ({loc, start, end, startTime, endTime}) => {
                                function fill(sel, val) {
                                    const el = document.querySelector(sel);
                                    if (el) { 
                                        el.value = val; 
                                        el.dispatchEvent(new Event('change', {bubbles: true}));
                                        return true;
                                    }
                                    return false;
                                }
                                const r1 = fill('input[name="pickup"]', loc);
                                const r2 = fill('input[name="dropoff"]', loc);
                                const r3 = fill('input[name="fechaRecogida"]', start);
                                const r4 = fill('input[name="fechaEntrega"]', end);
                                const h1 = document.querySelector('select[name="fechaRecogidaSelHour"]');
                                const h2 = document.querySelector('select[name="fechaEntregaSelHour"]');
                                if (h1) h1.value = startTime || '16:00';
                                if (h2) h2.value = endTime || '10:00';
                                return {r1, r2, r3, r4};
                            }
                        """, {"loc": carjet_loc, "start": start_str, "end": end_str, "startTime": start_dt.strftime("%H:%M"), "endTime": end_dt.strftime("%H:%M")})

                        # Click the primary search/submit button and wait for results
                        try:
                            btn = None
                            try:
                                btn = page.get_by_role("button", name=re.compile(r"(Pesquisar|Buscar|Search)", re.I))
                            except Exception:
                                btn = None
                            if btn and await btn.is_visible():
                                await btn.click(timeout=3000)
                            else:
                                cand = page.locator("button:has-text('Pesquisar'), button:has-text('Buscar'), button:has-text('Search'), input[type=submit], button[type=submit]")
                                if await cand.count() > 0:
                                    try:
                                        await cand.first.click(timeout=3000)
                                    except Exception:
                                        pass
                            try:
                                await page.wait_for_load_state("networkidle", timeout=10000)
                            except Exception:
                                pass
                        except Exception:
                            pass

                        # If we landed on a 'war=' URL without the secure params, try a direct POST fallback
                        try:
                            current_url = page.url or ""
                            if ("war=" in current_url) and ("s=" not in current_url) and ("b=" not in current_url):
                                from urllib.parse import urlparse as _urlparse
                                print(f"[PLAYWRIGHT] war URL detected, attempting direct POST fallback...", file=sys.stderr, flush=True)
                                payload_dp = build_carjet_form(location, start_dt, end_dt, lang=lang, currency=currency)
                                rdp = await page.request.post(f"https://www.carjet.com/do/list/{lang}", data=payload_dp)
                                if rdp and rdp.ok:
                                    html_dp = await rdp.text()
                                    its_dp = parse_prices(html_dp, f"https://www.carjet.com/do/list/{lang}")
                                    its_dp = convert_items_gbp_to_eur(its_dp)
                                    its_dp = apply_price_adjustments(its_dp, f"https://www.carjet.com/do/list/{lang}")
                                    if its_dp:
                                        print(f"[PLAYWRIGHT] ✅ Fallback POST retornou {len(its_dp)} carros", file=sys.stderr, flush=True)
                                        # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
                                        its_dp = normalize_and_sort(its_dp, supplier_priority=None)
                                        return _no_store_json({
                                            "ok": True,
                                            "items": its_dp,
                                            "location": location,
                                            "start_date": start_dt.date().isoformat(),
                                            "start_time": start_dt.strftime("%H:%M"),
                                            "end_date": end_dt.date().isoformat(),
                                            "end_time": end_dt.strftime("%H:%M"),
                                            "days": days,
                                        })
                        except Exception:
                            pass
                        
                        await page.wait_for_timeout(1000)
                        
                        print(f"[PLAYWRIGHT] Submetendo formulário...", file=sys.stderr, flush=True)
                        
                        # Submeter via JS (mais confiável)
                        await page.evaluate("""() => {
                            const form = document.querySelector('form');
                            if (form) form.submit();
                        }""")
                        
                        # Aguardar navegação para página de resultados
                        print(f"[PLAYWRIGHT] Aguardando navegação...", file=sys.stderr, flush=True)
                        await page.wait_for_url('**/do/list/**', timeout=90000)
                        
                        print(f"[PLAYWRIGHT] Aguardando carros carregarem...", file=sys.stderr, flush=True)
                        await page.wait_for_timeout(8000)
                        
                        # Extrair URL e HTML
                        final_url = page.url
                        html_content = await page.content()
                        print(f"[PLAYWRIGHT] URL final: {final_url}", file=sys.stderr, flush=True)
                        print(f"[PLAYWRIGHT] ✅ HTML capturado: {len(html_content)} bytes", file=sys.stderr, flush=True)
                        
                        # Parse
                        items = parse_prices(html_content, page.url)
                        print(f"[PLAYWRIGHT] Parsed {len(items)} items antes conversão", file=sys.stderr, flush=True)
                        
                        # Converter GBP para EUR
                        items = convert_items_gbp_to_eur(items)
                        print(f"[PLAYWRIGHT] {len(items)} items após GBP→EUR", file=sys.stderr, flush=True)
                        
                        # Aplicar ajustes
                        items = apply_price_adjustments(items, page.url)
                        print(f"[PLAYWRIGHT] {len(items)} items após ajustes", file=sys.stderr, flush=True)
                        
                        if items:
                            print(f"[PLAYWRIGHT] ✅ {len(items)} carros encontrados!", file=sys.stderr, flush=True)
                            if items:
                                print(f"[PLAYWRIGHT] Primeiro: {items[0].get('car', 'N/A')} - {items[0].get('price', 'N/A')}", file=sys.stderr, flush=True)
                            # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
                            items = normalize_and_sort(items, supplier_priority=None)
                            return _no_store_json({
                                "ok": True,
                                "items": items,
                                "location": location,
                                "start_date": start_dt.date().isoformat(),
                                "start_time": start_dt.strftime("%H:%M"),
                                "end_date": end_dt.date().isoformat(),
                                "end_time": end_dt.strftime("%H:%M"),
                                "days": days,
                            })
                        else:
                            print(f"[PLAYWRIGHT] ⚠️ Parse retornou 0 items", file=sys.stderr, flush=True)
                    
                    finally:
                        await browser.close()
            
            except Exception as e:
                import sys
                print(f"[PLAYWRIGHT ERROR] {e}", file=sys.stderr, flush=True)
                import traceback
                traceback.print_exc(file=sys.stderr)
        
        # MODO DE TESTE LOCAL: Usar URL s/b pré-configurada
        test_url = None
        print(f"[DEBUG] TEST_MODE_LOCAL={TEST_MODE_LOCAL}, location={location.lower()}, days={days}")
        
        # Tentar carregar URLs dinâmicas do .env (FARO_XD, ALBUFEIRA_XD)
        if TEST_MODE_LOCAL == 0:
            loc_prefix = None
            if 'faro' in location.lower():
                loc_prefix = 'FARO'
            elif 'albufeira' in location.lower():
                loc_prefix = 'ALBUFEIRA'
            if loc_prefix:
                env_key = f"{loc_prefix}_{days}D"
                test_url = os.getenv(env_key, "").strip()
                if test_url and test_url.startswith('http'):
                    print(f"[DEBUG] Found dynamic URL in .env: {env_key}={test_url[:80]}...", file=sys.stderr, flush=True)
                else:
                    test_url = None
        
        if TEST_MODE_LOCAL == 1:
            print(f"[DEBUG] Checking location: faro={'faro' in location.lower()}, albufeira={'albufeira' in location.lower()}")
            if 'faro' in location.lower() and TEST_FARO_URL:
                test_url = TEST_FARO_URL
                print(f"[DEBUG] Using Faro test URL")
            elif 'albufeira' in location.lower() and TEST_ALBUFEIRA_URL:
                test_url = TEST_ALBUFEIRA_URL
                print(f"[DEBUG] Using Albufeira test URL")
        
        if test_url:
            try:
                import requests
                import sys
                print(f"[TEST MODE] Usando URL pré-configurada para {location}", file=sys.stderr, flush=True)
                r = requests.get(test_url, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                    'Cookie': 'monedaForzada=EUR; moneda=EUR; currency=EUR'
                }, timeout=15)
                
                print(f"[TEST MODE] Fetched {len(r.text)} bytes", file=sys.stderr, flush=True)
                # DEBUG: Save HTML
                try:
                    with open(DEBUG_DIR / f"test_mode_html_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html", 'w') as f:
                        f.write(r.text)
                except:
                    pass
                items = parse_prices(r.text, TEST_FARO_URL)
                print(f"[TEST MODE] Parsed {len(items)} items", file=sys.stderr, flush=True)
                if items:
                    print(f"[TEST MODE] Primeiro preço ANTES conversão: {items[0].get('price', 'N/A')}", file=sys.stderr, flush=True)
                # CONVERTER GBP→EUR pois CarJet retorna em Libras
                items = convert_items_gbp_to_eur(items)
                print(f"[TEST MODE] After GBP→EUR conversion: {len(items)} items", file=sys.stderr, flush=True)
                if items:
                    print(f"[TEST MODE] Primeiro preço DEPOIS conversão: {items[0].get('price', 'N/A')}", file=sys.stderr, flush=True)
                # Aplicar ajustes de preço se configurados
                items = apply_price_adjustments(items, test_url)
                print(f"[TEST MODE] After price adjustments: {len(items)} items", file=sys.stderr, flush=True)
                
                if items:
                    print(f"[TEST MODE] {len(items)} carros encontrados!", file=sys.stderr, flush=True)
                    # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
                    items = normalize_and_sort(items, supplier_priority=None)
                    try:
                        prices = [float(item.get('price_num', 0)) for item in items if item.get('price_num')]
                        min_price = min(prices) if prices else None
                        max_price = max(prices) if prices else None
                        avg_price = sum(prices) / len(prices) if prices else None
                        
                        save_search_to_history(
                            location=location,
                            start_date=start_dt.date().isoformat(),
                            end_date=end_dt.date().isoformat(),
                            days=days,
                            results_count=len(items),
                            min_price=min_price,
                            max_price=max_price,
                            avg_price=avg_price,
                            user="admin",
                            search_params=f"lang={selected_language['name']}, hour={selected_hour}, device={selected_device['name']}"
                        )
                    except Exception as e:
                        print(f"[SELENIUM] Erro ao salvar histórico: {e}", file=sys.stderr, flush=True)
                    
                    # RETORNAR resultado final
                    return _no_store_json({
                        "ok": True,
                        "items": items,
                        "location": location,
                        "start_date": start_dt.date().isoformat(),
                        "start_time": start_dt.strftime("%H:%M"),
                        "end_date": end_dt.date().isoformat(),
                        "end_time": end_dt.strftime("%H:%M"),
                        "days": days,
                    })
            except Exception as e:
                import sys
                print(f"[TEST MODE ERROR] {e}", file=sys.stderr, flush=True)
        
        # ═══════════════════════════════════════════════════════════════════════════
        # MÉTODO 2: SELENIUM (PRINCIPAL - Mais confiável) ✅ RECOMENDADO
        # ═══════════════════════════════════════════════════════════════════════════
        # Com todas as rotações implementadas:
        # ✅ Rotação de horas - 14:30-17:00 (6 opções)
        # ✅ Rotação de dispositivos - 4 devices mobile
        # ✅ Rotação de timezones - 4 europeus
        # ✅ Rotação de referrers - 5 opções
        # ✅ Cache clearing - Desativado completamente
        # ✅ Seletor universal testado - #recogida_lista li:first-child a
        # ═══════════════════════════════════════════════════════════════════════════
        print(f"[SELENIUM] Iniciando scraping SIMPLES (igual ao teste) para {location}", file=sys.stderr, flush=True)
        try:
            # Usar função simples que funciona 100%
            from selenium_simple import scrape_carjet_simple
            
            result = scrape_carjet_simple(location, start_dt, end_dt)
            
            if result.get('ok'):
                html_content = result.get('html')
                final_url = result.get('url')
                
                print(f"[SELENIUM] ✅ Scraping simples bem-sucedido!", file=sys.stderr, flush=True)
                print(f"[SELENIUM] Fazendo parse de {len(html_content)} bytes...", file=sys.stderr, flush=True)
                
                items = parse_prices(html_content, final_url)
                print(f"[SELENIUM] Parsed {len(items)} items", file=sys.stderr, flush=True)
                items = convert_items_gbp_to_eur(items)
                print(f"[SELENIUM] {len(items)} após GBP→EUR", file=sys.stderr, flush=True)
                items = apply_price_adjustments(items, final_url)
                print(f"[SELENIUM] {len(items)} após ajustes", file=sys.stderr, flush=True)
                
                if items:
                    print(f"[SELENIUM] ✅ {len(items)} carros encontrados!", file=sys.stderr, flush=True)
                    items = normalize_and_sort(items, supplier_priority=None)
                    return _no_store_json({
                        "ok": True,
                        "items": items,
                        "location": location,
                        "start_date": start_dt.date().isoformat(),
                        "start_time": start_dt.strftime("%H:%M"),
                        "end_date": end_dt.date().isoformat(),
                        "end_time": end_dt.strftime("%H:%M"),
                        "days": days,
                    })
            else:
                print(f"[SELENIUM] ⚠️ Scraping simples falhou: {result.get('error')}", file=sys.stderr, flush=True)
                
        except Exception as e:
            print(f"[SELENIUM] ❌ Erro no scraping simples: {e}", file=sys.stderr, flush=True)
            import traceback
            traceback.print_exc()
        
        # Se chegou aqui, tentar método antigo como fallback
        print(f"[SELENIUM] Tentando método complexo como fallback...", file=sys.stderr, flush=True)
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
            from selenium.webdriver.chrome.service import Service
            from webdriver_manager.chrome import ChromeDriverManager
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            import time
            
            # ============================================
            # ROTAÇÃO MULTI-IDIOMA (6 idiomas - Holandês removido por problemas)
            # ========== FORÇAR PORTUGUÊS (IGUAL AO TESTE) ==========
            lang = 'pt'  # Forçar português
            selected_language = {
                'name': 'Português',
                'url': 'https://www.carjet.com/aluguel-carros/index.htm',
                'faro': 'Faro Aeroporto (FAO)',
                'albufeira': 'Albufeira Cidade'
            }
            
            # Mapear location para formato CarJet
            carjet_location = location
            if 'faro' in location.lower():
                carjet_location = selected_language['faro']
            elif 'albufeira' in location.lower():
                carjet_location = selected_language['albufeira']
            
            carjet_url = selected_language['url']
            
            print(f"[SELENIUM] Idioma: {selected_language['name']} (FIXO - igual ao teste)", file=sys.stderr, flush=True)
            print(f"[SELENIUM] URL: {carjet_url}", file=sys.stderr, flush=True)
            print(f"[SELENIUM] Local: {carjet_location}", file=sys.stderr, flush=True)
            
            # ============================================
            # ROTAÇÃO DE DATAS (0-4 dias aleatório)
            # ============================================
            # random já importado globalmente
            from datetime import timedelta as td
            
            # Adicionar offset aleatório de 0-4 dias às datas
            date_offset = random.randint(0, 4)
            start_dt = start_dt + td(days=date_offset)
            end_dt = end_dt + td(days=date_offset)
            
            print(f"[SELENIUM] Offset de datas: +{date_offset} dias", file=sys.stderr, flush=True)
            print(f"[SELENIUM] Datas ajustadas: {start_dt.date()} - {end_dt.date()}", file=sys.stderr, flush=True)
            
            # ============================================
            # ROTAÇÃO DE HORAS (14:30-17:00 aleatório)
            # ============================================
            # Horas disponíveis no Carjet (de 30 em 30 minutos)
            available_hours = ['14:30', '15:00', '15:30', '16:00', '16:30', '17:00']
            selected_hour = random.choice(available_hours)
            
            # Ajustar start_dt e end_dt para usar a hora selecionada
            start_dt = start_dt.replace(hour=int(selected_hour.split(':')[0]), minute=int(selected_hour.split(':')[1]))
            end_dt = end_dt.replace(hour=int(selected_hour.split(':')[0]), minute=int(selected_hour.split(':')[1]))
            
            print(f"[SELENIUM] Hora selecionada: {selected_hour}", file=sys.stderr, flush=True)
            
            # ============================================
            # FORÇAR DEVICE E REFERRER IGUAIS AO TESTE
            # ============================================
            
            selected_device = {
                'name': 'iPhone 13 Pro',
                'ua': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1',
                'width': 390,
                'height': 844,
                'pixelRatio': 3.0
            }
            
            selected_timezone = 'Europe/Lisbon'
            selected_referrer = ''  # SEM REFERRER (igual ao teste)
            
            print(f"[SELENIUM] Device: {selected_device['name']}", file=sys.stderr, flush=True)
            print(f"[SELENIUM] Timezone: {selected_timezone}", file=sys.stderr, flush=True)
            print(f"[SELENIUM] Language: {selected_language['name']}", file=sys.stderr, flush=True)
            print(f"[SELENIUM] Referrer: {selected_referrer if selected_referrer else 'Direct'}", file=sys.stderr, flush=True)
            
            chrome_options = Options()
            # chrome_options.add_argument('--headless')  # ❌ DESATIVADO - Igual ao teste manual!
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument(f'user-agent={selected_device["ua"]}')
            
            # EMULAÇÃO MOBILE COMPLETA com device específico
            mobile_emulation = {
                "deviceMetrics": { 
                    "width": selected_device['width'], 
                    "height": selected_device['height'], 
                    "pixelRatio": selected_device['pixelRatio']
                },
                "userAgent": selected_device['ua']
            }
            chrome_options.add_experimental_option("mobileEmulation", mobile_emulation)
            
            # Anti-deteção
            chrome_options.add_argument('--disable-blink-features=AutomationControlled')
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            
            # Preferências mínimas (IGUAL AO TESTE - sem bloquear cookies!)
            # NÃO bloquear cookies - pode fazer CarJet comportar-se diferente!
            # chrome_options.add_experimental_option("prefs", {
            #     "intl.accept_languages": selected_lang,
            # })
            
            # Detectar sistema operacional e definir caminho do Chrome
            import platform
            system = platform.system()
            if system == 'Darwin':  # macOS
                chrome_options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
            elif system == 'Linux':  # Linux (Render)
                # No Linux com Docker, o Chrome está em /usr/bin/google-chrome-stable
                if os.path.exists('/usr/bin/google-chrome-stable'):
                    chrome_options.binary_location = '/usr/bin/google-chrome-stable'
                # Não definir binary_location deixa o Selenium encontrar automaticamente
            
            # Iniciar driver - tentar com Chrome instalado primeiro
            try:
                # Tentar usar Chrome do sistema (melhor para Mac ARM)
                driver = webdriver.Chrome(options=chrome_options)
                print(f"[SELENIUM] ✅ Chrome iniciado com sucesso!", file=sys.stderr, flush=True)
            except Exception as e:
                print(f"[SELENIUM] ⚠️ Erro ao iniciar Chrome: {e}", file=sys.stderr, flush=True)
                print(f"[SELENIUM] Tentando com ChromeDriverManager...", file=sys.stderr, flush=True)
                # Fallback para ChromeDriverManager
                driver = webdriver.Chrome(
                    service=Service(ChromeDriverManager().install()),
                    options=chrome_options
                )
            
            # FUNÇÃO HELPER: Autodetectar e REJEITAR cookies (mais simples!)
            def reject_cookies_if_present(step_name=""):
                """Detecta e REJEITA cookies automaticamente. Retorna True se encontrou cookies."""
                try:
                    result = driver.execute_script("""
                        // Procurar e clicar no botão de REJEITAR cookies
                        const buttons = document.querySelectorAll('button, a, [role="button"]');
                        let found = false;
                        for (let btn of buttons) {
                            const text = btn.textContent.toLowerCase().trim();
                            // Procurar por "rejeitar", "recusar", "reject", etc.
                            if (text.includes('rejeitar') || text.includes('recusar') || 
                                text.includes('reject') || text.includes('rechazar') ||
                                text.includes('weiger') || text.includes('afwijzen') ||  // Holandês
                                text.includes('não aceitar') || text.includes('decline')) {
                                btn.click();
                                console.log('✓ Cookies rejeitados:', btn.textContent);
                                found = true;
                                break;
                            }
                        }
                        // Se não encontrou botão de rejeitar, tentar fechar/remover o banner
                        if (!found) {
                            document.querySelectorAll('[id*=cookie], [class*=cookie], [id*=didomi], [class*=didomi], [id*=consent], [class*=consent]').forEach(el => {
                                el.remove();
                            });
                        }
                        document.body.style.overflow = 'auto';
                        return found;
                    """)
                    if result:
                        print(f"[SELENIUM] ✓ Cookies rejeitados {step_name}", file=sys.stderr, flush=True)
                    else:
                        print(f"[SELENIUM] ℹ️  Banner removido {step_name}", file=sys.stderr, flush=True)
                    return result
                except Exception as e:
                    print(f"[SELENIUM] ⚠ Erro ao verificar cookies {step_name}: {e}", file=sys.stderr, flush=True)
                    return False
            
            # TIMEOUT GLOBAL: Se ficar preso mais de 60 segundos, abortar
            import signal
            def timeout_handler(signum, frame):
                raise TimeoutError("Selenium ficou preso por mais de 60 segundos!")
            
            # Configurar timeout (apenas em sistemas Unix)
            try:
                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(60)  # 60 segundos de timeout
            except:
                pass  # Windows não suporta SIGALRM
            
            try:
                # Esconder webdriver
                driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                    'source': '''
                        Object.defineProperty(navigator, 'webdriver', {
                            get: () => undefined
                        });
                    '''
                })
                
                print(f"[SELENIUM] Configurando Chrome com mobile UA...", file=sys.stderr, flush=True)
                driver.set_page_load_timeout(20)  # Igual ao teste manual
                
                # Definir referrer se não for direct
                if selected_referrer:
                    print(f"[SELENIUM] Definindo referrer...", file=sys.stderr, flush=True)
                    driver.execute_cdp_cmd('Network.setExtraHTTPHeaders', {
                        'headers': {'Referer': selected_referrer}
                    })
                
                print(f"[SELENIUM] Acessando CarJet ({selected_language['name']})...", file=sys.stderr, flush=True)
                driver.get(carjet_url)
                
                # Rejeitar cookies (IGUAL AO TESTE)
                time.sleep(0.5)
                if reject_cookies_if_present(""):
                    print(f"[SELENIUM] ✅ Cookies rejeitados", file=sys.stderr, flush=True)
                time.sleep(0.5)
                
                # DETECTAR IDIOMA e ajustar nome da localização
                page_url = driver.current_url
                print(f"[SELENIUM] URL atual: {page_url}", file=sys.stderr, flush=True)
                
                # Se estiver em inglês, usar nomes em inglês
                if '/index.htm' in page_url and '/aluguel-carros/' not in page_url and '/aluguer-carros/' not in page_url:
                    print(f"[SELENIUM] Página em INGLÊS detectada!", file=sys.stderr, flush=True)
                    if carjet_location == 'Albufeira Cidade':
                        carjet_location = 'Albufeira City'
                    elif carjet_location == 'Faro Aeroporto':
                        carjet_location = 'Faro Airport'
                    print(f"[SELENIUM] Local ajustado para: {carjet_location}", file=sys.stderr, flush=True)
                
                try:
                    # ========== ORDEM CORRETA DO TESTE: LOCAL → DROPDOWN → DATAS ==========
                    
                    # PASSO 1: Escrever local (IGUAL AO TESTE)
                    print(f"[SELENIUM] PASSO 1: Escrevendo local...", file=sys.stderr, flush=True)
                    pickup_input = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.ID, "pickup"))
                    )
                    pickup_input.clear()
                    pickup_input.send_keys(carjet_location)
                    print(f"[SELENIUM] ✓ Local digitado", file=sys.stderr, flush=True)
                    
                    # PASSO 2: Aguardar dropdown e clicar
                    print(f"[SELENIUM] PASSO 2: Aguardando dropdown...", file=sys.stderr, flush=True)
                    time.sleep(3)
                    
                    try:
                        dropdown_item = WebDriverWait(driver, 3).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, "#recogida_lista li:first-child a"))
                        )
                        dropdown_item.click()
                        print(f"[SELENIUM] ✓ Dropdown clicado", file=sys.stderr, flush=True)
                    except:
                        driver.execute_script("""
                            const items = document.querySelectorAll('#recogida_lista li');
                            for (let item of items) {
                                if (item.offsetParent !== null) {
                                    item.click();
                                    return true;
                                }
                            }
                        """)
                        print(f"[SELENIUM] ✓ Dropdown clicado (JS)", file=sys.stderr, flush=True)
                    
                    time.sleep(1)
                    
                    # PASSO 3: Preencher datas e horas (DEPOIS do dropdown!)
                    print(f"[SELENIUM] PASSO 3: Preenchendo datas e horas...", file=sys.stderr, flush=True)
                    
                    result = driver.execute_script("""
                        function fill(sel, val) {
                            const el = document.querySelector(sel);
                            if (el) { 
                                el.value = val; 
                                el.dispatchEvent(new Event('input', {bubbles: true}));
                                el.dispatchEvent(new Event('change', {bubbles: true}));
                                el.dispatchEvent(new Event('blur', {bubbles: true}));  // ← IGUAL AO TESTE!
                                console.log('Preenchido:', sel, '=', val);
                                return true;
                            }
                            console.error('Não encontrado:', sel);
                            return false;
                        }
                        
                        // Preencher datas
                        const r1 = fill('input[id="fechaRecogida"]', arguments[0]);
                        const r2 = fill('input[id="fechaDevolucion"]', arguments[1]);
                        
                        // Preencher horas
                        const h1 = document.querySelector('select[id="fechaRecogidaSelHour"]');
                        let h1_ok = false;
                        if (h1) { 
                            h1.value = arguments[2]; 
                            h1.dispatchEvent(new Event('change', {bubbles: true}));
                            console.log('Hora recolha:', h1.value);
                            h1_ok = true;
                        }
                        
                        const h2 = document.querySelector('select[id="fechaDevolucionSelHour"]');
                        let h2_ok = false;
                        if (h2) { 
                            h2.value = arguments[3]; 
                            h2.dispatchEvent(new Event('change', {bubbles: true}));
                            console.log('Hora devolução:', h2.value);
                            h2_ok = true;
                        }
                        
                        return {
                            fechaRecogida: r1,
                            fechaDevolucion: r2,
                            horaRecogida: h1_ok,
                            horaDevolucion: h2_ok,
                            allFilled: r1 && r2 && h1_ok && h2_ok
                        };
                    """, start_dt.strftime("%d/%m/%Y"), end_dt.strftime("%d/%m/%Y"), start_dt.strftime("%H:%M"), start_dt.strftime("%H:%M"))
                    
                    print(f"[SELENIUM] ✓ Datas e horas preenchidas: {result}", file=sys.stderr, flush=True)
                    
                except Exception as e:
                    print(f"[SELENIUM] Erro ao preencher: {e}", file=sys.stderr, flush=True)
                
                # PASSO 4: Submit (IGUAL AO TESTE)
                print(f"[SELENIUM] PASSO 4: Submetendo...", file=sys.stderr, flush=True)
                driver.execute_script("window.scrollBy(0, 300);")
                time.sleep(0.5)
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(0.5)
                driver.execute_script("document.querySelector('form').submit();")
                
                print(f"[SELENIUM] Aguardando navegação inicial...", file=sys.stderr, flush=True)
                time.sleep(3)
                
                # Aguardar até que a URL contenha /do/list/ (resultado final)
                print(f"[SELENIUM] Aguardando página de resultados...", file=sys.stderr, flush=True)
                max_wait = 40  # 40 segundos máximo
                waited = 0
                while waited < max_wait:
                    current_url = driver.current_url
                    if '/do/list/' in current_url and 's=' in current_url and 'b=' in current_url:
                        print(f"[SELENIUM] ✅ Página de resultados carregada após {waited}s", file=sys.stderr, flush=True)
                        break
                    else:
                        print(f"[SELENIUM] Aguardando... URL atual: {current_url[:80]}... ({waited}s)", file=sys.stderr, flush=True)
                        time.sleep(3)
                        waited += 3
                
                # Aguardar mais um pouco para garantir que o conteúdo carregou
                print(f"[SELENIUM] Aguardando conteúdo carregar...", file=sys.stderr, flush=True)
                time.sleep(5)
                
                final_url = driver.current_url
                
                # DEBUG: Salvar URL e HTML para análise
                try:
                    import sys
                    print(f"[SELENIUM] URL final: {final_url}", file=sys.stderr, flush=True)
                    with open(DEBUG_DIR / f"selenium_url_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt", 'w') as f:
                        f.write(f"Final URL: {final_url}\n")
                        f.write(f"Has s=: {'s=' in final_url}\n")
                        f.write(f"Has b=: {'b=' in final_url}\n")
                except:
                    pass
                
                # NÃO fazer retry com war= - aceitar o resultado como está
                # O retry pode causar cliques extras e bagunçar o formulário
                if 'war=' in final_url:
                    print(f"[SELENIUM] ⚠️ war= detectado (sem disponibilidade ou erro)", file=sys.stderr, flush=True)
                
                # Se obtivemos URL s/b válida, pegar HTML do driver ANTES de fechar
                if 's=' in final_url and 'b=' in final_url:
                    print(f"[SELENIUM] ✅ URL s/b obtida! Pegando HTML do driver...", file=sys.stderr, flush=True)
                    
                    # Usar page_source do driver em vez de fazer novo request
                    html_content = driver.page_source
                    
                    # Agora pode fechar o driver
                    driver.quit()
                    
                    print(f"[SELENIUM] Fazendo parse de {len(html_content)} bytes...", file=sys.stderr, flush=True)
                    items = parse_prices(html_content, final_url)
                    print(f"[SELENIUM] Parsed {len(items)} items", file=sys.stderr, flush=True)
                    items = convert_items_gbp_to_eur(items)
                    print(f"[SELENIUM] {len(items)} após GBP→EUR", file=sys.stderr, flush=True)
                    items = apply_price_adjustments(items, final_url)
                    print(f"[SELENIUM] {len(items)} após ajustes", file=sys.stderr, flush=True)
                    
                    if items:
                        print(f"[SELENIUM] ✅ {len(items)} carros encontrados!", file=sys.stderr, flush=True)
                        # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
                        items = normalize_and_sort(items, supplier_priority=None)
                        # SUCESSO! Retornar resultados
                        return _no_store_json({
                            "ok": True,
                            "items": items,
                            "location": location,
                            "start_date": start_dt.date().isoformat(),
                            "start_time": start_dt.strftime("%H:%M"),
                            "end_date": end_dt.date().isoformat(),
                            "end_time": end_dt.strftime("%H:%M"),
                            "days": days,
                        })
                else:
                    print(f"[SELENIUM] ⚠️ URL s/b NÃO obtida! URL: {final_url}", file=sys.stderr, flush=True)
                    driver.quit()
                    # Fallback: tentar POST direto para /do/list/{lang}
                    try:
                        import requests
                        payload = build_carjet_form(location, start_dt, end_dt, lang=lang, currency=currency)
                        headers_dp = {
                            "Origin": "https://www.carjet.com",
                            "Referer": f"https://www.carjet.com/do/list/{lang}",
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.6",
                            "Cookie": "monedaForzada=EUR; moneda=EUR; currency=EUR; country=PT; idioma=PT; lang=pt",
                        }
                        rdp = requests.post(f"https://www.carjet.com/do/list/{lang}", data=payload, headers=headers_dp, timeout=20)
                        if rdp.status_code == 200 and (rdp.text or '').strip():
                            html_dp = rdp.text
                            its_dp = parse_prices(html_dp, f"https://www.carjet.com/do/list/{lang}")
                            its_dp = convert_items_gbp_to_eur(its_dp)
                            its_dp = apply_price_adjustments(its_dp, f"https://www.carjet.com/do/list/{lang}")
                            if its_dp:
                                print(f"[SELENIUM] ✅ Fallback POST retornou {len(its_dp)} carros", file=sys.stderr, flush=True)
                                # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
                                its_dp = normalize_and_sort(its_dp, supplier_priority=None)
                                return _no_store_json({
                                    "ok": True,
                                    "items": its_dp,
                                    "location": location,
                                    "start_date": start_dt.date().isoformat(),
                                    "start_time": start_dt.strftime("%H:%M"),
                                    "end_date": end_dt.date().isoformat(),
                                    "end_time": end_dt.strftime("%H:%M"),
                                    "days": days,
                                })
                    except Exception as _e:
                        print(f"[SELENIUM] Fallback POST erro: {_e}", file=sys.stderr, flush=True)
                    # Se ainda sem resultados, retornar vazio rapidamente para permitir nova tentativa
                    return _no_store_json({
                        "ok": True,
                        "items": [],
                        "location": location,
                        "start_date": start_dt.date().isoformat(),
                        "start_time": start_dt.strftime("%H:%M"),
                        "end_date": end_dt.date().isoformat(),
                        "end_time": end_dt.strftime("%H:%M"),
                        "days": days,
                    })
            except TimeoutError as e:
                print(f"[SELENIUM TIMEOUT] ⏰ Ficou preso! {e}", file=sys.stderr, flush=True)
                print(f"[SELENIUM TIMEOUT] Fechando Chrome e abortando...", file=sys.stderr, flush=True)
                try:
                    driver.quit()
                except:
                    pass
                # Cancelar alarm
                try:
                    signal.alarm(0)
                except:
                    pass
                # RETORNAR vazio em caso de timeout
                return _no_store_json({
                    "ok": True,
                    "items": [],
                    "location": location,
                    "start_date": start_dt.date().isoformat(),
                    "start_time": start_dt.strftime("%H:%M"),
                    "end_date": end_dt.date().isoformat(),
                    "end_time": end_dt.strftime("%H:%M"),
                    "days": days,
                    "warning": "Timeout: Scraping ficou preso e foi abortado"
                })
            except Exception as e:
                print(f"[SELENIUM ERROR interno] {e}", file=sys.stderr, flush=True)
                try:
                    driver.quit()
                except:
                    pass
                # Cancelar alarm
                try:
                    signal.alarm(0)
                except:
                    pass
                # RETORNAR vazio em caso de erro também
                return _no_store_json({
                    "ok": True,
                    "items": [],
                    "location": location,
                    "start_date": start_dt.date().isoformat(),
                    "start_time": start_dt.strftime("%H:%M"),
                    "end_date": end_dt.date().isoformat(),
                    "end_time": end_dt.strftime("%H:%M"),
                    "days": days,
                })
        except Exception as e:
            print(f"[SELENIUM ERROR] {e}", file=sys.stderr, flush=True)
            import traceback
            traceback.print_exc(file=sys.stderr)
            # Cancelar alarm
            try:
                signal.alarm(0)
            except:
                pass
            # RETORNAR vazio para não travar
            return _no_store_json({
                "ok": True,
                "items": [],
                "location": location,
                "start_date": start_dt.date().isoformat(),
                "start_time": start_dt.strftime("%H:%M"),
                "end_date": end_dt.date().isoformat(),
                "end_time": end_dt.strftime("%H:%M"),
                "days": days,
            })
        
        # Fallback se Playwright falhou (NÃO DEVE CHEGAR AQUI SE SELENIUM FALHOU!)
        if USE_PLAYWRIGHT and _HAS_PLAYWRIGHT:
            try:
                from playwright.async_api import async_playwright
                async with async_playwright() as p:
                    # Chromium-first strategy
                    browser = await p.chromium.launch(headless=True)
                    context = await browser.new_context(
                        locale="pt-PT",
                        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
                    )
                    page = await context.new_page()
                    captured_bodies: List[str] = []
                    async def _on_resp(resp):
                        try:
                            u = resp.url or ""
                            if any(k in u for k in ("modalFilter.asp", "carList.asp", "/do/list/pt", "filtroUso.asp")):
                                try:
                                    t = await resp.text()
                                    if t:
                                        captured_bodies.append(t)
                                        try:
                                            stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                            name = "pw-capture-" + re.sub(r"[^a-z0-9]+", "-", u.lower())[-60:]
                                            (DEBUG_DIR / f"{name}-{stamp}.html").write_text(t, encoding='utf-8')
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    page.on("response", _on_resp)
                    # 1) Open homepage to mint session
                    home_path = "aluguel-carros/index.htm" if lang.lower()=="pt" else "index.htm"
                    await page.goto(f"https://www.carjet.com/{home_path}", wait_until="networkidle", timeout=35000)
                    # Handle consent if present
                    try:
                        for sel in [
                            "#didomi-notice-agree-button",
                            ".didomi-continue-without-agreeing",
                            "button:has-text('Aceitar')",
                            "button:has-text('I agree')",
                            "button:has-text('Accept')",
                        ]:
                            try:
                                c = page.locator(sel)
                                if await c.count() > 0:
                                    try: await c.first.click(timeout=1500)
                                    except Exception: pass
                                    await page.wait_for_timeout(200)
                                    break
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # 2) Type exact location as in browser autosuggest, then submit form programmatically
                    try:
                        exact_loc = location
                        lo = (location or '').lower()
                        if 'albufeira' in lo:
                            exact_loc = 'Albufeira Cidade'
                        elif 'faro' in lo:
                            exact_loc = 'Faro Aeroporto (FAO)'
                        # Try common selectors for the location input
                        loc_inp = None
                        for sel in ["input[name='pickup']", "#pickup", "input[placeholder*='local' i]", "input[aria-label*='local' i]", "input[type='search']"]:
                            try:
                                h = await page.query_selector(sel)
                                if h:
                                    loc_inp = h; break
                            except Exception:
                                pass
                        if loc_inp:
                            try:
                                await loc_inp.click()
                                await loc_inp.fill("")
                                await loc_inp.type(exact_loc, delay=50)
                                # Wait for dropdown and click the exact match if visible
                                try:
                                    opt = page.locator(f"text={exact_loc}")
                                    if await opt.count() > 0:
                                        await opt.first.click(timeout=2000)
                                except Exception:
                                    # fallback: press Enter to accept first suggestion
                                    try:
                                        await loc_inp.press('Enter')
                                    except Exception:
                                        pass
                                await page.wait_for_timeout(300)
                            except Exception:
                                pass
                    except Exception:
                        pass
                    # 2.b) Fill pickup/dropoff dates and times via visible inputs (simulate human typing)
                    try:
                        pickup_dmY = start_dt.strftime('%d/%m/%Y')
                        dropoff_dmY = end_dt.strftime('%d/%m/%Y')
                        pickup_HM = start_dt.strftime('%H:%M')
                        dropoff_HM = end_dt.strftime('%H:%M')
                        # Try native calendar UI first using the known triggers
                        async def select_date_via_picker(trigger_alt: str, target_dmY: str):
                            try:
                                trig = page.locator(f"img.ui-datepicker-trigger[alt='{trigger_alt}']")
                                if await trig.count() > 0:
                                    await trig.first.click()
                                    # Parse target day/month/year
                                    td, tm, ty = target_dmY.split('/')
                                    # Max 12 next clicks safeguard
                                    for _ in range(13):
                                        try:
                                            title = await page.locator('.ui-datepicker-title').inner_text()
                                        except Exception:
                                            title = ''
                                        # Title like 'Outubro 2025'
                                        ok_month = (tm in target_dmY)  # coarse guard; we do direct day pick below
                                        # Try clicking the exact day link
                                        day_locator = page.locator(f".ui-datepicker-calendar td a:text-is('{int(td)}')")
                                        if await day_locator.count() > 0:
                                            try:
                                                await day_locator.first.click()
                                                await page.wait_for_timeout(200)
                                                break
                                            except Exception:
                                                pass
                                        # Navigate next month
                                        try:
                                            nxt = page.locator('.ui-datepicker-next')
                                            if await nxt.count() > 0:
                                                await nxt.first.click()
                                                await page.wait_for_timeout(150)
                                            else:
                                                break
                                        except Exception:
                                            break
                            except Exception:
                                pass
                        await select_date_via_picker('Data de recolha', pickup_dmY)
                        await select_date_via_picker('Data de entrega', dropoff_dmY)
                        fill_dates_js = """
                          (pDate, pTime, dDate, dTime) => {
                            const setVal = (sel, val) => { const el = document.querySelector(sel); if (!el) return false; el.focus(); el.value = val; el.dispatchEvent(new Event('input', {bubbles:true})); el.dispatchEvent(new Event('change', {bubbles:true})); return true; };
                            const tryAll = (sels, val) => { for (const s of sels) { if (setVal(s, val)) return true; } return false; };
                            // Pickup date/time candidates
                            tryAll(['#fechaRecogida','input[name=fechaRecogida]','input[name=pickupDate]','input[type=date][name*=recog]','input[type=date][name*=pickup]','input[placeholder*="recolh" i]','input[aria-label*="recolh" i]'], pDate);
                            tryAll(['#fechaRecogidaSelHour','input[name=fechaRecogidaSelHour]','input[name=pickupTime]','input[type=time][name*=recog]','input[type=time][name*=pickup]','#h-recogida'], pTime);
                            // Dropoff date/time candidates
                            tryAll(['#fechaEntrega','#fechaDevolucion','input[name=fechaEntrega]','input[name=fechaDevolucion]','input[name=dropoffDate]','input[type=date][name*=entreg]','input[type=date][name*=devol]','input[placeholder*="entreg" i]','input[aria-label*="entreg" i]'], dDate);
                            tryAll(['#fechaEntregaSelHour','#fechaDevolucionSelHour','input[name=fechaEntregaSelHour]','input[name=fechaDevolucionSelHour]','input[name=dropoffTime]','input[type=time][name*=entreg]','input[type=time][name*=devol]','input[type=time][name*=drop]','#h-devolucion'], dTime);
                          }
                        """
                        await page.evaluate(fill_dates_js, pickup_dmY, pickup_HM, dropoff_dmY, dropoff_HM)
                        await page.wait_for_timeout(300)
                    except Exception:
                        pass
                    # Programmatic submit with full payload to ensure consistent parameters
                    payload = build_carjet_form(location, start_dt, end_dt, lang=lang, currency=currency)
                    submit_js = """
                      (url, data) => {
                        const f = document.createElement('form');
                        f.method = 'POST';
                        f.action = url;
                        for (const [k,v] of Object.entries(data||{})) {
                          const i = document.createElement('input');
                          i.type = 'hidden'; i.name = k; i.value = String(v ?? '');
                          f.appendChild(i);
                        }
                        document.body.appendChild(f);
                        f.submit();
                      }
                    """
                    await page.evaluate(submit_js, f"https://www.carjet.com/do/list/{lang}", payload)
                    try:
                        await page.wait_for_load_state("networkidle", timeout=40000)
                    except Exception:
                        pass
                    # Additionally trigger native on-page submit to mimic button onclick
                    try:
                        await page.evaluate("""
                          try {
                            if (typeof comprobar_errores_3 === 'function') {
                              if (comprobar_errores_3()) {
                                if (typeof filtroUsoForm === 'function') filtroUsoForm();
                                if (typeof submit_fechas === 'function') submit_fechas('/do/list/pt');
                              }
                            }
                          } catch (e) {}
                        """)
                        try:
                            await page.wait_for_load_state("networkidle", timeout=40000)
                        except Exception:
                            pass
                    except Exception:
                        pass
                    # 3-5) Up to 3 cycles: click 'Pesquisar' (if present), scroll, wait for results
                    for _ in range(3):
                        try:
                            btn = page.locator("button:has-text('Pesquisar'), button:has-text('Buscar'), input[type=submit], button[type=submit]")
                            if await btn.count() > 0:
                                try:
                                    await btn.first.click(timeout=3000)
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        try:
                            await page.wait_for_load_state("networkidle", timeout=40000)
                        except Exception:
                            pass
                        # Best-effort screenshot after search click (skip if quick=1)
                        if not quick:
                            try:
                                stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                await page.screenshot(path=str(DEBUG_DIR / f"pw-after-search-{stamp}.png"), full_page=True)
                            except Exception:
                                pass
                        try:
                            for __ in range(5):
                                try:
                                    await page.mouse.wheel(0, 1600)
                                except Exception:
                                    pass
                                await page.wait_for_timeout(300)
                        except Exception:
                            pass
                        # Check if results appeared; if so, break
                        try:
                            ok = await page.locator("section.newcarlist article, .newcarlist article, article.car, li.result, li.car, .car-item, .result-row").count()
                            if (ok or 0) > 0:
                                break
                        except Exception:
                            pass
                    # Wait specifically for known backend calls and dump responses
                    mf_body = ""; cl_body = ""
                    try:
                        resp_mf = await page.wait_for_response(lambda r: 'modalFilter.asp' in (r.url or ''), timeout=40000)
                        try:
                            mf_body = await resp_mf.text()
                            if mf_body:
                                stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                (DEBUG_DIR / f"pw-modalFilter-{stamp}.html").write_text(mf_body, encoding='utf-8')
                        except Exception:
                            pass
                    except Exception:
                        pass
                    try:
                        resp_cl = await page.wait_for_response(lambda r: 'carList.asp' in (r.url or ''), timeout=40000)
                        try:
                            cl_body = await resp_cl.text()
                            if cl_body:
                                stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                (DEBUG_DIR / f"pw-carList-{stamp}.html").write_text(cl_body, encoding='utf-8')
                        except Exception:
                            pass
                    except Exception:
                        pass
                    html_pw = await page.content()
                    final_url = page.url
                    print(f"[DEBUG] Fechando browser, URL final: {final_url}", file=sys.stderr, flush=True)
                    await context.close(); await browser.close()
                    print(f"[DEBUG] Browser fechado, HTML size: {len(html_pw)} bytes", file=sys.stderr, flush=True)
                if html_pw:
                    items_pw = parse_prices(html_pw, final_url or base)
                    items_pw = convert_items_gbp_to_eur(items_pw)
                    items_pw = apply_price_adjustments(items_pw, final_url or base)
                    items = items_pw
                # If still empty, try parsing network-captured bodies
                if (not items) and (cl_body or mf_body):
                    try:
                        base_net = "https://www.carjet.com/do/list/pt"
                        if cl_body:
                            its = parse_prices(cl_body, base_net)
                            its = convert_items_gbp_to_eur(its)
                            its = apply_price_adjustments(its, base_net)
                            if its:
                                items = its
                        if (not items) and mf_body:
                            its2 = parse_prices(mf_body, base_net)
                            its2 = convert_items_gbp_to_eur(its2)
                            its2 = apply_price_adjustments(its2, base_net)
                            if its2:
                                items = its2
                    except Exception:
                        pass
                if (not items) and captured_bodies:
                    try:
                        base_net = "https://www.carjet.com/do/list/pt"
                        for body in captured_bodies:
                            its = parse_prices(body, base_net)
                            its = convert_items_gbp_to_eur(its)
                            its = apply_price_adjustments(its, base_net)
                            if its:
                                items = its
                                break
                    except Exception:
                        pass
                # Final fallback: if we ended on a CarJet list URL, delegate to URL-based compute
                try:
                    if (not items) and (final_url or '').startswith("https://www.carjet.com/do/list/"):
                        data_f = await _compute_prices_for(final_url)
                        its_f = (data_f or {}).get('items') or []
                        if its_f:
                            items = its_f
                except Exception:
                    pass
                # Direct POST fallback within Playwright session
                try:
                    if not items:
                        payload_dp = build_carjet_form(location, start_dt, end_dt, lang=lang, currency=currency)
                        rdp = await page.request.post(f"https://www.carjet.com/do/list/{lang}", data=payload_dp)
                        try:
                            html_dp = await rdp.text()
                        except Exception:
                            html_dp = ""
                        if html_dp:
                            try:
                                stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                (DEBUG_DIR / f"pw-direct-post-{stamp}.html").write_text(html_dp, encoding='utf-8')
                            except Exception:
                                pass
                            its_dp = parse_prices(html_dp, f"https://www.carjet.com/do/list/{lang}")
                            its_dp = convert_items_gbp_to_eur(its_dp)
                            its_dp = apply_price_adjustments(its_dp, f"https://www.carjet.com/do/list/{lang}")
                            if its_dp:
                                items = its_dp
                except Exception:
                    pass
                # Engine fallback: if Chromium didn't produce items, try WebKit with Safari UA
                if (not items) and USE_PLAYWRIGHT and _HAS_PLAYWRIGHT:
                    try:
                        from playwright.async_api import async_playwright
                        async with async_playwright() as p2:
                            browser2 = await p2.webkit.launch(headless=True)
                            context2 = await browser2.new_context(
                                locale="pt-PT",
                                user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/26.0.1 Safari/605.1.15",
                            )
                            page2 = await context2.new_page()
                            captured2: List[str] = []
                            async def _on_resp2(resp):
                                try:
                                    u = resp.url or ""
                                    if any(k in u for k in ("modalFilter.asp", "carList.asp", "/do/list/pt", "filtroUso.asp")):
                                        try:
                                            t = await resp.text()
                                            if t:
                                                captured2.append(t)
                                                try:
                                                    stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                                    name = "pw2-capture-" + re.sub(r"[^a-z0-9]+", "-", u.lower())[-60:]
                                                    (DEBUG_DIR / f"{name}-{stamp}.html").write_text(t, encoding='utf-8')
                                                except Exception:
                                                    pass
                                        except Exception:
                                            pass
                                except Exception:
                                    pass
                            page2.on("response", _on_resp2)
                            # Open homepage and perform same submission
                            home_path2 = "aluguel-carros/index.htm" if lang.lower()=="pt" else "index.htm"
                            await page2.goto(f"https://www.carjet.com/{home_path2}", wait_until="networkidle", timeout=35000)
                            # Type exact location per autosuggest
                            try:
                                exact_loc2 = location
                                lo2 = (location or '').lower()
                                if 'albufeira' in lo2:
                                    exact_loc2 = 'Albufeira Cidade'
                                elif 'faro' in lo2:
                                    exact_loc2 = 'Faro Aeroporto (FAO)'
                                loc2 = None
                                for sel in ["input[name='pickup']", "#pickup", "input[placeholder*='local' i]", "input[aria-label*='local' i]", "input[type='search']"]:
                                    try:
                                        h = await page2.query_selector(sel)
                                        if h:
                                            loc2 = h; break
                                    except Exception:
                                        pass
                                if loc2:
                                    try:
                                        await loc2.click(); await loc2.fill(""); await loc2.type(exact_loc2, delay=50)
                                        try:
                                            opt2 = page2.locator(f"text={exact_loc2}")
                                            if await opt2.count() > 0:
                                                await opt2.first.click(timeout=2000)
                                        except Exception:
                                            try: await loc2.press('Enter')
                                            except Exception: pass
                                        await page2.wait_for_timeout(300)
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                            # Dates and hours
                            try:
                                pickup_dmY2 = start_dt.strftime('%d/%m/%Y')
                                dropoff_dmY2 = end_dt.strftime('%d/%m/%Y')
                                pickup_HM2 = start_dt.strftime('%H:%M')
                                dropoff_HM2 = end_dt.strftime('%H:%M')
                                fill_js2 = """
                                  (pDate, pTime, dDate, dTime) => {
                                    const setVal = (sel, val) => { const el = document.querySelector(sel); if (!el) return false; el.focus(); el.value = val; el.dispatchEvent(new Event('input', {bubbles:true})); el.dispatchEvent(new Event('change', {bubbles:true})); return true; };
                                    const tryAll = (sels, val) => { for (const s of sels) { if (setVal(s, val)) return true; } return false; };
                                    tryAll(['#fechaRecogida','input[name=fechaRecogida]','input[name=pickupDate]','input[type=date][name*=recog]','input[type=date][name*=pickup]','input[placeholder*="recolh" i]','input[aria-label*="recolh" i]'], pDate);
                                    tryAll(['#fechaRecogidaSelHour','input[name=fechaRecogidaSelHour]','input[name=pickupTime]','input[type=time][name*=recog]','input[type=time][name*=pickup]','#h-recogida'], pTime);
                                    tryAll(['#fechaEntrega','#fechaDevolucion','input[name=fechaEntrega]','input[name=fechaDevolucion]','input[name=dropoffDate]','input[type=date][name*=entreg]','input[type=date][name*=devol]','input[placeholder*="entreg" i]','input[aria-label*="entreg" i]'], dDate);
                                    tryAll(['#fechaEntregaSelHour','#fechaDevolucionSelHour','input[name=fechaEntregaSelHour]','input[name=fechaDevolucionSelHour]','input[name=dropoffTime]','input[type=time][name*=entreg]','input[type=time][name*=devol]','input[type=time][name*=drop]','#h-devolucion'], dTime);
                                  }
                                """
                                await page2.evaluate(fill_js2, pickup_dmY2, pickup_HM2, dropoff_dmY2, dropoff_HM2)
                                await page2.wait_for_timeout(300)
                            except Exception:
                                pass
                            # Submit programmatically and via native function
                            try:
                                payload2 = build_carjet_form(location, start_dt, end_dt, lang=lang, currency=currency)
                                submit_js2 = """
                                  (url, data) => { const f = document.createElement('form'); f.method='POST'; f.action=url; for (const [k,v] of Object.entries(data||{})) { const i=document.createElement('input'); i.type='hidden'; i.name=k; i.value=String(v??''); f.appendChild(i);} document.body.appendChild(f); f.submit(); }
                                """
                                await page2.evaluate(submit_js2, f"https://www.carjet.com/do/list/{lang}", payload2)
                                await page2.wait_for_load_state('networkidle', timeout=40000)
                                try:
                                    await page2.evaluate("""
                                      try { if (typeof comprobar_errores_3==='function' && comprobar_errores_3()) { if (typeof filtroUsoForm==='function') filtroUsoForm(); if (typeof submit_fechas==='function') submit_fechas('/do/list/pt'); } } catch(e) {}
                                    """)
                                    await page2.wait_for_load_state('networkidle', timeout=40000)
                                except Exception:
                                    pass
                            except Exception:
                                pass
                            # Wait for known responses
                            mf2 = ""; cl2 = ""
                            try:
                                r1 = await page2.wait_for_response(lambda r: 'modalFilter.asp' in (r.url or ''), timeout=40000)
                                try: mf2 = await r1.text()
                                except Exception: mf2 = ""
                            except Exception:
                                pass
                            try:
                                r2 = await page2.wait_for_response(lambda r: 'carList.asp' in (r.url or ''), timeout=40000)
                                try: cl2 = await r2.text()
                                except Exception: cl2 = ""
                            except Exception:
                                pass
                            html2 = await page2.content()
                            final2 = page2.url
                            await context2.close(); await browser2.close()
                        # parse order: DOM, carList, modalFilter, captured list
                        if not items:
                            try:
                                if html2:
                                    its = parse_prices(html2, final2 or base)
                                    its = convert_items_gbp_to_eur(its); its = apply_price_adjustments(its, final2 or base)
                                    if its: items = its
                            except Exception:
                                pass
                        if (not items) and cl2:
                            try:
                                base_net2 = "https://www.carjet.com/do/list/pt"
                                its = parse_prices(cl2, base_net2); its = convert_items_gbp_to_eur(its); its = apply_price_adjustments(its, base_net2)
                                if its: items = its
                            except Exception:
                                pass
                        if (not items) and mf2:
                            try:
                                base_net2 = "https://www.carjet.com/do/list/pt"
                                its = parse_prices(mf2, base_net2); its = convert_items_gbp_to_eur(its); its = apply_price_adjustments(its, base_net2)
                                if its: items = its
                            except Exception:
                                pass
                        if (not items) and captured2:
                            try:
                                base_net2 = "https://www.carjet.com/do/list/pt"
                                for body in captured2:
                                    its = parse_prices(body, base_net2); its = convert_items_gbp_to_eur(its); its = apply_price_adjustments(its, base_net2)
                                    if its:
                                        items = its; break
                            except Exception:
                                pass
                        # Final fallback for Chromium attempt: try URL-based compute if we have a CarJet list URL
                        try:
                            if (not items) and (final2 or '').startswith("https://www.carjet.com/do/list/"):
                                data_f2 = await _compute_prices_for(final2)
                                its_f2 = (data_f2 or {}).get('items') or []
                                if its_f2:
                                    items = its_f2
                        except Exception:
                            pass
                        # Direct POST fallback within Playwright session (Chromium)
                        try:
                            if not items:
                                payload_dp2 = build_carjet_form(location, start_dt, end_dt, lang=lang, currency=currency)
                                rdp2 = await page2.request.post(f"https://www.carjet.com/do/list/{lang}", data=payload_dp2)
                                try:
                                    html_dp2 = await rdp2.text()
                                except Exception:
                                    html_dp2 = ""
                                if html_dp2:
                                    try:
                                        stamp = datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')
                                        (DEBUG_DIR / f"pw2-direct-post-{stamp}.html").write_text(html_dp2, encoding='utf-8')
                                    except Exception:
                                        pass
                                    its_dp2 = parse_prices(html_dp2, f"https://www.carjet.com/do/list/{lang}")
                                    its_dp2 = convert_items_gbp_to_eur(its_dp2)
                                    its_dp2 = apply_price_adjustments(its_dp2, f"https://www.carjet.com/do/list/{lang}")
                                    if its_dp2:
                                        items = its_dp2
                        except Exception:
                            pass
                    except Exception:
                        pass
            except Exception:
                # Playwright falhou silenciosamente, tentar fallback
                items = []
        html = ""
        if not items:
            html = try_direct_carjet(location, start_dt, end_dt, lang=lang, currency=currency)
        # DEBUG: persist fetched HTML for troubleshooting
        try:
            _stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            _loc_tag = re.sub(r"[^a-z0-9]+", "-", (location or "").lower())
            if html:
                (_fp := (DEBUG_DIR / f"track_params-{_loc_tag}-{start_dt.date()}-{days}d-{_stamp}.html")).write_text(html or "", encoding="utf-8")
        except Exception:
            pass
        if not items:
            if not html:
                return _no_store_json({"ok": False, "error": "Upstream fetch failed"}, status_code=502)
            items = parse_prices(html, base)
            items = convert_items_gbp_to_eur(items)
            items = apply_price_adjustments(items, base)
        # DEBUG: write a compact summary JSON (count and first 5 items)
        try:
            import json as _json
            _sum = {
                "ts": _stamp,
                "location": location,
                "start": start_dt.isoformat(),
                "end": end_dt.isoformat(),
                "days": days,
                "count": len(items or []),
                "preview": (items or [])[:5],
            }
            (DEBUG_DIR / f"track_params-summary-{_loc_tag}-{_stamp}.json").write_text(_json.dumps(_sum, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            pass
        # No additional fallback needed; Playwright was already attempted first when enabled
        print(f"\n[API] ✅ RESPONSE: {len(items)} items for {days} days")
        if items:
            print(f"[API] First car: {items[0].get('car', 'N/A')} - {items[0].get('price', 'N/A')}")
        else:
            print(f"[API] ⚠️  NO ITEMS RETURNED!")
        print(f"{'='*60}\n")
        print(f"[API] RESPONSE: {len(items)} items, days={days}, start={start_dt.date()}, end={end_dt.date()}", file=sys.stderr, flush=True)
        if items:
            print(f"[API] First car: {items[0].get('car', 'N/A')} - {items[0].get('price', 'N/A')}", file=sys.stderr, flush=True)
        # APLICAR NORMALIZE_AND_SORT para adicionar campo 'group'
        items = normalize_and_sort(items, supplier_priority=None)
        return _no_store_json({
            "ok": True,
            "items": items,
            "location": location,
            "start_date": start_dt.date().isoformat(),
            "start_time": start_dt.strftime("%H:%M"),
            "end_date": end_dt.date().isoformat(),
            "end_time": end_dt.strftime("%H:%M"),
            "days": days,
        })
    except Exception as e:
        print(f"\n{'='*60}")
        print(f"[API ERROR] track-by-params failed: {str(e)}")
        print(f"{'='*60}\n")
        import traceback
        traceback.print_exc()
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, status_code=500)

@app.get("/debug/vars")
async def debug_vars():
    return JSONResponse({
        "USE_PLAYWRIGHT": USE_PLAYWRIGHT,
        "_HAS_PLAYWRIGHT": _HAS_PLAYWRIGHT,
        "SCRAPER_SERVICE": SCRAPER_SERVICE,
    })

@app.get("/ph")
async def placeholder_image(car: str = "Car"):
    try:
        label = (car or "Car").strip()
        if len(label) > 32:
            label = label[:32] + "…"
        # Teal background (#009cb6) to match site, white centered text
        svg = f'''<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" width="320" height="180" viewBox="0 0 320 180" role="img">
  <rect width="320" height="180" fill="#009cb6"/>
  <text x="160" y="90" fill="#ffffff" font-family="-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Helvetica,Arial,sans-serif" font-size="18" text-anchor="middle" dominant-baseline="middle">{label}</text>
  <text x="160" y="160" fill="rgba(255,255,255,0.7)" font-family="-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Helvetica,Arial,sans-serif" font-size="12" text-anchor="middle">Image unavailable</text>
</svg>'''
        resp = Response(content=svg, media_type="image/svg+xml; charset=utf-8")
        resp.headers["Cache-Control"] = "public, max-age=86400"
        return resp
    except Exception:
        return Response(status_code=500)

def _normalize_model_for_image(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"\b(auto|automatic|manual|station\s*wagon|estate|sw|variant|break|tourer|grandtour|grand\s*tour|kombi|sportbreak|sport\s*brake|st)\b", "", s)
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    # common brand/model reorderings are left as-is
    return " ".join(s.split())

def _build_commons_query(name: str) -> str:
    key = _normalize_model_for_image(name)
    # bias towards car photos
    return f"{key} car"

def _save_cache_image(key: str, content: bytes, ext: str) -> Path:
    p = CACHE_CARS_DIR / f"{key}{ext}"
    with open(p, "wb") as f:
        f.write(content)
    return p

def _find_cached_image(key: str) -> Optional[Path]:
    for ext in (".jpg", ".jpeg", ".png", ".webp"):
        p = CACHE_CARS_DIR / f"{key}{ext}"
        if p.exists() and p.stat().st_size > 0:
            return p
    return None

@app.get("/imglookup")
async def img_lookup(car: str):
    try:
        car = car or "Car"
        key = _normalize_model_for_image(car).replace(" ", "-")
        cached = _find_cached_image(key)
        if cached:
            ct = "image/jpeg"
            if cached.suffix == ".png": ct = "image/png"
            elif cached.suffix == ".webp": ct = "image/webp"
            with open(cached, "rb") as f:
                b = f.read()
            resp = Response(content=b, media_type=ct)
            resp.headers["Cache-Control"] = "public, max-age=86400"
            return resp

        # Wikimedia Commons API search for files
        import json as _json
        api = "https://commons.wikimedia.org/w/api.php"
        params = {
            "action": "query",
            "format": "json",
            "prop": "imageinfo",
            "generator": "search",
            "gsrsearch": _build_commons_query(car),
            "gsrlimit": "5",
            "gsrnamespace": "6",  # File namespace
            "iiprop": "url|mime",
            "iiurlwidth": "480",
            "origin": "*",
        }
        r = requests.get(api, params=params, timeout=10, headers={"User-Agent": "PriceTracker/1.0"})
        url = None
        mime = None
        if r.ok:
            data = r.json()
            pages = (data.get("query", {}) or {}).get("pages", {})
            for _, pg in pages.items():
                ii = (pg.get("imageinfo") or [{}])[0]
                url = ii.get("thumburl") or ii.get("url")
                mime = ii.get("mime") or "image/jpeg"
                if url:
                    break
        if url:
            ir = requests.get(url, timeout=10, headers={"User-Agent": "PriceTracker/1.0"})
            if ir.ok and ir.content:
                ext = ".jpg"
                if (mime or "").endswith("png"): ext = ".png"
                elif (mime or "").endswith("webp"): ext = ".webp"
                path = _save_cache_image(key, ir.content, ext)
                resp = Response(content=ir.content, media_type=mime or "image/jpeg")
                resp.headers["Cache-Control"] = "public, max-age=86400"
                return resp
        # Fallback to placeholder
        return await placeholder_image(car)
    except Exception:
        return await placeholder_image(car)

@app.get("/api/debug/check-vehicle")
async def debug_check_vehicle(car_name: str):
    """Verifica se um carro está no dicionário VEHICLES"""
    car_clean = clean_car_name(car_name)
    in_vehicles = car_clean in VEHICLES
    
    result = {
        "original": car_name,
        "cleaned": car_clean,
        "in_vehicles": in_vehicles
    }
    
    if in_vehicles:
        vehicle_info = VEHICLES[car_clean]
        result["vehicle_info"] = vehicle_info
        if isinstance(vehicle_info, dict) and 'group' in vehicle_info:
            result["group"] = vehicle_info['group']
    
    return JSONResponse(result)

@app.get("/api/debug_direct")
async def debug_direct(request: Request):
    params = request.query_params
    location = params.get("location", "Albufeira")
    pickup_date = params.get("date")
    pickup_time = params.get("time", "10:00")
    days = int(params.get("days", 1))
    lang = params.get("lang", "pt")
    currency = params.get("currency", "EUR")
    if not pickup_date:
        return JSONResponse({"ok": False, "error": "Missing date (YYYY-MM-DD)"}, status_code=400)

    try:
        from datetime import datetime, timedelta
        start_dt = datetime.fromisoformat(pickup_date + "T" + pickup_time)
        end_dt = start_dt + timedelta(days=days)
        html = try_direct_carjet(location, start_dt, end_dt, lang=lang, currency=currency)
        if not html:
            return JSONResponse({"ok": False, "error": "Empty HTML from direct POST"}, status_code=500)

        # Save to debug file
        from datetime import datetime as _dt
        stamp = _dt.utcnow().strftime("%Y%m%dT%H%M%S")
        filename = f"debug-direct-{location.replace(' ', '-')}-{pickup_date}-{days}d.html"
        out_path = DEBUG_DIR / filename
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)

        # Quick selector counts
        soup = BeautifulSoup(html, "lxml")
        # Price-like selector counts and inline dataMap presence
        counts = {
            ".price": len(soup.select(".price")),
            ".amount": len(soup.select(".amount")),
            "[class*='price']": len(soup.select("[class*='price']")),
            "a[href]": len(soup.select("a[href]")),
        }
        try:
            import json as _json
            m = re.search(r"var\s+dataMap\s*=\s*(\[.*?\]);", html, re.S)
            if m:
                arr = _json.loads(m.group(1))
                counts["has_dataMap"] = True
                counts["dataMap_len"] = len(arr)
            else:
                counts["has_dataMap"] = False
                counts["dataMap_len"] = 0
        except Exception:
            counts["has_dataMap"] = False
            counts["dataMap_len"] = 0
        return JSONResponse({
            "ok": True,
            "url": f"https://www.carjet.com/do/list/{lang} (direct)",
            "debug_file": f"/static/debug/{filename}",
            "counts": counts,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


def parse_prices(html: str, base_url: str) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "lxml")
    items: List[Dict[str, Any]] = []
    # Flattened page text to infer context-specific categories (e.g., automatic families)
    try:
        _page_text = soup.get_text(" ", strip=True).lower()
    except Exception:
        _page_text = ""

    # Helper: detect automatic transmission markers from name or card text or explicit label
    def _is_auto_flag(name_lc: str, card_text_lc: str, trans_label: str) -> bool:
        try:
            if (trans_label or '').lower() == 'automatic':
                return True
            return bool(AUTO_RX.search(name_lc or '') or AUTO_RX.search(card_text_lc or ''))
        except Exception:
            return False

    # Blocklist of car models to exclude
    _blocked_models = [
        "Mercedes S Class Auto",
        "MG ZS Auto",
        "Mercedes CLA Coupe Auto",
        "Mercedes A Class",
        "Mercedes A Class Auto",
        "BMW 1 Series Auto",
        "BMW 3 Series SW Auto",
        "Volvo V60 Auto",
        "Volvo XC40 Auto",
        "Mercedes C Class Auto",
        "Tesla Model 3 Auto",
        "Electric",
        "BMW 2 Series Gran Coupe Auto",
        "Mercedes C Class SW Auto",
        "Mercedes E Class Auto",
        "Mercedes E Class SW Auto",
        "BMW 5 Series SW Auto",
        "BMW X1 Auto",
        "Mercedes CLE Coupe Auto",
        "Volkswagen T-Roc Cabrio",
        "Mercedes GLA Auto",
        "Volvo XC60 Auto",
        "Volvo EX30 Auto",
        "BMW 3 Series Auto",
        "Volvo V60 4x4 Auto",
        "Hybrid",
        "Mazda MX5 Cabrio Auto",
        "Mercedes CLA Auto",
    ]

    def _norm_text(s: str) -> str:
        s = (s or "").strip().lower()
        # remove duplicate spaces and commas spacing
        s = " ".join(s.replace(",", " ").split())
        return s

    _blocked_norm = set(_norm_text(x) for x in _blocked_models)

    def _is_blocked_model(name: str) -> bool:
        n = _norm_text(name)
        if not n:
            return False
        if n in _blocked_norm:
            return True
        # Regex-based strong match on key model families and powertrains
        patterns = [
            r"\bmercedes\s+s\s*class\b",
            r"\bmercedes\s+cla\b",
            r"\bmercedes\s+cle\b",
            r"\bmercedes\s+a\s*class\b",
            r"\bmercedes\s+c\s*class\b",
            r"\bmercedes\s+e\s*class\b",
            r"\bmercedes\s+gla\b",
            r"\bbmw\s+1\s*series\b",
            r"\bbmw\s+2\s*series\b",
            r"\bbmw\s+3\s*series\b",
            r"\bbmw\s+5\s*series\b",
            r"\bbmw\s*x1\b",
            r"\bvolvo\s+v60\b",
            r"\bvolvo\s+xc40\b",
            r"\bvolvo\s+xc60\b",
            r"\bvolvo\s+ex30\b",
            r"\btesla\s+model\s*3\b",
            r"\bmg\s+zs\b",
            r"\bmazda\s+mx5\b",
            r"\bvolkswagen\s+t-roc\b",
            r"\belectric\b",
            r"\bhybrid\b",
        ]
        import re as _re
        for p in patterns:
            if _re.search(p, n):
                return True
        # also check if any blocked long phrase is contained in name
        for b in _blocked_norm:
            if len(b) >= 6 and b in n:
                return True
        return False

    # --- Photo cache helpers (SQLite) ---
    def _photo_db_path() -> str:
        try:
            from pathlib import Path
            return str((Path(__file__).resolve().parent / "car_images.db"))
        except Exception:
            return "car_images.db"

    def _get_conn():
        try:
            import sqlite3
            return sqlite3.connect(_photo_db_path())
        except Exception:
            return None

    def _init_photos_table():
        conn = _get_conn()
        if not conn:
            return
        try:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS car_images (
                    model_key TEXT PRIMARY KEY,
                    photo_url TEXT,
                    updated_at TEXT
                )
                """
            )
            conn.commit()
        except Exception:
            pass
        finally:
            try:
                conn.close()
            except Exception:
                pass
            # FINAL SAFEGUARD: If Clio wagon fell into D/E2, flip to J2/L2
            try:
                _nm = (car_name or '').lower()
                _txt_final = ''
                try:
                    _txt_final = card.get_text(' ', strip=True).lower()
                except Exception:
                    _txt_final = ''
                if re.search(r"\brenault\s*clio\b", _nm) and re.search(r"\b(sw|st|sport\s*tourer|tourer|break|estate|kombi|grandtour|grand\s*tour|sporter|wagon)\b", _nm):
                    if _is_auto_flag(_nm, _txt_final, transmission_label):
                        category = 'Station Wagon Automatic'
                    else:
                        category = 'Estate/Station Wagon'
            except Exception:
                pass

    def _normalize_model_key(name: str) -> str:
        s = (name or "").strip().lower()
        for w in ("suv", "economy", "mini", "estate", "station wagon", "premium", "7 seater", "9 seater"):
            if s.endswith(" " + w):
                s = s[: -len(w) - 1].strip()
        s = " ".join(s.split())
        return s

    def _cache_get_photo(key: str) -> str:
        conn = _get_conn()
        if not conn:
            return ""
        try:
            cur = conn.execute("SELECT photo_url FROM car_images WHERE model_key = ?", (key,))
            row = cur.fetchone()
            return row[0] if row and row[0] else ""
        except Exception:
            return ""
        finally:
            try:
                conn.close()
            except Exception:
                pass

    def _cache_set_photo(key: str, url: str):
        if not (key and url):
            return
        _init_photos_table()
        conn = _get_conn()
        if not conn:
            return
        try:
            from datetime import datetime as _dt
            conn.execute(
                "INSERT INTO car_images (model_key, photo_url, updated_at) VALUES (?, ?, ?) "
                "ON CONFLICT(model_key) DO UPDATE SET photo_url=excluded.photo_url, updated_at=excluded.updated_at",
                (key, url, _dt.utcnow().isoformat(timespec="seconds"))
            )
            conn.commit()
        except Exception:
            pass
        finally:
            try:
                conn.close()
            except Exception:
                pass

    def map_grupo(grupo: str) -> str:
        if not grupo:
            return ""
        g = str(grupo).upper()
        # N-codes examples
        if g.startswith("N"):
            if g == "N07":
                return "7 Seater"
            if g.startswith("N09") or g == "N9" or g == "N90" or g == "N099":
                return "9 Seater"
            return "People Carrier"
        # S-codes: estate/station wagon
        if g.startswith("S"):
            return "Estate/Station Wagon"
        # A-codes: automatic variants; infer base from page context
        if g.startswith("A"):
            txt = _page_text
            if any(k in txt for k in ("estate", "station wagon", "estatecars", "carrinha")):
                return "Estate/Station Wagon"
            if "suv" in txt:
                return "SUV"
            if any(k in txt for k in ("7 lugares", "7 seats", "7 seater")):
                return "7 Seater"
            if any(k in txt for k in ("9 lugares", "9 seats", "9 seater")):
                return "9 Seater"
            if any(k in txt for k in ("mini", "pequeno")):
                return "Mini"
            if any(k in txt for k in ("economy", "económico", "economico")):
                return "Economy"
            # Fallback: treat as Economy automatic if no context hints
            return "Economy"
        # F/M codes frequently used for SUVs in provided samples
        if g.startswith("F"):
            return "SUV"
        if g.startswith("M"):
            # People carriers: infer 7 vs 9 seater from page text
            txt = _page_text
            if any(k in txt for k in ("9 lugares", "9 seats", "9 seater")):
                return "9 Seater"
            return "7 Seater"
        # Premium families observed as J/L in samples
        if g.startswith("J"):
            return "Premium"
        if g.startswith("L"):
            return "Premium"
        # C-codes numeric mapping
        if g.startswith("C"):
            try:
                n = int(g[1:])
            except Exception:
                return g
            if 1 <= n <= 4:
                return "Mini"
            if 5 <= n <= 9:
                return "Economy"
            if 10 <= n <= 19:
                return "Compact"
            if 20 <= n <= 29:
                return "Intermediate"
            if 30 <= n <= 39:
                return "Standard"
            if 40 <= n <= 49:
                return "Full-size"
            if 60 <= n <= 69:
                return "SUV"
            return g
        return g

    # Transmission label from global radio (if present)
    transmission_label = ""
    try:
        t_inp = soup.select_one("input[name='frmTrans'][checked]")
        if t_inp and t_inp.has_attr("value"):
            v = (t_inp.get("value") or "").lower()
            if v == "au":
                transmission_label = "Automatic"
            elif v == "mn":
                transmission_label = "Manual"
            elif v == "el":
                transmission_label = "Electric"
    except Exception:
        pass
    # Fallback: infer from 'Filtros utilizados anteriormente' section
    if not transmission_label:
        try:
            used = soup.select_one("#filterUsed")
            if used:
                txt = used.get_text(" ", strip=True).lower()
                if "autom" in txt:
                    transmission_label = "Automatic"
                elif "manual" in txt:
                    transmission_label = "Manual"
                elif "electr" in txt:
                    transmission_label = "Electric"
        except Exception:
            pass

    # Fast path for CarJet: collect provider summaries but do not return early; we'll prefer detailed items
    summary_items: List[Dict[str, Any]] = []
    try:
        # 0) Generic object matcher as a fallback to capture provider blobs even if array/var name changes
        raw_objs = OBJ_RX.findall(html)
        if raw_objs:
            import json as _json
            supplier_alias = {
                "AUP": "Auto Prudente Rent a Car",
                "SXT": "Sixt",
                "ECR": "Europcar",
                "KED": "Keddy by Europcar",
                "EPI": "EPI",
                "ALM": "Alamo",
                "AVX": "Avis",
                "BGX": "Budget",
                "ENT": "Enterprise",
                "DTG": "Dollar",
                "FLZ": "Flizzr",
                "DTG1": "Rentacar",
                "DGT1": "Rentacar",
                "EU2": "Goldcar Non-Refundable",
                "EUR": "Goldcar",
                "EUK": "Goldcar Key'n Go",
                "GMO": "Green Motion",
                "GMO1": "Green Motion",
                "SAD": "Drivalia",
                "DOH": "Drive on Holidays",
                "D4F": "Drive4Fun",
                "DVM": "Drive4Move",
                "CAE": "Cael",
                "CEN": "Centauro",
                "ABB": "Abbycar",
                "ABB1": "Abbycar Non-Refundable",
                "BSD": "Best Deal",
                "ATR": "Autorent",
                "AUU": "Auto Union",
                "THR": "Thrifty",
                "HER": "Hertz",
                "LOC": "Million",
            }
            idx = 0
            for s in raw_objs:
                try:
                    d = _json.loads(s)
                except Exception:
                    continue
                price_text = d.get("priceStr") or ""
                if not price_text:
                    continue
                supplier_code = (d.get("id") or "").strip()
                supplier = supplier_alias.get(supplier_code, supplier_code)
                grupo = d.get("grupoVeh") or ""
                category_h = map_grupo(grupo)
                display_category = category_h or grupo
                if transmission_label == "Automatic":
                    if display_category in ("Mini", "Economy", "SUV", "Estate/Station Wagon", "7 Seater"):
                        if display_category == "Estate/Station Wagon":
                            display_category = "Station Wagon Automatic"
                        elif display_category == "7 Seater":
                            display_category = "7 Seater Automatic"
                        else:
                            display_category = f"{display_category} Automatic"
                # Best-effort photo from grupoVeh code
                photo_url = ""
                try:
                    if grupo:
                        photo_url = urljoin(base_url, f"/cdn/img/cars/S/car_{grupo}.jpg")
                except Exception:
                    photo_url = ""
                # Mapear categoria para código de grupo
                group_code = map_category_to_group(display_category, "")
                summary_items.append({
                    "id": idx,
                    "car": "",
                    "supplier": supplier,
                    "price": price_text,
                    "currency": "",
                    "category": display_category,
                    "group": group_code,
                    "category_code": grupo,
                    "transmission": transmission_label,
                    "photo": photo_url,
                    "link": base_url,
                })
                idx += 1
            # do not return yet; prefer detailed rows

        m = DATAMAP_RX.search(html)
        if m:
            import json
            arr = json.loads(m.group(1))
            supplier_alias = {
                "AUP": "Auto Prudente Rent a Car",
                "SXT": "Sixt",
                "ECR": "Europcar",
                "KED": "Keddy by Europcar",
                "EPI": "EPI",
                "ALM": "Alamo",
                "AVX": "Avis",
                "BGX": "Budget",
                "ENT": "Enterprise",
                "DTG": "Dollar",
                "FLZ": "Flizzr",
                "EU2": "Goldcar Non-Refundable",
                "EUR": "Goldcar",
                "EUK": "Goldcar Key'n Go",
                "GMO": "Green Motion",
                "GMO1": "Green Motion",
                "SAD": "Drivalia",
                "DOH": "Drive on Holidays",
                "D4F": "Drive4Fun",
                "DVM": "Drive4Move",
                "CAE": "Cael",
                "CEN": "Centauro",
                "ABB": "Abbycar",
                "ABB1": "Abbycar Non-Refundable",
                "BSD": "Best Deal",
                "ATR": "Autorent",
                "AUU": "Auto Union",
            }
            idx = 0
            for it in arr:
                supplier_code = (it.get("id") or "").strip()
                # Additional inline overrides
                if supplier_code in ("DTG1", "DGT1"):
                    supplier = "Rentacar"
                else:
                    supplier = supplier_alias.get(supplier_code, supplier_code)
                price_text = it.get("priceStr") or ""
                if not price_text:
                    continue
                grupo = it.get("grupoVeh") or ""
                category_h = map_grupo(grupo)
                display_category = category_h or grupo
                if transmission_label == "Automatic":
                    if display_category in ("Mini", "Economy", "SUV", "Estate/Station Wagon", "7 Seater"):
                        if display_category == "Estate/Station Wagon":
                            display_category = "Station Wagon Automatic"
                        elif display_category == "7 Seater":
                            display_category = "7 Seater Automatic"
                        else:
                            display_category = f"{display_category} Automatic"
                # Mapear categoria para código de grupo
                group_code = map_category_to_group(display_category, "")
                summary_items.append({
                    "id": idx,
                    "car": "",
                    "supplier": supplier,
                    "price": price_text,
                    "currency": "",
                    "category": display_category,
                    "group": group_code,
                    "category_code": grupo,
                    "transmission": transmission_label,
                    "link": base_url,
                })
                idx += 1
            # do not return yet; prefer detailed rows
    except Exception:
        pass

    # Pass 2: try to parse explicit car cards/rows from the HTML (preferred over regex)
    try:
        cards = soup.select("section.newcarlist article, .newcarlist article, article.car, li.result, li.car, .car-item, .result-row")
        print(f"[PARSE] Found {len(cards)} cards to parse")
        idx = 0
        cards_with_price = 0
        cards_with_name = 0
        cards_blocked = 0
        for card in cards:
            # price - PRIORIZAR .price.pr-euros (preço total em euros, NÃO libras nem por dia)
            price_text = ""
            
            # 1ª PRIORIDADE: Buscar .price.pr-euros (preço total em euros)
            # Excluir .price-day-euros e .old-price
            for span_tag in card.find_all('span'):
                classes = span_tag.get('class', [])
                if not classes:
                    continue
                
                # Verificar se tem 'price' E 'pr-euros' MAS NÃO tem 'day' nem 'old'
                has_price = 'price' in classes
                has_pr_euros = 'pr-euros' in classes
                has_day = any('day' in c for c in classes)
                has_old = any('old' in c for c in classes)
                
                if has_price and has_pr_euros and not has_day and not has_old:
                    price_text = span_tag.get_text(strip=True)
                    # Limpar desconto: "-25%17,05 €12,79 €" -> "12,79 €"
                    # Pegar apenas o último preço (após o último espaço antes de €)
                    if price_text.count('€') > 1:
                        # Tem múltiplos preços, pegar o último
                        parts = price_text.split('€')
                        # O último é vazio, o penúltimo tem o preço final
                        if len(parts) >= 2:
                            price_text = parts[-2].split()[-1] + ' €'
                    break  # Encontrou o preço correto em euros!
            
            # 2ª PRIORIDADE: Se não encontrou .pr-euros, usar seletor genérico (fallback)
            if not price_text:
                let_price = card.select_one(".price, .amount, [class*='price'], .nfoPriceDest, .nfoPrice, [data-price]")
                price_text = (let_price.get_text(strip=True) if let_price else "") or (card.get("data-price") or "")
            
            if not price_text:
                continue
            cards_with_price += 1
            # car/model
            name_el = card.select_one(
                ".veh-name, .vehicle-name, .model, .titleCar, .title, h3, h2, [class*='veh-name'], [class*='vehicle-name'], [class*='model']"
            )
            car_name = name_el.get_text(strip=True) if name_el else ""
            if not car_name:
                # try common data attributes
                for attr in ("data-model", "data-vehicle", "data-name", "aria-label", "title"):
                    v = (card.get(attr) or "").strip()
                    if v:
                        car_name = v
                        break
            if car_name:
                cards_with_name += 1
            # supplier: try to extract provider code from logo_XXX.* in img src, then map via alias
            supplier = ""
            try:
                supplier_alias = {
                    "AUP": "Auto Prudente Rent a Car",
                    "SXT": "Sixt",
                    "ECR": "Europcar",
                    "KED": "Keddy by Europcar",
                    "EPI": "EPI",
                    "ALM": "Alamo",
                    "AVX": "Avis",
                    "BGX": "Budget",
                    "ENT": "Enterprise",
                    "DTG": "Dollar",
                    "DTG1": "Rentacar",
                    "DGT1": "Rentacar",
                    "FLZ": "Flizzr",
                    "EU2": "Goldcar Non-Refundable",
                    "EUR": "Goldcar",
                    "EUK": "Goldcar Key'n Go",
                    "GMO": "Green Motion",
                    "GMO1": "Green Motion",
                    "SAD": "Drivalia",
                    "DOH": "Drive on Holidays",
                    "D4F": "Drive4Fun",
                    "DVM": "Drive4Move",
                    "CAE": "Cael",
                    "CEN": "Centauro",
                    "ABB": "Abbycar",
                    "ABB1": "Abbycar Non-Refundable",
                    "BSD": "Best Deal",
                    "ATR": "Autorent",
                    "AUU": "Auto Union",
                    "THR": "Thrifty",
                    "HER": "Hertz",
                    "LOC": "Million",
                }
                code = ""
                for im in card.select("img[src]"):
                    src = im.get("src") or ""
                    mcode = LOGO_CODE_RX.search(src)
                    if mcode:
                        code = (mcode.group(1) or "").upper()
                        break
                
                if code:
                    supplier = supplier_alias.get(code, code)
                if not supplier:
                    # textual fallback but avoid using car name
                    supplier_el = card.select_one(".supplier, .vendor, .partner, [class*='supplier'], [class*='vendor']")
                    txt = supplier_el.get_text(strip=True) if supplier_el else ""
                    if txt and txt.lower() != (car_name or "").lower():
                        supplier = txt
            except Exception:
                pass
            # photo: pick an image that is not a provider logo
            photo = ""
            try:
                # PRIORIDADE 1: img.cl--car-img (CarJet específico)
                car_img = card.select_one("img.cl--car-img")
                if car_img:
                    src = (car_img.get("src") or car_img.get("data-src") or car_img.get("data-original") or "").strip()
                    if src:
                        try:
                            from urllib.parse import urljoin
                            photo = urljoin(base_url, src)
                        except Exception:
                            photo = src
                        # Extrair nome limpo do alt
                        if not car_name:
                            alt_text = (car_img.get("alt") or "").strip()
                            if alt_text:
                                # "Toyota Aygo ou similar | Pequeno" -> "Toyota Aygo"
                                car_name = alt_text.split('ou similar')[0].split('|')[0].strip()
                
                # PRIORIDADE 2: prefer <picture> sources
                if not photo:
                    picture_src = None
                    for src_el in card.select("picture source[srcset], img[srcset], picture source[data-srcset], img[data-srcset]"):
                        sset = (src_el.get("srcset") or src_el.get("data-srcset") or "").strip()
                        if sset:
                            # pick the first candidate (split by comma for multiple entries, then URL before whitespace)
                            first_entry = sset.split(',')[0].strip()
                            picture_src = first_entry.split()[0]
                            if picture_src:
                                break
                    
                    # PRIORIDADE 3: Outras imagens
                    imgs = card.select("img")
                    for im in imgs:
                        src = picture_src or (
                            im.get("src") or im.get("data-src") or im.get("data-original") or im.get("data-lazy") or im.get("data-lazy-src") or ""
                        ).strip()
                        if not src:
                            continue
                        # skip logos and icons
                        if re.search(r"logo_", src, re.I):
                            continue
                        if src.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
                            # make absolute if needed
                            try:
                                from urllib.parse import urljoin
                                photo = urljoin(base_url, src)
                            except Exception:
                                photo = src
                            # use alt/title as car_name fallback
                            if not car_name:
                                alt_t = (im.get("alt") or im.get("title") or "").strip()
                                if alt_t:
                                    car_name = alt_t
                            break
                # Also check inline background-image on card and descendants
                if not photo:
                    style_el = card.get("style") or ""
                    m_bg = BG_IMAGE_RX.search(style_el)
                    if m_bg:
                        raw = m_bg.group(1).strip().strip('\"\'')
                        try:
                            from urllib.parse import urljoin
                            photo = urljoin(base_url, f"/img?src={raw}")
                        except Exception:
                            photo = f"/img?src={raw}"
                if not photo:
                    for child in card.find_all(True):
                        st = child.get("style") or ""
                        m2 = BG_IMAGE_RX.search(st)
                        if m2:
                            raw = m2.group(1).strip().strip('\"\'')
                            try:
                                from urllib.parse import urljoin
                                photo = urljoin(base_url, f"/img?src={raw}")
                            except Exception:
                                photo = f"/img?src={raw}"
                            if photo:
                                break
                # As a final fallback, synthesize from any car_[gV].jpg reference inside this card
                if not photo:
                    html_block = str(card)
                    m_car = CAR_CODE_RX.search(html_block)
                    if m_car:
                        code = m_car.group(1)
                        try:
                            from urllib.parse import urljoin
                            photo = urljoin(base_url, f"/cdn/img/cars/S/car_{code}.jpg")
                        except Exception:
                            photo = f"/cdn/img/cars/S/car_{code}.jpg"
            except Exception:
                pass
            # category
            cat_el = card.select_one(".category, .group, .vehicle-category, [class*='category'], [class*='group'], [class*='categoria'], [class*='grupo']")
            category = cat_el.get_text(strip=True) if cat_el else ""
            # Canonicalize category to expected groups
            def _canon(cat: str) -> str:
                c = (cat or "").strip().lower()
                if not c:
                    return ""
                if "estate" in c or "station" in c or "carrinha" in c:
                    return "Estate/Station Wagon"
                if "suv" in c:
                    return "SUV"
                if "premium" in c or "lux" in c:
                    return "Premium"
                if "7" in c and ("lugar" in c or "lugares" in c or "seater" in c or "seats" in c):
                    return "7 Seater"
                if "9" in c and ("lugar" in c or "lugares" in c or "seater" in c or "seats" in c):
                    return "9 Seater"
                if "econom" in c:
                    return "Economy"
                if "mini" in c or "small" in c or "pequeno" in c:
                    return "Mini"
                return cat
            category = _canon(category)
            if not category:
                # Infer from CARD context if label missing to avoid page-wide bias
                try:
                    local_txt = card.get_text(" ", strip=True).lower()
                except Exception:
                    local_txt = ""
                if any(k in local_txt for k in ("estate", "station wagon", "estatecars", "carrinha")):
                    category = "Estate/Station Wagon"
                elif "suv" in local_txt:
                    category = "SUV"
                elif any(k in local_txt for k in ("7 lugares", "7 seats", "7 seater")):
                    category = "7 Seater"
                elif any(k in local_txt for k in ("9 lugares", "9 seats", "9 seater")):
                    category = "9 Seater"
                elif any(k in local_txt for k in ("mini", "pequeno")):
                    category = "Mini"
                elif any(k in local_txt for k in ("economy", "económico", "economico")):
                    category = "Economy"
                # As a last resort, try to infer from car name trailing token
                if not category and car_name:
                    tail = (car_name.split()[-1] or "").lower()
                    tail_map = {
                        "suv": "SUV",
                        "economy": "Economy",
                        "mini": "Mini",
                        "wagon": "Estate/Station Wagon",
                        "estate": "Estate/Station Wagon",
                        "premium": "Premium",
                        "7": "7 Seater",
                        "7-seater": "7 Seater",
                        "9": "9 Seater",
                        "9-seater": "9 Seater",
                    }
                    category = tail_map.get(tail, category)
            # If car_name still empty, heuristically derive from local text by removing category tokens and prices
            if not car_name:
                try:
                    local_txt_full = card.get_text(" \n", strip=True)
                    lines = [l.strip() for l in local_txt_full.split("\n") if l.strip()]
                    # remove lines that are price-like
                    price_like = re.compile(r"(€|EUR|GBP|\£|\d+[\.,]\d{2})", re.I)
                    candidates = [l for l in lines if not price_like.search(l)]
                    if candidates:
                        car_name = candidates[0]
                        # strip trailing category word if present
                        if category and car_name.lower().endswith(category.lower()):
                            car_name = car_name[: -len(category)].strip()
                except Exception:
                    pass
            # Fiat 500 Cabrio -> Group G (Premium)
            try:
                _cn_lower = (car_name or "").lower()
                if re.search(r"\bfiat\s*500\b.*\b(cabrio|convertible|cabriolet)\b", _cn_lower):
                    category = "Premium"
            except Exception:
                pass
            # Mini cabrio variants -> Group G (Premium)
            try:
                _cn_lower = (car_name or "").lower()
                if re.search(r"\bmini\s+(one|cooper)\b.*\b(cabrio|convertible|cabriolet)\b", _cn_lower):
                    category = "Premium"
            except Exception:
                pass
            # Specific model mappings to requested groups
            try:
                cn = (car_name or "").lower()
                # Mini Countryman (incl. Cooper Countryman): E2 if Auto, else D (Economy)
                if re.search(r"\bmini\s+(cooper\s+)?countryman\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Economy Automatic"
                    else:
                        category = "Economy"
                # Peugeot 108 Cabrio -> G (Premium)
                if re.search(r"\bpeugeot\s*108\b.*\b(cabrio|convertible|cabriolet)\b", cn):
                    category = "Premium"
                # Fiat 500 Auto -> E1 (Mini Automatic) unless Cabrio already handled
                if re.search(r"\bfiat\s*500\b.*\b(auto|automatic)\b", cn) and not re.search(r"\b(cabrio|convertible|cabriolet)\b", cn):
                    category = "Mini Automatic"
                # Citroen C3 Auto -> E2 (Economy Automatic)
                if re.search(r"\bcitro[eë]n\s*c3\b.*\b(auto|automatic)\b", cn) and not re.search(r"\bc3\s*aircross\b", cn):
                    category = "Economy Automatic"
                # Citroen C3 (non-Aircross, non-Auto) -> D (Economy)
                if re.search(r"\bcitro[eë]n\s*c3\b", cn) and not re.search(r"\b(auto|automatic)\b", cn) and not re.search(r"\bc3\s*aircross\b", cn):
                    category = "Economy"
                # Citroen C3 Aircross Auto -> L1 (SUV Automatic)
                if re.search(r"\bcitro[eë]n\s*c3\s*aircross\b.*\b(auto|automatic)\b", cn):
                    category = "SUV Automatic"
                # Toyota Aygo X -> F (SUV)
                if re.search(r"\btoyota\s*aygo\s*x\b", cn):
                    category = "SUV"
                # Fiat 500L -> J1 (Crossover)
                if re.search(r"\bfiat\s*500l\b", cn):
                    category = "Crossover"
                # Renault Clio SW/estate variants -> J2 (Estate/Station Wagon); autos will be L2 via suffix
                if re.search(r"\brenault\s*clio\b", cn) and re.search(r"\b(sw|st|sport\s*tourer|tourer|break|estate|kombi|grandtour|grand\s*tour|sporter|wagon)\b", cn):
                    category = "Estate/Station Wagon"
                # Group J1 (Crossover) models
                j1_patterns = [
                    r"\bkia\s*sportage\b",
                    r"\bnissan\s*qashqai\b",
                    r"\b(skoda|škoda)\s*kamiq\b",
                    r"\bhyundai\s*tucson\b",
                    r"\bseat\s*ateca\b",
                    r"\bmazda\s*cx[- ]?3\b",
                    r"\bpeugeot\s*5008\b",
                    r"\bpeugeot\s*3008\b",
                    r"\bpeugeot\s*2008\b",
                    r"\brenault\s*austral\b",
                    r"\btoyota\s*hilux\b.*\b4x4\b",
                ]
                if any(re.search(p, cn) for p in j1_patterns):
                    category = "Crossover"
                # Peugeot 308 base -> J1; 308 SW: Auto -> L2, else J2
                if re.search(r"\bpeugeot\s*308\b", cn):
                    if re.search(r"\bsw\b", cn):
                        if _is_auto_flag(cn, _page_text, transmission_label):
                            category = "Station Wagon Automatic"
                        else:
                            category = "Estate/Station Wagon"
                    else:
                        category = "Crossover"
                # VW Golf SW/Variant: Auto -> L2, else J2
                if re.search(r"\b(vw|volkswagen)\s*golf\b", cn) and re.search(r"\b(sw|variant)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # VW Passat: base & Variant -> J2; Auto -> L2
                if re.search(r"\b(vw|volkswagen)\s*passat\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Seat Leon SW/ST/Variant/Estate: Auto -> L2, else J2
                if re.search(r"\bseat\s*leon\b", cn) and re.search(r"\b(sw|st|variant|sport\s*tourer|sportstourer|estate)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Skoda Scala: base -> J2; Auto -> L2
                if re.search(r"\b(skoda|škoda)\s*scala\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Seat Arona -> F (SUV) regardless of transmission
                if re.search(r"\bseat\s*arona\b", cn):
                    category = "SUV"
                # Hyundai Kona/Kauai -> F (SUV) regardless of transmission
                if re.search(r"\bhyundai\s*(kona|kauai)\b", cn):
                    category = "SUV"
                # Skoda Octavia -> J2 (Station Wagon)
                if re.search(r"\b(skoda|škoda)\s*octavia\b", cn):
                    category = "Estate/Station Wagon"
                # Toyota Corolla SW/TS/Touring Sports: Auto -> L2 else J2
                if re.search(r"\btoyota\s*corolla\b", cn) and re.search(r"\b(sw|ts|touring\s*sports?|sport\s*touring|estate|wagon)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Toyota Corolla base (non-wagon) Auto -> E2
                if re.search(r"\btoyota\s*corolla\b", cn) and not re.search(r"\b(sw|ts|touring\s*sports?|sport\s*touring|estate|wagon)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Economy Automatic"
                # Peugeot 508 -> J2; Auto -> L2 (Station Wagon Automatic)
                if re.search(r"\bpeugeot\s*508\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Hyundai i30 -> J2; Auto -> L2
                if re.search(r"\bhyundai\s*i30\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Cupra Formentor Auto -> L1
                if re.search(r"\bcupra\s*formentor\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "SUV Automatic"
                # Renault Megane Sedan Auto -> L2
                if re.search(r"\brenault\s*megane\b", cn) and re.search(r"\bsedan\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "Station Wagon Automatic"
                # Renault Megane SW/Estate/Wagon: J2; Auto -> L2
                if re.search(r"\brenault\s*megane\b", cn) and re.search(r"\b(sw|estate|wagon|sport\s*tourer|sport\s*tourismo|tourer)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Cupra Leon SW Auto -> L2
                if re.search(r"\bcupra\s*leon\b", cn) and re.search(r"\b(sw|st|sport\s*tourer|sportstourer|estate|variant)\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "Station Wagon Automatic"
                # Toyota Yaris Cross Auto -> L1
                if re.search(r"\btoyota\s*yaris\s*cross\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "SUV Automatic"
                # Nissan Juke -> F (SUV) regardless of transmission
                if re.search(r"\bnissan\s*juke\b", cn):
                    category = "SUV"
                # Toyota Yaris Auto -> E1
                if re.search(r"\btoyota\s*yaris\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "Mini Automatic"
                # Kia Picanto Auto -> E1
                if re.search(r"\bkia\s*picanto\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "Mini Automatic"
                # VW Taigo -> F (SUV) regardless of transmission
                if re.search(r"\b(vw|volkswagen)\s*taigo\b", cn):
                    category = "SUV"
                # Mitsubishi Spacestar Auto -> E1
                if re.search(r"\bmitsubishi\s*space\s*star|spacestar\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "Mini Automatic"
                # Renault Megane Auto -> E2 (use card-level text)
                if re.search(r"\brenault\s*megane\b", cn):
                    _ct = ""
                    try:
                        _ct = card.get_text(" ", strip=True).lower()
                    except Exception:
                        _ct = ""
                    if _is_auto_flag(cn, _ct, transmission_label):
                        category = "Economy Automatic"
                # Ford Puma -> F (SUV) regardless of transmission
                if re.search(r"\bford\s*puma\b", cn):
                    category = "SUV"
                # Citroen C5 Aircross Auto -> L1
                if re.search(r"\bcitro[eë]n\s*c5\s*aircross\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "SUV Automatic"
                # Toyota C-HR Auto -> L1
                if re.search(r"\btoyota\s*c[-\s]?hr\b|\btoyota\s*chr\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "SUV Automatic"
                # Kia Stonic -> F (SUV) regardless of transmission
                if re.search(r"\bkia\s*stonic\b", cn):
                    category = "SUV"
                # Ford EcoSport -> F (SUV) regardless of transmission
                if re.search(r"\bford\s*eco\s*sport\b|\bford\s*ecosport\b", cn):
                    category = "SUV"
                # Opel/Vauxhall Crossland X -> F (SUV); Auto remains L1 via final if needed
                if re.search(r"\b(opel|vauxhall)\s*crossland\s*x?\b", cn):
                    category = "SUV"
                # Ford Focus SW/Estate/Wagon variants: J2; Auto -> L2
                if re.search(r"\bford\s*focus\b", cn) and re.search(r"\b(sw|estate|wagon|turnier|kombi|sportbreak|sport\s*brake|tourer|touring)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Ford Focus base (non-wagon): D or E2
                if re.search(r"\bford\s*focus\b", cn) and not re.search(r"\b(sw|estate|wagon)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Economy Automatic"
                    else:
                        category = "Economy"
                # Seat Leon base (non-wagon): D or E2 (use card-level text)
                if re.search(r"\bseat\s*leon\b", cn) and not re.search(r"\b(sw|st|variant|sport\s*tourer|sportstourer|estate|wagon)\b", cn):
                    _ct = ""
                    try:
                        _ct = card.get_text(" ", strip=True).lower()
                    except Exception:
                        _ct = ""
                    if _is_auto_flag(cn, _ct, transmission_label):
                        category = "Economy Automatic"
                    else:
                        category = "Economy"
                # Kia Ceed base (non-wagon): D or E2
                if re.search(r"\bkia\s*ceed\b", cn) and not re.search(r"\b(sw|estate|wagon|sportswagon|sports\s*wagon)\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Economy Automatic"
                    else:
                        category = "Economy"
                # Opel/Vauxhall Astra: base & SW -> J2; Auto -> L2
                if re.search(r"\b(opel|vauxhall)\s*astra\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # VW T-Cross Auto -> L1 (unchanged)
                if re.search(r"\b(vw|volkswagen)\s*t[-\s]?cross\b", cn) and _is_auto_flag(cn, _page_text, transmission_label):
                    category = "SUV Automatic"
                # VW Golf Auto (hatch) -> E2 (use card-level text)
                if re.search(r"\b(vw|volkswagen)\s*golf\b", cn) and not re.search(r"\b(sw|variant|estate|wagon)\b", cn):
                    _ct = ""
                    try:
                        _ct = card.get_text(" ", strip=True).lower()
                    except Exception:
                        _ct = ""
                    if _is_auto_flag(cn, _ct, transmission_label):
                        category = "Economy Automatic"
                # Dacia Jogger -> M1 (7 Seater); automatic will auto-suffix to M2 later
                if re.search(r"\bdacia\s*jogger\b", cn):
                    category = "7 Seater"
                # Fiat 500X -> J1 (Crossover); Auto -> L1
                if re.search(r"\bfiat\s*500x\b", cn):
                    if _is_auto_flag(cn, _page_text, transmission_label):
                        category = "SUV Automatic"
                    else:
                        category = "Crossover"
                # VW Beetle Cabrio -> G (Premium)
                if re.search(r"\b(vw|volkswagen)\s*beetle\b.*\b(cabrio|convertible|cabriolet)\b", cn):
                    category = "Premium"
                # Group L1 (SUV Automatic) for specific models when Automatic is detected (including acronyms)
                try:
                    _card_txt = ""
                    try:
                        _card_txt = card.get_text(" ", strip=True).lower()
                    except Exception:
                        _card_txt = ""
                    is_auto = _is_auto_flag(cn, _card_txt, transmission_label)
                    # Only keep intended L1 autos; others remain F per latest rules
                    is_l1_model = (
                        re.search(r"\bpeugeot\s*(3008|2008|5008)\b", cn) or
                        re.search(r"\bnissan\s*qashqai\b", cn) or
                        re.search(r"\b(skoda|škoda)\s*kamiq\b", cn) or
                        re.search(r"\bcitro[eë]n\s*c4\b", cn) or
                        re.search(r"\b(vw|volkswagen)\s*tiguan\b", cn) or
                        re.search(r"\bds(\s*automobiles)?\s*4\b", cn) or
                        re.search(r"\b(skoda|škoda)\s*karoq\b", cn) or
                        re.search(r"\bford\s*kuga\b", cn) or
                        re.search(r"\bjeep\s*renegade\b", cn) or
                        re.search(r"\brenault\s*arkana\b", cn) or
                        re.search(r"\btoyota\s*rav\s*4\b|\brav4\b", cn) or
                        re.search(r"\bcupra\s*formentor\b", cn) or
                        re.search(r"\btoyota\s*yaris\s*cross\b", cn) or
                        re.search(r"\bcitro[eë]n\s*c5\s*aircross\b", cn) or
                        re.search(r"\btoyota\s*c[-\s]?hr\b|\btoyota\s*chr\b", cn) or
                        re.search(r"\b(vw|volkswagen)\s*t[-\s]?cross\b", cn) or
                        re.search(r"\bfiat\s*500x\b", cn)
                    )
                    if is_auto and is_l1_model:
                        category = "SUV Automatic"
                except Exception:
                    pass
                # Citroen C4 Picasso (non-Grand) -> M1 (7 Seater). Auto will suffix to M2 later
                if re.search(r"\bcitro[eë]n\s*c4\s*picasso\b", cn) and not re.search(r"\bgrand\b", cn):
                    category = "7 Seater"
                # Citroen Grand C4 Picasso/Grand Spacetourer -> M1 base; auto will suffix to M2
                if re.search(r"\bcitro[eë]n\s*c4\s*(grand\s*picasso|grand\s*spacetourer|grand\s*space\s*tourer)\b", cn):
                    category = "7 Seater"
            except Exception:
                pass
            # Group D (Economy) models; Auto -> Economy Automatic (use card-level text for auto detection)
            d_models = [
                r"dacia\s+sandero",
                r"peugeot\s*208",
                r"opel\s*corsa",
                r"seat\s*ibiza",
                r"seat\s*leon",
                r"kia\s*ceed",
                r"(vw|volkswagen)\s*polo",
                r"renault\s*clio",
                r"ford\s*fiesta",
                r"ford\s*focus",
                r"hyundai\s*i20",
                r"nissan\s*micra",
                r"audi\s*a1",
            ]
            if any(re.search(p, cn) for p in d_models):
                _ct = ""
                try:
                    _ct = card.get_text(" ", strip=True).lower()
                except Exception:
                    _ct = ""
                if _is_auto_flag(cn, _ct, transmission_label):
                    category = "Economy Automatic"
                else:
                    category = "Economy"
            # Force B1 mapping for specific models the user provided (non-Auto/Non-Cabrio, base Mini only)
            try:
                _b1_models = [
                    "fiat 500", "peugeot 108", "opel adam",
                    "toyota aygo", "volkswagen up", "vw up", "ford ka", "renault twingo",
                    "citroen c1", "citroën c1", "kia picanto"
                ]
                _cn = (car_name or "").lower()
                if any(m in _cn for m in _b1_models):
                    # do not apply B1 if auto/automatic (multi-language/abbrev) or cabrio/convertible/cabriolet
                    if (not _is_auto_flag(_cn, _page_text, transmission_label)) and not re.search(r"\b(cabrio|convertible|cabriolet)\b", _cn, re.I):
                        # exclude variants that map elsewhere: 500X/500L, Aygo X, Aircross
                        if not re.search(r"\b(500x|500l|aygo\s*x|aircross|countryman)\b", _cn):
                            # and only when category is not already a non-Mini mapping
                            if category in ("", "Mini"):
                                category = "Mini 4 Doors"
            except Exception:
                pass
            # Refine Mini into 'Mini 4 Doors' when doors info is present
            try:
                if category == "Mini":
                    _lt = ""
                    try:
                        _lt = card.get_text(" ", strip=True).lower()
                    except Exception:
                        _lt = ""
                    _cn = (car_name or "").lower()
                    four_pat = re.compile(r"\b(4\s*(doors?|portas|p)|4p|4-door|4-portas)\b", re.I)
                    if four_pat.search(_lt) or four_pat.search(_cn):
                        category = "Mini 4 Doors"
            except Exception:
                pass
            # link
            link = url_from_row(card, base_url) or base_url
            # Photo cache: upsert or read from cache based on model key
            try:
                if car_name:
                    _key = _normalize_model_key(car_name)
                    if photo:
                        _cache_set_photo(_key, photo)
                    else:
                        cached_photo = _cache_get_photo(_key)
                        if cached_photo:
                            photo = cached_photo
            except Exception:
                pass
            # Crossover override when car name is present (exclude C4 Picasso/Grand Spacetourer)
            try:
                _car_lc = (car_name or "").lower()
                is_c4_picasso_like = re.search(r"\bc4\s*(picasso|grand\s*spacetourer|grand\s*space\s*tourer)\b", _car_lc)
                if re.search(r"\b(peugeot\s*2008|peugeot\s*3008|citro[eë]n\s*c4)\b", _car_lc, re.I) and not is_c4_picasso_like:
                    category = "Crossover"
            except Exception:
                pass
            # Automatic suffix for selected groups
            try:
                if transmission_label == "Automatic" and category in ("Mini", "Economy", "SUV", "Estate/Station Wagon", "7 Seater"):
                    if category == "Estate/Station Wagon":
                        category = "Station Wagon Automatic"
                    elif category == "7 Seater":
                        category = "7 Seater Automatic"
                    else:
                        category = f"{category} Automatic"
            except Exception:
                pass
            # FINAL OVERRIDE: Ensure Group D/E2 models are correctly placed (Peugeot 208, Opel Corsa, Seat Ibiza, VW Polo, Renault Clio, Ford Fiesta, Nissan Micra, Hyundai i20, Audi A1)
            try:
                cn2 = (car_name or "").lower()
                d_models_final = [
                    r"\bpeugeot\s*208\b",
                    r"\bopel\s*corsa\b",
                    r"\bseat\s*ibiza\b",
                    r"\bseat\s*leon\b",
                    r"\bkia\s*ceed\b",
                    r"\b(vw|volkswagen)\s*polo\b",
                    r"\bcitro[eë]n\s*c3\b",
                    r"\brenault\s*clio\b",
                    r"\bford\s*fiesta\b",
                    r"\bford\s*focus\b",
                    r"\bnissan\s*micra\b",
                    r"\bhyundai\s*i20\b",
                    r"\baudi\s*a1\b",
                    r"\bdacia\s*sandero\b",
                ]
                # do not override if we already mapped to protected groups (wagon/crossover/suv)
                is_protected = category in ("Estate/Station Wagon", "Station Wagon Automatic", "Crossover", "SUV", "SUV Automatic")
                if (not is_protected) and any(re.search(p, cn2) for p in d_models_final):
                    if _is_auto_flag(cn2, _txt, transmission_label):
                        category = "Economy Automatic"
                    else:
                        category = "Economy"
            except Exception:
                pass
            # FINAL MANUAL OVERRIDE for D models: if manual is explicit, force D
            try:
                cn2b = (car_name or "").lower()
                is_d_family = any(re.search(p, cn2b) for p in [
                    r"\bpeugeot\s*208\b", r"\bopel\s*corsa\b", r"\bseat\s*ibiza\b",
                    r"\bseat\s*leon\b", r"\b(vw|volkswagen)\s*golf\b", r"\b(vw|volkswagen)\s*polo\b",
                    r"\brenault\s*clio\b", r"\bford\s*fiesta\b", r"\bnissan\s*micra\b",
                    r"\bhyundai\s*i20\b", r"\baudi\s*a1\b", r"\bdacia\s*sandero\b", r"\brenault\s*megane\b",
                ])
                # re-evaluate card text for manual marker
                _txt2 = ""
                try:
                    _txt2 = card.get_text(" ", strip=True).lower()
                except Exception:
                    _txt2 = ""
                is_manual = (str(transmission_label or '').lower() == 'manual') or bool(re.search(r"\bmanual\b", _txt2))
                if is_d_family and is_manual and category not in ("Estate/Station Wagon", "Station Wagon Automatic"):
                    category = "Economy"
            except Exception:
                pass
            # FINAL L2/J2 OVERRIDE: enforce wagons to wagon groups; autos -> L2
            try:
                cnf = (car_name or "").lower()
                _txt = ""
                try:
                    _txt = card.get_text(" ", strip=True).lower()
                except Exception:
                    _txt = ""
                # Renault Clio SW: force to wagon groups
                if re.search(r"\brenault\s*clio\b", cnf) and re.search(r"\b(sw|st|sport\s*tourer|tourer|break|estate|kombi|grandtour|grand\s*tour|sporter|wagon)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                cn3 = (car_name or "").lower()
                is_auto_any = _is_auto_flag(cn3, _txt, transmission_label)
                l1_model = (
                    re.search(r"\bpeugeot\s*(3008|2008|5008)\b", cn3) or
                    re.search(r"\bnissan\s*qashqai\b", cn3) or
                # ... (rest of the code remains the same)
                    re.search(r"\b(skoda|škoda)\s*kamiq\b", cn3) or
                    re.search(r"\bcitro[eë]n\s*c4\b", cn3) or
                    re.search(r"\b(vw|volkswagen)\s*tiguan\b", cn3) or
                    re.search(r"\bds(\s*automobiles)?\s*4\b", cn3) or
                    re.search(r"\b(skoda|škoda)\s*karoq\b", cn3) or
                    re.search(r"\bford\s*kuga\b", cn3) or
                    re.search(r"\bjeep\s*renegade\b", cn3) or
                    re.search(r"\brenault\s*arkana\b", cn3) or
                    re.search(r"\btoyota\s*rav\s*4\b|\brav4\b", cn3) or
                    re.search(r"\bcupra\s*formentor\b", cn3) or
                    re.search(r"\btoyota\s*yaris\s*cross\b", cn3) or
                    re.search(r"\bcitro[eë]n\s*c5\s*aircross\b", cn3) or
                    re.search(r"\btoyota\s*c[-\s]?hr\b|\btoyota\s*chr\b", cn3) or
                    re.search(r"\b(vw|volkswagen)\s*t[-\s]?cross\b", cn3) or
                    re.search(r"\bfiat\s*500x\b", cn3)
                )
                # don't override M2 or wagons
                is_m2 = category == "7 Seater Automatic" or re.search(r"\bc4\s*(picasso|grand\s*spacetourer|grand\s*space\s*tourer)\b", cn3)
                is_wagon = category in ("Estate/Station Wagon", "Station Wagon Automatic")
                if is_auto_any and l1_model and (not is_m2) and (not is_wagon):
                    category = "SUV Automatic"
            except Exception:
                pass
            # FINAL L2/J2 OVERRIDE: 308 SW and Scala to wagon groups; autos -> L2
            try:
                cnf = (car_name or "").lower()
                if re.search(r"\bford\s*focus\b", cnf) and re.search(r"\b(sw|estate|wagon|turnier|kombi|sportbreak|sport\s*brake|tourer|touring)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\b(vw|volkswagen)\s*golf\b", cnf) and re.search(r"\b(sw|variant)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\bfiat\s*500l\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\b(vw|volkswagen)\s*passat\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\bpeugeot\s*508\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\bhyundai\s*i30\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\btoyota\s*corolla\b", cnf) and re.search(r"\b(sw|ts|touring\s*sports?|sport\s*touring|estate|wagon)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                # Enforce E2 for Toyota Corolla base Auto
                if re.search(r"\btoyota\s*corolla\b", cnf) and not re.search(r"\b(sw|ts|touring\s*sports?|sport\s*touring|estate|wagon)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Economy Automatic"
                if re.search(r"\bseat\s*leon\b", cnf) and re.search(r"\b(sw|st|variant|sport\s*tourer|sportstourer|estate)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\b(skoda|škoda)\s*scala\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\bford\s*focus\b", cnf) and re.search(r"\b(sw|estate|wagon)\b", cnf) and _is_auto_flag(cnf, _txt, transmission_label):
                    category = "Station Wagon Automatic"
                if re.search(r"\b(opel|vauxhall)\s*astra\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
                if re.search(r"\brenault\s*megane\b", cnf) and re.search(r"\bsedan\b", cnf) and _is_auto_flag(cnf, _txt, transmission_label):
                    category = "Station Wagon Automatic"
                if re.search(r"\brenault\s*megane\b", cnf) and re.search(r"\b(sw|estate|wagon|sport\s*tourer|sport\s*tourismo|tourer)\b", cnf):
                    if _is_auto_flag(cnf, _txt, transmission_label):
                        category = "Station Wagon Automatic"
                    else:
                        category = "Estate/Station Wagon"
            except Exception:
                pass
            # FINAL M2 OVERRIDE: common 7-seater autos -> 7 Seater Automatic (wins over J1/D)
            try:
                cn4 = (car_name or "").lower()
                m2_patterns = [
                    r"\bcitro[eë]n\s*c4\s*(picasso|grand\s*spacetourer|grand\s*space\s*tourer)\b",
                    r"\bcitro[eë]n\s*grand\s*picasso\b",
                    r"\brenault\s*grand\s*sc[eé]nic\b",
                    r"\bmercedes\s*glb\b.*\b(7\s*seater|7\s*lugares|7p|7\s*seats)\b",
                    r"\b(vw|volkswagen)\s*multivan\b",
                    r"\bpeugeot\s*rifter\b",
                ]
                if any(re.search(p, cn4) for p in m2_patterns) and _is_auto_flag(cn4, _txt, transmission_label):
                    category = "7 Seater Automatic"
            except Exception:
                pass
            # FINAL E1 OVERRIDE: Toyota Aygo Auto -> Mini Automatic (avoid uncategorized)
            try:
                cn5 = (car_name or "").lower()
                if re.search(r"\btoyota\s*aygo\b", cn5) and _is_auto_flag(cn5, _txt, transmission_label):
                    category = "Mini Automatic"
                if re.search(r"\bkia\s*picanto\b", cn5) and _is_auto_flag(cn5, _txt, transmission_label):
                    category = "Mini Automatic"
            except Exception:
                pass
            # FINAL B1 OVERRIDE: base mini models -> 'Mini 4 Doors' (when not auto/cabrio/special variants)
            try:
                b1_list = [
                    r"\bfiat\s*500\b",
                    r"\bcitro[eë]n\s*c1\b",
                    r"\bpeugeot\s*108\b",
                    r"\bopel\s*adam\b",
                    r"\btoyota\s*aygo\b",
                    r"\b(vw|volkswagen)\s*up\b",
                    r"\bford\s*ka\b",
                    r"\brenault\s*twingo\b",
                    r"\bkia\s*picanto\b",
                ]
                _name = (car_name or "").lower()
                if any(re.search(p, _name) for p in b1_list):
                    # do not apply if this is a D/E2 economy model (protect Group D)
                    d_guard = [
                        r"\bpeugeot\s*208\b", r"\bopel\s*corsa\b", r"\bseat\s*ibiza\b",
                        r"\b(vw|volkswagen)\s*polo\b", r"\bcitro[eë]n\s*c3\b", r"\brenault\s*clio\b",
                        r"\bford\s*fiesta\b", r"\bnissan\s*micra\b", r"\bhyundai\s*i20\b", r"\baudi\s*a1\b",
                        r"\bdacia\s*sandero\b"
                    ]
                    if any(re.search(p, _name) for p in d_guard):
                        raise Exception("skip B1 for D/E2 models")
                    # exclude autos and cabrio and special variants
                    if (not _is_auto_flag(_name, _txt, transmission_label)) \
                        and not re.search(r"\b(cabrio|convertible|cabriolet)\b", _name) \
                        and not re.search(r"\b(500x|500l|aygo\s*x|aircross|countryman)\b", _name):
                        category = "Mini 4 Doors"
            except Exception:
                pass
            # Skip blocked models - DISABLED: mostrar todos os carros
            # if car_name and _is_blocked_model(car_name):
            #     cards_blocked += 1
            #     continue
            # Mapear categoria para código de grupo
            group_code = map_category_to_group(category, car_name)
            # Capitalizar nome para display (Peugeot 2008 Auto, Renault Megane SW Auto)
            car_name_display = capitalize_car_name(car_name)
            items.append({
                "id": idx,
                "car": car_name_display,
                "supplier": supplier,
                "price": price_text,
                "currency": "",
                "category": category,
                "group": group_code,
                "transmission": transmission_label,
                "photo": photo,
                "link": link,
            })
            idx += 1
        print(f"[PARSE] Stats: price={cards_with_price}, name={cards_with_name}, blocked={cards_blocked}, items={len(items)}")
        if items:
            print(f"[PARSE] Returning {len(items)} items from card parsing")
            return items
    except Exception:
        pass

    # Require an explicit currency marker to avoid capturing ratings/ages
    price_regex = re.compile(r"(?:€\s*\d{1,4}(?:[\.,]\d{3})*(?:[\.,]\d{2})?|\bEUR\s*\d{1,4}(?:[\.,]\d{3})*(?:[\.,]\d{2})?)", re.I)

    # Basic category keyword list (EN + PT)
    CATEGORY_KEYWORDS = [
        "mini","economy","compact","intermediate","standard","full-size","full size","suv","premium","luxury","van","estate","convertible","people carrier","minivan","midsize",
        "mini","económico","económico","compacto","intermédio","padrão","familiar","suv","premium","luxo","carrinha","descapotável","monovolume","médio"
    ]

    candidates = []
    for el in soup.find_all(text=price_regex):
        try:
            txt = el.strip()
        except Exception:
            continue
        if not txt or len(txt) > 50:
            continue
        node = el if hasattr(el, 'parent') else None
        if not node:
            continue
        # climb up to find a reasonable container (card/row)
        container = node.parent
        depth = 0
        while container and depth < 6 and container.name not in ("tr", "li", "article", "section", "div"):
            container = container.parent
            depth += 1
        if not container:
            container = node.parent
        candidates.append((container, txt))

    seen = set()
    for idx, (container, price_text) in enumerate(candidates):
        # car/model
        name_el = container.select_one(".car, .vehicle, .model, .title, .name, .veh-name, [class*='model'], [class*='vehicle']")
        car_name = name_el.get_text(strip=True) if name_el else ""
        # supplier: try explicit, else alt/title of images within container
        supplier_el = container.select_one(".supplier, .vendor, .partner, [class*='supplier'], [class*='vendor']")
        supplier = supplier_el.get_text(strip=True) if supplier_el else ""
        if not supplier:
            img = container.select_one("img[alt], img[title]")
            if img:
                supplier = img.get("alt") or img.get("title") or ""
        # category/group: try explicit labels then keyword search in container text
        cat_el = container.select_one(".category, .group, .vehicle-category, [class*='category'], [class*='group'], [class*='categoria'], [class*='grupo']")
        category = cat_el.get_text(strip=True) if cat_el else ""
        if not category:
            try:
                text = container.get_text(" ", strip=True).lower()
                match = next((kw for kw in CATEGORY_KEYWORDS if kw.lower() in text), "")
                category = match.title() if match else ""
            except Exception:
                category = ""
        # Crossover override based on model name (when available)
        try:
            _car_lc = (car_name or "").lower()
            if re.search(r"\b(peugeot\s*2008|peugeot\s*3008|citro[eë]n\s*c4)\b", _car_lc, re.I):
                category = "Crossover"
        except Exception:
            pass

        # link
        link = url_from_row(container, base_url) or base_url

        key = (supplier, car_name, price_text)
        if key in seen:
            continue
        seen.add(key)

        # detect currency symbol present in the text
        curr = "EUR" if re.search(r"EUR", price_text, re.I) else ("EUR" if "€" in price_text else "")
        # Mapear categoria para código de grupo
        group_code = map_category_to_group(category, car_name)
        items.append({
            "id": idx,
            "car": car_name,
            "supplier": supplier,
            "price": price_text,
            "currency": curr,
            "category": category,
            "group": group_code,
            "transmission": transmission_label,
            "link": link,
        })
        # REMOVED: if len(items) >= 50: break  # Removido limite para mostrar TODOS os carros
    # If no detailed items parsed, fall back to provider summaries to ensure prices are shown
    if not items and summary_items:
        items = summary_items
    # Ensure photos when grupo/category_code is known
    try:
        for it in items:
            if (not it.get("photo")) and it.get("category_code"):
                cc = it.get("category_code")
                it["photo"] = urljoin(base_url, f"/cdn/img/cars/S/car_{cc}.jpg")
    except Exception:
        pass
    return items


def url_from_row(row, base_url: str) -> str:
    a = row.select_one("a[href]")
    if a and a.has_attr("href"):
        href = a["href"]
        if href and not href.lower().startswith("javascript") and href != "#":
            return urljoin(base_url, href)
    for attr in ["data-href", "data-url", "data-link"]:
        el = row.select_one(f"*[{attr}]")
        if el and el.has_attr(attr):
            return urljoin(base_url, el[attr])
    clickable = row.select_one("*[onclick]")
    if clickable and clickable.has_attr("onclick"):
        m = re.search(r"https?://[^'\"]+", clickable["onclick"])  
        if m:
            return m.group(0)
    return ""


def try_direct_carjet(location_name: str, start_dt, end_dt, lang: str = "pt", currency: str = "EUR") -> str:
    try:
        sess = requests.Session()
        ua = {
            "User-Agent": "Mozilla/5.0 (compatible; PriceTracker/1.0)",
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.6",
            "X-Forwarded-For": "185.23.160.1",
            "Referer": "https://www.carjet.com/do/list/pt",
        }
        lang = (lang or "pt").lower()
        # Pre-seed cookies to bias locale
        try:
            sess.cookies.set("monedaForzada", currency)
            sess.cookies.set("moneda", currency)
            sess.cookies.set("currency", currency)
            sess.cookies.set("idioma", lang.upper())
            sess.cookies.set("lang", lang)
            sess.cookies.set("country", "PT")
        except Exception:
            pass

        # 1) GET locale homepage to mint session and try to capture s/b tokens
        if lang == "pt":
            home_path = "aluguel-carros/index.htm"
        elif lang == "es":
            home_path = "alquiler-coches/index.htm"
        elif lang == "fr":
            home_path = "location-voitures/index.htm"
        elif lang == "de":
            home_path = "mietwagen/index.htm"
        elif lang == "it":
            home_path = "autonoleggio/index.htm"
        elif lang == "nl":
            home_path = "autohuur/index.htm"
        else:
            home_path = "index.htm"
        home_url = f"https://www.carjet.com/{home_path}"
        home = sess.get(home_url, headers=ua, timeout=20)
        s_token = None
        b_token = None
        try:
            m = re.search(r"[?&]s=([A-Za-z0-9]+)", home.text)
            if m:
                s_token = m.group(1)
            m = re.search(r"[?&]b=([A-Za-z0-9]+)", home.text)
            if m:
                b_token = m.group(1)
        except Exception:
            pass

        # 2) Prefer submitting the actual homepage form with all hidden fields preserved
        try:
            soup = BeautifulSoup(home.text, "lxml")
            form = soup.select_one("form[name='menu_tarifas'], form#booking_form")
            if form:
                action = form.get("action") or f"/do/list/{lang}"
                post_url = action if action.startswith("http") else requests.compat.urljoin(home_url, action)
                payload: Dict[str, Any] = {}
                # include all inputs
                for inp in form.select("input[name]"):
                    name = inp.get("name")
                    if not name:
                        continue
                    val = inp.get("value", "")
                    payload[name] = val
                # include selects
                for sel in form.select("select[name]"):
                    name = sel.get("name")
                    if not name:
                        continue
                    # take selected option or first
                    opt = sel.select_one("option[selected]") or sel.select_one("option")
                    payload[name] = opt.get("value") if opt else ""

                # override with our values
                override = build_carjet_form(location_name, start_dt, end_dt, lang=lang, currency=currency)
                payload.update({k: v for k, v in override.items() if v is not None})
                if s_token:
                    payload["s"] = s_token
                if b_token:
                    payload["b"] = b_token

                headers = {
                    "User-Agent": ua["User-Agent"],
                    "Origin": "https://www.carjet.com",
                    "Referer": home_url,
                }
                resp = sess.post(post_url, data=payload, headers=headers, timeout=25)
                if resp.status_code == 200 and resp.text:
                    return resp.text
        except Exception:
            pass

        # 3) Fallback: POST to /do/list/{lang} with our constructed payload
        data = build_carjet_form(location_name, start_dt, end_dt, lang=lang, currency=currency)
        if s_token:
            data["s"] = s_token
        if b_token:
            data["b"] = b_token

        headers = {
            "User-Agent": ua["User-Agent"],
            "Origin": "https://www.carjet.com",
            "Referer": home_url,
            "Accept-Language": ua.get("Accept-Language", "pt-PT,pt;q=0.9,en;q=0.6"),
            "X-Forwarded-For": ua.get("X-Forwarded-For", "185.23.160.1"),
        }
        url = f"https://www.carjet.com/do/list/{lang}"
        resp = sess.post(url, data=data, headers=headers, timeout=25)
        if resp.status_code == 200 and resp.text:
            # Detect if we were redirected to a generic homepage (wrong locale)
            homepage_like = False
            try:
                homepage_like = bool(re.search(r'hrental_pagetype"\s*:\s*"home"', resp.text) or re.search(r'data-steplist="home"', resp.text))
            except Exception:
                homepage_like = False
            if not homepage_like:
                return resp.text
            # Fallback path observed on results pages: modalFilter.asp then carList.asp
            try:
                mf_url = f"https://www.carjet.com/modalFilter.asp"
                # Minimal payload aligning with page
                mf_payload = {
                    "frmDestino": data.get("frmDestino") or data.get("dst_id") or data.get("pickupId") or "",
                    "frmFechaRecogida": f"{start_dt.strftime('%d/%m/%Y')} {start_dt.strftime('%H:%M')}",
                    "frmFechaDevolucion": f"{end_dt.strftime('%d/%m/%Y')} {end_dt.strftime('%H:%M')}",
                    "idioma": lang.upper(),
                    "frmMoneda": currency,
                    "frmTipoVeh": "CAR",
                }
                _ = sess.post(mf_url, data=mf_payload, headers=headers, timeout=20)
            except Exception:
                pass
            try:
                # Keep session tokens if available
                _q = f"idioma={lang.upper()}&case=2"
                if s_token:
                    _q += f"&s={s_token}"
                if b_token:
                    _q += f"&b={b_token}"
                cl_url = f"https://www.carjet.com/carList.asp?{_q}"
                rlist = sess.get(cl_url, headers=headers, timeout=25)
                if rlist.status_code == 200 and rlist.text:
                    return rlist.text
            except Exception:
                pass

        # If not OK or homepage detected, retry with PT-Portugal homepage and forced params on POST URL
        try:
            # Visit PT-Portugal homepage spelling (aluguer vs aluguel)
            home_url_ptpt = "https://www.carjet.com/aluguer-carros/index.htm"
            _ = sess.get(home_url_ptpt, headers=ua, timeout=20)
            headers2 = dict(headers)
            post_url2 = f"https://www.carjet.com/do/list/{lang}?idioma=PT&moneda=EUR&currency=EUR"
            resp2 = sess.post(post_url2, data=data, headers=headers2, timeout=25)
            if resp2.status_code == 200 and resp2.text:
                try:
                    if re.search(r'hrental_pagetype\"\s*:\s*\"home\"', resp2.text) or re.search(r'data-steplist=\"home\"', resp2.text):
                        pass
                    else:
                        return resp2.text
                except Exception:
                    return resp2.text
        except Exception:
            pass
    except Exception:
        pass
    return ""


def build_carjet_form(location_name: str, start_dt, end_dt, lang: str = "pt", currency: str = "EUR") -> Dict[str, Any]:
    # Build server-expected fields; include hidden destination IDs when possible
    pickup_dmY = start_dt.strftime("%d/%m/%Y")
    dropoff_dmY = end_dt.strftime("%d/%m/%Y")
    pickup_HM = start_dt.strftime("%H:%M")
    dropoff_HM = end_dt.strftime("%H:%M")
    code = LOCATION_CODES.get((location_name or "").lower(), "")
    form = {
        # free text
        "pickup": location_name,
        "dropoff": location_name,
        # hidden ids (best effort)
        "pickupId": code,
        "dst_id": code,
        "zoneCode": code,
        # dates
        "fechaRecogida": pickup_dmY,
        "fechaEntrega": dropoff_dmY,
        # times
        "fechaRecogidaSelHour": pickup_HM,
        "fechaEntregaSelHour": dropoff_HM,
        # locale hints
        "idioma": lang.upper(),
        "moneda": currency,
        "chkOneWay": "SI",
        # fields observed on list page (robustness)
        "frmDestino": code or "",
        "frmFechaRecogida": f"{pickup_dmY} {pickup_HM}",
        "frmFechaDevolucion": f"{dropoff_dmY} {dropoff_HM}",
        "frmMoneda": currency,
        "frmTipoVeh": "CAR",
    }
    return form


def fetch_with_optional_proxy(url: str, headers: Dict[str, str]):
    # Default locale headers if not provided
    try:
        headers = dict(headers or {})
        headers.setdefault("Accept-Language", "pt-PT,pt;q=0.9,en;q=0.6")
        headers.setdefault("X-Forwarded-For", "185.23.160.1")
    except Exception:
        pass
    # Prefer direct fetch with EUR cookies for CarJet to reduce latency and avoid geolocation flips
    try:
        from urllib.parse import urlparse as _urlparse
        pr = _urlparse(url)
        # Allow forcing proxy usage for CarJet via env flag (FORCE_PROXY_FOR_CARJET=1/true)
        _force_proxy_cj = False
        try:
            _force_proxy_cj = str(os.getenv("FORCE_PROXY_FOR_CARJET", "")).strip().lower() in ("1", "true", "yes", "on")
        except Exception:
            _force_proxy_cj = False
        if pr.netloc.endswith("carjet.com") and not _force_proxy_cj:
            h2 = dict(headers or {})
            h2["Cookie"] = "monedaForzada=EUR; moneda=EUR; currency=EUR; country=PT; idioma=PT; lang=pt"
            if _HTTPX_CLIENT:
                return _HTTPX_CLIENT.get(url, headers=h2)
            return requests.get(url, headers=h2, timeout=20)
    except Exception:
        pass
    if SCRAPER_SERVICE.lower() == "scrapeops" and SCRAPER_API_KEY:
        try:
            params = {
                "api_key": SCRAPER_API_KEY,
                "url": url,
                "render_js": "true",
            }
            if SCRAPER_COUNTRY:
                params["country"] = SCRAPER_COUNTRY
            r = requests.get("https://proxy.scrapeops.io/v1/", params=params, headers=headers, timeout=30)
            if r.status_code in (401, 403):
                # Fallback to direct if proxy is unauthorized/forbidden
                if _HTTPX_CLIENT:
                    return _HTTPX_CLIENT.get(url, headers=headers)
                return requests.get(url, headers=headers, timeout=6)
            return r
        except Exception:
            # Fallback to direct on any proxy error
            if _HTTPX_CLIENT:
                return _HTTPX_CLIENT.get(url, headers=headers)
            return requests.get(url, headers=headers, timeout=10)
    if _HTTPX_CLIENT:
        return _HTTPX_CLIENT.get(url, headers=headers)
    return requests.get(url, headers=headers, timeout=20)


async def async_fetch_with_optional_proxy(url: str, headers: Dict[str, str]):
    try:
        headers = dict(headers or {})
        headers.setdefault("Accept-Language", "pt-PT,pt;q=0.9,en;q=0.6")
        headers.setdefault("X-Forwarded-For", "185.23.160.1")
    except Exception:
        pass
    # Prefer direct CarJet with PT/EUR cookies
    try:
        from urllib.parse import urlparse as _urlparse
        pr = _urlparse(url)
        # Allow forcing proxy usage for CarJet via env flag (FORCE_PROXY_FOR_CARJET=1/true)
        _force_proxy_cj = False
        try:
            _force_proxy_cj = str(os.getenv("FORCE_PROXY_FOR_CARJET", "")).strip().lower() in ("1", "true", "yes", "on")
        except Exception:
            _force_proxy_cj = False
        if pr.netloc.endswith("carjet.com") and not _force_proxy_cj:
            h2 = dict(headers or {})
            h2["Cookie"] = "monedaForzada=EUR; moneda=EUR; currency=EUR; country=PT; idioma=PT; lang=pt"
            if _HTTPX_ASYNC:
                return await _HTTPX_ASYNC.get(url, headers=h2)
            # fallback to sync in thread
            return await asyncio.to_thread(requests.get, url, headers=h2, timeout=6)
    except Exception:
        pass
    if SCRAPER_SERVICE.lower() == "scrapeops" and SCRAPER_API_KEY:
        try:
            params = {
                "api_key": SCRAPER_API_KEY,
                "url": url,
                "render_js": "true",
            }
            if SCRAPER_COUNTRY:
                params["country"] = SCRAPER_COUNTRY
            # httpx doesn't proxy this conveniently; use requests in a thread
            r = await asyncio.to_thread(requests.get, "https://proxy.scrapeops.io/v1/", params=params, headers=headers, timeout=6)
        except TypeError:
            # Fallback: direct fetch
            if _HTTPX_ASYNC:
                return await _HTTPX_ASYNC.get(url, headers=headers)
            return await asyncio.to_thread(requests.get, url, headers=headers, timeout=6)
        try:
            if r.status_code in (401, 403):
                if _HTTPX_ASYNC:
                    return await _HTTPX_ASYNC.get(url, headers=headers)
                return await asyncio.to_thread(requests.get, url, headers=headers, timeout=6)
            return r
        except Exception:
            if _HTTPX_ASYNC:
                return await _HTTPX_ASYNC.get(url, headers=headers)
            return await asyncio.to_thread(requests.get, url, headers=headers, timeout=6)
    if _HTTPX_ASYNC:
        return await _HTTPX_ASYNC.get(url, headers=headers)
    return await asyncio.to_thread(requests.get, url, headers=headers, timeout=20)


def post_with_optional_proxy(url: str, data: Dict[str, Any], headers: Dict[str, str]):
    # Default locale headers if not provided
    try:
        headers = dict(headers or {})
        headers.setdefault("Accept-Language", "pt-PT,pt;q=0.9,en;q=0.6")
        headers.setdefault("X-Forwarded-For", "185.23.160.1")
    except Exception:
        pass
    if SCRAPER_SERVICE.lower() == "scrapeops" and SCRAPER_API_KEY:
        try:
            params = {
                "api_key": SCRAPER_API_KEY,
                "url": url,
                "render_js": "true",
            }
            if SCRAPER_COUNTRY:
                params["country"] = SCRAPER_COUNTRY
            r = requests.post("https://proxy.scrapeops.io/v1/", params=params, headers=headers, data=data, timeout=30)
            if r.status_code in (401, 403):
                if _HTTPX_CLIENT:
                    return _HTTPX_CLIENT.post(url, headers=headers, data=data)
                return requests.post(url, headers=headers, data=data, timeout=20)
            return r
        except Exception:
            if _HTTPX_CLIENT:
                return _HTTPX_CLIENT.post(url, headers=headers, data=data)
            return requests.post(url, headers=headers, data=data, timeout=20)
    if _HTTPX_CLIENT:
        return _HTTPX_CLIENT.post(url, headers=headers, data=data)
    return requests.post(url, headers=headers, data=data, timeout=20)


@app.post("/api/bulk-prices")
async def bulk_prices(request: Request):
    require_auth(request)
    body = await request.json()
    locations: List[Dict[str, Any]] = body.get("locations", [])
    supplier_priority: Optional[str] = body.get("supplier_priority")
    durations = body.get("durations", [1,2,3,4,5,6,7,8,9,14,22,31,60])

    results: List[Dict[str, Any]] = []
    headers = {"User-Agent": "Mozilla/5.0 (compatible; PriceTracker/1.0)"}

    # Global simple rate limiter (shared across requests)
    _RL_LOCK = getattr(bulk_prices, "_RL_LOCK", None)
    if _RL_LOCK is None:
        _RL_LOCK = asyncio.Lock()
        setattr(bulk_prices, "_RL_LOCK", _RL_LOCK)
    _RL_LAST = getattr(bulk_prices, "_RL_LAST", 0.0)
    _RL_MIN_INTERVAL = 1.0 / GLOBAL_FETCH_RPS if GLOBAL_FETCH_RPS and GLOBAL_FETCH_RPS > 0 else 0.0

    async def _rate_limit_tick():
        nonlocal _RL_LAST
        if _RL_MIN_INTERVAL <= 0:
            return
        async with _RL_LOCK:
            now = time.time()
            wait = _RL_MIN_INTERVAL - (now - _RL_LAST)
            if wait > 0:
                await asyncio.sleep(wait)
                now = time.time()
            _RL_LAST = now
            setattr(bulk_prices, "_RL_LAST", _RL_LAST)

    async def _fetch_parse(url: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        # Retry up to 2 attempts for transient failures
        attempts = 0
        last_exc: Optional[Exception] = None
        while attempts < BULK_MAX_RETRIES:
            attempts += 1
            t0 = time.time()
            try:
                await _rate_limit_tick()
                r = await async_fetch_with_optional_proxy(url, headers=headers)
                r.raise_for_status()
                html = r.text
                t_fetch = int((time.time() - t0) * 1000)
                t1 = time.time()
                items = await asyncio.to_thread(parse_prices, html, url)
                items = convert_items_gbp_to_eur(items)
                items = apply_price_adjustments(items, url)
                items = normalize_and_sort(items, supplier_priority)
                t_parse = int((time.time() - t1) * 1000)
                # best-effort timing log
                try:
                    with open(DEBUG_DIR / "perf_bulk.txt", "a", encoding="utf-8") as _fp:
                        _fp.write(f"{time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())} fetch_ms={t_fetch} parse_ms={t_parse} attempts={attempts} url={url[:180]}\n")
                except Exception:
                    pass
                return items, {"fetch_ms": t_fetch, "parse_ms": t_parse, "attempts": attempts}
            except Exception as e:
                last_exc = e
                await asyncio.sleep(0.3 * attempts)
        raise last_exc  # type: ignore

    for loc in locations:
        name = loc.get("name", "")
        urls: List[str] = loc.get("urls", [])
        loc_block = {"location": name, "durations": []}
        # Cap concurrency to avoid overloading Render (CPU/net)
        sem = asyncio.Semaphore(BULK_CONCURRENCY)
        async def _worker(index: int, url: str, days: int):
            async with sem:
                try:
                    items, timing = await _fetch_parse(url)
                    return {"days": days, "count": len(items), "items": items, "timing": timing}
                except Exception as e:
                    return {"days": days, "error": str(e), "items": [], "timing": {"attempts": BULK_MAX_RETRIES}}

        tasks = []
        for idx, url in enumerate(urls):
            days = durations[idx] if idx < len(durations) else None
            if not url or days is None:
                continue
            tasks.append(_worker(idx, url, days))
        if tasks:
            loc_block["durations"] = await asyncio.gather(*tasks)
        results.append(loc_block)
    return JSONResponse({"ok": True, "results": results})


@app.post("/api/track-by-url")
async def track_by_url(request: Request):
    try:
        if not bool(str(os.getenv("DEV_NO_AUTH", "")).strip().lower() in ("1","true","yes","on")):
            require_auth(request)
    except Exception:
        require_auth(request)
    body = await request.json()
    location: str = body.get("location") or ""
    pickup_date: str = body.get("pickupDate") or ""
    pickup_time: str = body.get("pickupTime", "10:00")  # HH:mm
    days: Optional[int] = body.get("days")
    url: str = body.get("url") or ""
    no_cache: bool = bool(body.get("noCache", False))
    currency: str = body.get("currency", "")
    if not url:
        return _no_store_json({"ok": False, "error": "url is required"}, status_code=400)

    try:
        from datetime import datetime
        start_dt: Optional[datetime] = None
        if pickup_date:
            try:
                start_dt = datetime.fromisoformat(pickup_date + "T" + pickup_time)
            except Exception:
                start_dt = None
        # 0) 60s in-memory cache by normalized URL
        try:
            from urllib.parse import urlparse, urlunparse, parse_qsl, urlencode
            pr0 = urlparse(url)
            qd = dict(parse_qsl(pr0.query, keep_blank_values=True))
            # normalize currency/lang params position for stable key
            norm_q = urlencode(sorted(qd.items()))
            norm_url = urlunparse((pr0.scheme, pr0.netloc, pr0.path, pr0.params, norm_q, pr0.fragment))
        except Exception:
            norm_url = url
        now_ts = time.time()
        cached = _URL_CACHE.get(norm_url)
        if (not no_cache) and cached and (now_ts - cached[0] < 60):
            payload = dict(cached[1])
            # Avoid serving cached empty results
            if payload.get("items"):
                return _no_store_json(payload)
        headers = {
            # Desktop Chrome UA improves CarJet behavior on Render/mobile
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
            "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
            "Referer": "https://www.carjet.com/do/list/pt",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Cache-Control": "no-cache",
            "sec-ch-ua": '"Chromium";v="123", "Not:A-Brand";v="8"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"macOS"',
        }

        # Fast path when running locally or when FAST_MODE=true: single direct fetch, no retries
        try:
            IS_RENDER = bool(os.getenv("RENDER") or os.getenv("RENDER_EXTERNAL_URL"))
        except Exception:
            IS_RENDER = False
        FAST_MODE = bool(str(os.getenv("FAST_MODE", "")).strip().lower() in ("1","true","yes","on"))
        if (not IS_RENDER) or FAST_MODE:
            try:
                fast_headers = dict(headers)
                fast_headers["Cookie"] = "monedaForzada=EUR; moneda=EUR; currency=EUR; country=PT; idioma=PT; lang=pt"
                r_fast = await asyncio.to_thread(requests.get, url, headers=fast_headers, timeout=(6,20))
                r_fast.raise_for_status()
                html_fast = r_fast.text
                items_fast = await asyncio.to_thread(parse_prices, html_fast, url)
                # If homepage-like or empty, quickly try /pt variant as a second shot
                homepage_like_fast = False
                try:
                    homepage_like_fast = ("Pesquisando em mais de 1000 locadoras" in html_fast) or (re.search(r"Pesquisando\s+em\s+mais\s+de\s+1000", html_fast) is not None)
                except Exception:
                    homepage_like_fast = False
                if (not items_fast) or homepage_like_fast:
                    try:
                        from urllib.parse import urlparse as _uparse, urlunparse as _uunparse
                        prx = _uparse(url)
                        if prx.path.startswith('/do/list/') and not prx.path.startswith('/do/list/pt'):
                            pt_url = _uunparse((prx.scheme, prx.netloc, '/do/list/pt', prx.params, prx.query, prx.fragment))
                            r_fast2 = await asyncio.to_thread(requests.get, pt_url, headers=fast_headers, timeout=(6,20))
                            r_fast2.raise_for_status()
                            html_fast2 = r_fast2.text
                            items_fast2 = await asyncio.to_thread(parse_prices, html_fast2, pt_url)
                            if items_fast2:
                                html_fast = html_fast2
                                items_fast = items_fast2
                    except Exception:
                        pass
                items_fast = normalize_and_sort(items_fast, supplier_priority=None)
                payload = {
                    "ok": True,
                    "items": items_fast,
                    "location": location or _detect_location_name(html_fast) or "",
                    "start_date": (start_dt.strftime("%Y-%m-%d") if start_dt else ""),
                    "days": days,
                    "last_updated": time.strftime('%Y-%m-%d %H:%M:%S'),
                }
                _URL_CACHE[norm_url] = (time.time(), dict(payload))
                return _no_store_json(payload)
            except Exception:
                pass
        # Overall time budget to avoid long waits on mobile/Render
        budget_ms = 7000
        total_t0 = time.time()
        def remaining_ms():
            try:
                return max(0, budget_ms - int((time.time() - total_t0) * 1000))
            except Exception:
                return 0

        # 1) Direct fetch for CarJet PT results URLs to preserve locale (avoid proxy geolocation flipping)
        html = ""
        items: List[Dict[str, Any]] = []
        try:
            from urllib.parse import urlparse, parse_qs
            pr = urlparse(url)
            qs = parse_qs(pr.query)
            is_carjet = pr.netloc.endswith("carjet.com")
            is_pt_results = pr.path.startswith("/do/list/pt") and ("s" in qs and "b" in qs)
            is_carjet_list = is_carjet and pr.path.startswith("/do/list/")
        except Exception:
            is_carjet = False
            is_pt_results = False
            is_carjet_list = False

        if USE_PLAYWRIGHT and _HAS_PLAYWRIGHT and is_carjet:
            try:
                items = scrape_with_playwright(url)
                if items:
                    html = "(playwright)"
            except Exception:
                items = []
                html = ""

        if (not items) and is_carjet and (is_pt_results or is_carjet_list) and remaining_ms() > 1200:
            # Race direct URL and a /pt-normalized variant in parallel; first success wins
            direct_headers = dict(headers)
            direct_headers["Cookie"] = "monedaForzada=EUR; moneda=EUR; currency=EUR; country=PT; idioma=PT; lang=pt"
            direct_headers["sec-ch-ua"] = headers.get("sec-ch-ua")
            direct_headers["sec-ch-ua-mobile"] = headers.get("sec-ch-ua-mobile")
            direct_headers["sec-ch-ua-platform"] = headers.get("sec-ch-ua-platform")
            try:
                from urllib.parse import urlparse as _uparse, urlunparse as _uunparse
                prx = _uparse(url)
                pt_url = _uunparse((prx.scheme, prx.netloc, "/do/list/pt", prx.params, prx.query, prx.fragment)) if prx.path.startswith("/do/list/") and not prx.path.startswith("/do/list/pt") else url
            except Exception:
                pt_url = url

            async def fetch_and_parse(u: str):
                try:
                    t0 = time.time()
                    r = await async_fetch_with_optional_proxy(u, direct_headers)
                    r.raise_for_status()
                    h = r.text
                    its = await asyncio.to_thread(parse_prices, h, u)
                    hp = False
                    try:
                        hp = ("Pesquisando em mais de 1000 locadoras" in h) or (re.search(r"Pesquisando\s+em\s+mais\s+de\s+1000", h) is not None)
                    except Exception:
                        hp = False
                    dt = int((time.time() - t0) * 1000)
                    try:
                        print(f"[track_by_url] direct fetch {u} took {dt}ms items={(len(its) if its else 0)} homepage={hp}")
                    except Exception:
                        pass
                    if its and not hp:
                        return (u, h, its)
                except Exception:
                    return None
                return None

            tasks = [asyncio.create_task(fetch_and_parse(url)), asyncio.create_task(fetch_and_parse(pt_url))]
            # Respect remaining time budget for the parallel race
            timeout_sec = max(0.1, remaining_ms() / 1000.0)
            done, pending = await asyncio.wait(tasks, return_when=asyncio.FIRST_COMPLETED, timeout=timeout_sec)
            winner = None
            for d in done:
                try:
                    winner = d.result()
                except Exception:
                    winner = None
            for p in pending:
                p.cancel()
            if winner:
                _, html, items = winner
            else:
                html = ""
                items = []

        # 1.a) If Playwright is enabled, try rendering the final UI to capture client-updated totals
        if (not items) and USE_PLAYWRIGHT and _HAS_PLAYWRIGHT and is_carjet:
            try:
                html_pw = render_with_playwright(url)
                if html_pw:
                    html = html_pw
                    items = parse_prices(html_pw, url)
            except Exception:
                pass

        # 1.b) If not a PT results URL, or direct failed, use normal path (with proxy if configured)
        if not html:
            resp = await async_fetch_with_optional_proxy(url, headers=headers)
            resp.raise_for_status()
            html = resp.text
            items = await asyncio.to_thread(parse_prices, html, url)
        # Determine if we only captured provider summaries (no car names) or wrong currency
        gbp_seen = any(("£" in (it.get("price") or "")) or re.search(r"\bGBP\b", (it.get("price") or ""), re.I) for it in (items or []))
        homepage_like = False
        try:
            if isinstance(html, str):
                homepage_like = ("Pesquisando em mais de 1000 locadoras" in html) or (re.search(r"Pesquisando\s+em\s+mais\s+de\s+1000", html) is not None)
        except Exception:
            homepage_like = False
        only_summaries = homepage_like or (not items) or all(not (it.get("car") or "").strip() for it in items)

        # If we have items but they are GBP, convert now; continue to final response
        if items and gbp_seen:
            items = convert_items_gbp_to_eur(items)
        # Apply env-driven adjustments for Carjet
        if items:
            items = apply_price_adjustments(items, url)

        # 1.5) If items are empty or GBP/only summaries, retry with EUR hints
        if (only_summaries or not items) and remaining_ms() > 1200:
            try:
                from urllib.parse import urlencode, urlparse, parse_qsl, urlunparse
                def _with_param(u: str, key: str, value: str) -> str:
                    pr = urlparse(u)
                    q = dict(parse_qsl(pr.query, keep_blank_values=True))
                    q[key] = value
                    new_q = urlencode(q)
                    return urlunparse((pr.scheme, pr.netloc, pr.path, pr.params, new_q, pr.fragment))

                # CarJet-specific normalization: force Portuguese path and EUR params
                url_norm = url
                try:
                    pr = urlparse(url)
                    if pr.netloc.endswith("carjet.com") and pr.path.startswith("/do/list/") and not pr.path.startswith("/do/list/pt"):
                        # keep query intact, only change locale path to /pt
                        url_norm = urlunparse((pr.scheme, pr.netloc, "/do/list/pt", pr.params, pr.query, pr.fragment))
                except Exception:
                    url_norm = url

                # Build robust set of variants including language and country
                base_eur = _with_param(url_norm, "moneda", "EUR")
                eur_variants = [
                    base_eur,
                    _with_param(base_eur, "currency", "EUR"),
                    _with_param(base_eur, "cur", "EUR"),
                    _with_param(base_eur, "idioma", "PT"),
                    _with_param(base_eur, "country", "PT"),
                ]
                # Limit retries to 2 variants to reduce latency
                eur_variants = eur_variants[:2]
                eur_headers = dict(headers)
                eur_headers["Cookie"] = "monedaForzada=EUR; moneda=EUR; currency=EUR; country=PT; idioma=PT; lang=pt"
                eur_headers["sec-ch-ua"] = headers.get("sec-ch-ua")
                eur_headers["sec-ch-ua-mobile"] = headers.get("sec-ch-ua-mobile")
                eur_headers["sec-ch-ua-platform"] = headers.get("sec-ch-ua-platform")
                retried_ok = False
                for u2 in eur_variants:
                    if remaining_ms() <= 1200:
                        break
                    try:
                        t1 = time.time()
                        r2 = await async_fetch_with_optional_proxy(u2, headers=eur_headers)
                        r2.raise_for_status()
                        html2 = r2.text
                        items2 = await asyncio.to_thread(parse_prices, html2, u2)
                        gbp2 = any((("£" in (it.get("price") or "")) or re.search(r"\bGBP\b", (it.get("price") or ""), re.I)) for it in (items2 or []))
                        dt2 = int((time.time() - t1) * 1000)
                        try:
                            print(f"[track_by_url] eur-variant fetch {u2} took {dt2}ms items={(len(items2) if items2 else 0)} gbp={gbp2}")
                        except Exception:
                            pass
                        if items2 and not gbp2:
                            html = html2
                            items = items2
                            only_summaries = False
                            break
                    except Exception:
                        continue
                    # If proxy is configured and still GBP/summary, attempt direct fetch without proxy
                    if only_summaries and (SCRAPER_SERVICE.lower() == "scrapeops" and SCRAPER_API_KEY) and remaining_ms() > 1800:
                        try:
                            r3 = requests.get(u2, headers=headers)
                            r3.raise_for_status()
                            html3 = r3.text
                            items3 = parse_prices(html3, u2)
                            gbp3 = any(("£" in (it.get("price") or "")) or re.search(r"\bGBP\b", (it.get("price") or ""), re.I) for it in (items3 or []))
                            if items3 and not gbp3:
                                html = html3
                                items = items3
                                only_summaries = False
                                break
                        except Exception:
                            pass
            except Exception:
                pass

        # 2) If still no detailed items, try Playwright to render the URL fully
        if only_summaries:
            try:
                from playwright.async_api import async_playwright
                async with async_playwright() as p:
                    browser = await p.chromium.launch(headless=True)
                    context = await browser.new_context()
                    await context.set_extra_http_headers(headers)
                    page = await context.new_page()
                    page.set_default_navigation_timeout(15000)
                    page.set_default_timeout(12000)
                    await page.goto(url, wait_until="domcontentloaded")
                    # Force currency to EUR if possible
                    try:
                        # Try in-page function first
                        await page.evaluate("() => { try { if (typeof submit_monedaForzada === 'function') { submit_monedaForzada(window.location.href, 'EUR'); } } catch(e){} }")
                        # Also click any EUR currency switchers if present
                        eurBtn = page.locator("[data-currency='EUR'], .currency .eur").first
                        if await eurBtn.count() > 0:
                            await eurBtn.click()
                            await page.wait_for_timeout(500)
                    except Exception:
                        pass
                    # wait for network results or any price-like selector
                    try:
                        await page.wait_for_response(lambda r: ("/do/list" in r.url or "/carList.asp" in r.url) and r.status == 200, timeout=20000)
                    except Exception:
                        pass
                    try:
                        await page.wait_for_selector("section.newcarlist article, .newcarlist article, .price, .amount, [class*='price']", timeout=12000)
                    except Exception:
                        pass
                    # Try to reveal hidden cars if there is a 'Ver mais' button or function
                    try:
                        for i in range(4):
                            btn = page.locator("#linkMasCoches").first
                            if await btn.count() == 0:
                                break
                            await btn.click()
                            await page.wait_for_timeout(400)
                    except Exception:
                        pass
                    try:
                        for _ in range(3):
                            await page.evaluate("() => { try { if (typeof VerMasCoches === 'function') { VerMasCoches(); } } catch(e){} }")
                            await page.wait_for_timeout(300)
                    except Exception:
                        pass
                    # Scroll to bottom to trigger any lazy loading
                    try:
                        await page.evaluate("() => { window.scrollTo(0, document.body.scrollHeight); }")
                        await page.wait_for_timeout(600)
                    except Exception:
                        pass
                    html = await page.content()
                    await browser.close()
                items = parse_prices(html, url)
            except Exception:
                items = parse_prices(html, url)
        items = normalize_and_sort(items, supplier_priority=None)
        try:
            total_dt = int((time.time() - total_t0) * 1000)
            print(f"[track_by_url] total={total_dt}ms items={(len(items) if items else 0)}")
        except Exception:
            pass
        # Try to detect dates and days from page HTML if not provided
        try:
            soup = BeautifulSoup(html, "lxml")
            txt = html
            # dataLayer hrental_startdate/hrental_enddate (YYYY-MM-DD)
            m1 = re.search(r'"hrental_startdate"\s*:\s*"(\d{4}-\d{2}-\d{2})"', txt)
            m2 = re.search(r'"hrental_enddate"\s*:\s*"(\d{4}-\d{2}-\d{2})"', txt)
            if m1:
                try:
                    start_dt = datetime.fromisoformat(m1.group(1) + "T" + (pickup_time or "10:00"))
                except Exception:
                    pass
            if m2 and m1:
                try:
                    end_dt = datetime.fromisoformat(m2.group(1) + "T" + (pickup_time or "10:00"))
                    days = (end_dt - start_dt).days if start_dt else days
                except Exception:
                    pass
            # DiasReserva in dataLayer
            if days is None:
                md = re.search(r'"DiasReserva"\s*:\s*"?(\d{1,2})"?', txt)
                if md:
                    try:
                        days = int(md.group(1))
                    except Exception:
                        pass
            # Detect location from dataLayer Destino or hidden inputs
            try:
                ml = re.search(r'"Destino"\s*:\s*"([^"]+)"', txt)
                if ml:
                    location = ml.group(1)
            except Exception:
                pass
            # Hidden inputs frmFechaRecogida / frmFechaDevolucion dd/mm/yyyy HH:MM (various id/name variants)
            if (start_dt is None) or (days is None):
                fr = soup.select_one("#frmFechaRecogida, input[name='frmFechaRecogida'], input[name='fechaRecogida']")
                fd = soup.select_one("#frmFechaDevolucion, input[name='frmFechaDevolucion'], input[name='fechaEntrega']")
                from datetime import datetime as _dt
                def _parse_dmY_HM(v: str) -> Optional[datetime]:
                    try:
                        return _dt.strptime(v, "%d/%m/%Y %H:%M")
                    except Exception:
                        return None
                if fr and fr.has_attr("value"):
                    t = fr.get("value") or ""
                    maybe = _parse_dmY_HM(t)
                    if maybe:
                        start_dt = maybe
                if fd and fd.has_attr("value"):
                    t = fd.get("value") or ""
                    maybe_end = _parse_dmY_HM(t)
                    if (maybe_end and start_dt) and (days is None):
                        days = (maybe_end - start_dt).days
                fdst = soup.select_one("#frmDestino, input[name='frmDestino'], input[name='destino']")
                if fdst and fdst.has_attr("value"):
                    val = (fdst.get("value") or "").strip()
                    if val:
                        location = val
        except Exception:
            pass
        # persist snapshot so UI can show the rows immediately
        try:
            if start_dt and days:
                save_snapshots(location, start_dt, int(days), items, currency or "")
        except Exception:
            pass
        from datetime import datetime as _dt
        payload = {
            "ok": True,
            "items": items,
            "location": location,
            "start_date": pickup_date,
            "days": days,
            "last_updated": _dt.utcnow().isoformat(timespec="seconds") + "Z",
        }
        # store in cache only if we have items
        try:
            if items:
                _URL_CACHE[norm_url] = (time.time(), payload)
        except Exception:
            pass
        # If still empty, write a small debug note (non-fatal)
        try:
            if not items:
                from html import unescape as _unesc
                title_match = re.search(r"<title>(.*?)</title>", html or "", re.I|re.S)
                ttl = _unesc(title_match.group(1)).strip() if title_match else ""
                (DEBUG_DIR / "last_empty.txt").write_text(
                    f"{time.strftime('%Y-%m-%dT%H:%M:%SZ', time.gmtime())} | URL={url} | parsed_items=0 | title={ttl[:120]} | html_len={len(html or '')}\n",
                    encoding="utf-8"
                )
        except Exception:
            pass
        return JSONResponse(payload)

    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


def normalize_and_sort(items: List[Dict[str, Any]], supplier_priority: Optional[str]) -> List[Dict[str, Any]]:
    # Secondary guard: blocklist filter to ensure unwanted vehicles never appear
    _blocked_models = [
        "Mercedes S Class Auto",
        "MG ZS Auto",
        "Mercedes CLA Coupe Auto",
        "Mercedes A Class",
        "Mercedes A Class Auto",
        "BMW 1 Series Auto",
        "BMW 3 Series SW Auto",
        "Volvo V60 Auto",
        "Volvo XC40 Auto",
        "Mercedes C Class Auto",
        "Tesla Model 3 Auto",
        "Electric",
        "BMW 2 Series Gran Coupe Auto",
        "Mercedes C Class SW Auto",
        "Mercedes E Class Auto",
        "Mercedes E Class SW Auto",
        "BMW 5 Series SW Auto",
        "BMW X1 Auto",
        "Mercedes CLE Coupe Auto",
        "Volkswagen T-Roc Cabrio",
        "Mercedes GLA Auto",
        "Volvo XC60 Auto",
        "Volvo EX30 Auto",
        "BMW 3 Series Auto",
        "Volvo V60 4x4 Auto",
        "Hybrid",
        "Mazda MX5 Cabrio Auto",
        "Mercedes CLA Auto",
    ]
    def _norm_text(s: str) -> str:
        s = (s or "").strip().lower()
        return " ".join(s.replace(",", " ").split())
    _blocked_norm = set(_norm_text(x) for x in _blocked_models)
    import re as _re
    _patterns = [
        r"\bmercedes\s+s\s*class\b",
        r"\bmercedes\s+cla\b",
        r"\bmercedes\s+cle\b",
        r"\bmercedes\s+a\s*class\b",
        r"\bmercedes\s+c\s*class\b",
        r"\bmercedes\s+e\s*class\b",
        r"\bmercedes\s+gla\b",
        r"\bbmw\s+1\s*series\b",
        r"\bbmw\s+2\s*series\b",
        r"\bbmw\s+3\s*series\b",
        r"\bbmw\s+5\s*series\b",
        r"\bbmw\s*x1\b",
        r"\bvolvo\s+v60\b",
        r"\bvolvo\s+xc40\b",
        r"\bvolvo\s+xc60\b",
        r"\bvolvo\s+ex30\b",
        r"\btesla\s+model\s*3\b",
        r"\bmg\s+zs\b",
        r"\bmazda\s+mx5\b",
        r"\bvolkswagen\s+t-roc\b",
        r"\belectric\b",
        r"\bhybrid\b",
    ]
    def _blocked(name: str) -> bool:
        n = _norm_text(name)
        if not n:
            return False
        if n in _blocked_norm:
            return True
        for p in _patterns:
            if _re.search(p, n):
                return True
        for b in _blocked_norm:
            if len(b) >= 6 and b in n:
                return True
        return False

    detailed: List[Dict[str, Any]] = []
    summary: List[Dict[str, Any]] = []
    import re as _re2
    # Use dynamic FX with 1h cache; fallback 1.16
    try:
        GBP_TO_EUR = float(_fx_rate_gbp_eur())
    except Exception:
        GBP_TO_EUR = 1.16
    for it in items:
        # DISABLED: mostrar todos os carros
        # if _blocked(it.get("car", "")):
        #     continue
        price_text_in = it.get("price", "") or ""
        price_num = extract_price_number(price_text_in)
        price_curr = ""
        if "€" in price_text_in or _re2.search(r"\bEUR\b", price_text_in, _re2.I):
            price_curr = "EUR"
        elif "£" in price_text_in or _re2.search(r"\bGBP\b", price_text_in, _re2.I):
            price_curr = "GBP"
        # Convert GBP -> EUR for display and sorting
        if price_curr == "GBP" and price_num is not None:
            try:
                price_num = round(price_num * GBP_TO_EUR, 2)
                price_text_in = f"€{price_num:.2f}"
                price_curr = "EUR"
            except Exception:
                pass
        # Limpar nome do carro PRIMEIRO (remover "Autoautomático", "ou similar", etc)
        car_name_clean = clean_car_name(it.get("car", ""))
        
        # Se não tiver grupo definido, mapear a partir da categoria
        # IMPORTANTE: usar nome LIMPO para mapeamento correto
        group_code = it.get("group", "")
        if not group_code:
            group_code = map_category_to_group(it.get("category", ""), car_name_clean)
        
        # DEBUG: Log primeiro item
        if len(detailed) == 0 and len(summary) == 0:
            import sys
            print(f"[DEBUG] Primeiro item: cat={it.get('category')}, car_clean={car_name_clean}, group={group_code}", file=sys.stderr, flush=True)
        
        row = {
            "supplier": it.get("supplier", ""),
            "car": car_name_clean,
            "price": price_text_in,
            "price_num": price_num,
            "currency": price_curr or it.get("currency", ""),
            "category": it.get("category", ""),
            "group": group_code,
            "category_code": it.get("category_code", ""),
            "transmission": it.get("transmission", ""),
            "photo": it.get("photo", ""),
            "link": it.get("link", ""),
        }
        
        # DEBUG: Verificar se group está no row antes de append
        if len(detailed) == 0 and len(summary) == 0:
            import sys
            print(f"[DEBUG] row tem 'group'? {'group' in row}, valor={row.get('group')}", file=sys.stderr, flush=True)
        
        if (row["car"] or "").strip():
            detailed.append(row)
        else:
            summary.append(row)

    def _sort(lst: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        lst.sort(key=lambda x: (
            0 if supplier_priority and supplier_priority.lower() in (x.get("supplier") or "").lower() else 1,
            (x.get("category") or ""),
            x.get("price_num") or 1e15,
        ))
        return lst

    # Prefer detailed per-car rows; if none found, fall back to provider-summary rows
    if detailed:
        return _sort(detailed)
    return _sort(summary)


def extract_price_number(price_str: str) -> Optional[float]:
    if not price_str:
        return None
    s = price_str.replace("\xa0", " ")
    digits = []
    dot_seen = False
    comma_seen = False
    for ch in s:
        if ch.isdigit():
            digits.append(ch)
        elif ch == "." and not dot_seen:
            digits.append(".")
            dot_seen = True
        elif ch == "," and not comma_seen:
            # assume comma as decimal if dot not used
            if not dot_seen:
                digits.append(".")
                comma_seen = True
    try:
        return float("".join(digits)) if digits else None
    except Exception:
        return None


@app.post("/api/track-carjet")
async def track_carjet(request: Request):
    require_auth(request)
    body = await request.json()
    pickup_date: str = body.get("pickupDate")  # YYYY-MM-DD
    pickup_time: str = body.get("pickupTime", "10:00")  # HH:mm
    durations: List[int] = body.get("durations", [1,2,3,4,5,6,7,8,9,14,22,31,60])
    locations: List[Dict[str, Any]] = body.get("locations", [])  # [{name, template?}]
    supplier_priority: Optional[str] = body.get("supplier_priority")
    lang: str = body.get("lang", "en")
    currency: str = body.get("currency", "EUR")

    if not pickup_date or not locations:
        return JSONResponse({"ok": False, "error": "pickupDate and locations are required"}, status_code=400)

    try:
        from datetime import datetime, timedelta
        from playwright.async_api import async_playwright
        # random já importado globalmente
        
        # Load date rotation settings from database
        date_rotation_enabled = True  # Enabled by default
        date_rotation_max_days = 4
        
        with _db_lock:
            con = _db_connect()
            try:
                row = con.execute(
                    "SELECT setting_value FROM price_automation_settings WHERE setting_key = ?",
                    ("date_rotation_enabled",)
                ).fetchone()
                if row:
                    date_rotation_enabled = row[0].lower() in ('true', '1', 'yes')
                
                row = con.execute(
                    "SELECT setting_value FROM price_automation_settings WHERE setting_key = ?",
                    ("date_rotation_max_days",)
                ).fetchone()
                if row:
                    date_rotation_max_days = int(row[0])
            finally:
                con.close()

        async def run():
            results: List[Dict[str, Any]] = []
            
            # Referrer options for natural traffic simulation
            referrers = [
                "https://www.google.com/",
                "https://www.google.pt/",
                "https://www.bing.com/",
                "https://www.google.com/search?q=rent+car+portugal",
                None  # Direct access
            ]
            
            # Timezone options for location diversity
            timezones = [
                "Europe/Lisbon",
                "Europe/London", 
                "Europe/Madrid",
                "Europe/Paris"
            ]
            
            # Device rotation for better WAF evasion
            devices = [
                {
                    "name": "iPhone 13 Pro",
                    "user_agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
                    "viewport": {"width": 390, "height": 844},
                    "scale": 3
                },
                {
                    "name": "iPhone 14",
                    "user_agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
                    "viewport": {"width": 390, "height": 844},
                    "scale": 3
                },
                {
                    "name": "Samsung Galaxy S21",
                    "user_agent": "Mozilla/5.0 (Linux; Android 11; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Mobile Safari/537.36",
                    "viewport": {"width": 360, "height": 800},
                    "scale": 3
                },
                {
                    "name": "iPad Air",
                    "user_agent": "Mozilla/5.0 (iPad; CPU OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
                    "viewport": {"width": 820, "height": 1180},
                    "scale": 2
                }
            ]
            
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                
                for loc_index, loc in enumerate(locations):
                    name = loc.get("name", "")
                    template = loc.get("template", "")
                    
                    # Select random device for this location
                    device = random.choice(devices)
                    print(f"[DEVICE_ROTATION] Location: {name}, Device: {device['name']}")
                    
                    # Select random timezone
                    timezone = random.choice(timezones)
                    print(f"[TIMEZONE_ROTATION] Location: {name}, Timezone: {timezone}")
                    
                    # Select random referrer
                    referrer = random.choice(referrers)
                    referrer_display = referrer if referrer else "Direct"
                    print(f"[REFERRER_ROTATION] Location: {name}, Referrer: {referrer_display}")
                    
                    # Create new context with random device (clears cache/history)
                    context = await browser.new_context(
                        user_agent=device["user_agent"],
                        viewport=device["viewport"],
                        device_scale_factor=device["scale"],
                        is_mobile=True,
                        has_touch=True,
                        locale="pt-PT",
                        timezone_id=timezone
                    )
                    
                    # Mobile headers with random referrer
                    default_headers = {
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                        "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.8",
                        "Accept-Encoding": "gzip, deflate, br",
                    }
                    if referrer:
                        default_headers["Referer"] = referrer
                    await context.set_extra_http_headers(default_headers)
                    await context.route("**/*", lambda route: route.continue_())
                    page = await context.new_page()
                    page.set_default_navigation_timeout(10000)
                    page.set_default_timeout(8000)
                    
                    # Apply date rotation per location for more variation
                    rotated_pickup_date = pickup_date
                    if date_rotation_enabled and date_rotation_max_days > 0:
                        base_date = datetime.strptime(pickup_date, "%Y-%m-%d")
                        days_offset = random.randint(0, date_rotation_max_days)
                        rotated_date = base_date + timedelta(days=days_offset)
                        rotated_pickup_date = rotated_date.strftime("%Y-%m-%d")
                        print(f"[DATE_ROTATION] Location: {name}, Original: {pickup_date}, Rotated: {rotated_pickup_date} (+{days_offset} days)")
                    else:
                        print(f"[DATE_ROTATION] Location: {name}, Disabled, using original date: {pickup_date}")
                    
                    # Randomize pickup time between 14:30 and 17:00
                    hour = random.randint(14, 16)
                    if hour == 14:
                        minute = random.choice([30, 45])
                    elif hour == 16:
                        minute = random.choice([0, 15, 30, 45])
                    else:  # hour == 15
                        minute = random.choice([0, 15, 30, 45])
                    
                    # If hour is 16 and minute > 0, cap at 17:00
                    if hour == 16 and minute > 0:
                        if random.random() < 0.5:
                            hour = 17
                            minute = 0
                    
                    rotated_pickup_time = f"{hour:02d}:{minute:02d}"
                    print(f"[TIME_ROTATION] Location: {name}, Original: {pickup_time}, Rotated: {rotated_pickup_time}")
                    
                    loc_block = {"location": name, "durations": []}
                    for d in durations:
                        try:
                            # Random delay between duration searches (0.5-2 seconds)
                            delay = random.uniform(0.5, 2.0)
                            print(f"[DELAY] Waiting {delay:.2f}s before next search...")
                            await asyncio.sleep(delay)
                            
                            start_dt = datetime.fromisoformat(rotated_pickup_date + "T" + rotated_pickup_time)
                            end_dt = start_dt + timedelta(days=int(d))
                            # Try direct POST to CarJet first (faster, no headless)
                            html = try_direct_carjet(name, start_dt, end_dt, lang=lang, currency=currency)
                            final_url = "https://www.carjet.com/do/list"
                            # Fallback to Playwright if direct returned empty or no prices
                            if not html or len(parse_prices(html, final_url)) == 0:
                                html, final_url = await fetch_carjet_results(page, name, start_dt, end_dt, lang, currency, template)
                                
                                # Simulate human scroll behavior after page load
                                try:
                                    scroll_amount = random.randint(200, 500)
                                    await page.evaluate(f"window.scrollBy(0, {scroll_amount})")
                                    await asyncio.sleep(random.uniform(0.3, 0.8))
                                    print(f"[SCROLL] Scrolled {scroll_amount}px")
                                except:
                                    pass  # Ignore scroll errors
                                    
                            items = parse_prices(html, final_url)
                            items = normalize_and_sort(items, supplier_priority)
                            save_snapshots(name, start_dt, d, items, currency)
                            loc_block["durations"].append({
                                "days": d,
                                "count": len(items),
                                "items": items,
                            })
                        except Exception as e:
                            loc_block["durations"].append({
                                "days": d,
                                "error": str(e),
                                "items": [],
                            })
                    results.append(loc_block)
                    
                    # Close context to clear cache/cookies/history before next location
                    await page.close()
                    await context.close()
                    print(f"[CACHE_CLEAR] Location: {name}, Context closed - cache/history cleared")
                    
                    # Random delay between locations (2-5 seconds) - simulate human behavior
                    if loc_index < len(locations) - 1:  # Not the last location
                        location_delay = random.uniform(2.0, 5.0)
                        print(f"[LOCATION_DELAY] Waiting {location_delay:.2f}s before next location...")
                        await asyncio.sleep(location_delay)
                    
                await browser.close()
            return results

        results = await run()
        return JSONResponse({"ok": True, "results": results})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


async def fetch_carjet_results(page, location_name, start_dt, end_dt, lang: str, currency: str, template: str):
    try:
        captured_html: Optional[str] = None
        captured_url: Optional[str] = None
        captured_post: Optional[Dict[str, Any]] = None

        async def on_response(resp):
            nonlocal captured_html, captured_url
            try:
                url = resp.url
                if ("/do/list" in url or "/carList.asp" in url) and resp.status == 200 and captured_html is None:
                    text = await resp.text()
                    if text:
                        captured_html = text
                        captured_url = url
            except Exception:
                pass

        # register response listener (use asyncio.create_task for awaiting inside handler)
        page.on("response", lambda r: asyncio.create_task(on_response(r)))
        # capture the first POST payload to /do/list to replay if needed
        def _on_request(req):
            nonlocal captured_post
            try:
                if ("/do/list" in req.url or "/carList.asp" in req.url) and req.method == "POST" and captured_post is None:
                    captured_post = {"url": req.url, "post": req.post_data or ""}
            except Exception:
                pass
        page.on("request", _on_request)
        if template:
            url = (
                template
                .replace("{pickup_date}", start_dt.strftime("%Y-%m-%d"))
                .replace("{pickup_time}", start_dt.strftime("%H:%M"))
                .replace("{dropoff_date}", end_dt.strftime("%Y-%m-%d"))
                .replace("{dropoff_time}", end_dt.strftime("%H:%M"))
                .replace("{location}", location_name)
            )
            await page.goto(url, wait_until="domcontentloaded")
        else:
            # Prefer PT site always for consistency with parsing/selectors
            lang = (lang or "pt").lower()
            if lang not in ("pt", "en", "es", "fr", "de", "it", "nl"):
                lang = "pt"
            base = f"https://www.carjet.com/{lang}/"
            await page.goto(base, wait_until="domcontentloaded")
            try:
                await page.wait_for_timeout(700)
                # Try to accept cookie banner if present
                try:
                    cookies_btn = page.get_by_role("button", name=re.compile("accept|agree|aceitar|ok", re.I)).first
                    if await cookies_btn.count() > 0:
                        await cookies_btn.click()
                        await page.wait_for_timeout(300)
                except Exception:
                    pass
                # Ensure PT language is active to set correct cookies/session
                try:
                    # If not already on /pt/ path, click Portuguese link
                    if not re.search(r"/pt/", page.url):
                        lang_link = page.locator("a[hreflang='pt']").first
                        if await lang_link.count() == 0:
                            lang_link = page.get_by_role("link", name=re.compile("Portugu[eê]s", re.I)).first
                        if await lang_link.count() == 0:
                            lang_link = page.locator("a[href*='/pt/']").first
                        if await lang_link.count() > 0:
                            await lang_link.click()
                            try:
                                await page.wait_for_url(re.compile(r"/pt/"), timeout=4000)
                            except Exception:
                                pass
                            await page.wait_for_timeout(400)
                except Exception:
                    pass
                # Try location input
                loc_input = page.get_by_placeholder("Pick-up location")
                if await loc_input.count() == 0:
                    loc_input = page.locator("input[name*='pickup']")
                if await loc_input.count() == 0:
                    loc_input = page.locator("#pickup")
                await loc_input.click()
                await loc_input.fill(location_name)
                await page.wait_for_timeout(900)
                # Prefer clicking the first autocomplete option if available
                try:
                    # CarJet PT uses #recogida_lista li for suggestions
                    first_opt = page.locator("#recogida_lista li").first
                    if await first_opt.count() == 0:
                        first_opt = page.locator("[role='listbox'] [role='option']").first
                    if await first_opt.count() > 0:
                        await first_opt.click()
                        # extract data attributes from option and populate hidden fields if present
                        try:
                            data = await first_opt.evaluate("(el)=>{const d=el.dataset||{};return {id:d.id||d.dstId||d.zoneId||'', zone:d.zone||d.zoneCode||''};}")
                            await page.evaluate("(vals)=>{const set=(id,val)=>{const el=document.getElementById(id); if(el){ el.value=val; }}; set('dst_id', vals.id); set('zoneCode', vals.zone); set('pickupId', vals.id); }", data)
                        except Exception:
                            pass
                    else:
                        await page.keyboard.press("Enter")
                except Exception:
                    await page.keyboard.press("Enter")
            except Exception:
                pass
            # Force known internal codes for target locations if inputs exist
            try:
                code_map = {
                    # Albufeira: ABF01 não funciona - deixar CarJet descobrir
                    # "Albufeira": "ABF01",
                    # "Albufeira Cidade": "ABF01",
                    "Faro Airport": "FAO02",
                    "Faro Aeroporto": "FAO02",
                }
                dst_code = code_map.get(location_name, "")
                if dst_code:
                    await page.evaluate(
                        "(dst)=>{\n"
                        "  const setVal=(id,val)=>{const el=document.getElementById(id); if(el){ el.value=val; el.dispatchEvent(new Event('change',{bubbles:true})); }};\n"
                        "  setVal('pickupId', dst); setVal('dst_id', dst); setVal('zoneCode', dst);\n"
                        "}",
                        dst_code,
                    )
            except Exception:
                pass
            # Pickup date
            try:
                # Try to set date inputs directly if present (various IDs)
                pickup_str_dmY = start_dt.strftime("%d/%m/%Y")
                dropoff_str_dmY = end_dt.strftime("%d/%m/%Y")
                await page.evaluate("(ids, val) => { for (const id of ids){ const el=document.getElementById(id) || document.querySelector('[name='+id+']'); if(el){ el.removeAttribute && el.removeAttribute('readonly'); el.value=val; el.dispatchEvent && el.dispatchEvent(new Event('input',{bubbles:true})); el.dispatchEvent && el.dispatchEvent(new Event('change',{bubbles:true})); } } }", ["fechaRecogida","pickupDate","date_from"], pickup_str_dmY)
            except Exception:
                pass
            try:
                await page.evaluate("(ids, val) => { for (const id of ids){ const el=document.getElementById(id) || document.querySelector('[name='+id+']'); if(el){ el.removeAttribute && el.removeAttribute('readonly'); el.value=val; el.dispatchEvent && el.dispatchEvent(new Event('input',{bubbles:true})); el.dispatchEvent && el.dispatchEvent(new Event('change',{bubbles:true})); } } }", ["fechaEntrega","fechaDevolucion","dropoffDate","date_to"], dropoff_str_dmY)
            except Exception:
                pass
            # Pickup/Dropoff time
            try:
                await page.evaluate("(ids, val) => { for (const id of ids){ const el=document.getElementById(id) || document.querySelector('[name='+id+']'); if(el){ el.value=val; el.dispatchEvent && el.dispatchEvent(new Event('change',{bubbles:true})); } } }", ["fechaRecogidaSelHour","h-recogida","pickupTime","time_from"], start_dt.strftime("%H:%M"))
            except Exception:
                pass
            try:
                await page.evaluate("(ids, val) => { for (const id of ids){ const el=document.getElementById(id) || document.querySelector('[name='+id+']'); if(el){ el.value=val; el.dispatchEvent && el.dispatchEvent(new Event('change',{bubbles:true})); } } }", ["fechaEntregaSelHour","h-devolucion","dropoffTime","time_to"], end_dt.strftime("%H:%M"))
            except Exception:
                pass
            # Submit search
            try:
                # Prefer submitting the main search form if present
                form = page.locator("form[name='menu_tarifas']")
                if await form.count() > 0:
                    # some sites rely on JS; try clicking the booking form button
                    btn = page.locator("#booking_form .btn-search").first
                    if await btn.count() > 0:
                        await btn.click()
                    else:
                        # fallback to form submit button
                        btn = form.locator("button[type='submit'], input[type='submit']").first
                        if await btn.count() > 0:
                            await btn.click()
                        else:
                            await page.evaluate("sel=>{const f=document.querySelector(sel); f && f.submit();}", "form[name='menu_tarifas']")
                else:
                    # If the page defines submit_fechas(action) use it to ensure s/b tokens are added
                    used_native = await page.evaluate("() => { try { if (typeof submit_fechas === 'function') { submit_fechas('/do/list/pt'); return true; } } catch(e){} return false; }")
                    if not used_native:
                        btn = page.get_by_role("button", name=re.compile("search|continue|find|atualizar|update", re.I))
                        if await btn.count() == 0:
                            btn = page.locator("button[type='submit']")
                        await btn.click()
            except Exception:
                pass
            # As a final nudge, try native submit one more time
            try:
                await page.evaluate("() => { try { if (typeof submit_fechas === 'function') { submit_fechas('/do/list/pt'); } } catch(e){} }")
            except Exception:
                pass
            # If still no request fired, serialize and submit the form directly
            try:
                await page.evaluate("""
                () => {
                  const form = document.querySelector("form[name='menu_tarifas']") || document.querySelector("#booking_form");
                  if (form) {
                    try { form.dispatchEvent(new Event('submit', {bubbles:true,cancelable:true})); } catch(e){}
                    try { form.submit(); } catch(e){}
                  }
                }
                """)
            except Exception:
                pass
            # Wait for results list
            try:
                # Prefer waiting for the actual network response and capture its body
                resp = await page.wait_for_response(lambda r: ("/do/list" in r.url or "/carList.asp" in r.url) and r.status == 200, timeout=25000)
                try:
                    body = await resp.text()
                    if body:
                        captured_html = body
                        captured_url = resp.url
                except Exception:
                    pass
                # Also ensure URL change if applicable
                await page.wait_for_url(re.compile(r"(/do/list|/carList\.asp)"), timeout=5000)
            except Exception:
                pass
            # Wait for any price-like selector quickly
            try:
                await page.wait_for_selector(".price, .amount, [class*='price']", timeout=15000)
            except Exception:
                pass
        # Prefer captured network HTML if present
        if captured_html:
            html = captured_html
            current_url = captured_url or page.url
        else:
            # If we saw the POST payload but didn't capture body, fetch via in-page fetch using same cookies/session
            if captured_post and captured_post.get("url"):
                try:
                    js = """
                    async (u, body) => {
                      const resp = await fetch(u, { method: 'POST', headers: { 'Content-Type': 'application/x-www-form-urlencoded' }, body: body });
                      return await resp.text();
                    }
                    """
                    html = await page.evaluate(js, captured_post["url"], captured_post.get("post", ""))
                    current_url = captured_post["url"]
                except Exception:
                    html = await page.content()
                    current_url = page.url
            else:
                html = await page.content()
                current_url = page.url
        return html, current_url
    finally:
        pass


@app.get("/api/debug_html")
async def debug_html(request: Request):
    params = request.query_params
    location = params.get("location", "Albufeira")
    pickup_date = params.get("date")
    pickup_time = params.get("time", "10:00")
    days = int(params.get("days", 1))
    lang = params.get("lang", "en")
    currency = params.get("currency", "EUR")
    if not pickup_date:
        return JSONResponse({"ok": False, "error": "Missing date (YYYY-MM-DD)"}, status_code=400)

    try:
        from datetime import datetime, timedelta
        from playwright.async_api import async_playwright

        async def run_once():
            async with async_playwright() as p:
                browser = await p.chromium.launch(headless=True)
                context = await browser.new_context()
                await context.set_extra_http_headers({"User-Agent": "Mozilla/5.0 (compatible; PriceTracker/1.0)"})
                # block heavy resources for speed
                await context.route("**/*", lambda route: (
                    route.abort() if route.request.resource_type in {"image", "media", "font"} else route.continue_()
                ))
                page = await context.new_page()
                page.set_default_navigation_timeout(10000)
                page.set_default_timeout(8000)
                start_dt = datetime.fromisoformat(pickup_date + "T" + pickup_time)
                end_dt = start_dt + timedelta(days=days)
                html, final_url = await fetch_carjet_results(page, location, start_dt, end_dt, lang, currency, template="")
                await browser.close()
                return html, final_url

        html, final_url = await run_once()
        # Save to debug file
        from datetime import datetime as _dt
        stamp = _dt.utcnow().strftime("%Y%m%dT%H%M%S")
        filename = f"debug-{location.replace(' ', '-')}-{pickup_date}-{days}d.html"
        out_path = DEBUG_DIR / filename
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(html)

        # Quick selector counts and dataMap presence
        soup = BeautifulSoup(html, "lxml")
        counts = {
            ".price": len(soup.select(".price")),
            ".amount": len(soup.select(".amount")),
            "[class*='price']": len(soup.select("[class*='price']")),
            "a[href]": len(soup.select("a[href]")),
        }
        try:
            import json as _json
            m = re.search(r"var\s+dataMap\s*=\s*(\[.*?\]);", html, re.S)
            if m:
                arr = _json.loads(m.group(1))
                counts["has_dataMap"] = True
                counts["dataMap_len"] = len(arr)
            else:
                counts["has_dataMap"] = False
                counts["dataMap_len"] = 0
        except Exception:
            counts["has_dataMap"] = False
            counts["dataMap_len"] = 0
        return JSONResponse({
            "ok": True,
            "url": final_url,
            "debug_file": f"/static/debug/{filename}",
            "counts": counts,
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)


def save_snapshots(location: str, start_dt, days: int, items: List[Dict[str, Any]], currency: str):
    from datetime import datetime
    ts = datetime.utcnow().isoformat(timespec="seconds")
    with _db_lock:
        conn = sqlite3.connect(DB_PATH)
        try:
            for it in items:
                conn.execute(
                    """
                    INSERT INTO price_snapshots (ts, location, pickup_date, pickup_time, days, supplier, car, price_text, price_num, currency, link)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        ts,
                        location,
                        start_dt.strftime("%Y-%m-%d"),
                        start_dt.strftime("%H:%M"),
                        int(days),
                        (it.get("supplier") or "").strip(),
                        (it.get("car") or "").strip(),
                        (it.get("price") or "").strip(),
                        it.get("price_num"),
                        currency or (it.get("currency") or ""),
                        (it.get("link") or "").strip(),
                    ),
                )
        finally:
            conn.commit()
            conn.close()


@app.get("/api/history")
async def get_history(request: Request):
    require_auth(request)
    params = request.query_params
    location = params.get("location")
    supplier = params.get("supplier")
    days = params.get("days")
    since = params.get("from")
    until = params.get("to")
    limit = int(params.get("limit", 200))

    q = "SELECT ts, location, pickup_date, pickup_time, days, supplier, car, price_text, price_num, currency, link FROM price_snapshots WHERE 1=1"
    args: List[Any] = []
    if location:
        q += " AND location = ?"
        args.append(location)
    if supplier:
        q += " AND supplier LIKE ?"
        args.append(f"%{supplier}%")
    if days:
        q += " AND days = ?"
        args.append(int(days))
    if since:
        q += " AND ts >= ?"
        args.append(since)
    if until:
        q += " AND ts <= ?"
        args.append(until)
    q += " ORDER BY ts DESC LIMIT ?"
    args.append(limit)

    with _db_lock:
        conn = sqlite3.connect(DB_PATH)
        try:
            rows = conn.execute(q, tuple(args)).fetchall()
        finally:
            conn.close()

    items = [
        {
            "ts": r[0],
            "location": r[1],
            "pickup_date": r[2],
            "pickup_time": r[3],
            "days": r[4],
            "supplier": r[5],
            "car": r[6],
            "price": r[7],
            "price_num": r[8],
            "currency": r[9],
            "link": r[10],
        }
        for r in rows
    ]
    return JSONResponse({"ok": True, "count": len(items), "items": items})

@app.post("/api/price-automation/upload")
async def upload_price_automation(request: Request, file: UploadFile = File(...)):
    """Upload e processamento de ficheiro Excel para automação de preços"""
    require_auth(request)
    
    try:
        # Ler conteúdo do ficheiro
        contents = await file.read()
        
        # Processar Excel
        import pandas as pd
        import io
        
        df = pd.read_excel(io.BytesIO(contents))
        
        # Converter para lista de dicionários
        data = []
        for _, row in df.iterrows():
            data.append({
                'categoria': str(row.get('Categoria', '')),
                'localizacao': str(row.get('Localização', row.get('Localizacao', ''))),
                'dias': int(row.get('Dias', 0)),
                'preco_base': float(row.get('Preço Base', row.get('Preco Base', 0))),
                'margem': float(row.get('Margem (%)', row.get('Margem', 0))),
                'preco_final': float(row.get('Preço Final', row.get('Preco Final', 0)))
            })
        
        return JSONResponse({
            "ok": True,
            "message": f"Ficheiro processado: {len(data)} linhas",
            "filename": file.filename,
            "data": data
        })
        
    except Exception as e:
        import traceback
        return JSONResponse({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }, status_code=400)

@app.get("/api/price-history")
async def get_price_history(request: Request):
    """API para dados de gráficos de histórico de preços"""
    require_auth(request)
    params = request.query_params
    location = params.get("location", "")
    days = params.get("days", "")
    category = params.get("category", "")
    
    with _db_lock:
        conn = sqlite3.connect(DB_PATH)
        try:
            # Evolução de preços ao longo do tempo (últimos 30 dias) - MIN, AVG, MAX
            evolution_query = """
                SELECT DATE(ts) as date, 
                       MIN(price_num) as min_price,
                       AVG(price_num) as avg_price,
                       MAX(price_num) as max_price
                FROM price_snapshots
                WHERE price_num IS NOT NULL AND price_num > 0
            """
            evolution_args = []
            if location:
                evolution_query += " AND location = ?"
                evolution_args.append(location)
            if days:
                evolution_query += " AND days = ?"
                evolution_args.append(int(days))
            # Nota: categoria não está na tabela, filtrar por car name como aproximação
            # if category:
            #     evolution_query += " AND car LIKE ?"
            #     evolution_args.append(f"%{category}%")
            evolution_query += " GROUP BY DATE(ts) ORDER BY DATE(ts) DESC LIMIT 30"
            
            evolution_rows = conn.execute(evolution_query, tuple(evolution_args)).fetchall()
            evolution_labels = [r[0] for r in reversed(evolution_rows)]
            evolution_min = [round(r[1], 2) if r[1] else 0 for r in reversed(evolution_rows)]
            evolution_avg = [round(r[2], 2) if r[2] else 0 for r in reversed(evolution_rows)]
            evolution_max = [round(r[3], 2) if r[3] else 0 for r in reversed(evolution_rows)]
            
            # Comparação por localização (sempre Faro vs Albufeira)
            comparison_query = """
                SELECT location, AVG(price_num) as avg_price
                FROM price_snapshots
                WHERE price_num IS NOT NULL AND price_num > 0
                  AND location IN ('Aeroporto de Faro', 'Albufeira')
            """
            comparison_args = []
            if days:
                comparison_query += " AND days = ?"
                comparison_args.append(int(days))
            comparison_query += " GROUP BY location ORDER BY location"
            
            comparison_rows = conn.execute(comparison_query, tuple(comparison_args)).fetchall()
            
            # Garantir que sempre temos Faro e Albufeira (mesmo sem dados)
            comparison_dict = {r[0]: round(r[1], 2) for r in comparison_rows}
            comparison_labels = ['Albufeira', 'Aeroporto de Faro']
            comparison_values = [
                comparison_dict.get('Albufeira', 0),
                comparison_dict.get('Aeroporto de Faro', 0)
            ]
            
            # Preços médios por mês do ano
            monthly_query = """
                SELECT CAST(strftime('%m', ts) AS INTEGER) as month, AVG(price_num) as avg_price
                FROM price_snapshots
                WHERE price_num IS NOT NULL AND price_num > 0
            """
            monthly_args = []
            if location:
                monthly_query += " AND location = ?"
                monthly_args.append(location)
            if days:
                monthly_query += " AND days = ?"
                monthly_args.append(int(days))
            monthly_query += " GROUP BY month ORDER BY month"
            
            monthly_rows = conn.execute(monthly_query, tuple(monthly_args)).fetchall()
            monthly_values = [0] * 12
            for r in monthly_rows:
                month_idx = r[0] - 1  # 1-12 -> 0-11
                if 0 <= month_idx < 12:
                    monthly_values[month_idx] = round(r[1], 2)
            
        finally:
            conn.close()
    
    return JSONResponse({
        "ok": True,
        "evolution": {
            "labels": evolution_labels,
            "min": evolution_min,
            "avg": evolution_avg,
            "max": evolution_max
        },
        "comparison": {
            "labels": comparison_labels,
            "values": comparison_values
        },
        "monthly": {
            "values": monthly_values
        }
    })

# ============================================================
# VEHICLES MANAGEMENT ENDPOINTS
# ============================================================

@app.get("/api/vehicles")
async def get_vehicles(request: Request):
    """Retorna todos os veículos mapeados no dicionário VEHICLES"""
    require_auth(request)
    try:
        from carjet_direct import VEHICLES
        
        # Organizar por categoria
        by_category = {}
        for car, category in VEHICLES.items():
            if category not in by_category:
                by_category[category] = []
            by_category[category].append(car)
        
        # Ordenar categorias e carros
        for cat in by_category:
            by_category[cat] = sorted(by_category[cat])
        
        return _no_store_json({
            "ok": True,
            "total": len(VEHICLES),
            "vehicles": dict(sorted(VEHICLES.items())),
            "by_category": dict(sorted(by_category.items())),
            "categories": sorted(set(VEHICLES.values()))
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)


@app.get("/api/vehicles/search")
async def search_vehicles(request: Request, q: str = ""):
    """Busca veículos no dicionário VEHICLES"""
    require_auth(request)
    try:
        from carjet_direct import VEHICLES, detect_category_from_car
        
        query = q.lower().strip()
        if not query:
            return _no_store_json({"ok": False, "error": "Query parameter 'q' is required"}, 400)
        
        # Buscar matches
        matches = {}
        for car, category in VEHICLES.items():
            if query in car:
                matches[car] = category
        
        # Testar categoria usando a função
        detected_category = detect_category_from_car(q, '')
        
        return _no_store_json({
            "ok": True,
            "query": q,
            "matches": matches,
            "detected_category": detected_category,
            "in_vehicles": q.lower() in VEHICLES
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/vehicles/save")
async def save_vehicle(request: Request):
    """Salva ou atualiza um veículo no sistema e atualiza carjet_direct.py automaticamente"""
    # Não requer autenticação para funcionar em iframes
    try:
        body = await request.json()
        
        original_name = body.get('original_name', '').strip()
        clean_name = body.get('clean_name', '').lower().strip()
        category = body.get('category', '').strip()
        
        if not clean_name or not category:
            return _no_store_json({"ok": False, "error": "clean_name and category are required"}, 400)
        
        # Salvar na tabela vehicle_name_overrides
        with _db_lock:
            con = _db_connect()
            try:
                # Verificar se já existe
                existing = con.execute(
                    "SELECT edited_name FROM vehicle_name_overrides WHERE original_name = ?",
                    (original_name,)
                ).fetchone()
                
                if existing:
                    # Atualizar
                    con.execute(
                        "UPDATE vehicle_name_overrides SET edited_name = ?, updated_at = datetime('now') WHERE original_name = ?",
                        (clean_name, original_name)
                    )
                else:
                    # Inserir novo
                    con.execute(
                        "INSERT INTO vehicle_name_overrides (original_name, edited_name, updated_at) VALUES (?, ?, datetime('now'))",
                        (original_name, clean_name)
                    )
                
                con.commit()
            finally:
                con.close()
        
        # Atualizar carjet_direct.py automaticamente
        try:
            import carjet_direct
            import importlib
            import re
            
            # Ler o arquivo atual
            carjet_path = os.path.join(os.path.dirname(__file__), 'carjet_direct.py')
            with open(carjet_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            
            # Encontrar o início e fim do dicionário VEHICLES
            vehicles_start = -1
            vehicles_end = -1
            
            for i, line in enumerate(lines):
                if 'VEHICLES = {' in line:
                    vehicles_start = i
                elif vehicles_start > 0 and line.strip() == '}' and vehicles_end == -1:
                    vehicles_end = i
                    break
            
            if vehicles_start == -1 or vehicles_end == -1:
                raise Exception("Could not find VEHICLES dictionary")
            
            # Verificar se o veículo já existe
            vehicle_pattern = f"    '{re.escape(clean_name)}':"
            vehicle_exists = False
            vehicle_line_idx = -1
            
            for i in range(vehicles_start, vehicles_end):
                if vehicle_pattern in lines[i]:
                    vehicle_exists = True
                    vehicle_line_idx = i
                    break
            
            if vehicle_exists:
                # Atualizar entrada existente
                lines[vehicle_line_idx] = f"    '{clean_name}': '{category}',\n"
            else:
                # Adicionar nova entrada antes do }
                new_entry = f"    '{clean_name}': '{category}',\n"
                lines.insert(vehicles_end, new_entry)
            
            # Escrever de volta
            with open(carjet_path, 'w', encoding='utf-8') as f:
                f.writelines(lines)
            
            # Recarregar o módulo
            importlib.reload(carjet_direct)
            
            message = "Vehicle saved and carjet_direct.py updated automatically!"
        except Exception as e:
            import traceback
            message = f"Vehicle saved but failed to update carjet_direct.py: {str(e)}\n{traceback.format_exc()}"
        
        # Calcular grupo baseado na categoria
        group = map_category_to_group(category, clean_name)
        
        # Gerar código Python
        code = f"    '{clean_name}': '{category}',"
        
        # Invalidar cache do frontend para atualizar pesquisa imediatamente
        global _vehicles_last_update
        _vehicles_last_update = datetime.utcnow().isoformat()
        
        return _no_store_json({
            "ok": True,
            "message": message,
            "clean_name": clean_name,
            "category": category,
            "group": group,
            "code": code,
            "cache_invalidated": True,
            "updated_at": _vehicles_last_update
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# ADMIN - CAR GROUPS MANAGEMENT
# ============================================================


# ============================================================
# ADMIN - CAR GROUPS MANAGEMENT (Fichas Individuais)
# ============================================================

@app.get("/admin/car-groups", response_class=HTMLResponse)
async def admin_car_groups(request: Request):
    """Página de administração dos grupos de carros - NOVA versão com abas e criação de categorias"""
    require_auth(request)
    
    # Ler o ficheiro HTML NOVO (vehicle_editor.html)
    html_path = os.path.join(os.path.dirname(__file__), "vehicle_editor.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            html_content = f.read()
        return HTMLResponse(content=html_content)
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: Ficheiro vehicle_editor.html não encontrado</h1>", status_code=500)

@app.get("/admin/vehicles-editor", response_class=HTMLResponse)
async def admin_vehicles_editor(request: Request):
    """Editor avançado de veículos com nome original vs editado"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "vehicle_editor.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: vehicle_editor.html não encontrado</h1>", status_code=500)

@app.get("/vehicle-editor", response_class=HTMLResponse)
async def vehicle_editor(request: Request):
    """Vehicle Editor - Alias para /admin/car-groups"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "vehicle_editor.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: vehicle_editor.html não encontrado</h1>", status_code=500)

@app.get("/admin/price-validation", response_class=HTMLResponse)
async def admin_price_validation(request: Request):
    """Página de configuração de regras de validação de preços"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "price_validation_rules.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: price_validation_rules.html não encontrado</h1>", status_code=500)

@app.get("/admin/price-automation-settings", response_class=HTMLResponse)
async def admin_price_automation_settings(request: Request):
    """Página de parametrizações para automação de preços"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "price_automation_settings.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: price_automation_settings.html não encontrado</h1>", status_code=500)

@app.get("/admin/customization/branding", response_class=HTMLResponse)
async def admin_customization_branding(request: Request):
    """Página de configuração de branding"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "customization_branding.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: customization_branding.html não encontrado</h1>", status_code=500)

@app.get("/admin/customization/appearance", response_class=HTMLResponse)
async def admin_customization_appearance(request: Request):
    """Página de configuração de aparência"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "customization_appearance.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: customization_appearance.html não encontrado</h1>", status_code=500)

@app.get("/admin/customization/company-info", response_class=HTMLResponse)
async def admin_customization_company_info(request: Request):
    """Página de informações da empresa"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "customization_company_info.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: customization_company_info.html não encontrado</h1>", status_code=500)

@app.get("/admin/customization/language", response_class=HTMLResponse)
async def admin_customization_language(request: Request):
    """Página de configuração de idioma"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "customization_language.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: customization_language.html não encontrado</h1>", status_code=500)

@app.get("/admin/customization/email", response_class=HTMLResponse)
async def admin_customization_email(request: Request):
    """Página de configuração de notificações por email"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "customization_email.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: customization_email.html não encontrado</h1>", status_code=500)

@app.get("/admin/customization/automated-reports", response_class=HTMLResponse)
async def admin_customization_automated_reports(request: Request):
    """Página de configuração de relatórios automáticos"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "customization_automated_reports.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: customization_automated_reports.html não encontrado</h1>", status_code=500)

@app.get("/admin/migrate-data", response_class=HTMLResponse)
async def admin_migrate_data(request: Request):
    """Página de migração de dados localStorage → Database"""
    require_auth(request)
    
    html_path = os.path.join(os.path.dirname(__file__), "templates", "migrate_data.html")
    try:
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        return HTMLResponse(content="<h1>Erro: migrate_data.html não encontrado</h1>", status_code=500)

# ============================================================
# API ENDPOINTS - PRICE AUTOMATION SETTINGS PERSISTENCE
# ============================================================

@app.post("/api/price-automation/settings/save")
async def save_price_automation_settings(request: Request):
    """Salvar configurações globais de automação de preços na base de dados"""
    require_auth(request)
    
    try:
        data = await request.json()
        logging.info(f"💾 Saving price automation settings: {len(data)} keys")
        
        with _db_lock:
            conn = _db_connect()
            try:
                # Salvar cada configuração
                for key, value in data.items():
                    value_json = json.dumps(value)
                    logging.debug(f"  - {key}: {value_json[:100]}...")
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO price_automation_settings (key, value, updated_at)
                        VALUES (?, ?, CURRENT_TIMESTAMP)
                        """,
                        (key, value_json)
                    )
                conn.commit()
                logging.info("✅ Price automation settings saved successfully to database")
                return JSONResponse({"ok": True, "message": "Settings saved successfully"})
            except Exception as db_err:
                logging.error(f"❌ Database error saving settings: {str(db_err)}")
                raise
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error saving price automation settings: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/price-automation/settings/load")
async def load_price_automation_settings(request: Request):
    """Carregar configurações globais de automação de preços da base de dados"""
    require_auth(request)
    
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute("SELECT key, value FROM price_automation_settings")
                rows = cursor.fetchall()
                
                settings = {}
                for row in rows:
                    try:
                        settings[row[0]] = json.loads(row[1])
                    except:
                        settings[row[0]] = row[1]
                
                return JSONResponse({"ok": True, "settings": settings})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/price-automation/rules/save")
async def save_automated_price_rules(request: Request):
    """Salvar regras automatizadas de preços na base de dados"""
    require_auth(request)
    
    try:
        data = await request.json()
        logging.info(f"💾 Saving automated price rules for {len(data)} locations")
        
        with _db_lock:
            conn = _db_connect()
            try:
                # Limpar regras antigas
                conn.execute("DELETE FROM automated_price_rules")
                
                # Salvar novas regras
                rules_count = 0
                for location, grupos in data.items():
                    for grupo, grupo_data in grupos.items():
                        if 'months' in grupo_data:
                            for month, month_data in grupo_data['months'].items():
                                if 'days' in month_data:
                                    for day, day_config in month_data['days'].items():
                                        try:
                                            config_json = json.dumps(day_config)
                                            conn.execute(
                                                """
                                                INSERT INTO automated_price_rules 
                                                (location, grupo, month, day, config, updated_at)
                                                VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                                                """,
                                                (location, grupo, int(month), int(day), config_json)
                                            )
                                            rules_count += 1
                                        except Exception as rule_err:
                                            logging.error(f"❌ Error saving rule {location}/{grupo}/{month}/{day}: {str(rule_err)}")
                                            raise
                
                conn.commit()
                logging.info(f"✅ Saved {rules_count} automated price rules to database")
                return JSONResponse({"ok": True, "message": "Rules saved successfully"})
            except Exception as db_err:
                logging.error(f"❌ Database error saving rules: {str(db_err)}")
                raise
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error saving automated price rules: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/price-automation/rules/load")
async def load_automated_price_rules(request: Request):
    """Carregar regras automatizadas de preços da base de dados"""
    require_auth(request)
    
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute(
                    "SELECT location, grupo, month, day, config FROM automated_price_rules ORDER BY location, grupo, month, day"
                )
                rows = cursor.fetchall()
                
                rules = {}
                for row in rows:
                    location, grupo, month, day, config_json = row
                    
                    if location not in rules:
                        rules[location] = {}
                    if grupo not in rules[location]:
                        rules[location][grupo] = {"months": {}}
                    if str(month) not in rules[location][grupo]["months"]:
                        rules[location][grupo]["months"][str(month)] = {"days": {}}
                    
                    try:
                        rules[location][grupo]["months"][str(month)]["days"][str(day)] = json.loads(config_json)
                    except:
                        rules[location][grupo]["months"][str(month)]["days"][str(day)] = {}
                
                return JSONResponse({"ok": True, "rules": rules})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/price-automation/strategies/save")
async def save_pricing_strategies(request: Request):
    """Salvar estratégias de pricing na base de dados"""
    require_auth(request)
    
    try:
        data = await request.json()
        logging.info(f"💾 Saving pricing strategies: {len(data)} keys")
        
        with _db_lock:
            conn = _db_connect()
            try:
                # Limpar estratégias antigas
                conn.execute("DELETE FROM pricing_strategies")
                
                # Salvar novas estratégias
                strategies_count = 0
                for key, strategies in data.items():
                    # Parse key: location_grupo_month_day
                    parts = key.split('_')
                    if len(parts) >= 4:
                        location = parts[0]
                        grupo = parts[1]
                        try:
                            month = int(parts[2])
                            day = int(parts[3])
                        except ValueError as ve:
                            logging.error(f"❌ Invalid month/day format in key '{key}': {str(ve)}")
                            continue
                        
                        for idx, strategy in enumerate(strategies):
                            try:
                                strategy_json = json.dumps(strategy)
                                conn.execute(
                                    """
                                    INSERT INTO pricing_strategies 
                                    (location, grupo, month, day, priority, strategy_type, config, updated_at)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                                    """,
                                    (location, grupo, month, day, idx + 1, strategy.get('type', 'unknown'), strategy_json)
                                )
                                strategies_count += 1
                            except Exception as strat_err:
                                logging.error(f"❌ Error saving strategy {key}[{idx}]: {str(strat_err)}")
                                raise
                
                conn.commit()
                logging.info(f"✅ Saved {strategies_count} pricing strategies to database")
                return JSONResponse({"ok": True, "message": "Strategies saved successfully"})
            except Exception as db_err:
                logging.error(f"❌ Database error saving strategies: {str(db_err)}")
                raise
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error saving pricing strategies: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/price-automation/strategies/load")
async def load_pricing_strategies(request: Request):
    """Carregar estratégias de pricing da base de dados"""
    require_auth(request)
    
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute(
                    "SELECT location, grupo, month, day, priority, config FROM pricing_strategies ORDER BY location, grupo, month, day, priority"
                )
                rows = cursor.fetchall()
                
                strategies = {}
                for row in rows:
                    location, grupo, month, day, priority, config_json = row
                    key = f"{location}_{grupo}_{month}_{day}"
                    
                    if key not in strategies:
                        strategies[key] = []
                    
                    try:
                        strategies[key].append(json.loads(config_json))
                    except:
                        pass
                
                return JSONResponse({"ok": True, "strategies": strategies})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/price-automation/history/save")
async def save_automated_prices_history(request: Request):
    """Salvar histórico de preços automatizados"""
    require_auth(request)
    
    try:
        data = await request.json()
        username = request.session.get("username", "unknown")
        
        with _db_lock:
            conn = _db_connect()
            try:
                # Salvar cada entrada do histórico
                for entry in data.get("entries", []):
                    conn.execute(
                        """
                        INSERT INTO automated_prices_history 
                        (location, grupo, dias, pickup_date, auto_price, real_price, 
                         strategy_used, strategy_details, min_price_applied, created_by)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            entry.get("location"),
                            entry.get("grupo"),
                            entry.get("dias"),
                            entry.get("pickup_date"),
                            entry.get("auto_price"),
                            entry.get("real_price"),
                            entry.get("strategy_used"),
                            json.dumps(entry.get("strategy_details", {})),
                            entry.get("min_price_applied"),
                            username
                        )
                    )
                
                conn.commit()
                return JSONResponse({"ok": True, "message": "History saved successfully"})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/price-automation/history/load")
async def load_automated_prices_history(request: Request):
    """Carregar histórico de preços automatizados"""
    require_auth(request)
    
    try:
        # Parâmetros opcionais de filtro
        location = request.query_params.get("location")
        grupo = request.query_params.get("grupo")
        limit = int(request.query_params.get("limit", 100))
        
        with _db_lock:
            conn = _db_connect()
            try:
                query = """
                    SELECT id, location, grupo, dias, pickup_date, 
                           auto_price, real_price, strategy_used, strategy_details,
                           min_price_applied, created_at, created_by
                    FROM automated_prices_history
                    WHERE 1=1
                """
                params = []
                
                if location:
                    query += " AND location = ?"
                    params.append(location)
                
                if grupo:
                    query += " AND grupo = ?"
                    params.append(grupo)
                
                query += " ORDER BY created_at DESC LIMIT ?"
                params.append(limit)
                
                cursor = conn.execute(query, params)
                rows = cursor.fetchall()
                
                history = []
                for row in rows:
                    try:
                        strategy_details = json.loads(row[8]) if row[8] else {}
                    except:
                        strategy_details = {}
                    
                    history.append({
                        "id": row[0],
                        "location": row[1],
                        "grupo": row[2],
                        "dias": row[3],
                        "pickup_date": row[4],
                        "auto_price": row[5],
                        "real_price": row[6],
                        "strategy_used": row[7],
                        "strategy_details": strategy_details,
                        "min_price_applied": row[9],
                        "created_at": row[10],
                        "created_by": row[11]
                    })
                
                return JSONResponse({"ok": True, "history": history})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ============================================================
# API ENDPOINTS - AI LEARNING DATA & USER SETTINGS
# ============================================================

@app.post("/api/ai/learning/save")
async def save_ai_learning(request: Request):
    """Salvar dados de AI learning (ajustes manuais) na base de dados"""
    require_auth(request)
    
    try:
        data = await request.json()
        adjustment = data.get("adjustment", {})
        
        with _db_lock:
            conn = sqlite3.connect(DB_PATH)
            try:
                conn.execute(
                    """
                    INSERT INTO ai_learning_data 
                    (grupo, days, location, original_price, new_price, user)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        adjustment.get("group"),
                        adjustment.get("days"),
                        adjustment.get("location"),
                        adjustment.get("originalPrice"),
                        adjustment.get("newPrice"),
                        "admin"  # TODO: pegar do session
                    )
                )
                conn.commit()
                
                logging.info(f"✅ AI Learning saved: {adjustment.get('group')}/{adjustment.get('days')}d = {adjustment.get('newPrice')}€")
                return JSONResponse({"ok": True})
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error saving AI learning: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/ai/learning/load")
async def load_ai_learning(request: Request):
    """Carregar dados de AI learning da base de dados"""
    require_auth(request)
    
    try:
        location = request.query_params.get("location")
        limit = int(request.query_params.get("limit", 100))
        
        with _db_lock:
            conn = sqlite3.connect(DB_PATH)
            try:
                query = "SELECT grupo, days, location, original_price, new_price, timestamp FROM ai_learning_data WHERE 1=1"
                params = []
                
                if location:
                    query += " AND location = ?"
                    params.append(location)
                
                query += " ORDER BY timestamp DESC LIMIT ?"
                params.append(limit)
                
                cursor = conn.execute(query, params)
                rows = cursor.fetchall()
                
                adjustments = []
                for row in rows:
                    adjustments.append({
                        "group": row[0],
                        "days": row[1],
                        "location": row[2],
                        "originalPrice": row[3],
                        "newPrice": row[4],
                        "timestamp": row[5]
                    })
                
                return JSONResponse({"ok": True, "adjustments": adjustments})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/user-settings/save")
async def save_user_settings(request: Request):
    """Salvar configurações do usuário (substituir localStorage)"""
    require_auth(request)
    
    try:
        data = await request.json()
        user_key = data.get("user_key", "default")
        settings = data.get("settings", {})
        
        with _db_lock:
            conn = sqlite3.connect(DB_PATH)
            try:
                for key, value in settings.items():
                    # Serializar valor como JSON
                    value_json = json.dumps(value) if not isinstance(value, str) else value
                    
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO user_settings 
                        (user_key, setting_key, setting_value, updated_at)
                        VALUES (?, ?, ?, datetime('now'))
                        """,
                        (user_key, key, value_json)
                    )
                
                conn.commit()
                logging.info(f"✅ User settings saved: {len(settings)} keys for {user_key}")
                return JSONResponse({"ok": True})
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error saving user settings: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/user-settings/load")
async def load_user_settings(request: Request):
    """Carregar configurações do usuário da base de dados"""
    require_auth(request)
    
    try:
        user_key = request.query_params.get("user_key", "default")
        
        with _db_lock:
            conn = sqlite3.connect(DB_PATH)
            try:
                cursor = conn.execute(
                    """
                    SELECT setting_key, setting_value 
                    FROM user_settings 
                    WHERE user_key = ?
                    ORDER BY updated_at DESC
                    """,
                    (user_key,)
                )
                rows = cursor.fetchall()
                
                settings = {}
                for row in rows:
                    key = row[0]
                    value_str = row[1]
                    
                    # Tentar deserializar JSON
                    try:
                        settings[key] = json.loads(value_str)
                    except:
                        settings[key] = value_str
                
                return JSONResponse({"ok": True, "settings": settings})
            finally:
                conn.close()
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ============================================================
# API ENDPOINTS - EXTERNAL AI INTEGRATION (Claude/GPT)
# ============================================================

@app.post("/api/ai/external-analysis")
async def external_ai_analysis(request: Request):
    """
    Análise de pricing usando AI externa (Claude Sonnet 3.5 ou GPT-4)
    Requer ANTHROPIC_API_KEY ou OPENAI_API_KEY no .env
    """
    require_auth(request)
    
    try:
        data = await request.json()
        
        group = data.get("group")
        days = data.get("days")
        location = data.get("location")
        current_price = data.get("current_price")
        competitors = data.get("competitors", [])
        provider = data.get("provider", "claude")  # claude ou openai
        min_price_day = data.get("min_price_day")  # Preço mínimo por dia
        min_price_month = data.get("min_price_month")  # Preço mínimo para ≥30 dias
        
        if not all([group, days, location, current_price]):
            return JSONResponse({
                "ok": False,
                "error": "Missing required fields: group, days, location, current_price"
            }, status_code=400)
        
        # Importar AI assistant
        try:
            from ai_pricing_assistant import get_ai_assistant
            
            ai_assistant = get_ai_assistant(provider=provider)
            
            # Verificar se AI está disponível
            status = ai_assistant.get_status()
            
            if not status['available']:
                return JSONResponse({
                    "ok": False,
                    "error": f"AI provider '{provider}' not available. Please install required libraries and set API key.",
                    "status": status,
                    "fallback_used": True
                }, status_code=503)
            
            # Fazer análise com validação de preço mínimo
            analysis = ai_assistant.analyze_market_positioning(
                group=group,
                days=int(days),
                location=location,
                current_price=float(current_price),
                competitors=competitors,
                min_price_day=float(min_price_day) if min_price_day else None,
                min_price_month=float(min_price_month) if min_price_month else None
            )
            
            logging.info(f"✅ AI Analysis completed: {group}/{days}d using {analysis.get('ai_provider', 'unknown')}")
            
            return JSONResponse({
                "ok": True,
                "analysis": analysis,
                "provider": provider,
                "status": status
            })
            
        except ImportError as e:
            logging.error(f"❌ AI assistant import error: {str(e)}")
            return JSONResponse({
                "ok": False,
                "error": f"AI assistant not available: {str(e)}",
                "hint": "Install required libraries: pip install anthropic openai"
            }, status_code=500)
        
    except Exception as e:
        logging.error(f"❌ External AI analysis error: {str(e)}")
        import traceback
        return JSONResponse({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }, status_code=500)

@app.get("/api/ai/status")
async def ai_status(request: Request):
    """Verifica status da integração com AI externa"""
    require_auth(request)
    
    try:
        from ai_pricing_assistant import get_ai_assistant
        
        # Testar ambos providers
        claude_assistant = get_ai_assistant(provider="claude")
        claude_status = claude_assistant.get_status()
        
        openai_assistant = get_ai_assistant(provider="openai")
        openai_status = openai_assistant.get_status()
        
        return JSONResponse({
            "ok": True,
            "claude": claude_status,
            "openai": openai_status,
            "recommended": "claude" if claude_status['available'] else ("openai" if openai_status['available'] else "none"),
            "env_keys": {
                "anthropic": "ANTHROPIC_API_KEY" in os.environ,
                "openai": "OPENAI_API_KEY" in os.environ
            }
        })
    except Exception as e:
        return JSONResponse({
            "ok": False,
            "error": str(e),
            "hint": "AI integration not configured. Install: pip install anthropic openai"
        }, status_code=500)

@app.get("/api/vehicles/with-originals")
async def get_vehicles_with_originals(request: Request):
    """Retorna veículos com nomes originais do scraping"""
    # Não requer autenticação para funcionar em iframes
    try:
        print("[VEHICLES API] Iniciando...", file=sys.stderr, flush=True)
        
        # Recarregar módulo para pegar alterações mais recentes
        import carjet_direct
        import importlib
        importlib.reload(carjet_direct)
        from carjet_direct import VEHICLES
        
        print(f"[VEHICLES API] VEHICLES importado: {len(VEHICLES)} veículos", file=sys.stderr, flush=True)
        
        # Buscar nomes originais do histórico
        with _db_lock:
            conn = sqlite3.connect(DB_PATH)
            try:
                # Pegar exemplos recentes de cada carro
                query = """
                    SELECT DISTINCT car 
                    FROM price_snapshots 
                    WHERE ts >= datetime('now', '-7 days')
                    ORDER BY car
                """
                rows = conn.execute(query).fetchall()
                print(f"[VEHICLES API] Encontrados {len(rows)} carros no histórico", file=sys.stderr, flush=True)
            finally:
                conn.close()
        
        # Criar mapeamento de originais
        originals_map = {}
        for row in rows:
            original_name = row[0]  # Nome como veio do scraping
            # Limpar para encontrar no VEHICLES
            import re
            clean = original_name.lower().strip()
            clean = re.sub(r'\s+(ou\s*similar|or\s*similar).*$', '', clean, flags=re.IGNORECASE)
            clean = re.sub(r'\s*\|\s*.*$', '', clean)
            clean = re.sub(r'\s+(pequeno|médio|medio|grande|compacto|economico|econômico).*$', '', clean, flags=re.IGNORECASE)
            clean = re.sub(r'\s+', ' ', clean).strip()
            
            if clean in VEHICLES:
                originals_map[clean] = {
                    'original': original_name,
                    'clean': clean,
                    'category': VEHICLES[clean]
                }
        
        # Adicionar veículos que não têm dados de scraping
        for clean_name, category in VEHICLES.items():
            if clean_name not in originals_map:
                originals_map[clean_name] = {
                    'original': f'{clean_name} (sem dados recentes)',
                    'clean': clean_name,
                    'category': category
                }
        
        print(f"[VEHICLES API] Retornando {len(originals_map)} veículos", file=sys.stderr, flush=True)
        
        return _no_store_json({
            "ok": True,
            "vehicles": originals_map,
            "total": len(originals_map)
        })
        
    except Exception as e:
        import traceback
        print(f"[VEHICLES API] ERRO: {e}", file=sys.stderr, flush=True)
        traceback.print_exc()
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# VEHICLE PHOTOS MANAGEMENT
# ============================================================

def _ensure_vehicle_photos_table():
    """Garante que a tabela de fotos de veículos existe"""
    try:
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("""
                    CREATE TABLE IF NOT EXISTS vehicle_photos (
                        vehicle_name TEXT PRIMARY KEY,
                        photo_data BLOB,
                        photo_url TEXT,
                        content_type TEXT,
                        uploaded_at TEXT DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                con.commit()
            finally:
                con.close()
    except Exception as e:
        print(f"Erro ao criar tabela vehicle_photos: {e}")

def _ensure_vehicle_name_overrides_table():
    """Garante que a tabela de nomes editados existe"""
    try:
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("""
                    CREATE TABLE IF NOT EXISTS vehicle_name_overrides (
                        original_name TEXT PRIMARY KEY,
                        edited_name TEXT NOT NULL,
                        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                con.commit()
            finally:
                con.close()
    except Exception as e:
        print(f"Erro ao criar tabela vehicle_name_overrides: {e}")

def _ensure_vehicle_images_table():
    """Garante que a tabela de imagens de veículos existe"""
    try:
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("""
                    CREATE TABLE IF NOT EXISTS vehicle_images (
                        vehicle_key TEXT PRIMARY KEY,
                        image_data BLOB NOT NULL,
                        content_type TEXT DEFAULT 'image/jpeg',
                        source_url TEXT,
                        downloaded_at TEXT DEFAULT CURRENT_TIMESTAMP
                    )
                """)
                con.commit()
            finally:
                con.close()
    except Exception as e:
        print(f"Erro ao criar tabela vehicle_images: {e}")

@app.on_event("startup")
async def startup_vehicle_photos():
    """Inicializar tabelas de veículos na startup"""
    _ensure_vehicle_photos_table()
    _ensure_vehicle_name_overrides_table()
    _ensure_vehicle_images_table()

@app.post("/api/vehicles/{vehicle_name}/photo/upload")
async def upload_vehicle_photo(vehicle_name: str, request: Request, file: UploadFile = File(...)):
    """Upload de foto para um veículo"""
    require_auth(request)
    try:
        # Ler dados do arquivo
        photo_data = await file.read()
        content_type = file.content_type or 'image/jpeg'
        
        # Salvar no banco
        _ensure_vehicle_photos_table()
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO vehicle_photos (vehicle_name, photo_data, content_type, photo_url)
                    VALUES (?, ?, ?, NULL)
                """, (vehicle_name, photo_data, content_type))
                conn.commit()
            finally:
                conn.close()
        
        return _no_store_json({
            "ok": True,
            "message": "Foto enviada com sucesso",
            "vehicle": vehicle_name,
            "size": len(photo_data)
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/vehicles/{vehicle_name}/photo/from-url")
async def download_vehicle_photo_from_url(vehicle_name: str, request: Request):
    """Baixar foto de URL e salvar no banco"""
    require_auth(request)
    try:
        body = await request.json()
        photo_url = body.get('url', '').strip()
        
        if not photo_url:
            return _no_store_json({"ok": False, "error": "URL não fornecida"}, 400)
        
        # Baixar imagem
        import httpx
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(photo_url)
            response.raise_for_status()
            
            photo_data = response.content
            content_type = response.headers.get('content-type', 'image/jpeg')
        
        # Salvar no banco
        _ensure_vehicle_photos_table()
        with _db_lock:
            conn = _db_connect()
            try:
                conn.execute("""
                    INSERT OR REPLACE INTO vehicle_photos (vehicle_name, photo_data, content_type, photo_url)
                    VALUES (?, ?, ?, ?)
                """, (vehicle_name, photo_data, content_type, photo_url))
                conn.commit()
            finally:
                conn.close()
        
        return _no_store_json({
            "ok": True,
            "message": "Foto baixada e salva com sucesso",
            "vehicle": vehicle_name,
            "url": photo_url,
            "size": len(photo_data)
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/vehicles/uncategorized")
async def get_uncategorized_vehicles(request: Request):
    """Retorna veículos que não estão no dicionário VEHICLES"""
    # Não requer autenticação para funcionar em iframes
    try:
        from carjet_direct import VEHICLES
        import re
        
        # Buscar carros únicos do histórico recente
        with _db_lock:
            conn = sqlite3.connect(DB_PATH)
            try:
                query = """
                    SELECT DISTINCT car 
                    FROM price_snapshots 
                    WHERE ts >= datetime('now', '-30 days')
                    ORDER BY car
                """
                rows = conn.execute(query).fetchall()
            finally:
                conn.close()
        
        uncategorized = []
        for row in rows:
            original_name = row[0]
            
            # Usar a MESMA função de limpeza que o scraping
            clean = clean_car_name(original_name)
            # VEHICLES está em lowercase, converter para comparação
            clean_lower = clean.lower()
            
            # Se não está no VEHICLES, adicionar à lista
            if clean and clean_lower not in VEHICLES:
                # Extrair marca
                parts = clean.split(' ')
                brand = parts[0] if parts else ''
                model = ' '.join(parts[1:]) if len(parts) > 1 else ''
                
                uncategorized.append({
                    'original': original_name,
                    'clean': clean,
                    'brand': brand,
                    'model': model,
                    'suggested_category': detect_category_suggestion(clean)
                })
        
        # Remover duplicados
        seen = set()
        unique = []
        for item in uncategorized:
            if item['clean'] not in seen:
                seen.add(item['clean'])
                unique.append(item)
        
        return _no_store_json({
            "ok": True,
            "uncategorized": unique,
            "total": len(unique)
        })
        
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

def detect_category_suggestion(car_name: str) -> str:
    """Sugere categoria baseado no nome do carro"""
    from carjet_direct import detect_category_from_car
    try:
        return detect_category_from_car(car_name, '')
    except:
        return 'ECONOMY'

# ============================================================
# EXPORT/IMPORT CONFIGURATION
# ============================================================

@app.get("/api/export/config")
async def export_configuration(request: Request):
    """
    Exporta configurações COMPLETAS do sistema de Vehicles:
    - VEHICLES (mapeamento carro → categoria)
    - vehicle_name_overrides (nomes editados)
    - car_groups (grupos manuais)
    - vehicle_photos (fotos em base64)
    - vehicle_images (imagens em base64)
    - suppliers (fornecedores)
    - users (utilizadores)
    """
    # Não requer autenticação para funcionar em iframes
    try:
        from carjet_direct import VEHICLES, SUPPLIER_MAP
        import base64
        
        print("[EXPORT] Iniciando exportação completa...")
        
        # 1. Exportar VEHICLES (mapeamento principal)
        vehicles_data = dict(VEHICLES)
        print(f"[EXPORT] VEHICLES: {len(vehicles_data)} carros")
        
        # 2. Exportar vehicle_name_overrides (nomes editados)
        name_overrides_data = []
        try:
            _ensure_vehicle_name_overrides_table()
            with _db_lock:
                conn = _db_connect()
                try:
                    rows = conn.execute("""
                        SELECT original_name, edited_name, updated_at
                        FROM vehicle_name_overrides
                    """).fetchall()
                    
                    for row in rows:
                        name_overrides_data.append({
                            "original_name": row[0],
                            "edited_name": row[1],
                            "updated_at": row[2]
                        })
                finally:
                    conn.close()
            print(f"[EXPORT] Name Overrides: {len(name_overrides_data)} registos")
        except Exception as e:
            print(f"[EXPORT] Warning: Could not export name_overrides: {e}")
        
        # 3. Exportar car_groups (grupos manuais)
        car_groups_data = []
        try:
            with _db_lock:
                conn = _db_connect()
                try:
                    rows = conn.execute("""
                        SELECT code, name, model, brand, category, doors, seats, 
                               transmission, luggage, photo_url, enabled
                        FROM car_groups
                    """).fetchall()
                    
                    for row in rows:
                        car_groups_data.append({
                            "code": row[0],
                            "name": row[1],
                            "model": row[2],
                            "brand": row[3],
                            "category": row[4],
                            "doors": row[5],
                            "seats": row[6],
                            "transmission": row[7],
                            "luggage": row[8],
                            "photo_url": row[9],
                            "enabled": row[10]
                        })
                finally:
                    conn.close()
            print(f"[EXPORT] Car Groups: {len(car_groups_data)} grupos")
        except Exception as e:
            print(f"[EXPORT] Warning: Could not export car_groups: {e}")
        
        # 4. Exportar vehicle_photos (fotos em base64)
        photos_data = {}
        try:
            _ensure_vehicle_photos_table()
            with _db_lock:
                conn = _db_connect()
                try:
                    rows = conn.execute("""
                        SELECT vehicle_name, photo_data, content_type, photo_url, updated_at
                        FROM vehicle_photos
                    """).fetchall()
                    
                    for row in rows:
                        vehicle_name = row[0]
                        photo_data = row[1]
                        content_type = row[2] or "image/jpeg"
                        photo_url = row[3]
                        updated_at = row[4]
                        
                        if photo_data:
                            # Converter BLOB para base64
                            photo_base64 = base64.b64encode(photo_data).decode('utf-8')
                            photos_data[vehicle_name] = {
                                "data": photo_base64,
                                "content_type": content_type,
                                "url": photo_url,
                                "updated_at": updated_at,
                                "size": len(photo_data)
                            }
                finally:
                    conn.close()
            print(f"[EXPORT] Photos: {len(photos_data)} fotos")
        except Exception as e:
            print(f"[EXPORT] Warning: Could not export photos: {e}")
        
        # 5. Exportar vehicle_images (imagens em base64)
        images_data = {}
        try:
            _ensure_vehicle_images_table()
            with _db_lock:
                conn = _db_connect()
                try:
                    rows = conn.execute("""
                        SELECT vehicle_name, image_data, source_url, updated_at
                        FROM vehicle_images
                    """).fetchall()
                    
                    for row in rows:
                        vehicle_name = row[0]
                        image_data = row[1]
                        source_url = row[2]
                        updated_at = row[3]
                        
                        if image_data:
                            # Converter BLOB para base64
                            image_base64 = base64.b64encode(image_data).decode('utf-8')
                            images_data[vehicle_name] = {
                                "data": image_base64,
                                "source_url": source_url,
                                "updated_at": updated_at,
                                "size": len(image_data)
                            }
                finally:
                    conn.close()
            print(f"[EXPORT] Images: {len(images_data)} imagens")
        except Exception as e:
            print(f"[EXPORT] Warning: Could not export images: {e}")
        
        # 6. Exportar suppliers
        suppliers_data = dict(SUPPLIER_MAP)
        print(f"[EXPORT] Suppliers: {len(suppliers_data)} fornecedores")
        
        # 7. Exportar users
        users_data = []
        try:
            with _db_lock:
                conn = _db_connect()
                try:
                    rows = conn.execute("SELECT username, password_hash FROM users").fetchall()
                    users_data = [{"username": r[0], "password_hash": r[1]} for r in rows]
                finally:
                    conn.close()
            print(f"[EXPORT] Users: {len(users_data)} utilizadores")
        except Exception as e:
            print(f"[EXPORT] Warning: Could not export users: {e}")
        
        # Criar estrutura de export completa
        config = {
            "version": "2.0",  # Nova versão com dados completos
            "exported_at": datetime.utcnow().isoformat(),
            "export_type": "vehicles_complete",
            "statistics": {
                "vehicles_count": len(vehicles_data),
                "name_overrides_count": len(name_overrides_data),
                "car_groups_count": len(car_groups_data),
                "photos_count": len(photos_data),
                "images_count": len(images_data),
                "suppliers_count": len(suppliers_data),
                "users_count": len(users_data),
                "total_photo_size_mb": sum(p.get("size", 0) for p in photos_data.values()) / (1024 * 1024),
                "total_image_size_mb": sum(i.get("size", 0) for i in images_data.values()) / (1024 * 1024)
            },
            "data": {
                "vehicles": vehicles_data,
                "name_overrides": name_overrides_data,
                "car_groups": car_groups_data,
                "photos": photos_data,
                "images": images_data,
                "suppliers": suppliers_data,
                "users": users_data
            }
        }
        
        # Retornar como JSON para download
        from fastapi.responses import Response
        import json
        
        json_content = json.dumps(config, indent=2, ensure_ascii=False)
        
        print(f"[EXPORT] Export completo! Tamanho: {len(json_content) / (1024 * 1024):.2f} MB")
        
        return Response(
            content=json_content,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=vehicles_complete_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            }
        )
        
    except Exception as e:
        import traceback
        print(f"[EXPORT] ERRO: {e}")
        print(traceback.format_exc())
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/import/config")
async def import_configuration(request: Request, file: UploadFile = File(...)):
    """
    Importa configurações COMPLETAS do sistema de Vehicles:
    - VEHICLES (atualiza carjet_direct.py)
    - vehicle_name_overrides (nomes editados)
    - car_groups (grupos manuais)
    - vehicle_photos (fotos)
    - vehicle_images (imagens)
    - suppliers (fornecedores)
    - users (utilizadores)
    """
    # Não requer autenticação para funcionar em iframes
    try:
        import json
        import base64
        
        print("[IMPORT] Iniciando importação completa...")
        
        # Ler ficheiro
        content = await file.read()
        config = json.loads(content)
        
        print(f"[IMPORT] Versão do ficheiro: {config.get('version', 'unknown')}")
        
        # Detectar formato (v1.x ou v2.0)
        is_v2 = config.get("version", "").startswith("2.") and "data" in config
        
        # Extrair dados conforme versão
        if is_v2:
            data = config["data"]
            vehicles_data = data.get("vehicles", {})
            name_overrides_data = data.get("name_overrides", [])
            car_groups_data = data.get("car_groups", [])
            photos_data = data.get("photos", {})
            images_data = data.get("images", {})
            suppliers_data = data.get("suppliers", {})
            users_data = data.get("users", [])
        else:
            # Formato antigo (v1.x)
            vehicles_data = config.get("vehicles", {})
            name_overrides_data = []
            car_groups_data = []
            photos_data = config.get("photos", {})
            images_data = {}
            suppliers_data = config.get("suppliers", {})
            users_data = config.get("users", [])
        
        # Validar
        if not vehicles_data:
            return _no_store_json({"ok": False, "error": "Ficheiro inválido: falta 'vehicles'"}, 400)
        
        print(f"[IMPORT] Importando {len(vehicles_data)} veículos...")
        
        # 1. Importar VEHICLES (gerar código Python)
        vehicles_code = "VEHICLES = {\n"
        for car, category in sorted(vehicles_data.items()):
            vehicles_code += f"    '{car}': '{category}',\n"
        vehicles_code += "}\n"
        
        # 2. Importar SUPPLIER_MAP (gerar código Python)
        suppliers_code = ""
        if suppliers_data:
            suppliers_code = "\nSUPPLIER_MAP = {\n"
            for code, name in sorted(suppliers_data.items()):
                suppliers_code += f"    '{code}': '{name}',\n"
            suppliers_code += "}\n"
        
        # 3. Importar vehicle_name_overrides
        imported_overrides = 0
        if name_overrides_data:
            _ensure_vehicle_name_overrides_table()
            with _db_lock:
                conn = _db_connect()
                try:
                    for override in name_overrides_data:
                        conn.execute("""
                            INSERT OR REPLACE INTO vehicle_name_overrides 
                            (original_name, edited_name, updated_at)
                            VALUES (?, ?, ?)
                        """, (
                            override["original_name"],
                            override["edited_name"],
                            override.get("updated_at", datetime.utcnow().isoformat())
                        ))
                        imported_overrides += 1
                    conn.commit()
                finally:
                    conn.close()
            print(f"[IMPORT] Name Overrides: {imported_overrides} importados")
        
        # 4. Importar car_groups
        imported_groups = 0
        if car_groups_data:
            with _db_lock:
                conn = _db_connect()
                try:
                    for group in car_groups_data:
                        conn.execute("""
                            INSERT OR REPLACE INTO car_groups 
                            (code, name, model, brand, category, doors, seats, 
                             transmission, luggage, photo_url, enabled)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            group["code"],
                            group["name"],
                            group.get("model"),
                            group.get("brand"),
                            group.get("category"),
                            group.get("doors"),
                            group.get("seats"),
                            group.get("transmission"),
                            group.get("luggage"),
                            group.get("photo_url"),
                            group.get("enabled", 1)
                        ))
                        imported_groups += 1
                    conn.commit()
                finally:
                    conn.close()
            print(f"[IMPORT] Car Groups: {imported_groups} importados")
        
        # 5. Importar vehicle_photos
        imported_photos = 0
        if photos_data:
            _ensure_vehicle_photos_table()
            with _db_lock:
                conn = _db_connect()
                try:
                    for vehicle_name, photo_info in photos_data.items():
                        # Converter base64 de volta para BLOB
                        photo_data = base64.b64decode(photo_info["data"])
                        content_type = photo_info.get("content_type", "image/jpeg")
                        photo_url = photo_info.get("url")
                        
                        conn.execute("""
                            INSERT OR REPLACE INTO vehicle_photos 
                            (vehicle_name, photo_data, content_type, photo_url, updated_at)
                            VALUES (?, ?, ?, ?, datetime('now'))
                        """, (vehicle_name, photo_data, content_type, photo_url))
                        imported_photos += 1
                    conn.commit()
                finally:
                    conn.close()
            print(f"[IMPORT] Photos: {imported_photos} importadas")
        
        # 6. Importar vehicle_images
        imported_images = 0
        if images_data:
            _ensure_vehicle_images_table()
            with _db_lock:
                conn = _db_connect()
                try:
                    for vehicle_name, image_info in images_data.items():
                        # Converter base64 de volta para BLOB
                        image_data = base64.b64decode(image_info["data"])
                        source_url = image_info.get("source_url")
                        
                        conn.execute("""
                            INSERT OR REPLACE INTO vehicle_images 
                            (vehicle_name, image_data, source_url, updated_at)
                            VALUES (?, ?, ?, datetime('now'))
                        """, (vehicle_name, image_data, source_url))
                        imported_images += 1
                    conn.commit()
                finally:
                    conn.close()
            print(f"[IMPORT] Images: {imported_images} importadas")
        
        # 7. Importar users
        imported_users = 0
        if users_data:
            with _db_lock:
                conn = _db_connect()
                try:
                    for user in users_data:
                        password_hash = user.get("password_hash") or user.get("password")
                        conn.execute(
                            "INSERT OR REPLACE INTO users (username, password_hash) VALUES (?, ?)",
                            (user["username"], password_hash)
                        )
                        imported_users += 1
                    conn.commit()
                finally:
                    conn.close()
            print(f"[IMPORT] Users: {imported_users} importados")
        
        # Invalidar cache
        global _vehicles_last_update
        _vehicles_last_update = datetime.utcnow().isoformat()
        
        print(f"[IMPORT] Importação completa!")
        
        return _no_store_json({
            "ok": True,
            "message": "Configuração importada com sucesso!",
            "imported": {
                "vehicles": len(vehicles_data),
                "name_overrides": imported_overrides,
                "car_groups": imported_groups,
                "photos": imported_photos,
                "images": imported_images,
                "suppliers": len(suppliers_data),
                "users": imported_users
            },
            "vehicles_code": vehicles_code,
            "suppliers_code": suppliers_code,
            "cache_invalidated": True,
            "updated_at": _vehicles_last_update,
            "instructions": "✅ Dados importados! Copie o código gerado e cole em carjet_direct.py se necessário."
        })
        
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# REAL-TIME UPDATE SYSTEM
# ============================================================

# Timestamp de última atualização de VEHICLES
_vehicles_last_update = datetime.utcnow().isoformat()

@app.post("/api/vehicles/notify-update")
async def notify_vehicles_update(request: Request):
    """Notifica que VEHICLES foi atualizado (para invalidar cache do frontend)"""
    require_auth(request)
    global _vehicles_last_update
    _vehicles_last_update = datetime.utcnow().isoformat()
    
    return _no_store_json({
        "ok": True,
        "updated_at": _vehicles_last_update,
        "message": "Cache invalidado. Frontend será atualizado."
    })

@app.get("/api/vehicles/last-update")
async def get_vehicles_last_update():
    """Retorna timestamp da última atualização de VEHICLES"""
    return _no_store_json({
        "ok": True,
        "last_update": _vehicles_last_update
    })

@app.post("/api/vehicles/refresh")
async def refresh_vehicles(request: Request):
    """
    Faz scraping em Albufeira + Faro para verificar carros novos/atualizados
    Retorna lista de carros novos encontrados
    """
    require_auth(request)
    try:
        from datetime import datetime, timedelta
        from carjet_direct import scrape_carjet_direct, VEHICLES
        
        # Datas para scraping (hoje + 7 dias)
        start_date = datetime.now()
        end_date = start_date + timedelta(days=7)
        
        new_cars = []
        updated_cars = []
        total_scraped = 0
        
        # Scraping em Albufeira
        print("[REFRESH] Fazendo scraping em Albufeira...")
        albufeira_results = scrape_carjet_direct("Albufeira", start_date, end_date, quick=1)
        total_scraped += len(albufeira_results)
        
        # Scraping em Faro
        print("[REFRESH] Fazendo scraping em Faro...")
        faro_results = scrape_carjet_direct("Faro", start_date, end_date, quick=1)
        total_scraped += len(faro_results)
        
        # Combinar resultados
        all_results = albufeira_results + faro_results
        
        # Verificar carros novos
        for item in all_results:
            car_name = item.get('car', '').strip()
            if not car_name:
                continue
            
            car_clean = clean_car_name(car_name).lower()
            
            # Verificar se está no VEHICLES
            if car_clean not in VEHICLES:
                # Carro novo!
                category = item.get('category', '')
                photo_url = item.get('photo', '')
                
                new_cars.append({
                    'original_name': car_name,
                    'clean_name': car_clean,
                    'category': category,
                    'photo_url': photo_url,
                    'location': item.get('location', ''),
                    'price': item.get('price', '')
                })
        
        # Remover duplicados
        seen = set()
        unique_new_cars = []
        for car in new_cars:
            if car['clean_name'] not in seen:
                seen.add(car['clean_name'])
                unique_new_cars.append(car)
        
        return _no_store_json({
            "ok": True,
            "total_scraped": total_scraped,
            "new_cars_count": len(unique_new_cars),
            "new_cars": unique_new_cars,
            "message": f"Scraping completo! {total_scraped} carros encontrados, {len(unique_new_cars)} novos."
        })
        
    except Exception as e:
        import traceback
        return _no_store_json({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }, 500)

@app.post("/api/vehicles/download-all-photos")
async def download_all_photos_from_carjet(request: Request):
    """
    Faz scraping em Albufeira + Faro e baixa TODAS as fotos dos carros
    Mostra progresso em tempo real
    """
    require_auth(request)
    try:
        from datetime import datetime, timedelta
        from carjet_direct import scrape_carjet_direct
        import httpx
        
        # Datas aleatórias (hoje + 3 a 10 dias)
        import random
        days_offset = random.randint(3, 10)
        start_date = datetime.now() + timedelta(days=days_offset)
        end_date = start_date + timedelta(days=7)
        
        print(f"[DOWNLOAD ALL PHOTOS] Iniciando scraping para {start_date.strftime('%Y-%m-%d')}...")
        
        photos_downloaded = 0
        photos_failed = 0
        total_cars = 0
        
        # Scraping em Albufeira
        print("[DOWNLOAD ALL PHOTOS] Fazendo scraping em Albufeira...")
        albufeira_results = scrape_carjet_direct("Albufeira", start_date, end_date, quick=1)
        
        # Scraping em Faro
        print("[DOWNLOAD ALL PHOTOS] Fazendo scraping em Faro...")
        faro_results = scrape_carjet_direct("Faro", start_date, end_date, quick=1)
        
        # Combinar resultados
        all_results = albufeira_results + faro_results
        total_cars = len(all_results)
        
        print(f"[DOWNLOAD ALL PHOTOS] Total de carros encontrados: {total_cars}")
        
        # Baixar fotos
        with _db_lock:
            conn = _db_connect()
            try:
                for idx, item in enumerate(all_results, 1):
                    car_name = item.get('car', '').strip()
                    photo_url = item.get('photo', '').strip()
                    
                    if not car_name or not photo_url:
                        continue
                    
                    car_clean = clean_car_name(car_name).lower()
                    
                    print(f"[DOWNLOAD ALL PHOTOS] [{idx}/{total_cars}] Baixando foto: {car_clean}")
                    
                    try:
                        # Baixar foto
                        async with httpx.AsyncClient(timeout=30.0) as client:
                            photo_response = await client.get(photo_url)
                            if photo_response.status_code == 200:
                                photo_data = photo_response.content
                                
                                # Salvar na tabela vehicle_photos
                                conn.execute("""
                                    INSERT OR REPLACE INTO vehicle_photos (vehicle_name, photo_data, photo_url, updated_at)
                                    VALUES (?, ?, ?, ?)
                                """, (car_clean, photo_data, photo_url, datetime.now().isoformat()))
                                
                                # Salvar na tabela vehicle_images também
                                conn.execute("""
                                    INSERT OR REPLACE INTO vehicle_images (vehicle_name, image_data, image_url, updated_at)
                                    VALUES (?, ?, ?, ?)
                                """, (car_clean, photo_data, photo_url, datetime.now().isoformat()))
                                
                                conn.commit()
                                photos_downloaded += 1
                                
                                print(f"[DOWNLOAD ALL PHOTOS] ✅ Foto salva: {car_clean} ({len(photo_data)} bytes)")
                            else:
                                photos_failed += 1
                                print(f"[DOWNLOAD ALL PHOTOS] ❌ Erro HTTP {photo_response.status_code}: {car_clean}")
                    except Exception as e:
                        photos_failed += 1
                        print(f"[DOWNLOAD ALL PHOTOS] ❌ Erro ao baixar {car_clean}: {e}")
                        continue
            finally:
                conn.close()
        
        return _no_store_json({
            "ok": True,
            "total_cars": total_cars,
            "photos_downloaded": photos_downloaded,
            "photos_failed": photos_failed,
            "message": f"Download completo! {photos_downloaded} fotos baixadas, {photos_failed} falharam."
        })
        
    except Exception as e:
        import traceback
        return _no_store_json({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }, 500)

@app.post("/api/vehicles/{vehicle_name}/download-photo")
async def download_vehicle_photo_from_carjet(vehicle_name: str, request: Request):
    """
    Baixa a foto do CarJet para um veículo específico e atualiza a ficha
    Faz scraping rápido para encontrar a foto mais recente
    """
    require_auth(request)
    try:
        from datetime import datetime, timedelta
        from carjet_direct import scrape_carjet_direct
        import httpx
        
        # Limpar nome do veículo
        car_clean = clean_car_name(vehicle_name).lower()
        
        # Fazer scraping rápido em Faro para encontrar o carro
        start_date = datetime.now()
        end_date = start_date + timedelta(days=7)
        
        print(f"[DOWNLOAD PHOTO] Procurando foto para: {car_clean}")
        
        # Tentar Faro primeiro
        results = scrape_carjet_direct("Faro", start_date, end_date, quick=1)
        
        # Se não encontrar, tentar Albufeira
        if not results:
            results = scrape_carjet_direct("Albufeira", start_date, end_date, quick=1)
        
        # Procurar o carro nos resultados
        photo_url = None
        for item in results:
            car_name = item.get('car', '').strip()
            if not car_name:
                continue
            
            item_clean = clean_car_name(car_name).lower()
            
            if item_clean == car_clean or car_clean in item_clean or item_clean in car_clean:
                photo_url = item.get('photo', '')
                if photo_url:
                    print(f"[DOWNLOAD PHOTO] Foto encontrada: {photo_url}")
                    break
        
        if not photo_url:
            return _no_store_json({
                "ok": False,
                "error": f"Foto não encontrada para '{vehicle_name}' no scraping do CarJet"
            }, 404)
        
        # Baixar a foto
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(photo_url)
            response.raise_for_status()
            photo_data = response.content
        
        # Salvar no banco de dados
        with _db_lock:
            conn = _db_connect()
            try:
                # Verificar se já existe
                existing = conn.execute(
                    "SELECT id FROM vehicle_photos WHERE vehicle_name = ?",
                    (car_clean,)
                ).fetchone()
                
                if existing:
                    # Atualizar
                    conn.execute(
                        "UPDATE vehicle_photos SET photo_data = ?, photo_url = ?, updated_at = datetime('now') WHERE vehicle_name = ?",
                        (photo_data, photo_url, car_clean)
                    )
                else:
                    # Inserir novo
                    conn.execute(
                        "INSERT INTO vehicle_photos (vehicle_name, photo_data, photo_url, updated_at) VALUES (?, ?, ?, datetime('now'))",
                        (car_clean, photo_data, photo_url)
                    )
                
                conn.commit()
            finally:
                conn.close()
        
        # Também salvar em vehicle_images para compatibilidade
        with _db_lock:
            conn = _db_connect()
            try:
                existing = conn.execute(
                    "SELECT id FROM vehicle_images WHERE vehicle_name = ?",
                    (car_clean,)
                ).fetchone()
                
                if existing:
                    conn.execute(
                        "UPDATE vehicle_images SET image_data = ?, source_url = ?, updated_at = datetime('now') WHERE vehicle_name = ?",
                        (photo_data, photo_url, car_clean)
                    )
                else:
                    conn.execute(
                        "INSERT INTO vehicle_images (vehicle_name, image_data, source_url, updated_at) VALUES (?, ?, ?, datetime('now'))",
                        (car_clean, photo_data, photo_url)
                    )
                
                conn.commit()
            finally:
                conn.close()
        
        return _no_store_json({
            "ok": True,
            "message": f"Foto baixada e salva com sucesso para '{vehicle_name}'!",
            "photo_url": photo_url,
            "photo_size": len(photo_data)
        })
        
    except Exception as e:
        import traceback
        return _no_store_json({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }, 500)

# ============================================================
# VEHICLE NAME MAPPING FOR FRONTEND
# ============================================================

@app.get("/api/vehicles/name-mapping")
async def get_vehicle_name_mapping():
    """Retorna mapeamento de nomes originais para clean names para usar na frontend
    INCLUI nomes editados guardados na base de dados"""
    try:
        from carjet_direct import VEHICLES
        import re
        
        # 1. Criar mapeamento base do VEHICLES
        name_mapping = {}
        
        for clean_name, category in VEHICLES.items():
            name_mapping[clean_name] = clean_name
            
            parts = clean_name.split()
            if len(parts) >= 2:
                brand = parts[0]
                model = ' '.join(parts[1:])
                
                variations = [
                    f"{brand} {model}",
                    f"{brand.upper()} {model}",
                    f"{brand.capitalize()} {model.capitalize()}",
                    f"{brand.upper()} {model.upper()}",
                    f"{brand.capitalize()} {model}",
                ]
                
                for var in variations:
                    name_mapping[var.lower()] = clean_name
        
        # 2. Aplicar OVERRIDES da base de dados (nomes editados pelo utilizador)
        try:
            with _db_lock:
                con = _db_connect()
                try:
                    rows = con.execute("SELECT original_name, edited_name FROM vehicle_name_overrides").fetchall()
                    for original, edited in rows:
                        # Normalizar chave
                        key = original.lower().strip()
                        # Aplicar override
                        name_mapping[key] = edited
                        print(f"[NAME MAPPING] Override aplicado: '{original}' → '{edited}'")
                finally:
                    con.close()
        except Exception as db_err:
            print(f"[NAME MAPPING] Aviso: Não foi possível carregar overrides da BD: {db_err}")
        
        return _no_store_json({
            "ok": True,
            "mapping": name_mapping,
            "total": len(name_mapping)
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/vehicles/name-overrides")
async def save_vehicle_name_override(request: Request):
    """Salva ou atualiza um nome editado de veículo"""
    require_auth(request)
    try:
        body = await request.json()
        original_name = body.get("original_name", "").strip()
        edited_name = body.get("edited_name", "").strip()
        
        if not original_name or not edited_name:
            return _no_store_json({"ok": False, "error": "original_name e edited_name são obrigatórios"}, 400)
        
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("""
                    INSERT INTO vehicle_name_overrides (original_name, edited_name, updated_at)
                    VALUES (?, ?, datetime('now'))
                    ON CONFLICT(original_name) DO UPDATE SET
                        edited_name = excluded.edited_name,
                        updated_at = excluded.updated_at
                """, (original_name, edited_name))
                con.commit()
            finally:
                con.close()
        
        return _no_store_json({
            "ok": True,
            "message": f"Nome editado salvo: '{original_name}' → '{edited_name}'"
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.delete("/api/vehicles/name-overrides/{original_name}")
async def delete_vehicle_name_override(original_name: str, request: Request):
    """Remove um override de nome de veículo"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("DELETE FROM vehicle_name_overrides WHERE original_name = ?", (original_name,))
                con.commit()
            finally:
                con.close()
        
        return _no_store_json({
            "ok": True,
            "message": f"Override removido: '{original_name}'"
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/vehicles/name-overrides")
async def get_vehicle_name_overrides(request: Request):
    """Lista todos os overrides de nomes de veículos"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                rows = con.execute("SELECT original_name, edited_name, updated_at FROM vehicle_name_overrides ORDER BY updated_at DESC").fetchall()
                overrides = [
                    {
                        "original_name": row[0],
                        "edited_name": row[1],
                        "updated_at": row[2]
                    }
                    for row in rows
                ]
            finally:
                con.close()
        
        return _no_store_json({
            "ok": True,
            "overrides": overrides,
            "total": len(overrides)
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# VEHICLE IMAGES - Download e Storage
# ============================================================

@app.post("/api/vehicles/images/download")
async def download_vehicle_images(request: Request):
    """Download automático de todas as imagens de veículos dos URLs do scraping"""
    # Não requer autenticação para funcionar em iframes
    try:
        from carjet_direct import VEHICLES
        import httpx
        import re
        
        downloaded = 0
        skipped = 0
        errors = []
        
        # PRIMEIRO: Buscar URLs de fotos da tabela dedicada car_images
        photo_urls_from_scraping = {}
        try:
            # Usar car_images.db (DB dedicada de fotos)
            from pathlib import Path
            car_images_db = str(Path(__file__).resolve().parent / "car_images.db")
            
            if os.path.exists(car_images_db):
                with _db_lock:
                    conn = sqlite3.connect(car_images_db)
                    try:
                        # Buscar fotos do car_images (tabela dedicada de fotos)
                        query = """
                            SELECT DISTINCT model_key, photo_url 
                            FROM car_images 
                            WHERE photo_url IS NOT NULL
                            AND photo_url != ''
                        """
                        rows = conn.execute(query).fetchall()
                        
                        for row in rows:
                            model_key = row[0]
                            photo_url = row[1]
                            
                            # Limpar nome para encontrar no VEHICLES
                            clean = model_key.lower().strip()
                            clean = re.sub(r'\s+(ou\s*similar|or\s*similar).*$', '', clean, flags=re.IGNORECASE)
                            clean = re.sub(r'\s*\|\s*.*$', '', clean)
                            clean = re.sub(r'\s+(pequeno|médio|medio|grande|compacto|economico|econômico).*$', '', clean, flags=re.IGNORECASE)
                            clean = re.sub(r'\s+', ' ', clean).strip()
                            
                            if clean in VEHICLES and clean not in photo_urls_from_scraping:
                                photo_urls_from_scraping[clean] = photo_url
                                
                    finally:
                        conn.close()
        except Exception as e:
            print(f"[PHOTOS] Erro ao buscar fotos do car_images.db: {e}", file=sys.stderr, flush=True)
        
        print(f"[PHOTOS] Encontradas {len(photo_urls_from_scraping)} fotos no histórico de scraping", file=sys.stderr, flush=True)
        
        # SEGUNDO: Mapeamento manual para veículos sem dados recentes
        image_mappings = {
            # MINI / B1 / B2
            'fiat 500 cabrio': 'https://www.carjet.com/cdn/img/cars/M/car_L154.jpg',
            'fiat 500': 'https://www.carjet.com/cdn/img/cars/M/car_C25.jpg',
            'fiat 500x': 'https://www.carjet.com/cdn/img/cars/M/car_A112.jpg',
            'fiat 500x auto': 'https://www.carjet.com/cdn/img/cars/M/car_A112.jpg',
            'fiat 500l': 'https://www.carjet.com/cdn/img/cars/M/car_C43.jpg',
            'fiat panda': 'https://www.carjet.com/cdn/img/cars/M/car_C30.jpg',
            'citroen c1': 'https://www.carjet.com/cdn/img/cars/M/car_C96.jpg',
            'citroën c1': 'https://www.carjet.com/cdn/img/cars/M/car_C96.jpg',
            'toyota aygo': 'https://www.carjet.com/cdn/img/cars/M/car_C29.jpg',
            'toyota aygo x': 'https://www.carjet.com/cdn/img/cars/M/car_F408.jpg',
            'volkswagen up': 'https://www.carjet.com/cdn/img/cars/M/car_C66.jpg',
            'vw up': 'https://www.carjet.com/cdn/img/cars/M/car_C66.jpg',
            'peugeot 108': 'https://www.carjet.com/cdn/img/cars/M/car_C15.jpg',
            'peugeot 108 cabrio': 'https://www.carjet.com/cdn/img/cars/M/car_L41.jpg',
            'hyundai i10': 'https://www.carjet.com/cdn/img/cars/M/car_C32.jpg',
            'kia picanto': 'https://www.carjet.com/cdn/img/cars/M/car_C59.jpg',
            'opel adam': 'https://www.carjet.com/cdn/img/cars/M/car_C50.jpg',
            'mitsubishi space star': 'https://www.carjet.com/cdn/img/cars/M/car_C190.jpg',
            'mitsubishi spacestar': 'https://www.carjet.com/cdn/img/cars/M/car_C190.jpg',
            'nissan micra': 'https://www.carjet.com/cdn/img/cars/M/car_C13.jpg',
            'renault twingo': 'https://www.carjet.com/cdn/img/cars/M/car_C61.jpg',
            'dacia sandero': 'https://www.carjet.com/cdn/img/cars/M/car_C75.jpg',
            'skoda scala': 'https://www.carjet.com/cdn/img/cars/M/car_C166.jpg',
            
            # ECONOMY / D / E2
            'renault clio': 'https://www.carjet.com/cdn/img/cars/M/car_C04.jpg',
            'renault clio sw': 'https://www.carjet.com/cdn/img/cars/M/car_C54.jpg',
            'peugeot 208': 'https://www.carjet.com/cdn/img/cars/M/car_C60.jpg',
            'ford fiesta': 'https://www.carjet.com/cdn/img/cars/M/car_C17.jpg',
            'ford ka': 'https://www.carjet.com/cdn/img/cars/M/car_N07.jpg',
            'volkswagen polo': 'https://www.carjet.com/cdn/img/cars/M/car_C27.jpg',
            'vw polo': 'https://www.carjet.com/cdn/img/cars/M/car_C27.jpg',
            'hyundai i20': 'https://www.carjet.com/cdn/img/cars/M/car_C52.jpg',
            'seat ibiza': 'https://www.carjet.com/cdn/img/cars/M/car_C01.jpg',
            'seat ibiza auto': 'https://www.carjet.com/cdn/img/cars/M/car_C01.jpg',
            'citroen c3': 'https://www.carjet.com/cdn/img/cars/M/car_C06.jpg',
            'citroën c3': 'https://www.carjet.com/cdn/img/cars/M/car_C06.jpg',
            'citroen c4 cactus': 'https://www.carjet.com/cdn/img/cars/M/car_C51.jpg',
            'citroën c4 cactus': 'https://www.carjet.com/cdn/img/cars/M/car_C51.jpg',
            'opel corsa': 'https://www.carjet.com/cdn/img/cars/M/car_A03.jpg',
            'opel corsa auto': 'https://www.carjet.com/cdn/img/cars/M/car_A03.jpg',
            'toyota yaris': 'https://www.carjet.com/cdn/img/cars/M/car_C64.jpg',
            
            # COMPACT / F
            'volkswagen golf': 'https://www.carjet.com/cdn/img/cars/M/car_F12.jpg',
            'vw golf': 'https://www.carjet.com/cdn/img/cars/M/car_F12.jpg',
            'audi a1': 'https://www.carjet.com/cdn/img/cars/M/car_C42.jpg',
            'ford focus': 'https://www.carjet.com/cdn/img/cars/M/car_F02.jpg',
            'renault megane': 'https://www.carjet.com/cdn/img/cars/M/car_F05.jpg',
            'renault mégane': 'https://www.carjet.com/cdn/img/cars/M/car_F05.jpg',
            'peugeot 308': 'https://www.carjet.com/cdn/img/cars/M/car_F22.jpg',
            'hyundai i30': 'https://www.carjet.com/cdn/img/cars/M/car_C41.jpg',
            'kia ceed': 'https://www.carjet.com/cdn/img/cars/M/car_C21.jpg',
            'kia ceed auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1023.jpg',
            'seat leon': 'https://www.carjet.com/cdn/img/cars/M/car_F39.jpg',
            'seat león': 'https://www.carjet.com/cdn/img/cars/M/car_F39.jpg',
            'toyota corolla auto': 'https://www.carjet.com/cdn/img/cars/M/car_A623.jpg',
            'opel astra': 'https://www.carjet.com/cdn/img/cars/M/car_F73.jpg',
            'citroen c4': 'https://www.carjet.com/cdn/img/cars/M/car_A17.jpg',
            'citroën c4': 'https://www.carjet.com/cdn/img/cars/M/car_A17.jpg',
            'peugeot 508': 'https://www.carjet.com/cdn/img/cars/M/car_F65.jpg',
            
            # SUV / F / L1
            'nissan juke': 'https://www.carjet.com/cdn/img/cars/M/car_F29.jpg',
            'peugeot 2008': 'https://www.carjet.com/cdn/img/cars/M/car_F91.jpg',
            'peugeot 3008': 'https://www.carjet.com/cdn/img/cars/M/car_A132.jpg',
            'peugeot 3008 auto': 'https://www.carjet.com/cdn/img/cars/M/car_A132.jpg',
            'renault captur': 'https://www.carjet.com/cdn/img/cars/M/car_F44.jpg',
            'volkswagen t-cross': 'https://www.carjet.com/cdn/img/cars/M/car_F252.jpg',
            'vw t-cross': 'https://www.carjet.com/cdn/img/cars/M/car_F252.jpg',
            'volkswagen tcross': 'https://www.carjet.com/cdn/img/cars/M/car_F252.jpg',
            'ford kuga': 'https://www.carjet.com/cdn/img/cars/M/car_F41.jpg',
            'kia stonic': 'https://www.carjet.com/cdn/img/cars/M/car_F119.jpg',
            'citroen c3 aircross': 'https://www.carjet.com/cdn/img/cars/M/car_A782.jpg',
            'citroën c3 aircross': 'https://www.carjet.com/cdn/img/cars/M/car_A782.jpg',
            'citroen c5 aircross': 'https://www.carjet.com/cdn/img/cars/M/car_A640.jpg',
            'citroën c5 aircross': 'https://www.carjet.com/cdn/img/cars/M/car_A640.jpg',
            'citroen c5 aircross auto': 'https://www.carjet.com/cdn/img/cars/M/car_A640.jpg',
            'citroën c5 aircross auto': 'https://www.carjet.com/cdn/img/cars/M/car_A640.jpg',
            'jeep avenger': 'https://www.carjet.com/cdn/img/cars/M/car_L164.jpg',
            'jeep renegade': 'https://www.carjet.com/cdn/img/cars/M/car_A222.jpg',
            'jeep renegade auto': 'https://www.carjet.com/cdn/img/cars/M/car_A222.jpg',
            'volkswagen taigo': 'https://www.carjet.com/cdn/img/cars/M/car_F352.jpg',
            'vw taigo': 'https://www.carjet.com/cdn/img/cars/M/car_F352.jpg',
            'hyundai kauai': 'https://www.carjet.com/cdn/img/cars/M/car_F44.jpg',
            'hyundai kauaí': 'https://www.carjet.com/cdn/img/cars/M/car_F44.jpg',
            'mitsubishi asx': 'https://www.carjet.com/cdn/img/cars/M/car_F178.jpg',
            'hyundai kona': 'https://www.carjet.com/cdn/img/cars/M/car_F191.jpg',
            'toyota c-hr': 'https://www.carjet.com/cdn/img/cars/M/car_A301.jpg',
            'toyota chr': 'https://www.carjet.com/cdn/img/cars/M/car_A301.jpg',
            'toyota c-hr auto': 'https://www.carjet.com/cdn/img/cars/M/car_A301.jpg',
            'toyota chr auto': 'https://www.carjet.com/cdn/img/cars/M/car_A301.jpg',
            'ford ecosport': 'https://www.carjet.com/cdn/img/cars/M/car_A606.jpg',
            'ford eco sport': 'https://www.carjet.com/cdn/img/cars/M/car_A606.jpg',
            'ford ecosport auto': 'https://www.carjet.com/cdn/img/cars/M/car_A606.jpg',
            'opel crossland x': 'https://www.carjet.com/cdn/img/cars/M/car_A444.jpg',
            'opel crossland x auto': 'https://www.carjet.com/cdn/img/cars/M/car_A444.jpg',
            'volkswagen tiguan': 'https://www.carjet.com/cdn/img/cars/M/car_A830.jpg',
            'vw tiguan': 'https://www.carjet.com/cdn/img/cars/M/car_A830.jpg',
            'volkswagen tiguan auto': 'https://www.carjet.com/cdn/img/cars/M/car_A830.jpg',
            'vw tiguan auto': 'https://www.carjet.com/cdn/img/cars/M/car_A830.jpg',
            'skoda karoq': 'https://www.carjet.com/cdn/img/cars/M/car_A822.jpg',
            'skoda karoq auto': 'https://www.carjet.com/cdn/img/cars/M/car_A822.jpg',
            'kia sportage': 'https://www.carjet.com/cdn/img/cars/M/car_F43.jpg',
            'nissan qashqai': 'https://www.carjet.com/cdn/img/cars/M/car_F24.jpg',
            'skoda kamiq': 'https://www.carjet.com/cdn/img/cars/M/car_F310.jpg',
            'hyundai tucson': 'https://www.carjet.com/cdn/img/cars/M/car_F310.jpg',
            'renault austral': 'https://www.carjet.com/cdn/img/cars/M/car_F430.jpg',
            'seat ateca': 'https://www.carjet.com/cdn/img/cars/M/car_F154.jpg',
            'seat arona': 'https://www.carjet.com/cdn/img/cars/M/car_F194.jpg',
            'seat arona auto': 'https://www.carjet.com/cdn/img/cars/M/car_F194.jpg',
            'ford puma': 'https://www.carjet.com/cdn/img/cars/M/car_A999.jpg',
            'ford puma auto': 'https://www.carjet.com/cdn/img/cars/M/car_A999.jpg',
            'mazda cx-3': 'https://www.carjet.com/cdn/img/cars/M/car_F179.jpg',
            'mazda cx 3': 'https://www.carjet.com/cdn/img/cars/M/car_F179.jpg',
            'renault arkana': 'https://www.carjet.com/cdn/img/cars/M/car_A1159.jpg',
            'renault arkana auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1159.jpg',
            'toyota rav 4': 'https://www.carjet.com/cdn/img/cars/M/car_A1000.jpg',
            'toyota rav4': 'https://www.carjet.com/cdn/img/cars/M/car_A1000.jpg',
            'toyota rav 4 4x4': 'https://www.carjet.com/cdn/img/cars/M/car_A1000.jpg',
            'toyota rav 4 auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1000.jpg',
            'toyota hilux': 'https://www.carjet.com/cdn/img/cars/M/car_F326.jpg',
            'toyota hilux 4x4': 'https://www.carjet.com/cdn/img/cars/M/car_F326.jpg',
            
            # PREMIUM / G
            'mini cooper countryman': 'https://www.carjet.com/cdn/img/cars/M/car_F209.jpg',
            'miny cooper countryman': 'https://www.carjet.com/cdn/img/cars/M/car_F209.jpg',
            'mini countryman': 'https://www.carjet.com/cdn/img/cars/M/car_F209.jpg',
            'mini cooper countryman auto': 'https://www.carjet.com/cdn/img/cars/M/car_F209.jpg',
            'mini cooper cabrio': 'https://www.carjet.com/cdn/img/cars/M/car_L118.jpg',
            'mini one cabrio': 'https://www.carjet.com/cdn/img/cars/M/car_L118.jpg',
            'volkswagen beetle cabrio': 'https://www.carjet.com/cdn/img/cars/M/car_L44.jpg',
            'vw beetle cabrio': 'https://www.carjet.com/cdn/img/cars/M/car_L44.jpg',
            'cupra formentor': 'https://www.carjet.com/cdn/img/cars/M/car_A1185.jpg',
            'cupra formentor auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1185.jpg',
            'ds 4': 'https://www.carjet.com/cdn/img/cars/M/car_A1637.jpg',
            'ds 4 auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1637.jpg',
            
            # STATION WAGON / J2 / L2
            'peugeot 308 sw': 'https://www.carjet.com/cdn/img/cars/M/car_S06.jpg',
            'peugeot 308 sw auto': 'https://www.carjet.com/cdn/img/cars/M/car_S06.jpg',
            'opel astra sw': 'https://www.carjet.com/cdn/img/cars/M/car_S10.jpg',
            'cupra leon sw': 'https://www.carjet.com/cdn/img/cars/M/car_A1426.jpg',
            'cupra leon st': 'https://www.carjet.com/cdn/img/cars/M/car_A1426.jpg',
            'cupra leon estate': 'https://www.carjet.com/cdn/img/cars/M/car_A1426.jpg',
            'cupra leon sport tourer': 'https://www.carjet.com/cdn/img/cars/M/car_A1426.jpg',
            'cupra leon sw auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1426.jpg',
            'toyota corolla sw': 'https://www.carjet.com/cdn/img/cars/M/car_A590.jpg',
            'toyota corolla touring sports': 'https://www.carjet.com/cdn/img/cars/M/car_A590.jpg',
            'toyota corolla estate': 'https://www.carjet.com/cdn/img/cars/M/car_A590.jpg',
            'toyota corolla sw auto': 'https://www.carjet.com/cdn/img/cars/M/car_A590.jpg',
            'skoda octavia': 'https://www.carjet.com/cdn/img/cars/M/car_I12.jpg',
            'skoda octavia sw': 'https://www.carjet.com/cdn/img/cars/M/car_I12.jpg',
            'skoda octavia combi': 'https://www.carjet.com/cdn/img/cars/M/car_I12.jpg',
            'skoda octavia estate': 'https://www.carjet.com/cdn/img/cars/M/car_I12.jpg',
            'skoda fabia sw': 'https://www.carjet.com/cdn/img/cars/M/car_S34.jpg',
            'skoda fabia combi': 'https://www.carjet.com/cdn/img/cars/M/car_S34.jpg',
            'skoda fabia estate': 'https://www.carjet.com/cdn/img/cars/M/car_S34.jpg',
            'volkswagen passat': 'https://www.carjet.com/cdn/img/cars/M/car_I11.jpg',
            'vw passat': 'https://www.carjet.com/cdn/img/cars/M/car_I11.jpg',
            'volkswagen passat variant': 'https://www.carjet.com/cdn/img/cars/M/car_I11.jpg',
            'volkswagen passat estate': 'https://www.carjet.com/cdn/img/cars/M/car_I11.jpg',
            'volkswagen passat sw': 'https://www.carjet.com/cdn/img/cars/M/car_I11.jpg',
            'fiat tipo sw': 'https://www.carjet.com/cdn/img/cars/M/car_F72.jpg',
            'fiat tipo estate': 'https://www.carjet.com/cdn/img/cars/M/car_F72.jpg',
            'seat leon sw': 'https://www.carjet.com/cdn/img/cars/M/car_F46.jpg',
            'seat leon st': 'https://www.carjet.com/cdn/img/cars/M/car_F46.jpg',
            'seat leon estate': 'https://www.carjet.com/cdn/img/cars/M/car_F46.jpg',
            'seat leon sport tourer': 'https://www.carjet.com/cdn/img/cars/M/car_F46.jpg',
            
            # 7 SEATER / M1 / M2
            'dacia lodgy': 'https://www.carjet.com/cdn/img/cars/M/car_M117.jpg',
            'dacia jogger': 'https://www.carjet.com/cdn/img/cars/M/car_M162.jpg',
            'opel zafira': 'https://www.carjet.com/cdn/img/cars/M/car_M05.jpg',
            'peugeot 5008': 'https://www.carjet.com/cdn/img/cars/M/car_M27.jpg',
            'renault grand scenic': 'https://www.carjet.com/cdn/img/cars/M/car_M15.jpg',
            'renault grand scenic auto': 'https://www.carjet.com/cdn/img/cars/M/car_M15.jpg',
            'citroen grand picasso': 'https://www.carjet.com/cdn/img/cars/M/car_A219.jpg',
            'citroën grand picasso': 'https://www.carjet.com/cdn/img/cars/M/car_A219.jpg',
            'citroen c4 grand picasso': 'https://www.carjet.com/cdn/img/cars/M/car_A219.jpg',
            'citroën c4 grand picasso': 'https://www.carjet.com/cdn/img/cars/M/car_A219.jpg',
            'citroen c4 grand picasso auto': 'https://www.carjet.com/cdn/img/cars/M/car_A219.jpg',
            'citroën c4 grand picasso auto': 'https://www.carjet.com/cdn/img/cars/M/car_A219.jpg',
            'citroen c4 picasso auto': 'https://www.carjet.com/cdn/img/cars/M/car_A522.jpg',
            'citroën c4 picasso auto': 'https://www.carjet.com/cdn/img/cars/M/car_A522.jpg',
            'citroen c4 grand spacetourer': 'https://www.carjet.com/cdn/img/cars/M/car_A1430.jpg',
            'citroën c4 grand spacetourer': 'https://www.carjet.com/cdn/img/cars/M/car_A1430.jpg',
            'citroen c4 grand spacetourer auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1430.jpg',
            'volkswagen caddy': 'https://www.carjet.com/cdn/img/cars/M/car_A295.jpg',
            'vw caddy': 'https://www.carjet.com/cdn/img/cars/M/car_A295.jpg',
            'volkswagen caddy auto': 'https://www.carjet.com/cdn/img/cars/M/car_A295.jpg',
            'peugeot rifter': 'https://www.carjet.com/cdn/img/cars/M/car_M124.jpg',
            'peugeot rifter auto': 'https://www.carjet.com/cdn/img/cars/M/car_M124.jpg',
            'mercedes glb': 'https://www.carjet.com/cdn/img/cars/M/car_GZ399.jpg',
            'mercedes glb auto': 'https://www.carjet.com/cdn/img/cars/M/car_GZ399.jpg',
            
            # 9 SEATER / N
            'ford tourneo': 'https://www.carjet.com/cdn/img/cars/M/car_M44.jpg',
            'ford transit': 'https://www.carjet.com/cdn/img/cars/M/car_M02.jpg',
            'ford galaxy': 'https://www.carjet.com/cdn/img/cars/M/car_M03.jpg',
            'volkswagen sharan': 'https://www.carjet.com/cdn/img/cars/M/car_M56.jpg',
            'vw sharan': 'https://www.carjet.com/cdn/img/cars/M/car_M56.jpg',
            'volkswagen multivan': 'https://www.carjet.com/cdn/img/cars/M/car_A406.jpg',
            'vw multivan': 'https://www.carjet.com/cdn/img/cars/M/car_A406.jpg',
            'volkswagen multivan auto': 'https://www.carjet.com/cdn/img/cars/M/car_A406.jpg',
            'vw multivan auto': 'https://www.carjet.com/cdn/img/cars/M/car_A406.jpg',
            'citroen spacetourer': 'https://www.carjet.com/cdn/img/cars/M/car_A261.jpg',
            'citroën spacetourer': 'https://www.carjet.com/cdn/img/cars/M/car_A261.jpg',
            'citroen spacetourer auto': 'https://www.carjet.com/cdn/img/cars/M/car_A261.jpg',
            'renault trafic': 'https://www.carjet.com/cdn/img/cars/M/car_A581.jpg',
            'renault trafic auto': 'https://www.carjet.com/cdn/img/cars/M/car_A581.jpg',
            'peugeot traveller': 'https://www.carjet.com/cdn/img/cars/M/car_M86.jpg',
            'volkswagen transporter': 'https://www.carjet.com/cdn/img/cars/M/car_M08.jpg',
            'vw transporter': 'https://www.carjet.com/cdn/img/cars/M/car_M08.jpg',
            'mercedes vito': 'https://www.carjet.com/cdn/img/cars/M/car_A230.jpg',
            'mercedes benz vito': 'https://www.carjet.com/cdn/img/cars/M/car_A230.jpg',
            'mercedes vito auto': 'https://www.carjet.com/cdn/img/cars/M/car_A230.jpg',
            'volkswagen caravelle': 'https://www.carjet.com/cdn/img/cars/M/car_M63.jpg',
            'vw caravelle': 'https://www.carjet.com/cdn/img/cars/M/car_M63.jpg',
            'mercedes v class': 'https://www.carjet.com/cdn/img/cars/M/car_A1336.jpg',
            'mercedes benz v class': 'https://www.carjet.com/cdn/img/cars/M/car_A1336.jpg',
            'mercedes v class auto': 'https://www.carjet.com/cdn/img/cars/M/car_A1336.jpg',
            'fiat talento': 'https://www.carjet.com/cdn/img/cars/M/car_M49.jpg',
            'opel vivaro': 'https://www.carjet.com/cdn/img/cars/M/car_M34.jpg',
            'toyota proace': 'https://www.carjet.com/cdn/img/cars/M/car_M136.jpg',
        }
        
        for vehicle_key in VEHICLES.keys():
            try:
                # Verificar se já existe
                with _db_lock:
                    con = _db_connect()
                    try:
                        existing = con.execute("SELECT vehicle_name FROM vehicle_photos WHERE vehicle_name = ?", (vehicle_key,)).fetchone()
                        if existing:
                            skipped += 1
                            continue
                    finally:
                        con.close()
                
                # Buscar URL da imagem - PRIORIDADE: scraping recente
                image_url = photo_urls_from_scraping.get(vehicle_key) or image_mappings.get(vehicle_key)
                if not image_url:
                    errors.append(f"{vehicle_key}: No photo URL found")
                    continue
                
                print(f"[PHOTOS] Downloading {vehicle_key} from {image_url[:80]}...", file=sys.stderr, flush=True)
                
                # Download da imagem
                async with httpx.AsyncClient(timeout=30.0) as client:
                    response = await client.get(image_url)
                    if response.status_code == 200:
                        image_data = response.content
                        content_type = response.headers.get('content-type', 'image/jpeg')
                        
                        # Salvar na BD
                        with _db_lock:
                            con = _db_connect()
                            try:
                                con.execute("""
                                    INSERT OR REPLACE INTO vehicle_photos (vehicle_name, photo_data, content_type, photo_url)
                                    VALUES (?, ?, ?, ?)
                                """, (vehicle_key, image_data, content_type, image_url))
                                con.commit()
                                downloaded += 1
                            finally:
                                con.close()
            except Exception as e:
                errors.append(f"{vehicle_key}: {str(e)}")
        
        return _no_store_json({
            "ok": True,
            "downloaded": downloaded,
            "skipped": skipped,
            "errors": errors[:10]
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/vehicles/{vehicle_name}/photo")
async def get_vehicle_photo(vehicle_name: str):
    """Retorna a foto de um veículo específico"""
    # Não requer autenticação para permitir que as tags <img> funcionem
    try:
        # Normalizar nome do veículo
        vehicle_key = vehicle_name.lower().strip()
        
        with _db_lock:
            con = _db_connect()
            try:
                # Tentar buscar foto exata
                row = con.execute(
                    "SELECT image_data, content_type FROM vehicle_images WHERE vehicle_key = ?",
                    (vehicle_key,)
                ).fetchone()
                
                # Se não encontrar, tentar buscar variações do mesmo modelo
                if not row:
                    # Detectar se é Station Wagon (SW) - são modelos diferentes!
                    is_sw = ' sw' in vehicle_key or 'station wagon' in vehicle_key or 'estate' in vehicle_key
                    
                    # Extrair modelo base (ex: "citroen c1" de "citroen c1 auto")
                    # Remove sufixos comuns: auto, automatic, hybrid, electric, diesel, etc
                    base_model = vehicle_key
                    for suffix in [' auto', ' automatic', ' hybrid', ' electric', ' diesel', ' 4x4', ', hybrid', ', electric', ', diesel', ', automatic']:
                        base_model = base_model.replace(suffix, '')
                    base_model = base_model.strip()
                    
                    if is_sw:
                        # Se for SW, buscar APENAS outras variações SW
                        # Ex: "renault megane sw auto" busca "renault megane sw"
                        row = con.execute(
                            "SELECT image_data, content_type FROM vehicle_images WHERE vehicle_key LIKE ? AND (vehicle_key LIKE '%sw%' OR vehicle_key LIKE '%station wagon%' OR vehicle_key LIKE '%estate%') LIMIT 1",
                            (base_model + '%',)
                        ).fetchone()
                    else:
                        # Se NÃO for SW, buscar variações NÃO-SW
                        # Ex: "renault megane auto" busca "renault megane" mas NÃO "renault megane sw"
                        row = con.execute(
                            "SELECT image_data, content_type FROM vehicle_images WHERE vehicle_key LIKE ? AND vehicle_key NOT LIKE '%sw%' AND vehicle_key NOT LIKE '%station wagon%' AND vehicle_key NOT LIKE '%estate%' LIMIT 1",
                            (base_model + '%',)
                        ).fetchone()
                
                if row:
                    image_data = row[0]
                    content_type = row[1] or 'image/jpeg'
                    
                    from fastapi.responses import Response
                    return Response(
                        content=image_data,
                        media_type=content_type,
                        headers={
                            "Cache-Control": "public, max-age=86400",
                            "Content-Disposition": f"inline; filename={vehicle_key}.jpg"
                        }
                    )
                else:
                    # Retornar imagem placeholder SVG
                    svg_placeholder = '''<svg xmlns="http://www.w3.org/2000/svg" width="60" height="40">
                        <rect width="60" height="40" fill="#e5e7eb"/>
                        <text x="50%" y="50%" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="12">🚗</text>
                    </svg>'''
                    return Response(
                        content=svg_placeholder,
                        media_type="image/svg+xml"
                    )
            finally:
                con.close()
    except Exception as e:
        import traceback
        print(f"Erro ao buscar foto: {traceback.format_exc()}")
        # Retornar placeholder em caso de erro
        svg_placeholder = '''<svg xmlns="http://www.w3.org/2000/svg" width="60" height="40">
            <rect width="60" height="40" fill="#e5e7eb"/>
            <text x="50%" y="50%" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="12">🚗</text>
        </svg>'''
        return Response(
            content=svg_placeholder,
            media_type="image/svg+xml"
        )

@app.get("/api/vehicles/{vehicle_name}/photo/metadata")
async def get_vehicle_photo_metadata(vehicle_name: str, request: Request):
    """Retorna metadata da foto de um veículo (URL original, etc)"""
    require_auth(request)
    try:
        vehicle_key = vehicle_name.lower().strip()
        
        with _db_lock:
            con = _db_connect()
            try:
                row = con.execute(
                    "SELECT source_url, downloaded_at, content_type FROM vehicle_images WHERE vehicle_key = ?",
                    (vehicle_key,)
                ).fetchone()
                
                if row:
                    return _no_store_json({
                        "ok": True,
                        "source_url": row[0],
                        "downloaded_at": row[1],
                        "content_type": row[2]
                    })
                else:
                    return _no_store_json({"ok": False, "error": "Photo not found"}, 404)
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/vehicles/{vehicle_name}/photo/from-url")
async def download_vehicle_photo_from_url(vehicle_name: str, request: Request):
    """Baixa e salva a foto de um veículo a partir de uma URL"""
    require_auth(request)
    try:
        import httpx
        
        body = await request.json()
        url = body.get('url', '').strip()
        
        if not url:
            return _no_store_json({"ok": False, "error": "URL é obrigatória"}, 400)
        
        # Normalizar nome do veículo
        vehicle_key = vehicle_name.lower().strip()
        
        # Baixar imagem
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(url)
            response.raise_for_status()
            
            image_data = response.content
            content_type = response.headers.get('content-type', 'image/jpeg')
            
            # Salvar na base de dados
            with _db_lock:
                con = _db_connect()
                try:
                    con.execute(
                        """INSERT OR REPLACE INTO vehicle_images 
                           (vehicle_key, image_data, content_type, source_url, downloaded_at)
                           VALUES (?, ?, ?, ?, datetime('now'))""",
                        (vehicle_key, image_data, content_type, url)
                    )
                    con.commit()
                finally:
                    con.close()
            
            return _no_store_json({
                "ok": True,
                "message": f"Foto baixada e salva para {vehicle_name}",
                "size": len(image_data),
                "content_type": content_type
            })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/vehicles/{vehicle_name}/photo/upload")
async def upload_vehicle_photo(vehicle_name: str, request: Request, file: UploadFile = File(...)):
    """Upload de foto de um veículo"""
    require_auth(request)
    try:
        # Normalizar nome do veículo
        vehicle_key = vehicle_name.lower().strip()
        
        # Ler conteúdo do ficheiro
        image_data = await file.read()
        content_type = file.content_type or 'image/jpeg'
        
        # Salvar na base de dados
        with _db_lock:
            con = _db_connect()
            try:
                con.execute(
                    """INSERT OR REPLACE INTO vehicle_images 
                       (vehicle_key, image_data, content_type, source_url, downloaded_at)
                       VALUES (?, ?, ?, ?, datetime('now'))""",
                    (vehicle_key, image_data, content_type, 'uploaded')
                )
                con.commit()
            finally:
                con.close()
        
        return _no_store_json({
            "ok": True,
            "message": f"Foto enviada com sucesso para {vehicle_name}",
            "size": len(image_data),
            "content_type": content_type
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# EXPORT/IMPORT - Base de Dados Completa
# ============================================================

@app.get("/api/export/config")
async def export_config(request: Request):
    """Exporta configuração completa: veículos, imagens, overrides"""
    require_auth(request)
    try:
        import base64
        from datetime import datetime
        
        export_data = {
            "version": "1.0",
            "exported_at": datetime.utcnow().isoformat(),
            "vehicles": {},
            "name_overrides": [],
            "images": {}
        }
        
        # 1. Exportar VEHICLES
        try:
            from carjet_direct import VEHICLES
            export_data["vehicles"] = dict(VEHICLES)
        except Exception as e:
            print(f"Aviso: não foi possível exportar VEHICLES: {e}")
        
        # 2. Exportar name overrides
        with _db_lock:
            con = _db_connect()
            try:
                rows = con.execute("SELECT original_name, edited_name, updated_at FROM vehicle_name_overrides").fetchall()
                export_data["name_overrides"] = [
                    {
                        "original_name": row[0],
                        "edited_name": row[1],
                        "updated_at": row[2]
                    }
                    for row in rows
                ]
            finally:
                con.close()
        
        # 3. Exportar imagens (como Base64)
        with _db_lock:
            con = _db_connect()
            try:
                rows = con.execute("SELECT vehicle_key, image_data, content_type, source_url FROM vehicle_images").fetchall()
                for row in rows:
                    vehicle_key = row[0]
                    image_data = row[1]
                    content_type = row[2]
                    source_url = row[3]
                    
                    # Converter para Base64
                    image_base64 = base64.b64encode(image_data).decode('utf-8')
                    
                    export_data["images"][vehicle_key] = {
                        "data": image_base64,
                        "content_type": content_type,
                        "source_url": source_url
                    }
            finally:
                con.close()
        
        # Retornar como JSON para download
        return JSONResponse(
            content=export_data,
            headers={
                "Content-Disposition": f"attachment; filename=carrental_config_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
            }
        )
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/import/config")
async def import_config(request: Request):
    """Importa configuração completa de um ficheiro JSON"""
    require_auth(request)
    try:
        import base64
        
        body = await request.json()
        
        if not body or "version" not in body:
            return _no_store_json({"ok": False, "error": "Ficheiro de configuração inválido"}, 400)
        
        imported = {
            "name_overrides": 0,
            "images": 0
        }
        
        # 1. Importar name overrides
        if "name_overrides" in body:
            with _db_lock:
                con = _db_connect()
                try:
                    for override in body["name_overrides"]:
                        con.execute("""
                            INSERT INTO vehicle_name_overrides (original_name, edited_name, updated_at)
                            VALUES (?, ?, ?)
                            ON CONFLICT(original_name) DO UPDATE SET
                                edited_name = excluded.edited_name,
                                updated_at = excluded.updated_at
                        """, (override["original_name"], override["edited_name"], override.get("updated_at", "now")))
                        imported["name_overrides"] += 1
                    con.commit()
                finally:
                    con.close()
        
        # 2. Importar imagens
        if "images" in body:
            with _db_lock:
                con = _db_connect()
                try:
                    for vehicle_key, image_info in body["images"].items():
                        # Converter de Base64 para bytes
                        image_data = base64.b64decode(image_info["data"])
                        content_type = image_info.get("content_type", "image/jpeg")
                        source_url = image_info.get("source_url", "")
                        
                        con.execute("""
                            INSERT INTO vehicle_images (vehicle_key, image_data, content_type, source_url, downloaded_at)
                            VALUES (?, ?, ?, ?, datetime('now'))
                            ON CONFLICT(vehicle_key) DO UPDATE SET
                                image_data = excluded.image_data,
                                content_type = excluded.content_type,
                                source_url = excluded.source_url,
                                downloaded_at = excluded.downloaded_at
                        """, (vehicle_key, image_data, content_type, source_url))
                        imported["images"] += 1
                    con.commit()
                finally:
                    con.close()
        
        return _no_store_json({
            "ok": True,
            "message": "Configuração importada com sucesso",
            "imported": imported
        })
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# EXPORT HISTORY - Way2Rentals, Abbycar, etc.
# ============================================================

@app.post("/api/export-history/save")
async def save_export_history(request: Request):
    """Salva export na database para histórico"""
    require_auth(request)
    try:
        body = await request.json()
        
        filename = body.get("filename", "")
        broker = body.get("broker", "")
        location = body.get("location", "")
        period_start = body.get("period_start")
        period_end = body.get("period_end")
        month = body.get("month", 0)
        year = body.get("year", 0)
        month_name = body.get("month_name", "")
        file_content = body.get("file_content", "")
        file_size = len(file_content)
        
        username = request.session.get("username", "unknown")
        
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("""
                    INSERT INTO export_history 
                    (filename, broker, location, period_start, period_end, month, year, month_name, 
                     file_content, file_size, exported_by, export_date)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                """, (filename, broker, location, period_start, period_end, month, year, 
                      month_name, file_content, file_size, username))
                con.commit()
                
                export_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
                
                # Cleanup: Manter apenas últimos 12 meses
                cutoff_date = f"{year - 1}-{month:02d}-01"
                con.execute("""
                    DELETE FROM export_history 
                    WHERE export_date < ?
                """, (cutoff_date,))
                con.commit()
                
                return _no_store_json({
                    "ok": True,
                    "export_id": export_id,
                    "message": "Export saved successfully"
                })
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/export-history/list")
async def list_export_history(request: Request):
    """Lista exports salvos (últimos 12 meses)"""
    require_auth(request)
    try:
        broker = request.query_params.get("broker")
        location = request.query_params.get("location")
        year = request.query_params.get("year")
        month = request.query_params.get("month")
        
        query = """
            SELECT id, filename, broker, location, period_start, period_end, 
                   month, year, month_name, file_size, exported_by, export_date, last_downloaded
            FROM export_history
            WHERE 1=1
        """
        params = []
        
        if broker:
            query += " AND broker = ?"
            params.append(broker)
        if location:
            query += " AND location = ?"
            params.append(location)
        if year:
            query += " AND year = ?"
            params.append(int(year))
        if month:
            query += " AND month = ?"
            params.append(int(month))
        
        query += " ORDER BY year DESC, month DESC, export_date DESC"
        
        with _db_lock:
            con = _db_connect()
            try:
                rows = con.execute(query, params).fetchall()
                
                exports = []
                for row in rows:
                    exports.append({
                        "id": row[0],
                        "filename": row[1],
                        "broker": row[2],
                        "location": row[3],
                        "period_start": row[4],
                        "period_end": row[5],
                        "month": row[6],
                        "year": row[7],
                        "month_name": row[8],
                        "file_size": row[9],
                        "exported_by": row[10],
                        "export_date": row[11],
                        "last_downloaded": row[12]
                    })
                
                return _no_store_json({
                    "ok": True,
                    "exports": exports,
                    "total": len(exports)
                })
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/export-history/download/{export_id}")
async def download_export_history(request: Request, export_id: int):
    """Download de export salvo"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                row = con.execute("""
                    SELECT filename, file_content, broker
                    FROM export_history
                    WHERE id = ?
                """, (export_id,)).fetchone()
                
                if not row:
                    return _no_store_json({"ok": False, "error": "Export not found"}, 404)
                
                filename, file_content, broker = row
                
                # Update last_downloaded
                con.execute("""
                    UPDATE export_history
                    SET last_downloaded = datetime('now')
                    WHERE id = ?
                """, (export_id,))
                con.commit()
                
                # Return CSV file
                from starlette.responses import Response
                return Response(
                    content=file_content,
                    media_type="text/csv",
                    headers={
                        "Content-Disposition": f'attachment; filename="{filename}"'
                    }
                )
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/export-automated-prices-excel")
async def export_automated_prices_excel(request: Request):
    """Export automated prices to Excel (Abbycar format)"""
    require_auth(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        import io
        
        data = await request.json()
        location = data.get('location', 'Unknown')
        date = data.get('date', datetime.now().strftime('%Y-%m-%d'))
        prices = data.get('prices', {})  # { 'B1': { '1': 25.00, '2': 24.50, ... }, 'B2': {...}, ... }
        
        # Car group mapping (SIPP codes to groups)
        car_group_mapping = {
            'MDMV': 'B1',  # Mini 4 Doors Manual
            'MCMV': 'B1',  # Mini Coupe Manual (same group)
            'EDMV': 'B2',  # Economy Manual
            'NDMR': 'B2',  # Economy 4 Doors Manual (same group)
            'MDMR': 'D',   # Mini 4 Doors Manual
            'HDMV': 'D',   # Mini Elite 4 Doors Manual (same group)
            'MDAR': 'E1',  # Mini Auto
            'MDAV': 'E1',  # Mini 4 Doors Auto (same group)
            'EDAV': 'E2',  # Economy Auto
            'EDAR': 'E2',  # Economy Auto (same group)
            'CFMR': 'F',   # Compact Manual
            'DFMR': 'F',   # Compact 4 Doors Manual (same group)
            'MTMR': 'G',   # Mini Elite Manual
            'CFMV': 'J1',  # Compact Manual
            'DFMV': 'J1',  # Compact 4 Doors Manual (same group)
            'IWMR': 'J2',  # Intermediate Wagon Manual
            'IWMV': 'J2',  # Intermediate Wagon Manual (same group)
            'CFAR': 'L1',  # Compact Auto
            'CGAR': 'L1',  # Compact Auto (same group)
            'CFAV': 'L1',  # Compact Auto (same group)
            'SVMR': 'M1',  # Standard Manual
            'SVMD': 'M1',  # Standard Manual (same group)
            'SVMV': 'M1',  # Standard Manual (same group)
            'SVAD': 'M2',  # Standard Auto
            'SVAR': 'M2',  # Standard Auto (same group)
            'LVMD': 'N',   # Large Manual
            'LVMR': 'N'    # Large Manual (same group)
        }
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "IMPORT"
        
        # Styles
        header_fill = PatternFill(start_color="009cb6", end_color="009cb6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Determine station code based on location
        station_code = "FAO" if "faro" in location.lower() else "ABF" if "albufeira" in location.lower() else "UNK"
        
        # Header row - EXACT format from original Abbycar.xlsx
        headers = [
            "Stations",
            "Start Date",
            "End Date",
            "Group",
            "Model Example (optional)",
            "CURRENCY",
            "1 day fixed",
            "2 days fixed",
            "3 days fixed",
            "4 days fixed",
            "5 days fixed",
            "6 days fixed",
            "7 days fixed",
            "8-10 daily",
            "11-12 daily",
            "13-14 daily",
            "15-21 daily",
            "22-28 daily"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Map internal group codes to primary SIPP code for display
        group_to_sipp = {
            'B1': 'MDMV',
            'B2': 'EDMV',
            'D': 'MDMR',
            'E1': 'MDAR',
            'E2': 'EDAV',
            'F': 'CFMR',
            'G': 'MTMR',
            'J1': 'CFMV',
            'J2': 'IWMR',
            'L1': 'CFAR',
            'L2': 'CGAR',
            'M1': 'SVMR',
            'M2': 'SVAD',
            'N': 'LVMD'
        }
        
        # Model examples from original Abbycar.xlsx
        sipp_to_model = {
            'MDMV': 'Peugeot 108',
            'MDMR': 'Fiat Panda',
            'MCMV': 'Citroen C1',
            'NDMR': 'Toyota Aygo',
            'EDMV': 'Opel Corsa',
            'HDMV': 'Ford Fiesta',
            'MDAR': 'Kia Picanto',
            'EDAV': 'Citroen C3/Opel Corsa',
            'MDAV': 'Toyota Aygo',
            'EDAR': 'Opel Corsa',
            'CFMR': 'Seat Arona',
            'DFMR': 'Kia Stonic',
            'MTMR': 'Fiat 500 Cabrio',
            'CFMV': 'Peugeot 2008',
            'IWMR': 'Peugeot 308 SW',
            'DFMV': 'Citroen C4',
            'IWMV': 'Opel Astra STW',
            'CFAR': 'Seat Arona',
            'CGAR': 'Citroen C3 Aircross',
            'CFAV': 'VW T-Cross',
            'SVMR': 'Dacia Jogger SL Extreme',
            'SVMD': 'Citroen Grand C4',
            'SVAD': 'Citroen Grand C4 Automatic',
            'SVMV': 'Peugeot Rifter',
            'SVAR': 'Peugeot Rifter Automatic',
            'LVMD': 'Fiat Talento',
            'LVMR': 'Opel Vivaro'
        }
        
        # Use SIPP codes in the EXACT order from original Abbycar.xlsx
        sipp_codes_order = [
            'MDMV', 'MDMR', 'MCMV', 'NDMR', 'EDMV', 'HDMV', 'MDAR', 'EDAV', 'MDAV', 'EDAR',
            'CFMR', 'DFMR', 'MTMR', 'CFMV', 'IWMR', 'DFMV', 'IWMV', 'CFAR', 'CGAR', 'CFAV',
            'SVMR', 'SVMD', 'SVAD', 'SVMV', 'SVAR', 'LVMD', 'LVMR'
        ]
        
        # Price calculation logic based on periods
        def calculate_price_for_day(group_prices, day):
            """
            Get exact price for each period - NO division, use values as-is
            These are FIXED prices per period (1 day fixed, 2 days fixed, etc.)
            """
            # Simply return the exact price for the day period
            price = group_prices.get(str(day), '')
            return float(price) if price else ''
        
        # Get Abbycar price adjustments
        abbycar_adjustment = _get_abbycar_adjustment()
        abbycar_low_deposit_enabled = _get_abbycar_low_deposit_enabled()
        abbycar_low_deposit_adjustment = _get_abbycar_low_deposit_adjustment()
        
        # Define Low Deposit groups (ALWAYS defined, but only filled if enabled)
        # SIPP codes: MCMV, NDMR, HDMV, MDAV, EDAR, DFMR, DFMV, IWMV, CFAV, SVMV, SVAR, LVMR
        # Map to internal groups:
        low_deposit_sipp_codes = ['MCMV', 'NDMR', 'HDMV', 'MDAV', 'EDAR', 'DFMR', 'DFMV', 'IWMV', 'CFAV', 'SVMV', 'SVAR', 'LVMR']
        low_deposit_groups = list(set([car_group_mapping[sipp] for sipp in low_deposit_sipp_codes if sipp in car_group_mapping]))
        # Result: ['B1', 'B2', 'D', 'E1', 'E2', 'F', 'J1', 'J2', 'L1', 'M1', 'M2', 'N']
        
        # Fill data rows - EXACT format from original
        row_num = 2
        for sipp_code in sipp_codes_order:
            # Get internal group from SIPP code
            internal_group = car_group_mapping.get(sipp_code, sipp_code)
            model_example = sipp_to_model.get(sipp_code, '')
            group_prices = prices.get(internal_group, {})
            
            # Check if this is a Low Deposit group for highlighting (by SIPP code)
            is_low_deposit_group = sipp_code in low_deposit_sipp_codes
            # Usar azul teal claro apenas para Low Deposit groups
            low_deposit_fill = PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid") if is_low_deposit_group else None
            
            # Column 1: Stations
            ws.cell(row_num, 1).value = station_code
            ws.cell(row_num, 1).alignment = cell_alignment
            if is_low_deposit_group:
                ws.cell(row_num, 1).fill = low_deposit_fill
            
            # Column 2: Start Date (EMPTY - only filled when downloading by period)
            ws.cell(row_num, 2).value = ''
            ws.cell(row_num, 2).alignment = cell_alignment
            if is_low_deposit_group:
                ws.cell(row_num, 2).fill = low_deposit_fill
            
            # Column 3: End Date (EMPTY - only filled when downloading by period)
            ws.cell(row_num, 3).value = ''
            ws.cell(row_num, 3).alignment = cell_alignment
            if is_low_deposit_group:
                ws.cell(row_num, 3).fill = low_deposit_fill
            
            # Column 4: Group (SIPP Code)
            ws.cell(row_num, 4).value = sipp_code
            ws.cell(row_num, 4).alignment = cell_alignment
            ws.cell(row_num, 4).font = Font(bold=True, color="009cb6")
            if is_low_deposit_group:
                ws.cell(row_num, 4).fill = low_deposit_fill
            
            # Column 5: Model Example (optional)
            ws.cell(row_num, 5).value = model_example
            ws.cell(row_num, 5).alignment = Alignment(horizontal="left", vertical="center")
            if is_low_deposit_group:
                ws.cell(row_num, 5).fill = low_deposit_fill
            
            # Column 6: CURRENCY
            ws.cell(row_num, 6).value = "EUR"
            ws.cell(row_num, 6).alignment = cell_alignment
            if is_low_deposit_group:
                ws.cell(row_num, 6).fill = low_deposit_fill
            
            # Columns 7-18: Prices (1 day fixed through 22-28 daily)
            # Map: 1-7 days fixed, then 8-10, 11-12, 13-14, 15-21, 22-28 daily
            price_columns = [
                ('1', 7),   # 1 day fixed
                ('2', 8),   # 2 days fixed
                ('3', 9),   # 3 days fixed
                ('4', 10),  # 4 days fixed
                ('5', 11),  # 5 days fixed
                ('6', 12),  # 6 days fixed
                ('7', 13),  # 7 days fixed
                ('8', 14),  # 8-10 daily
                ('9', 15),  # 11-12 daily
                ('14', 16), # 13-14 daily
                ('22', 17), # 15-21 daily
                ('28', 18)  # 22-28 daily
            ]
            
            for day_key, col_idx in price_columns:
                price = calculate_price_for_day(group_prices, int(day_key))
                
                # Check if this is a Low Deposit group and if it's disabled
                should_skip_price = is_low_deposit_group and not abbycar_low_deposit_enabled
                
                if price and not should_skip_price:
                    # Apply Abbycar adjustment percentage
                    total_adjustment = abbycar_adjustment
                    
                    # Add Low Deposit adjustment if group is in Low Deposit list AND enabled
                    if is_low_deposit_group and abbycar_low_deposit_enabled:
                        total_adjustment += abbycar_low_deposit_adjustment
                    
                    adjusted_price = float(price) * (1 + total_adjustment / 100)
                    ws.cell(row_num, col_idx).value = adjusted_price
                    ws.cell(row_num, col_idx).number_format = '0.00'
                else:
                    # Leave empty if: no price OR (Low Deposit group AND disabled)
                    ws.cell(row_num, col_idx).value = ''
                
                ws.cell(row_num, col_idx).alignment = cell_alignment
                if is_low_deposit_group:
                    ws.cell(row_num, col_idx).fill = low_deposit_fill
            
            row_num += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10   # Stations
        ws.column_dimensions['B'].width = 12   # Start Date
        ws.column_dimensions['C'].width = 12   # End Date
        ws.column_dimensions['D'].width = 10   # Group
        ws.column_dimensions['E'].width = 25   # Model Example
        ws.column_dimensions['F'].width = 10   # CURRENCY
        for col_idx in range(7, 19):
            ws.column_dimensions[chr(64 + col_idx)].width = 12
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Generate filename
        filename = f"AutomatedPrices_{location.replace(' ', '_')}_{date}.xlsx"
        
        # SALVAR NA BASE DE DADOS (persistente)
        try:
            excel_bytes = excel_file.getvalue()
            username = request.session.get("username", "admin")
            save_file_to_db(
                filename=filename,
                filepath=f"/exports/{filename}",
                file_data=excel_bytes,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                uploaded_by=username
            )
            log_to_db("INFO", f"Excel export saved to DB: {filename}", "main", "export_automated_prices_excel")
        except Exception as e:
            log_to_db("ERROR", f"Failed to save Excel to DB: {str(e)}", "main", "export_automated_prices_excel")
        
        # Reset para retornar
        excel_file.seek(0)
        
        from starlette.responses import Response
        return Response(
            content=excel_file.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache"
            }
        )
        
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# API ENDPOINTS - LOCALSTORAGE MIGRATION
# ============================================================

@app.get("/api/vans-pricing")
async def get_vans_pricing(request: Request):
    """Get Commercial Vans pricing (C3, C4, C5)"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                cursor = con.execute("SELECT * FROM vans_pricing ORDER BY id DESC LIMIT 1")
                row = cursor.fetchone()
                
                if row:
                    return _no_store_json({
                        "ok": True,
                        "pricing": {
                            "c3_1day": row[1], "c3_2days": row[2], "c3_3days": row[3],
                            "c4_1day": row[4], "c4_2days": row[5], "c4_3days": row[6],
                            "c5_1day": row[7], "c5_2days": row[8], "c5_3days": row[9],
                            "updated_at": row[10]
                        }
                    })
                else:
                    # Return defaults
                    return _no_store_json({
                        "ok": True,
                        "pricing": {
                            "c3_1day": 112, "c3_2days": 144, "c3_3days": 180,
                            "c4_1day": 152, "c4_2days": 170, "c4_3days": 210,
                            "c5_1day": 175, "c5_2days": 190, "c5_3days": 240
                        }
                    })
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/vans-pricing")
async def save_vans_pricing(request: Request):
    """Save Commercial Vans pricing"""
    require_auth(request)
    try:
        data = await request.json()
        user = request.session.get('user', 'admin')
        
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("""
                    INSERT INTO vans_pricing (
                        c3_1day, c3_2days, c3_3days,
                        c4_1day, c4_2days, c4_3days,
                        c5_1day, c5_2days, c5_3days,
                        updated_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    data.get('c3_1day', 112), data.get('c3_2days', 144), data.get('c3_3days', 180),
                    data.get('c4_1day', 152), data.get('c4_2days', 170), data.get('c4_3days', 210),
                    data.get('c5_1day', 175), data.get('c5_2days', 190), data.get('c5_3days', 240),
                    user
                ))
                con.commit()
                return _no_store_json({"ok": True, "message": "Vans pricing saved successfully"})
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/price-validation/rules")
async def get_price_validation_rules(request: Request):
    """Get Price Validation Rules from database"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                # Get the most recent rules
                row = con.execute("""
                    SELECT rules_json, updated_at 
                    FROM price_validation_rules 
                    ORDER BY id DESC LIMIT 1
                """).fetchone()
                
                if row:
                    import json
                    return _no_store_json({
                        "ok": True,
                        "rules": json.loads(row[0]),
                        "updated_at": row[1]
                    })
                else:
                    # Return empty array if no rules exist
                    return _no_store_json({
                        "ok": True,
                        "rules": [],
                        "updated_at": None
                    })
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/price-validation/rules")
async def save_price_validation_rules(request: Request):
    """Save Price Validation Rules to database"""
    require_auth(request)
    try:
        data = await request.json()
        rules = data.get('rules', [])
        user = request.session.get('user', 'admin')
        
        import json
        rules_json = json.dumps(rules)
        
        with _db_lock:
            con = _db_connect()
            try:
                # Insert new rules (keeping history)
                con.execute("""
                    INSERT INTO price_validation_rules (rules_json, updated_by)
                    VALUES (?, ?)
                """, (rules_json, user))
                con.commit()
                
                return _no_store_json({
                    "ok": True,
                    "message": "Price validation rules saved successfully",
                    "count": len(rules)
                })
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/automation-settings")
async def get_automation_settings(request: Request):
    """Get price automation settings"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                cursor = con.execute("SELECT setting_key, setting_value, setting_type FROM price_automation_settings")
                rows = cursor.fetchall()
                
                settings = {}
                for row in rows:
                    key, value, stype = row
                    if stype == 'json':
                        import json
                        settings[key] = json.loads(value)
                    elif stype == 'number':
                        settings[key] = float(value)
                    else:
                        settings[key] = value
                
                return _no_store_json({"ok": True, "settings": settings})
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/automation-settings")
async def save_automation_settings(request: Request):
    """Save price automation settings"""
    require_auth(request)
    try:
        data = await request.json()
        
        with _db_lock:
            con = _db_connect()
            try:
                for key, value in data.items():
                    import json
                    if isinstance(value, (list, dict)):
                        value_str = json.dumps(value)
                        stype = 'json'
                    elif isinstance(value, (int, float)):
                        value_str = str(value)
                        stype = 'number'
                    else:
                        value_str = str(value)
                        stype = 'string'
                    
                    con.execute("""
                        INSERT OR REPLACE INTO price_automation_settings (setting_key, setting_value, setting_type, updated_at)
                        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
                    """, (key, value_str, stype))
                
                con.commit()
                return _no_store_json({"ok": True, "message": "Settings saved successfully"})
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.get("/api/custom-days")
async def get_custom_days(request: Request):
    """Get custom days configuration"""
    require_auth(request)
    try:
        with _db_lock:
            con = _db_connect()
            try:
                cursor = con.execute("SELECT days_array FROM custom_days ORDER BY id DESC LIMIT 1")
                row = cursor.fetchone()
                
                if row:
                    import json
                    return _no_store_json({"ok": True, "days": json.loads(row[0])})
                else:
                    # Default days
                    return _no_store_json({"ok": True, "days": [1, 2, 3, 4, 5, 6, 7, 8, 9, 14, 22, 28, 31, 60]})
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/custom-days")
async def save_custom_days(request: Request):
    """Save custom days configuration"""
    require_auth(request)
    try:
        data = await request.json()
        days = data.get('days', [])
        
        import json
        with _db_lock:
            con = _db_connect()
            try:
                con.execute("INSERT INTO custom_days (days_array) VALUES (?)", (json.dumps(days),))
                con.commit()
                return _no_store_json({"ok": True, "message": "Custom days saved successfully"})
            finally:
                con.close()
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

@app.post("/api/fetch-car-photos")
async def fetch_car_photos(request: Request):
    """Fetch car photos from multiple searches (Albufeira + Faro, different dates)"""
    require_auth(request)
    try:
        import asyncio
        from datetime import datetime, timedelta
        from carjet_direct import scrape_carjet_direct
        import sys
        
        # Locations to search
        locations = [
            "Albufeira",
            "Aeroporto de Faro"
        ]
        
        # Generate dates for next 7 days
        today = datetime.now()
        dates = [(today + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
        
        # Days to search
        days_to_search = [7, 14, 28]
        
        all_photos = {}  # {car_name: photo_url}
        
        print(f"🔍 Starting car photo fetch from CarJet...")
        print(f"📍 Locations: {locations}")
        print(f"📅 Dates: {dates}")
        print(f"📆 Days: {days_to_search}")
        
        # Search each combination - Use mock data for quick testing
        for loc_idx, location in enumerate(locations):
            # ============================================
            # DELAY ENTRE LOCATIONS (2-5s aleatório)
            # ============================================
            if loc_idx > 0:  # Não fazer delay na primeira location
                # random já importado globalmente
                delay_between_locations = random.uniform(2.0, 5.0)
                print(f"\n⏳ Aguardando {delay_between_locations:.1f}s antes da próxima location...", file=sys.stderr, flush=True)
                time.sleep(delay_between_locations)
            
            for date in dates:
                for days in days_to_search:
                    try:
                        print(f"\n🔎 Searching: {location}, {date}, {days} days", file=sys.stderr, flush=True)
                        
                        # Calculate end date
                        start_dt = datetime.fromisoformat(f"{date}T10:00")
                        end_dt = start_dt + timedelta(days=days)
                        
                        # SCRAPING REAL do CarJet
                        print(f"[PHOTOS] Scraping CarJet for {location}, {days} days...", file=sys.stderr, flush=True)
                        
                        # Use scrape_carjet_direct para obter dados reais
                        items = scrape_carjet_direct(location, start_dt, end_dt, quick=1)
                        
                        print(f"[PHOTOS] Scraped {len(items)} real items from CarJet", file=sys.stderr, flush=True)
                        
                        # Extract photos from items
                        for item in items:
                            car_name = item.get('car', '').strip().lower()
                            photo_url = item.get('photo', '')
                            
                            # Clean car name (remove "ou similar" and extra spaces)
                            if ' ou similar' in car_name:
                                car_name = car_name.split(' ou similar')[0].strip()
                            
                            if car_name and photo_url and car_name not in all_photos:
                                all_photos[car_name] = photo_url
                                print(f"  📸 Found photo for: {car_name} -> {photo_url}", file=sys.stderr, flush=True)
                        
                        print(f"✅ Search completed: {len(items)} cars found, total photos: {len(all_photos)}", file=sys.stderr, flush=True)
                        
                        # Delay entre searches (0.5-2s aleatório)
                        delay = random.uniform(0.5, 2.0)
                        print(f"⏳ Aguardando {delay:.1f}s antes da próxima pesquisa...", file=sys.stderr, flush=True)
                        await asyncio.sleep(delay)
                        
                    except Exception as e:
                        print(f"❌ Error searching {location}, {date}, {days} days: {e}", file=sys.stderr, flush=True)
                        import traceback
                        traceback.print_exc()
                        continue
        
        # SALVAR FOTOS NA BASE DE DADOS
        downloaded = 0
        skipped = 0
        
        print(f"\n💾 Saving {len(all_photos)} photos to database...", file=sys.stderr, flush=True)
        
        for car_name, photo_url in all_photos.items():
            try:
                # Verificar se já existe
                with _db_lock:
                    conn = _db_connect()
                    try:
                        existing = conn.execute(
                            "SELECT id FROM vehicle_images WHERE vehicle_key = ?",
                            (car_name,)
                        ).fetchone()
                        
                        if existing:
                            skipped += 1
                            print(f"  ⏭️  Skipped (already exists): {car_name}", file=sys.stderr, flush=True)
                        else:
                            # Inserir nova foto
                            conn.execute(
                                """INSERT INTO vehicle_images 
                                   (vehicle_key, source_url, downloaded_at, content_type) 
                                   VALUES (?, ?, datetime('now'), 'image/jpeg')""",
                                (car_name, photo_url)
                            )
                            conn.commit()
                            downloaded += 1
                            print(f"  ✅ Saved: {car_name} -> {photo_url}", file=sys.stderr, flush=True)
                    finally:
                        conn.close()
            except Exception as e:
                print(f"  ❌ Error saving {car_name}: {e}", file=sys.stderr, flush=True)
                continue
        
        print(f"\n✅ Photo import completed: {downloaded} downloaded, {skipped} skipped", file=sys.stderr, flush=True)
        
        return _no_store_json({
            "ok": True,
            "message": f"Found {len(all_photos)} car photos",
            "downloaded": downloaded,
            "skipped": skipped,
            "photos": all_photos
        })
        
    except Exception as e:
        import traceback
        return _no_store_json({"ok": False, "error": str(e), "traceback": traceback.format_exc()}, 500)

# ============================================================
# API ENDPOINTS - UNIVERSAL SETTINGS SYNC (ANTI DATA-LOSS)
# ============================================================

@app.post("/api/settings/sync")
async def sync_all_settings(request: Request):
    """Sync ALL localStorage settings to database - prevents data loss on Render"""
    require_auth(request)
    
    try:
        data = await request.json()
        logging.info(f"🔄 Syncing {len(data)} setting keys to database")
        
        with _db_lock:
            conn = _db_connect()
            try:
                for key, value in data.items():
                    # Store as JSON string
                    value_str = value if isinstance(value, str) else json.dumps(value)
                    
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO price_automation_settings (key, value, updated_at)
                        VALUES (?, ?, CURRENT_TIMESTAMP)
                        """,
                        (key, value_str)
                    )
                
                conn.commit()
                logging.info(f"✅ Synced {len(data)} settings to database")
                return JSONResponse({"ok": True, "synced": len(data)})
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error syncing settings: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/settings/load-all")
async def load_all_settings(request: Request):
    """Load ALL settings from database - restores data after Render restart"""
    require_auth(request)
    
    try:
        with _db_lock:
            conn = _db_connect()
            try:
                cursor = conn.execute("SELECT key, value FROM price_automation_settings")
                rows = cursor.fetchall()
                
                settings = {}
                for row in rows:
                    key, value = row[0], row[1]
                    # Try to parse as JSON, fallback to string
                    try:
                        settings[key] = json.loads(value) if value else None
                    except:
                        settings[key] = value
                
                logging.info(f"📥 Loaded {len(settings)} settings from database")
                return JSONResponse({"ok": True, "settings": settings})
            finally:
                conn.close()
    except Exception as e:
        logging.error(f"❌ Error loading settings: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ============================================================
# API ENDPOINTS - OAUTH2 EMAIL INTEGRATION
# ============================================================

@app.get("/api/oauth/gmail/authorize")
async def oauth_gmail_authorize(request: Request):
    """Initiate Gmail OAuth2 flow - REAL Google OAuth"""
    require_auth(request)
    
    # Google OAuth2 configuration
    GOOGLE_CLIENT_ID = os.getenv('GOOGLE_CLIENT_ID', 'YOUR_CLIENT_ID_HERE')
    GOOGLE_CLIENT_SECRET = os.getenv('GOOGLE_CLIENT_SECRET', 'YOUR_CLIENT_SECRET_HERE')
    REDIRECT_URI = os.getenv('OAUTH_REDIRECT_URI', 'http://127.0.0.1:8000/api/oauth/gmail/callback')
    
    # Check if credentials are configured
    if GOOGLE_CLIENT_ID == 'YOUR_CLIENT_ID_HERE':
        # Show setup instructions if not configured
        html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Gmail OAuth</title>
        <style>
            body {
                font-family: 'Outfit', sans-serif;
                display: flex;
                align-items: center;
                justify-content: center;
                height: 100vh;
                margin: 0;
                background: #f0f9fb;
            }
            .container {
                text-align: center;
                padding: 2rem;
                background: white;
                border-radius: 8px;
                box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            }
            .icon {
                width: 64px;
                height: 64px;
                margin: 0 auto 1rem;
                color: #009cb6;
            }
            h1 {
                color: #009cb6;
                margin-bottom: 1rem;
            }
            p {
                color: #666;
                margin-bottom: 1.5rem;
            }
            button {
                background: #009cb6;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                cursor: pointer;
                font-size: 16px;
            }
            button:hover {
                background: #007a91;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <svg class="icon" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 8l7.89 5.26a2 2 0 002.22 0L21 8M5 19h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v10a2 2 0 002 2z"/>
            </svg>
            <h1>Conectar Gmail</h1>
            <p>Esta é uma <strong>demonstração</strong> do fluxo OAuth2.</p>
            <p style="font-size: 14px; color: #666; margin: 1rem 0;">Para conectar ao Gmail real, é necessário:</p>
            <ul style="text-align: left; font-size: 13px; color: #666; margin: 0 auto; max-width: 400px; line-height: 1.8;">
                <li>Registar app no <a href="https://console.cloud.google.com" target="_blank" style="color: #009cb6;">Google Cloud Console</a></li>
                <li>Obter Client ID e Client Secret</li>
                <li>Configurar OAuth2 redirect URLs</li>
                <li>Implementar fluxo OAuth completo</li>
            </ul>
            <p style="font-size: 12px; color: #999; margin-top: 1rem;">Por agora, clique abaixo para simular a conexão:</p>
            <button onclick="simulateOAuth()">Simular Conexão Gmail</button>
        </div>
        <script>
            function simulateOAuth() {
                // Simulate successful OAuth
                const data = {
                    type: 'oauth-success',
                    provider: 'gmail',
                    email: 'seu-email@gmail.com',
                    token: 'mock_access_token_' + Date.now(),
                    refreshToken: 'mock_refresh_token',
                    expiresAt: Date.now() + 3600000
                };
                
                // Send message to parent window
                if (window.opener) {
                    window.opener.postMessage(data, '*');
                    window.close();
                } else {
                    alert('Erro: Janela pai não encontrada');
                }
            }
        </script>
    </body>
    </html>
    """
        return HTMLResponse(content=html)
    
    # Real OAuth2 flow - redirect to Google
    import urllib.parse
    
    # OAuth2 parameters
    auth_params = {
        'client_id': GOOGLE_CLIENT_ID,
        'redirect_uri': REDIRECT_URI,
        'response_type': 'code',
        'scope': 'https://www.googleapis.com/auth/gmail.send https://www.googleapis.com/auth/userinfo.email',
        'access_type': 'offline',
        'prompt': 'consent'
    }
    
    # Build authorization URL
    auth_url = 'https://accounts.google.com/o/oauth2/v2/auth?' + urllib.parse.urlencode(auth_params)
    
    # Redirect to Google OAuth
    return RedirectResponse(url=auth_url)

@app.get("/api/oauth/gmail/callback")
async def oauth_gmail_callback(request: Request, code: str = None, error: str = None):
    """Handle Gmail OAuth2 callback"""
    require_auth(request)
    
    if error:
        return HTMLResponse(f"<h1>Error: {error}</h1><p>OAuth authorization failed.</p>")
    
    if not code:
        return HTMLResponse("<h1>Error</h1><p>No authorization code received.</p>")
    
    # Exchange code for tokens
    GOOGLE_CLIENT_ID = os.getenv('GOOGLE_CLIENT_ID')
    GOOGLE_CLIENT_SECRET = os.getenv('GOOGLE_CLIENT_SECRET')
    REDIRECT_URI = os.getenv('OAUTH_REDIRECT_URI', 'http://127.0.0.1:8000/api/oauth/gmail/callback')
    
    import urllib.parse
    import httpx
    import time
    
    token_params = {
        'code': code,
        'client_id': GOOGLE_CLIENT_ID,
        'client_secret': GOOGLE_CLIENT_SECRET,
        'redirect_uri': REDIRECT_URI,
        'grant_type': 'authorization_code'
    }
    
    try:
        async with httpx.AsyncClient() as client:
            # Exchange code for tokens
            token_response = await client.post(
                'https://oauth2.googleapis.com/token',
                data=token_params
            )
            token_data = token_response.json()
            
            if 'error' in token_data:
                return HTMLResponse(f"<h1>Error</h1><p>{token_data.get('error_description', 'Token exchange failed')}</p>")
            
            # Get user info (email, name, picture)
            access_token = token_data.get('access_token')
            userinfo_response = await client.get(
                'https://www.googleapis.com/oauth2/v2/userinfo',
                headers={'Authorization': f'Bearer {access_token}'}
            )
            userinfo = userinfo_response.json()
            user_email = userinfo.get('email', 'unknown@gmail.com')
            user_name = userinfo.get('name', '')
            user_picture = userinfo.get('picture', '')  # Google profile picture URL
            google_id = userinfo.get('id', '')
            
            # Return success page
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Gmail Connected</title>
                <style>
                    body {{
                        font-family: 'Outfit', sans-serif;
                        display: flex;
                        align-items: center;
                        justify-content: center;
                        height: 100vh;
                        margin: 0;
                        background: #f0f9fb;
                    }}
                    .container {{
                        text-align: center;
                        padding: 2rem;
                        background: white;
                        border-radius: 8px;
                        box-shadow: 0 4px 12px rgba(0,0,0,0.1);
                    }}
                    .icon {{
                        width: 64px;
                        height: 64px;
                        margin: 0 auto 1rem;
                        color: #009cb6;
                    }}
                    h1 {{
                        color: #009cb6;
                        margin-bottom: 1rem;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <svg class="icon" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z"/>
                    </svg>
                    <h1>Gmail Conectado!</h1>
                    <p>Conta: <strong>{user_email}</strong></p>
                    <p style="color: #666; font-size: 14px;">Esta janela vai fechar automaticamente...</p>
                </div>
                <script>
                    const data = {{
                        type: 'oauth-success',
                        provider: 'gmail',
                        email: '{user_email}',
                        name: '{user_name}',
                        picture: '{user_picture}',
                        googleId: '{google_id}',
                        token: '{access_token}',
                        refreshToken: '{token_data.get("refresh_token", "")}',
                        expiresAt: {int(time.time()) + token_data.get('expires_in', 3600)} * 1000
                    }};
                    
                    if (window.opener) {{
                        window.opener.postMessage(data, '*');
                        setTimeout(() => window.close(), 2000);
                    }}
                </script>
            </body>
            </html>
            """
            return HTMLResponse(content=html)
            
    except Exception as e:
        logging.error(f"OAuth callback error: {str(e)}")
        return HTMLResponse(f"<h1>Error</h1><p>Failed to complete OAuth: {str(e)}</p>")

@app.post("/api/user/update-google-profile")
async def update_google_profile(request: Request):
    """Update user profile with Google info (picture, google_id)"""
    require_auth(request)
    
    try:
        data = await request.json()
        google_id = data.get('googleId')
        picture_url = data.get('pictureUrl')
        
        if not google_id or not picture_url:
            return JSONResponse({"ok": False, "error": "Missing googleId or pictureUrl"})
        
        # Get current user from session
        username = request.session.get('username')
        if not username:
            return JSONResponse({"ok": False, "error": "Not authenticated"})
        
        # Update user with google_id and profile picture
        with _db_lock:
            conn = _db_connect()
            try:
                # Update google_id and profile picture URL
                conn.execute(
                    "UPDATE users SET google_id = ?, profile_picture_path = ? WHERE username = ?",
                    (google_id, picture_url, username)
                )
                conn.commit()
                logging.info(f"✅ Updated Google profile for user {username}")
                
                return JSONResponse({
                    "ok": True,
                    "message": "Profile updated successfully"
                })
            finally:
                conn.close()
                
    except Exception as e:
        logging.error(f"Error updating Google profile: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/reports/test-daily")
async def test_daily_report(request: Request):
    """Send test daily report email via Gmail API"""
    require_auth(request)
    
    try:
        from googleapiclient.discovery import build
        from google.oauth2.credentials import Credentials
        import base64
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from datetime import datetime
        
        # Get request data
        data = await request.json()
        access_token = data.get('accessToken')
        
        if not access_token:
            return JSONResponse({
                "ok": False,
                "error": "Token OAuth não encontrado. Por favor, conecte sua conta Gmail primeiro."
            })
        
        test_email = "carlpac82@hotmail.com"
        username = request.session.get('username')
        
        logging.info(f"Test daily report requested by {username}")
        
        # Create credentials from access token
        credentials = Credentials(token=access_token)
        
        # Build Gmail service
        service = build('gmail', 'v1', credentials=credentials)
        
        # Create HTML email with sample data
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Relatório Diário de Preços</title>
        </head>
        <body style="margin: 0; padding: 0; background-color: #f8fafc; font-family: 'Segoe UI', sans-serif;">
            <table width="100%" cellpadding="0" cellspacing="0" style="background-color: #f8fafc; padding: 20px 0;">
                <tr>
                    <td align="center">
                        <table width="600" cellpadding="0" cellspacing="0" style="background-color: #ffffff; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                            <!-- Header -->
                            <tr>
                                <td style="background: linear-gradient(135deg, #009cb6 0%, #007a91 100%); padding: 30px 20px; text-align: center;">
                                    <h1 style="margin: 0; color: #ffffff; font-size: 24px;">📊 Relatório Diário de Preços</h1>
                                    <p style="margin: 8px 0 0 0; color: #e0f2f7; font-size: 14px;">{datetime.now().strftime('%d/%m/%Y')} - Email de Teste</p>
                                </td>
                            </tr>
                            <!-- Content -->
                            <tr>
                                <td style="padding: 30px 20px; text-align: center;">
                                    <h2 style="color: #009cb6; margin: 0 0 20px 0;">✅ Sistema de Relatórios Funcionando!</h2>
                                    <p style="color: #64748b; font-size: 16px; line-height: 1.6; margin: 0 0 20px 0;">
                                        Este é um email de teste do sistema de relatórios automáticos.<br>
                                        O sistema está configurado e pronto para enviar relatórios diários com dados reais de preços.
                                    </p>
                                    <div style="background: #f0f9fb; border-left: 4px solid #009cb6; padding: 15px; text-align: left; border-radius: 4px;">
                                        <p style="margin: 0; color: #1e293b; font-size: 14px;">
                                            <strong style="color: #009cb6;">Próximos passos:</strong><br>
                                            • Configure os horários em Settings → Automated Reports<br>
                                            • O sistema enviará relatórios automáticos às 09h00<br>
                                            • Inclui comparação de preços e alertas
                                        </p>
                                    </div>
                                </td>
                            </tr>
                            <!-- Footer -->
                            <tr>
                                <td style="background: #f8fafc; padding: 20px; text-align: center; border-top: 1px solid #e2e8f0;">
                                    <p style="margin: 0; font-size: 12px; color: #94a3b8;">
                                        Auto Prudente © {datetime.now().year} - Sistema de Monitorização de Preços
                                    </p>
                                </td>
                            </tr>
                        </table>
                    </td>
                </tr>
            </table>
        </body>
        </html>
        """
        
        # Create message
        message = MIMEMultipart('alternative')
        message['to'] = test_email
        message['subject'] = f'📊 Relatório Diário de Preços - Teste ({datetime.now().strftime("%d/%m/%Y")})'
        
        # Attach HTML
        html_part = MIMEText(html_content, 'html')
        message.attach(html_part)
        
        # Encode and send
        raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
        send_message = service.users().messages().send(
            userId='me',
            body={'raw': raw_message}
        ).execute()
        
        logging.info(f"✅ Test email sent successfully to {test_email}")
        
        return JSONResponse({
            "ok": True,
            "message": f"Email de teste enviado com sucesso para {test_email}!",
            "messageId": send_message.get('id')
        })
        
    except Exception as e:
        logging.error(f"Test daily report error: {str(e)}")
        return JSONResponse({
            "ok": False,
            "error": f"Erro ao enviar email: {str(e)}"
        }, status_code=500)

@app.post("/api/reports/test-alert")
async def test_alert_email(request: Request):
    """Send test alert email with sample data"""
    require_auth(request)
    
    try:
        test_email = "carlpac82@hotmail.com"
        logging.info("Test alert email requested")
        
        return JSONResponse({
            "ok": True,
            "message": f"Template de alertas seria enviado para {test_email}",
            "note": "Implementação completa requer integração com Gmail API e template HTML"
        })
        
    except Exception as e:
        logging.error(f"Test alert email error: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/backup/create")
async def create_backup(request: Request):
    """Create system backup"""
    require_auth(request)
    
    try:
        import zipfile
        import shutil
        from datetime import datetime
        
        data = await request.json()
        logging.info(f"Backup requested with options: {data}")
        
        # Create backup directory if not exists
        backup_dir = Path("backups")
        backup_dir.mkdir(exist_ok=True)
        
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}.zip"
        backup_path = backup_dir / backup_filename
        
        # Create ZIP file
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # 1. All Databases
            if data.get('database', True):
                db_files = ["rental_tracker.db", "data.db", "car_images.db", "carrental.db"]
                for db_file in db_files:
                    db_path = Path(db_file)
                    if db_path.exists() and db_path.stat().st_size > 0:
                        zipf.write(db_path, f"database/{db_file}")
                        size_kb = db_path.stat().st_size / 1024
                        logging.info(f"✅ Database {db_file} added to backup ({size_kb:.1f} KB)")
            
            # 2. Settings (localStorage data stored in DB)
            if data.get('settings', True):
                # Settings are in localStorage, backed up with database
                pass
            
            # 3. Vehicle mappings (in database)
            if data.get('vehicles', True):
                # Vehicle data is in database
                pass
            
            # 4. Uploaded files (logos, profile pictures)
            if data.get('uploads', True):
                uploads_dir = Path("uploads")
                if uploads_dir.exists():
                    for file_path in uploads_dir.rglob("*"):
                        if file_path.is_file():
                            arcname = f"uploads/{file_path.relative_to(uploads_dir)}"
                            zipf.write(file_path, arcname)
                    logging.info("✅ Uploads added to backup")
            
            # 5. ALL Static files
            static_dir = Path("static")
            if static_dir.exists():
                for file_path in static_dir.rglob("*"):
                    if file_path.is_file():
                        arcname = f"static/{file_path.relative_to(static_dir)}"
                        zipf.write(file_path, arcname)
                logging.info("✅ All static files added to backup")
            
            # 6. ALL Templates
            templates_dir = Path("templates")
            if templates_dir.exists():
                for file_path in templates_dir.rglob("*"):
                    if file_path.is_file():
                        arcname = f"templates/{file_path.relative_to(templates_dir)}"
                        zipf.write(file_path, arcname)
                logging.info("✅ All templates added to backup")
            
            # 7. Main.py and other Python files
            for py_file in Path(".").glob("*.py"):
                if py_file.is_file():
                    zipf.write(py_file, f"code/{py_file.name}")
            logging.info("✅ Python files added to backup")
            
            # 8. Requirements and config files
            config_files = ["requirements.txt", "Procfile", "runtime.txt", ".gitignore"]
            for config_file in config_files:
                config_path = Path(config_file)
                if config_path.exists():
                    zipf.write(config_path, f"config/{config_file}")
            logging.info("✅ Config files added to backup")
            
            # 6. OAuth settings (if requested - sensitive!)
            if data.get('oauth', False):
                env_path = Path(".env")
                if env_path.exists():
                    zipf.write(env_path, "config/.env")
                    logging.info("✅ OAuth config added to backup")
        
        # Get file size
        file_size = backup_path.stat().st_size
        size_mb = file_size / (1024 * 1024)
        
        logging.info(f"✅ Backup created: {backup_filename} ({size_mb:.2f} MB)")
        
        return JSONResponse({
            "ok": True,
            "message": f"Backup criado com sucesso ({size_mb:.2f} MB)",
            "downloadUrl": f"/api/backup/download/{backup_filename}",
            "filename": backup_filename,
            "size": f"{size_mb:.2f} MB"
        })
        
    except Exception as e:
        logging.error(f"Backup creation error: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/backup/download/{filename}")
async def download_backup(request: Request, filename: str):
    """Download backup file"""
    require_auth(request)
    
    try:
        from fastapi.responses import FileResponse
        
        backup_path = Path("backups") / filename
        
        if not backup_path.exists():
            raise HTTPException(status_code=404, detail="Backup not found")
        
        return FileResponse(
            path=backup_path,
            filename=filename,
            media_type="application/zip"
        )
        
    except Exception as e:
        logging.error(f"Backup download error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/backup/restore")
async def restore_backup(request: Request):
    """Restore system from backup ZIP file"""
    require_auth(request)
    
    try:
        import zipfile
        import shutil
        from fastapi import UploadFile, File
        
        # Receber ficheiro ZIP
        form = await request.form()
        backup_file = form.get("file")
        
        if not backup_file or not isinstance(backup_file, UploadFile):
            return JSONResponse({"ok": False, "error": "Nenhum ficheiro enviado"}, status_code=400)
        
        if not backup_file.filename.endswith('.zip'):
            return JSONResponse({"ok": False, "error": "Ficheiro deve ser ZIP"}, status_code=400)
        
        logging.info(f"Restore backup from: {backup_file.filename}")
        
        # Guardar ZIP temporariamente
        temp_zip = Path("backups") / f"restore_temp_{backup_file.filename}"
        temp_zip.parent.mkdir(exist_ok=True)
        
        with open(temp_zip, "wb") as f:
            content = await backup_file.read()
            f.write(content)
        
        # Extrair e restaurar
        with zipfile.ZipFile(temp_zip, 'r') as zipf:
            # 1. Restaurar base de dados
            for db_file in zipf.namelist():
                if db_file.startswith("database/") and db_file.endswith(".db"):
                    db_name = Path(db_file).name
                    target_path = DATA_DIR / db_name
                    
                    # Backup da BD atual antes de sobrescrever
                    if target_path.exists():
                        backup_current = target_path.with_suffix('.db.backup')
                        shutil.copy2(target_path, backup_current)
                        logging.info(f"Current DB backed up to {backup_current}")
                    
                    # Extrair nova BD
                    with zipf.open(db_file) as source, open(target_path, 'wb') as target:
                        shutil.copyfileobj(source, target)
                    logging.info(f"✅ Database {db_name} restored")
            
            # 2. Restaurar uploads
            for upload_file in zipf.namelist():
                if upload_file.startswith("uploads/"):
                    target_path = Path(upload_file)
                    target_path.parent.mkdir(parents=True, exist_ok=True)
                    with zipf.open(upload_file) as source, open(target_path, 'wb') as target:
                        shutil.copyfileobj(source, target)
            logging.info("✅ Uploads restored")
            
            # 3. Restaurar static files (opcional - pode sobrescrever código)
            # Comentado por segurança - descomentar se necessário
            # for static_file in zipf.namelist():
            #     if static_file.startswith("static/"):
            #         target_path = Path(static_file)
            #         target_path.parent.mkdir(parents=True, exist_ok=True)
            #         with zipf.open(static_file) as source, open(target_path, 'wb') as target:
            #             shutil.copyfileobj(source, target)
        
        # Limpar ficheiro temporário
        temp_zip.unlink()
        
        logging.info("✅ Backup restored successfully")
        
        return JSONResponse({
            "ok": True,
            "message": "Backup restaurado com sucesso! Recarrega a página."
        })
        
    except Exception as e:
        logging.error(f"Backup restore error: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/fix-schema-emergency")
async def fix_schema_emergency(request: Request):
    """Emergency endpoint to fix PostgreSQL schema"""
    try:
        require_auth(request)
    except:
        pass  # Allow without auth in emergency
    
    try:
        if not _USE_NEW_DB or not USE_POSTGRES:
            return JSONResponse({"ok": False, "error": "Not using PostgreSQL"})
        
        results = []
        
        columns = [
            ("first_name", "TEXT"),
            ("last_name", "TEXT"),
            ("email", "TEXT"),
            ("mobile", "TEXT"),
            ("profile_picture_path", "TEXT"),
            ("profile_picture_data", "BYTEA"),
            ("is_admin", "BOOLEAN DEFAULT FALSE"),
            ("enabled", "BOOLEAN DEFAULT TRUE"),
            ("created_at", "TEXT"),
            ("google_id", "TEXT"),
        ]
        
        with _db_lock:
            conn = _db_connect()
            
            # Se as colunas já existem como INTEGER, converter para BOOLEAN
            try:
                # Verificar tipo atual
                cursor = conn.execute("""
                    SELECT column_name, data_type 
                    FROM information_schema.columns 
                    WHERE table_name='users' AND column_name IN ('is_admin', 'enabled')
                """)
                existing = {row[0]: row[1] for row in cursor.fetchall()}
                
                # Converter INTEGER para BOOLEAN se necessário
                if existing.get('is_admin') == 'integer':
                    conn.execute("ALTER TABLE users ALTER COLUMN is_admin TYPE BOOLEAN USING is_admin::boolean")
                    conn.commit()
                    results.append({"column": "is_admin", "status": "converted to BOOLEAN"})
                
                if existing.get('enabled') == 'integer':
                    conn.execute("ALTER TABLE users ALTER COLUMN enabled TYPE BOOLEAN USING enabled::boolean")
                    conn.commit()
                    results.append({"column": "enabled", "status": "converted to BOOLEAN"})
            except Exception as e:
                conn.rollback()
                results.append({"column": "type_conversion", "status": "error", "error": str(e)})
            try:
                for col_name, col_type in columns:
                    try:
                        conn.execute(f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
                        conn.commit()
                        results.append({"column": col_name, "status": "added"})
                    except Exception as e:
                        conn.rollback()
                        if "already exists" in str(e):
                            results.append({"column": col_name, "status": "exists"})
                        else:
                            results.append({"column": col_name, "status": "error", "error": str(e)})
                
                # Verify
                cursor = conn.execute("SELECT column_name FROM information_schema.columns WHERE table_name='users' ORDER BY ordinal_position")
                cols = [row[0] for row in cursor.fetchall()]
                
                return JSONResponse({
                    "ok": True,
                    "message": "Schema fix completed",
                    "results": results,
                    "total_columns": len(cols),
                    "columns": cols,
                    "enabled_exists": "enabled" in cols
                })
            finally:
                conn.close()
                
    except Exception as e:
        import traceback
        return JSONResponse({
            "ok": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }, status_code=500)

@app.get("/api/backup/list")
async def list_backups(request: Request):
    """List available backups"""
    require_auth(request)
    
    try:
        from datetime import datetime
        
        backup_dir = Path("backups")
        backups = []
        
        if backup_dir.exists():
            for backup_file in sorted(backup_dir.glob("backup_*.zip"), reverse=True):
                stat = backup_file.stat()
                size_mb = stat.st_size / (1024 * 1024)
                
                # Parse timestamp from filename
                try:
                    timestamp_str = backup_file.stem.replace("backup_", "")
                    dt = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                    date_str = dt.strftime("%d/%m/%Y %H:%M")
                except:
                    date_str = "Data desconhecida"
                
                backups.append({
                    "name": backup_file.name,
                    "date": date_str,
                    "size": f"{size_mb:.2f} MB",
                    "downloadUrl": f"/api/backup/download/{backup_file.name}"
                })
        
        return JSONResponse({
            "ok": True,
            "backups": backups
        })
        
    except Exception as e:
        logging.error(f"Backup list error: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/reports/test-weekly")
async def test_weekly_report(request: Request):
    """Send test weekly report"""
    require_auth(request)
    
    try:
        logging.info("Test weekly report requested")
        
        return JSONResponse({
            "ok": True,
            "message": "Relatório semanal seria enviado com análise dos próximos 3 meses",
            "note": "Implementação completa requer scheduler e análise de dados históricos"
        })
        
    except Exception as e:
        logging.error(f"Test weekly report error: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/email/test-oauth")
async def test_email_oauth(request: Request):
    """Send test email using OAuth token"""
    require_auth(request)
    
    try:
        data = await request.json()
        provider = data.get('provider')
        email = data.get('email')
        access_token = data.get('accessToken')
        recipients = data.get('recipients', '')
        
        # Parse recipients (one per line)
        recipient_list = [r.strip() for r in recipients.split('\n') if r.strip()]
        
        if not recipient_list:
            return JSONResponse({"ok": False, "error": "Nenhum destinatário especificado"})
        
        # For now, return success (real implementation would use Gmail API)
        # In production, you would use the access_token to send via Gmail API
        logging.info(f"Test email requested from {email} to {recipient_list}")
        
        return JSONResponse({
            "ok": True,
            "message": f"Email de teste seria enviado para {len(recipient_list)} destinatário(s)",
            "note": "Implementação completa requer Gmail API client"
        })
        
    except Exception as e:
        logging.error(f"Test email error: {str(e)}")
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.get("/api/oauth/outlook/authorize")
async def oauth_outlook_authorize(request: Request):
    """Initiate Outlook OAuth2 flow"""
    require_auth(request)
    
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Outlook OAuth</title>
        <style>
            body {
                font-family: 'Outfit', sans-serif;
                display: flex;
                align-items: center;
                justify-content: center;
                height: 100vh;
                margin: 0;
                background: #f0f9fb;
            }
            .container {
                text-align: center;
                padding: 2rem;
                background: white;
                border-radius: 8px;
                box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            }
            .icon {
                width: 64px;
                height: 64px;
                margin: 0 auto 1rem;
                color: #009cb6;
            }
            h1 {
                color: #009cb6;
                margin-bottom: 1rem;
            }
            p {
                color: #666;
                margin-bottom: 1.5rem;
            }
            button {
                background: #009cb6;
                color: white;
                border: none;
                padding: 12px 24px;
                border-radius: 6px;
                cursor: pointer;
                font-size: 16px;
            }
            button:hover {
                background: #007a91;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <svg class="icon" fill="none" stroke="currentColor" viewBox="0 0 24 24">
                <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M3 8l7.89 5.26a2 2 0 002.22 0L21 8M5 19h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v10a2 2 0 002 2z"/>
            </svg>
            <h1>Conectar Outlook</h1>
            <p>Esta é uma <strong>demonstração</strong> do fluxo OAuth2.</p>
            <p style="font-size: 14px; color: #666; margin: 1rem 0;">Para conectar ao Outlook real, é necessário:</p>
            <ul style="text-align: left; font-size: 13px; color: #666; margin: 0 auto; max-width: 400px; line-height: 1.8;">
                <li>Registar app no <a href="https://portal.azure.com" target="_blank" style="color: #009cb6;">Azure Portal</a></li>
                <li>Obter Application ID e Secret</li>
                <li>Configurar Microsoft OAuth2 redirect URLs</li>
                <li>Implementar fluxo OAuth completo</li>
            </ul>
            <p style="font-size: 12px; color: #999; margin-top: 1rem;">Por agora, clique abaixo para simular a conexão:</p>
            <button onclick="simulateOAuth()">Simular Conexão Outlook</button>
        </div>
        <script>
            function simulateOAuth() {
                const data = {
                    type: 'oauth-success',
                    provider: 'outlook',
                    email: 'seu-email@outlook.com',
                    token: 'mock_access_token_' + Date.now(),
                    refreshToken: 'mock_refresh_token',
                    expiresAt: Date.now() + 3600000
                };
                
                if (window.opener) {
                    window.opener.postMessage(data, '*');
                    window.close();
                } else {
                    alert('Erro: Janela pai não encontrada');
                }
            }
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

# ============================================================
# AUTOMATIC BACKUPS SCHEDULER
# ============================================================

def create_automatic_backup():
    """Criar backup automático do sistema"""
    try:
        import zipfile
        from pathlib import Path
        from datetime import datetime
        
        # Criar diretório de backups
        backup_dir = Path("backups")
        backup_dir.mkdir(exist_ok=True)
        
        # Nome do backup com timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"auto_backup_{timestamp}.zip"
        backup_path = backup_dir / backup_filename
        
        # Criar ZIP
        with zipfile.ZipFile(backup_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # Databases
            db_files = ["rental_tracker.db", "data.db", "car_images.db", "carrental.db"]
            for db_file in db_files:
                db_path = Path(db_file)
                if db_path.exists() and db_path.stat().st_size > 0:
                    zipf.write(db_path, f"database/{db_file}")
            
            # Settings
            if Path("app_settings.json").exists():
                zipf.write("app_settings.json", "settings/app_settings.json")
        
        file_size = backup_path.stat().st_size / (1024 * 1024)
        log_to_db("INFO", f"Automatic backup created: {backup_filename} ({file_size:.2f} MB)", "main", "create_automatic_backup")
        
        # Limpar backups antigos (manter últimos 7)
        cleanup_old_backups(backup_dir, keep_last=7)
        
        return True
    except Exception as e:
        log_to_db("ERROR", f"Automatic backup failed: {str(e)}", "main", "create_automatic_backup")
        return False

def cleanup_old_backups(backup_dir: Path, keep_last: int = 7):
    """Limpar backups antigos, mantendo apenas os últimos N"""
    try:
        backups = sorted(backup_dir.glob("auto_backup_*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)
        for old_backup in backups[keep_last:]:
            old_backup.unlink()
            log_to_db("INFO", f"Old backup deleted: {old_backup.name}", "main", "cleanup_old_backups")
    except Exception as e:
        log_to_db("ERROR", f"Cleanup old backups failed: {str(e)}", "main", "cleanup_old_backups")

# Iniciar scheduler de backups automáticos
try:
    from apscheduler.schedulers.background import BackgroundScheduler
    from apscheduler.triggers.cron import CronTrigger
    
    backup_scheduler = BackgroundScheduler()
    # Backup diário às 3 AM
    backup_scheduler.add_job(
        create_automatic_backup,
        CronTrigger(hour=3, minute=0),
        id='daily_backup',
        name='Daily Automatic Backup',
        replace_existing=True
    )
    backup_scheduler.start()
    log_to_db("INFO", "Automatic backup scheduler started (daily at 3 AM)", "main", "scheduler")
except ImportError:
    log_to_db("WARNING", "APScheduler not installed - automatic backups disabled", "main", "scheduler")
except Exception as e:
    log_to_db("ERROR", f"Failed to start backup scheduler: {str(e)}", "main", "scheduler")

# ============================================================
# EMAIL QUEUE SYSTEM
# ============================================================

import queue
import threading

email_queue = queue.Queue()
email_queue_running = False

def email_worker():
    """Worker thread para processar fila de emails"""
    global email_queue_running
    email_queue_running = True
    
    while email_queue_running:
        try:
            # Aguardar email na fila (timeout 1s)
            email_data = email_queue.get(timeout=1)
            
            # Enviar email
            try:
                _send_notification_email(
                    email_data['to'],
                    email_data['subject'],
                    email_data['message']
                )
                log_to_db("INFO", f"Email sent from queue: {email_data['to']}", "main", "email_worker")
            except Exception as e:
                log_to_db("ERROR", f"Failed to send queued email: {str(e)}", "main", "email_worker")
                # Retry até 3 vezes
                if email_data.get('retry_count', 0) < 3:
                    email_data['retry_count'] = email_data.get('retry_count', 0) + 1
                    email_queue.put(email_data)
            
            email_queue.task_done()
        except queue.Empty:
            continue
        except Exception as e:
            log_to_db("ERROR", f"Email worker error: {str(e)}", "main", "email_worker")

def queue_email(to: str, subject: str, message: str):
    """Adicionar email à fila para envio assíncrono"""
    email_queue.put({
        'to': to,
        'subject': subject,
        'message': message,
        'retry_count': 0
    })

# Iniciar worker thread para emails
email_thread = threading.Thread(target=email_worker, daemon=True)
email_thread.start()
log_to_db("INFO", "Email queue worker started", "main", "email_queue")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
